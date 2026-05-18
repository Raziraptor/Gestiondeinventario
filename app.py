# ==============================================================================
# 1. IMPORTACIONES
# ==============================================================================

# --- Núcleo de Python ---
import os
import io
import csv
import json
import secrets
import hashlib
from functools import wraps
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo
from collections import defaultdict
from threading import Thread

# --- Variables de entorno ---
from dotenv import load_dotenv
load_dotenv()

# --- Flask y Extensiones ---
from flask import (Flask, render_template, request, redirect, url_for, flash,
                   send_file, make_response, jsonify, current_app)
from flask.cli import with_appcontext
import click
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileAllowed
from flask_mail import Mail
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# --- Formularios (WTForms) ---
from wtforms import StringField, PasswordField, SubmitField, BooleanField
from wtforms.validators import DataRequired, Length, EqualTo, ValidationError, Email

# --- Seguridad y Tokens ---
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from itsdangerous.url_safe import URLSafeTimedSerializer

# --- Base de Datos ---
from sqlalchemy import extract, Date, text
from sqlalchemy.exc import IntegrityError
from decimal import Decimal

# --- Imágenes y QR ---
from PIL import Image, ImageDraw, ImageFont
import qrcode

# --- Reportes (PDF) ---
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import inch, mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, Image as ReportLabImage)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing

# --- Reportes (Excel) ---
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, NamedStyle, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XlImage

# --- HTTP ---
import requests


# ==============================================================================
# 2. CONFIGURACIÓN DE LA APLICACIÓN
# ==============================================================================

basedir = os.path.abspath(os.path.dirname(__file__))

_TZ_MX = ZoneInfo('America/Mexico_City')

def now_mx() -> datetime:
    """Hora actual en zona horaria de México (naive, lista para guardar en DB)."""
    return datetime.now(_TZ_MX).replace(tzinfo=None)

CATEGORIAS_GASTO = ['Servicios', 'Nómina', 'Mantenimiento', 'Insumos', 'Inventario', 'Otros']
MESES_ES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

app = Flask(__name__)

# Serializador JSON que maneja Decimal (necesario tras migrar Float → Numeric)
from flask.json.provider import DefaultJSONProvider
class _JSONProvider(DefaultJSONProvider):
    @staticmethod
    def default(o):
        if isinstance(o, Decimal):
            return float(o)
        return DefaultJSONProvider.default(o)
app.json_provider_class = _JSONProvider
app.json = _JSONProvider(app)

csrf = CSRFProtect(app)
app.jinja_env.add_extension('jinja2.ext.do') # Para la lógica de 'set' en bucles

# --- Configuración de Variables de Entorno ---
_secret_key = os.environ.get('SECRET_KEY')
if not _secret_key:
    print("ADVERTENCIA: SECRET_KEY no está definida en el entorno. "
          "Las sesiones se invalidarán en cada reinicio del servidor.")
    _secret_key = secrets.token_hex(32)
app.secret_key = _secret_key

db_url = os.environ.get('DATABASE_URL')
if db_url:
    app.config['SQLALCHEMY_DATABASE_URI'] = db_url
else:
    print("ADVERTENCIA: No se encontró DATABASE_URL. Usando SQLite local.")
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'inventario.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['UPLOAD_FOLDER'] = os.path.join(basedir, 'static/uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# --- Seguridad de Cookies de Sesión ---
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') != 'development'
app.config['REMEMBER_COOKIE_HTTPONLY'] = True
app.config['REMEMBER_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') != 'development'

# TWA (Play Store): completar con package_name y sha256 tras correr bubblewrap
# Ver instrucciones en twa/README.md
app.config['ASSETLINKS'] = []

db = SQLAlchemy(app)
login_manager = LoginManager(app)
mail = Mail(app)
_limiter_storage = os.environ.get("REDIS_URL", "memory://")
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri=_limiter_storage,
)

login_manager.login_view = 'login'
login_manager.login_message = 'Por favor, inicia sesión para acceder a esta página.'
login_manager.login_message_category = 'info'

s = URLSafeTimedSerializer(app.config['SECRET_KEY'])

def _flash_err(user_msg: str, exc: Exception | None = None) -> None:
    """Muestra un mensaje de error seguro al usuario y loguea la excepción real al servidor."""
    if exc is not None:
        app.logger.error("%s: %s", user_msg, exc, exc_info=True)
    flash(user_msg, 'danger')

# ==============================================================================
# 3. COMANDOS CLI (Para Despliegue)
# ==============================================================================

@app.cli.command("create-db")
@with_appcontext
def create_db_command():
    """Crea todas las tablas de la base de datos."""
    db.create_all()
    print("¡Base de datos y tablas creadas exitosamente!")

@app.cli.command("fix-db-proyectos")
@with_appcontext
def fix_db_proyectos():
    """Corrige columnas de proyecto_oc_detalle: agrega faltantes y amplía límites de texto."""
    from sqlalchemy import text
    stmts = [
        # Columnas que pueden no existir
        "ALTER TABLE proyecto_oc_detalle ADD COLUMN IF NOT EXISTS enlace_proveedor VARCHAR(500)",
        "ALTER TABLE proyecto_oc_detalle ADD COLUMN IF NOT EXISTS comentarios_detalle TEXT",
        # Ampliar columnas con límite demasiado corto (causa StringDataRightTruncation)
        "ALTER TABLE proyecto_oc_detalle ALTER COLUMN descripcion_nuevo TYPE TEXT",
        "ALTER TABLE proyecto_oc_detalle ALTER COLUMN proveedor_sugerido TYPE VARCHAR(255)",
    ]
    with db.engine.connect() as conn:
        for stmt in stmts:
            try:
                conn.execute(text(stmt))
                conn.commit()
                print(f"OK: {stmt}")
            except Exception as e:
                conn.rollback()
                print(f"Omitido: {e}")
    print("fix-db-proyectos completado.")

@app.cli.command("fix-enlace-text")
@with_appcontext
def fix_enlace_text():
    """Amplía enlace_proveedor a TEXT en producto, orden_compra_detalle y proyecto_oc_detalle."""
    from sqlalchemy import text
    stmts = [
        "ALTER TABLE producto ALTER COLUMN enlace_proveedor TYPE TEXT",
        "ALTER TABLE orden_compra_detalle ALTER COLUMN enlace_proveedor TYPE TEXT",
        "ALTER TABLE proyecto_oc_detalle ALTER COLUMN enlace_proveedor TYPE TEXT",
    ]
    with db.engine.connect() as conn:
        for stmt in stmts:
            try:
                conn.execute(text(stmt))
                conn.commit()
                print(f"OK: {stmt}")
            except Exception as e:
                conn.rollback()
                print(f"Omitido: {e}")
    print("fix-enlace-text completado.")

@app.cli.command("fix-float-to-numeric")
@with_appcontext
def fix_float_to_numeric():
    """Migra columnas de montos de DOUBLE PRECISION a NUMERIC(10,2). Idempotente y sin pérdida de datos."""
    columnas = [
        ("producto",             "precio_unitario"),
        ("orden_compra_detalle", "costo_unitario_estimado"),
        ("gasto",                "monto"),
        ("pago_servicio",        "monto"),
        ("factura_proveedor",    "monto"),
        ("proyecto_oc_detalle",  "costo_unitario"),
    ]
    with db.engine.connect() as conn:
        for tabla, columna in columnas:
            chk = text("""
                SELECT data_type FROM information_schema.columns
                WHERE table_name = :t AND column_name = :c
            """)
            row = conn.execute(chk, {"t": tabla, "c": columna}).fetchone()
            if not row:
                print(f"OMITIDO: {tabla}.{columna} no encontrada")
                continue
            if row[0] != 'double precision':
                print(f"OMITIDO: {tabla}.{columna} ya es {row[0]}")
                continue
            try:
                conn.execute(text(
                    f"ALTER TABLE {tabla} ALTER COLUMN {columna} "
                    f"TYPE NUMERIC(10,2) USING ROUND({columna}::NUMERIC, 2)"
                ))
                conn.commit()
                print(f"OK: {tabla}.{columna} → NUMERIC(10,2)")
            except Exception as e:
                conn.rollback()
                print(f"ERROR {tabla}.{columna}: {e}")
    print("fix-float-to-numeric completado.")

@app.cli.command("fix-add-token-usado")
@with_appcontext
def fix_add_token_usado():
    """Crea la tabla token_usado para la blocklist de tokens de reset. Idempotente."""
    with db.engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS token_usado (
                id         SERIAL PRIMARY KEY,
                token_hash VARCHAR(64) NOT NULL UNIQUE,
                usado_en   TIMESTAMP NOT NULL DEFAULT NOW(),
                expira_en  TIMESTAMP NOT NULL
            )
        """))
        conn.execute(text(
            "CREATE INDEX IF NOT EXISTS ix_token_usado_token_hash ON token_usado (token_hash)"
        ))
        conn.commit()
    print("fix-add-token-usado completado.")

@app.cli.command("limpiar-tokens-expirados")
@with_appcontext
def limpiar_tokens_expirados():
    """Elimina registros de token_usado cuya fecha de expiración ya pasó."""
    with db.engine.connect() as conn:
        result = conn.execute(text("DELETE FROM token_usado WHERE expira_en < NOW()"))
        conn.commit()
        print(f"limpiar-tokens-expirados: {result.rowcount} registros eliminados.")

@app.cli.command("migrate-fase-c")
@with_appcontext
def migrate_fase_c():
    """Agrega centro_costo_id a gasto, pago_servicio y factura_proveedor (seguro, idempotente)."""
    from sqlalchemy import text
    stmts = [
        "ALTER TABLE gasto ADD COLUMN IF NOT EXISTS centro_costo_id INTEGER REFERENCES centro_costo(id) ON DELETE SET NULL",
        "ALTER TABLE pago_servicio ADD COLUMN IF NOT EXISTS centro_costo_id INTEGER REFERENCES centro_costo(id) ON DELETE SET NULL",
        "ALTER TABLE factura_proveedor ADD COLUMN IF NOT EXISTS centro_costo_id INTEGER REFERENCES centro_costo(id) ON DELETE SET NULL",
    ]
    with db.engine.connect() as conn:
        for stmt in stmts:
            try:
                conn.execute(text(stmt))
                conn.commit()
                print(f"OK: {stmt[:70]}")
            except Exception as e:
                conn.rollback()
                print(f"Omitido (ya existe?): {e}")

@app.cli.command("limpiar-push-subs")
@with_appcontext
def limpiar_push_subs():
    """Elimina TODAS las suscripciones push de la BD (usar tras cambiar claves VAPID)."""
    n = PushSubscription.query.delete()
    db.session.commit()
    print(f"limpiar-push-subs: {n} suscripciones eliminadas. Los usuarios deberán re-suscribirse.")

@app.cli.command("gen-vapid")
@with_appcontext
def gen_vapid_command():
    """Genera un par de claves VAPID para Web Push Notifications."""
    try:
        from py_vapid import Vapid
        v = Vapid()
        v.generate_keys()
        pub  = v.public_key_urlsafe_b64
        priv = v.private_key_urlsafe_b64
        if isinstance(pub,  bytes): pub  = pub.decode()
        if isinstance(priv, bytes): priv = priv.decode()
        print("\n=== CLAVES VAPID GENERADAS ===")
        print(f"VAPID_PUBLIC_KEY={pub}")
        print(f"VAPID_PRIVATE_KEY={priv}")
        print(f"VAPID_CLAIMS_EMAIL=notifications@tudominio.com")
        print("\nAgrega estas líneas a tu archivo .env en el servidor.")
    except ImportError:
        print("Error: ejecuta 'pip install pywebpush' primero.")

@app.context_processor
def inject_vapid_key():
    return {'vapid_public_key': os.environ.get('VAPID_PUBLIC_KEY', '')}

@app.context_processor
def inject_aprobaciones_badge():
    """Inyecta conteo de solicitudes de aprobación pendientes para admins."""
    if (current_user.is_authenticated and current_user.organizacion_id
            and current_user.rol in ['super_admin', 'admin']):
        try:
            count = SolicitudAprobacion.query.filter_by(
                organizacion_id=current_user.organizacion_id,
                estado='pendiente'
            ).count()
            return {'aprobaciones_badge': count}
        except Exception:
            return {'aprobaciones_badge': 0}
    return {'aprobaciones_badge': 0}

@app.context_processor
def inject_servicios_badge():
    """Inyecta conteo de pagos de servicios urgentes/vencidos en todos los templates."""
    if current_user.is_authenticated and current_user.organizacion_id:
        try:
            hoy = now_mx().date()
            count = PagoServicio.query.join(Servicio).filter(
                Servicio.organizacion_id == current_user.organizacion_id,
                PagoServicio.estado.in_(['pendiente', 'vencido']),
                PagoServicio.fecha_vencimiento <= hoy + timedelta(days=7)
            ).count()
            return {'servicios_badge': count}
        except Exception:
            return {'servicios_badge': 0}
    return {'servicios_badge': 0}

@app.cli.command("make-super-admin")
@with_appcontext
@click.argument("username")
def make_super_admin_command(username):
    """Asigna el rol 'super_admin' a un usuario existente."""
    user = User.query.filter_by(username=username).first()
    if user:
        user.rol = 'super_admin'
        user.organizacion_id = None 
        db.session.commit()
        print(f"¡Éxito! El usuario '{username}' ahora es Super Admin.")
    else:
        print(f"Error: No se encontró al usuario '{username}'.")

# ==============================================================================
# 4. MODELOS DE BASE DE DATOS
# ==============================================================================

# --- Modelos Principales (Padres) ---

class Organizacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(120), unique=True, nullable=False)
    codigo_invitacion = db.Column(db.String(10), unique=True, nullable=True)
    
    # --- CAMPOS DE PERSONALIZACIÓN ---
    logo_url          = db.Column(db.String(255), nullable=True)
    header_titulo     = db.Column(db.String(150), nullable=True)
    header_subtitulo  = db.Column(db.String(200), nullable=True)
    color_primario    = db.Column(db.String(7),  default='#333333')
    color_secundario  = db.Column(db.String(7),  default='#f1f5f9')
    tipo_letra        = db.Column(db.String(50), default='Helvetica')
    direccion         = db.Column(db.Text,       nullable=True)
    telefono          = db.Column(db.String(20), nullable=True)
    rfc               = db.Column(db.String(20), nullable=True)
    correo_empresa    = db.Column(db.String(120),nullable=True)
    footer_texto      = db.Column(db.Text,       nullable=True)
    pdf_mostrar_qr    = db.Column(db.Boolean,    default=False)
    whatsapp_notify   = db.Column(db.String(25), nullable=True, default=None)

    # --- ETIQUETAS ---
    etiqueta_fuente       = db.Column(db.String(50), default='Inter')
    etiqueta_color_fondo  = db.Column(db.String(7),  default='#FFFFFF')
    etiqueta_color_texto  = db.Column(db.String(7),  default='#1a1a1a')
    etiqueta_color_sku    = db.Column(db.String(7),  default='#1f4e79')
    etiqueta_estilo       = db.Column(db.String(20), default='moderno')
    etiqueta_mostrar_logo = db.Column(db.Boolean,    default=True)

    # --- EXCEL ---
    excel_color_header   = db.Column(db.String(7),  default='#1f4e79')
    excel_color_accent   = db.Column(db.String(7),  default='#dbeafe')
    excel_fuente         = db.Column(db.String(30), default='Calibri')
    excel_mostrar_logo   = db.Column(db.Boolean,    default=True)
    excel_mostrar_id     = db.Column(db.Boolean,    default=True)
    excel_mostrar_oc     = db.Column(db.Boolean,    default=True)
    excel_mostrar_origen = db.Column(db.Boolean,    default=True)

    usuarios = db.relationship('User', backref='organizacion', lazy=True)
    productos = db.relationship('Producto', backref='organizacion', lazy=True)
    categorias = db.relationship('Categoria', backref='organizacion', lazy=True)
    proveedores = db.relationship('Proveedor', backref='organizacion', lazy=True)
    ordenes_compra = db.relationship('OrdenCompra', backref='organizacion', lazy=True)
    salidas = db.relationship('Salida', backref='organizacion', lazy=True)
    gastos = db.relationship('Gasto', backref='organizacion', lazy=True)
    movimientos = db.relationship('Movimiento', backref='organizacion', lazy=True)
    proyectos_oc = db.relationship('ProyectoOC', backref='organizacion', lazy=True)

    # --- AÑADIDO ---
    almacenes = db.relationship('Almacen', backref='organizacion', lazy=True)

    def __repr__(self):
        return f'<Organizacion {self.nombre}>'

class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    image_file = db.Column(db.String(20), nullable=False, default='default.jpg')
    password_hash = db.Column(db.String(255), nullable=False) 
    
    rol = db.Column(db.String(20), nullable=False, default='user')
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=True)
    
    # --- PERMISOS GRANULARES ---
    perm_view_dashboard = db.Column(db.Boolean, nullable=False, default=False)
    perm_view_management = db.Column(db.Boolean, nullable=False, default=False)
    perm_edit_management = db.Column(db.Boolean, nullable=False, default=False)
    perm_create_oc_standard = db.Column(db.Boolean, nullable=False, default=False)
    perm_create_oc_proyecto = db.Column(db.Boolean, nullable=False, default=False)
    perm_do_salidas = db.Column(db.Boolean, nullable=False, default=False)
    perm_view_gastos = db.Column(db.Boolean, nullable=False, default=False)
    
    # Relaciones de Auditoría
    ordenes_creadas = db.relationship('OrdenCompra', foreign_keys='OrdenCompra.creador_id', backref='creador', lazy=True)
    ordenes_canceladas = db.relationship('OrdenCompra', foreign_keys='OrdenCompra.cancelado_por_id', backref='cancelado_por', lazy=True)
    salidas_creadas = db.relationship('Salida', foreign_keys='Salida.creador_id', backref='creador', lazy=True)
    salidas_canceladas = db.relationship('Salida', foreign_keys='Salida.cancelado_por_id', backref='cancelado_por', lazy=True)
    proyectos_oc_creados = db.relationship('ProyectoOC', foreign_keys='ProyectoOC.creador_id', backref='creador', lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'<User {self.username}>'

# --- Modelos Secundarios (Hijos) ---

class Proveedor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False, unique=True)
    contacto_email = db.Column(db.String(100))
    contacto_telefono = db.Column(db.String(50), nullable=True)
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)

class Categoria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False, unique=True)
    descripcion = db.Column(db.String(255), nullable=True)
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)

# --- MODELO 'Producto' MODIFICADO ---
class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(255), nullable=False)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    # --- CAMPOS DE STOCK ELIMINADOS ---
    precio_unitario = db.Column(db.Numeric(10, 2), default=0)
    imagen_url = db.Column(db.String(255), nullable=True)
    
    # --- NUEVO CAMPO ---
    enlace_proveedor = db.Column(db.Text, nullable=True)
    
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria.id'), nullable=True)
    categoria = db.relationship('Categoria', backref='productos', lazy=True)
    
    proveedor_id = db.Column(db.Integer, db.ForeignKey('proveedor.id'), nullable=True)
    proveedor = db.relationship('Proveedor', backref='productos', lazy=True)
    
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)
    unidades_por_caja = db.Column(db.Integer, default=1)
    
    # --- NUEVA RELACIÓN ---
    stocks = db.relationship('Stock', backref='producto', lazy='dynamic', cascade="all, delete-orphan")
    
    def get_stock_total(self):
        """ Suma el stock de este producto en TODOS los almacenes. """
        return db.session.query(db.func.sum(Stock.cantidad)).filter_by(producto_id=self.id).scalar() or 0


# --- NUEVO MODELO 'Almacen' ---
class Almacen(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    ubicacion = db.Column(db.String(255), nullable=True) # ej. "Pasillo 5, Estante A"
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)
    
    stocks = db.relationship('Stock', backref='almacen', lazy='dynamic', cascade="all, delete-orphan")

    def __repr__(self):
        return f'<Almacen {self.nombre}>'

# --- NUEVO MODELO 'Stock' ---
class Stock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    almacen_id = db.Column(db.Integer, db.ForeignKey('almacen.id'), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False, default=0)
    
    stock_minimo = db.Column(db.Integer, nullable=False, default=5)
    stock_maximo = db.Column(db.Integer, nullable=False, default=100)
    
    # --- NUEVO CAMPO ---
    ubicacion = db.Column(db.String(100), nullable=True) 
    
    __table_args__ = (db.UniqueConstraint('producto_id', 'almacen_id', name='_producto_almacen_uc'),)
    
    @property
    def estado_stock(self):
        if self.cantidad < self.stock_minimo:
            return 'bajo'
        elif self.cantidad > self.stock_maximo:
            return 'exceso'
        return 'ok'
    def __repr__(self):
        return f'<Stock ProdID: {self.producto_id} AlmID: {self.almacen_id} Qty: {self.cantidad}>'

# --- MODELO 'OrdenCompra' MODIFICADO ---
class OrdenCompra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha_creacion = db.Column(db.DateTime, nullable=False, default=now_mx)
    fecha_recepcion = db.Column(db.DateTime, nullable=True)
    estado = db.Column(db.String(20), nullable=False, default='borrador')
    
    proveedor_id = db.Column(db.Integer, db.ForeignKey('proveedor.id'), nullable=False)
    # --- LÍNEA RESTAURADA ---
    proveedor = db.relationship('Proveedor', backref='ordenes_compra', lazy=True)
    # ------------------------

    almacen_id = db.Column(db.Integer, db.ForeignKey('almacen.id'), nullable=False)
    almacen = db.relationship('Almacen')
    
    detalles = db.relationship('OrdenCompraDetalle', backref='orden', lazy=True, cascade="all, delete-orphan")
    
    creador_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    cancelado_por_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)
    
    @property
    def costo_total(self):
        return sum(detalle.subtotal for detalle in self.detalles)

class OrdenCompraDetalle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    orden_id = db.Column(db.Integer, db.ForeignKey('orden_compra.id'), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    producto = db.relationship('Producto')
    cantidad_solicitada = db.Column(db.Integer, nullable=False, default=1)
    cajas = db.Column(db.Float, nullable=True, default=0.0)
    costo_unitario_estimado = db.Column(db.Numeric(10, 2), nullable=True, default=0)
    enlace_proveedor = db.Column(db.Text, nullable=True)

    @property
    def subtotal(self):
        return self.cantidad_solicitada * (self.costo_unitario_estimado or 0)

class Gasto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    descripcion = db.Column(db.String(255), nullable=False)
    monto = db.Column(db.Numeric(10, 2), nullable=False, default=0)
    categoria = db.Column(db.String(50), nullable=True)
    fecha = db.Column(db.DateTime, nullable=False, default=now_mx)
    
    orden_compra_id = db.Column(db.Integer, db.ForeignKey('orden_compra.id'), nullable=True)
    orden_compra = db.relationship('OrdenCompra', backref='gastos_asociados', lazy=True)

    centro_costo_id = db.Column(db.Integer, db.ForeignKey('centro_costo.id'), nullable=True)
    centro_costo = db.relationship('CentroCosto', backref='gastos', lazy=True)

    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)

    def __repr__(self):
        return f'<Gasto {self.descripcion} - ${self.monto}>'
    
class Salida(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False, default=now_mx().date)
    estado = db.Column(db.String(20), nullable=False, default='abierta')
    
    creador_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    cancelado_por_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True) # (Para auditoría futura) 
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)

    # --- LÍNEA AÑADIDA ---
    almacen_id = db.Column(db.Integer, db.ForeignKey('almacen.id'), nullable=False)
    almacen = db.relationship('Almacen') # Para fácil acceso
    
    movimientos = db.relationship('Movimiento', backref='salida', lazy='dynamic', cascade="all, delete-orphan")

class Movimiento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    producto = db.relationship('Producto', backref='movimientos', lazy=True)
    
    cantidad = db.Column(db.Integer, nullable=False) 
    tipo = db.Column(db.String(20), nullable=False) 
    fecha = db.Column(db.DateTime, nullable=False, default=now_mx)
    
    motivo = db.Column(db.String(255), nullable=False) 
    
    orden_compra_id = db.Column(db.Integer, db.ForeignKey('orden_compra.id'), nullable=True)
    salida_id = db.Column(db.Integer, db.ForeignKey('salida.id'), nullable=True)

    almacen_id = db.Column(db.Integer, db.ForeignKey('almacen.id'), nullable=True)
    almacen    = db.relationship('Almacen', foreign_keys=[almacen_id], backref='movimientos', lazy=True)

    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)

    def __repr__(self):
        return f'<Movimiento {self.producto_id} ({self.cantidad})>'

# --- MODELO 'AuditLog' ---
class AuditLog(db.Model):
    __tablename__ = 'audit_log'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, nullable=False, default=now_mx, index=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    usuario = db.relationship('User', foreign_keys=[usuario_id])
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False, index=True)
    accion = db.Column(db.String(30), nullable=False)
    entidad = db.Column(db.String(50), nullable=False)
    entidad_id = db.Column(db.Integer, nullable=True)
    descripcion = db.Column(db.String(500), nullable=False)

    def __repr__(self):
        return f'<AuditLog {self.accion} {self.entidad} #{self.entidad_id}>'

# --- MODELO 'PushSubscription' ---
class PushSubscription(db.Model):
    __tablename__ = 'push_subscription'
    id                = db.Column(db.Integer, primary_key=True)
    user_id           = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    organizacion_id   = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)
    endpoint          = db.Column(db.Text, nullable=False, unique=True)
    subscription_json = db.Column(db.Text, nullable=False)
    creada_en         = db.Column(db.DateTime, default=now_mx)
    user              = db.relationship('User', backref='push_subscriptions')

# --- MODELO 'ProyectoOC' MODIFICADO ---
            
class ProyectoOC(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre_proyecto = db.Column(db.String(255), nullable=False)
    fecha_creacion  = db.Column(db.DateTime, nullable=False, default=now_mx)
    estado          = db.Column(db.String(20), nullable=False, default='borrador')

    creador_id      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)

    almacen_id      = db.Column(db.Integer, db.ForeignKey('almacen.id'), nullable=True)
    almacen         = db.relationship('Almacen', foreign_keys=[almacen_id])

    # Trazabilidad de estados
    fecha_envio      = db.Column(db.DateTime, nullable=True)
    fecha_recepcion  = db.Column(db.DateTime, nullable=True)
    recibido_por_id  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)

    recibido_por = db.relationship('User', foreign_keys=[recibido_por_id], overlaps="proyectos_oc_creados")

    detalles = db.relationship('ProyectoOCDetalle', backref='proyecto_oc', lazy=True, cascade="all, delete-orphan")

    @property
    def costo_total(self):
        return sum(detalle.subtotal for detalle in self.detalles)

class ProyectoOCDetalle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    proyecto_oc_id = db.Column(db.Integer, db.ForeignKey('proyecto_oc.id'), nullable=False)
    
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=True)
    producto = db.relationship('Producto')
    
    descripcion_nuevo = db.Column(db.Text, nullable=True)
    proveedor_sugerido = db.Column(db.String(255), nullable=True)
    
    cantidad = db.Column(db.Integer, nullable=False, default=1)
    costo_unitario = db.Column(db.Numeric(10, 2), nullable=False, default=0)

    # --- NUEVOS CAMPOS AÑADIDOS ---
    enlace_proveedor = db.Column(db.Text, nullable=True)
    comentarios_detalle = db.Column(db.Text, nullable=True)

    @property
    def subtotal(self):
        return self.cantidad * self.costo_unitario

class Servicio(db.Model):
    __tablename__ = 'servicio'
    id               = db.Column(db.Integer, primary_key=True)
    nombre           = db.Column(db.String(100), nullable=False)
    tipo             = db.Column(db.String(30),  default='otro')
    proveedor_nombre = db.Column(db.String(80),  nullable=True)
    numero_contrato  = db.Column(db.String(60),  nullable=True)
    dia_vencimiento  = db.Column(db.Integer,     nullable=True)  # día del mes 1-31
    dias_aviso       = db.Column(db.Integer,     default=5)
    notas            = db.Column(db.Text,        nullable=True)
    activo           = db.Column(db.Boolean,     default=True)
    organizacion_id  = db.Column(db.Integer, db.ForeignKey('organizacion.id'), nullable=False)
    creado_en        = db.Column(db.DateTime,    default=now_mx)
    pagos            = db.relationship('PagoServicio', backref='servicio', lazy=True,
                                       order_by='PagoServicio.fecha_vencimiento.desc()',
                                       cascade='all, delete-orphan')


class PagoServicio(db.Model):
    __tablename__ = 'pago_servicio'
    id                = db.Column(db.Integer,    primary_key=True)
    servicio_id       = db.Column(db.Integer,    db.ForeignKey('servicio.id'), nullable=False)
    monto             = db.Column(db.Numeric(10, 2), nullable=False)
    fecha_vencimiento = db.Column(db.Date,       nullable=False)
    fecha_pago        = db.Column(db.Date,       nullable=True)
    estado            = db.Column(db.String(20), default='pendiente')  # pendiente | pagado | vencido
    notas             = db.Column(db.Text,       nullable=True)
    comprobante_url   = db.Column(db.String(300),nullable=True)
    registrado_por_id = db.Column(db.Integer,    db.ForeignKey('user.id'), nullable=True)
    creado_en         = db.Column(db.DateTime,   default=now_mx)
    centro_costo_id   = db.Column(db.Integer,    db.ForeignKey('centro_costo.id'), nullable=True)
    centro_costo      = db.relationship('CentroCosto', backref='pagos_servicio', lazy=True)


# --- MODELO 'FacturaProveedor' — Cuentas por Pagar ---
class FacturaProveedor(db.Model):
    __tablename__ = 'factura_proveedor'
    id                = db.Column(db.Integer,    primary_key=True)
    numero_factura    = db.Column(db.String(80),  nullable=False)
    proveedor_id      = db.Column(db.Integer,    db.ForeignKey('proveedor.id'), nullable=False)
    proveedor         = db.relationship('Proveedor', backref='facturas')
    orden_compra_id   = db.Column(db.Integer,    db.ForeignKey('orden_compra.id'), nullable=True)
    orden_compra      = db.relationship('OrdenCompra', backref='facturas')
    monto             = db.Column(db.Numeric(10, 2), nullable=False)
    fecha_emision     = db.Column(db.Date,       nullable=False)
    fecha_vencimiento = db.Column(db.Date,       nullable=False)
    estado            = db.Column(db.String(20), nullable=False, default='pendiente')  # pendiente | pagado | vencido
    notas             = db.Column(db.Text,       nullable=True)
    registrado_por_id = db.Column(db.Integer,    db.ForeignKey('user.id'), nullable=True)
    registrado_por    = db.relationship('User',  foreign_keys=[registrado_por_id])
    creado_en         = db.Column(db.DateTime,   default=now_mx)
    organizacion_id   = db.Column(db.Integer,    db.ForeignKey('organizacion.id'), nullable=False)
    centro_costo_id   = db.Column(db.Integer,    db.ForeignKey('centro_costo.id'), nullable=True)
    centro_costo      = db.relationship('CentroCosto', backref='facturas', lazy=True)

    @property
    def dias_vencimiento(self):
        return (self.fecha_vencimiento - date.today()).days

    @property
    def esta_vencida(self):
        return self.estado == 'pendiente' and self.fecha_vencimiento < date.today()

    @property
    def bucket_aging(self):
        if self.estado == 'pagado':
            return 'pagado'
        dias = (date.today() - self.fecha_vencimiento).days
        if dias <= 0:
            return 'vigente'
        elif dias <= 30:
            return '1-30'
        elif dias <= 60:
            return '31-60'
        elif dias <= 90:
            return '61-90'
        return '90+'


# --- MODELO 'CentroCosto' — Fase C ERP ---
class CentroCosto(db.Model):
    __tablename__ = 'centro_costo'
    id              = db.Column(db.Integer,    primary_key=True)
    nombre          = db.Column(db.String(120), nullable=False)
    descripcion     = db.Column(db.Text,       nullable=True)
    presupuesto     = db.Column(db.Float,      nullable=True)   # None = sin límite
    estado          = db.Column(db.String(20), nullable=False, default='activo')  # activo | cerrado
    creado_en       = db.Column(db.DateTime,   default=now_mx)
    organizacion_id = db.Column(db.Integer,    db.ForeignKey('organizacion.id'), nullable=False)
    creador_id      = db.Column(db.Integer,    db.ForeignKey('user.id'), nullable=True)
    creador         = db.relationship('User',  foreign_keys=[creador_id])

    @property
    def total_gastado(self):
        g = sum(x.monto for x in self.gastos)
        p = sum(x.monto for x in self.pagos_servicio if x.estado == 'pagado')
        f = sum(x.monto for x in self.facturas)
        return g + p + f

    @property
    def pct_presupuesto(self):
        if not self.presupuesto or self.presupuesto <= 0:
            return None
        return min(round(self.total_gastado / self.presupuesto * 100, 1), 100)


class Presupuesto(db.Model):
    __tablename__ = 'presupuesto'
    id              = db.Column(db.Integer,    primary_key=True)
    categoria       = db.Column(db.String(50), nullable=False)
    anio            = db.Column(db.Integer,    nullable=False)
    mes             = db.Column(db.Integer,    nullable=True)   # None = presupuesto anual
    monto           = db.Column(db.Float,      nullable=False)
    organizacion_id = db.Column(db.Integer,    db.ForeignKey('organizacion.id'), nullable=False)
    creado_en       = db.Column(db.DateTime,   default=now_mx)

    __table_args__ = (
        db.UniqueConstraint('organizacion_id', 'categoria', 'anio', 'mes',
                            name='uq_presupuesto_cat_periodo'),
    )

    @property
    def periodo_label(self):
        if self.mes:
            return f"{MESES_ES[self.mes]} {self.anio}"
        return f"Anual {self.anio}"

class SolicitudAprobacion(db.Model):
    __tablename__ = 'solicitud_aprobacion'
    id              = db.Column(db.Integer,    primary_key=True)
    entidad_tipo    = db.Column(db.String(30), nullable=False)   # 'proyecto_oc'
    entidad_id      = db.Column(db.Integer,    nullable=False)
    solicitante_id  = db.Column(db.Integer,    db.ForeignKey('user.id'), nullable=False)
    aprobador_id    = db.Column(db.Integer,    db.ForeignKey('user.id'), nullable=True)
    estado          = db.Column(db.String(20), nullable=False, default='pendiente')  # pendiente|aprobado|rechazado
    comentario      = db.Column(db.Text,       nullable=True)
    creado_en       = db.Column(db.DateTime,   default=now_mx)
    resuelto_en     = db.Column(db.DateTime,   nullable=True)
    organizacion_id = db.Column(db.Integer,    db.ForeignKey('organizacion.id'), nullable=False)

    solicitante = db.relationship('User', foreign_keys=[solicitante_id])
    aprobador   = db.relationship('User', foreign_keys=[aprobador_id])

class TokenUsado(db.Model):
    """Blocklist de tokens de reset de contraseña ya utilizados (SHA-256)."""
    __tablename__ = 'token_usado'
    id          = db.Column(db.Integer,     primary_key=True)
    token_hash  = db.Column(db.String(64),  unique=True, nullable=False, index=True)
    usado_en    = db.Column(db.DateTime,    nullable=False, default=datetime.utcnow)
    expira_en   = db.Column(db.DateTime,    nullable=False)

# ==============================================================================
# 5. CARGADOR DE USUARIO (FLASK-LOGIN)
# ==============================================================================

@login_manager.user_loader
def load_user(user_id):
    """Callback para recargar el objeto User desde el ID de la sesión."""
    return User.query.get(int(user_id))

# ==============================================================================
# 6. FORMULARIOS (FLASK-WTF)
# ==============================================================================

class RegistrationForm(FlaskForm):
    username = StringField('Usuario', validators=[DataRequired(), Length(min=4, max=80)])
    email = StringField('E-mail', validators=[DataRequired(), Email(message='E-mail no válido.')])
    password = PasswordField('Contraseña', validators=[DataRequired(), Length(min=6)])
    confirm_password = PasswordField('Confirmar Contraseña', 
                                     validators=[DataRequired(), EqualTo('password', message='Las contraseñas deben coincidir.')])
    
    # --- LÍNEA AÑADIDA ---
    codigo_invitacion = StringField('Código de Invitación (Opcional)')
    
    submit = SubmitField('Registrarse')
    def validate_username(self, username):
        user = User.query.filter_by(username=username.data).first()
        if user:
            raise ValidationError('Ese nombre de usuario ya existe. Por favor, elige otro.')
            
    def validate_email(self, email):
        user = User.query.filter_by(email=email.data).first()
        if user:
            raise ValidationError('Ese e-mail ya está registrado. Por favor, usa otro.')

class LoginForm(FlaskForm):
    username = StringField('Usuario', validators=[DataRequired()])
    password = PasswordField('Contraseña', validators=[DataRequired()])
    submit = SubmitField('Iniciar Sesión')

class UpdateAccountForm(FlaskForm):
    username = StringField('Usuario', validators=[DataRequired(), Length(min=4, max=80)])
    email = StringField('E-mail', validators=[DataRequired(), Email(message='E-mail no válido.')])
    picture = FileField('Actualizar Foto de Perfil', validators=[FileAllowed(['jpg', 'png', 'jpeg'])])
    submit_account = SubmitField('Actualizar Datos')

    def validate_username(self, username):
        if username.data != current_user.username:
            user = User.query.filter_by(username=username.data).first()
            if user:
                raise ValidationError('Ese nombre de usuario ya existe. Por favor, elige otro.')
            
    def validate_email(self, email):
        if email.data != current_user.email:
            user = User.query.filter_by(email=email.data).first()
            if user:
                raise ValidationError('Ese e-mail ya está registrado. Por favor, usa otro.')

class ChangePasswordForm(FlaskForm):
    old_password = PasswordField('Contraseña Actual', validators=[DataRequired()])
    password = PasswordField('Nueva Contraseña', validators=[DataRequired(), Length(min=6)])
    confirm_password = PasswordField('Confirmar Nueva Contraseña', 
                                     validators=[DataRequired(), EqualTo('password', message='Las contraseñas deben coincidir.')])
    submit_password = SubmitField('Cambiar Contraseña')

class RequestResetForm(FlaskForm):
    email = StringField('E-mail', validators=[DataRequired(), Email()])
    submit = SubmitField('Solicitar Reseteo de Contraseña')

class ResetPasswordForm(FlaskForm):
    password = PasswordField('Nueva Contraseña', validators=[DataRequired(), Length(min=6)])
    confirm_password = PasswordField('Confirmar Nueva Contraseña', 
                                     validators=[DataRequired(), EqualTo('password', message='Las contraseñas deben coincidir.')])
    submit = SubmitField('Restablecer Contraseña')

class AdminPermissionForm(FlaskForm):
    perm_view_dashboard = BooleanField('Ver Inventario')
    perm_view_management = BooleanField('Ver Gestión (Cat/Prov)')
    perm_edit_management = BooleanField('Editar Gestión (Cat/Prov/Prod)')
    perm_create_oc_standard = BooleanField('Crear OC Normal')
    perm_create_oc_proyecto = BooleanField('Crear OC Proyecto')
    perm_do_salidas = BooleanField('Registrar Salidas')
    perm_view_gastos = BooleanField('Ver/Crear Gastos')
    submit = SubmitField('Guardar Permisos')

# ==============================================================================
# 7. FUNCIONES AUXILIARES (Decoradores, Subida de Imágenes)
# ==============================================================================

def super_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol != 'super_admin':
            flash('Acceso denegado. Se requieren privilegios de Super Administrador.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    """Verifica si la extensión del archivo es válida."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_picture(form_picture):
    """Guarda y redimensiona la foto de perfil subida."""
    random_hex = secrets.token_hex(8)
    _, f_ext = os.path.splitext(form_picture.filename)
    picture_fn = random_hex + f_ext
    picture_path = os.path.join(app.root_path, 'static/uploads/profile_pics', picture_fn)

    output_size = (125, 125)
    i = Image.open(form_picture)
    i.thumbnail(output_size)
    i.save(picture_path)

    return picture_fn

def check_org_permission(f):
    """
    Decorador para verificar que un usuario (no super_admin)
    pertenece a una organización antes de crear/ver datos.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.rol != 'super_admin' and not current_user.organizacion_id:
            flash('No puedes realizar esta acción. Primero debes ser asignado a una organización por un Super Admin.', 'warning')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def get_item_or_404(model, item_id):
    """
    Función de seguridad que obtiene un item Y verifica
    que pertenece a la organización del usuario.
    """
    if current_user.rol == 'super_admin':
        query = model.query
    else:
        query = model.query.filter_by(organizacion_id=current_user.organizacion_id)
    
    item = query.filter_by(id=item_id).first_or_404()
    return item

def admin_required(f):
    """
    Decorador personalizado para verificar que el usuario
    sea 'admin' o 'super_admin'.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.rol not in ['super_admin', 'admin']:
            flash('No tienes permiso para acceder a esta página.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def check_permission(permission_name):
    """
    Decorador personalizado que verifica si un usuario tiene un permiso específico.
    Los 'admin' y 'super_admin' siempre tienen permiso.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if current_user.rol in ['super_admin', 'admin']:
                return f(*args, **kwargs)
            
            if not getattr(current_user, permission_name, False):
                flash('No tienes permiso para acceder a esta función.', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator
  
# ==============================================================================
# 🛡️ CABECERAS DE SEGURIDAD (SECURITY HEADERS)
# ==============================================================================
@app.after_request
def add_security_headers(response):
    # 1. Prevenir Clickjacking
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    
    # 2. Prevenir MIME-sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'
    
    # 3. Forzar conexiones seguras HTTPS (HSTS)
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    
    # 4. Política de Referencia
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    
    # 5. Desactivar hardware innecesario (cámara permitida para escáner QR)
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=()'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    
    # 6. Content Security Policy (CSP)
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' cdn.jsdelivr.net fonts.googleapis.com; "
        "font-src 'self' data: cdn.jsdelivr.net fonts.gstatic.com; "
        "img-src 'self' data: blob: https:; "
        "connect-src 'self';"
    )
    response.headers['Content-Security-Policy'] = csp
    
    return response
# ==============================================================================
# 8. RUTAS DE LA APLICACIÓN
# ==============================================================================

# -----------------------------------------------------------------
# GESTIÓN DE USUARIOS Y CONTRASEÑAS (ADMIN)
# -----------------------------------------------------------------

@app.route('/usuarios')
@login_required
def lista_usuarios():
    """Muestra la lista de todos los usuarios registrados (Solo Admins)."""
    if current_user.rol not in ['super_admin', 'admin']:
        flash('Acceso restringido a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    usuarios = User.query.order_by(User.username).all()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/admin/usuario/<int:id>/reset_password', methods=['POST'])
@login_required
def admin_reset_password(id):
    """
    Acción para que un Admin fuerce el cambio de contraseña de otro usuario.
    """
    # 1. Seguridad: Solo Admins
    if current_user.rol not in ['super_admin', 'admin']:
        flash('No tienes permisos para realizar esta acción.', 'danger')
        return redirect(url_for('dashboard'))

    # 2. Buscar usuario
    usuario_objetivo = User.query.get_or_404(id)
    
    # 3. Obtener nueva contraseña del form
    nueva_password = request.form.get('new_password')
    
    if not nueva_password or len(nueva_password) < 8:
        flash('La contraseña es muy corta (mínimo 8 caracteres).', 'warning')
        return redirect(url_for('lista_usuarios'))

    try:
        # 4. Sobrescribir contraseña
        usuario_objetivo.password_hash = generate_password_hash(nueva_password)
        db.session.commit()
        
        flash(f'✅ Contraseña actualizada correctamente para: {usuario_objetivo.username}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar: {e}', 'danger')

    return redirect(url_for('lista_usuarios'))

def enviar_correo_api(destinatario, reset_url):
    """
    Envía correo de recuperación via Brevo API.
    Devuelve (True, None) en éxito o (False, mensaje_error) en fallo.
    """
    import logging
    API_KEY = os.environ.get("BREVO_API_KEY")
    SENDER_EMAIL = os.environ.get("BREVO_SENDER_EMAIL", "deinventarioc@gmail.com")
    SENDER_NAME  = os.environ.get("BREVO_SENDER_NAME",  "Soporte Inventario")

    if not API_KEY:
        logging.error("[BREVO] Falta BREVO_API_KEY en el entorno del servidor.")
        return False, "BREVO_API_KEY no configurada"

    payload = {
        "sender": {"name": SENDER_NAME, "email": SENDER_EMAIL},
        "to": [{"email": destinatario}],
        "subject": "Restablecimiento de Contraseña — Gestor de Inventario",
        "htmlContent": f"""
            <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto;
                        border:1px solid #e2e8f0;border-radius:12px;overflow:hidden;">
                <div style="background:#4f46e5;padding:24px;text-align:center;">
                    <h2 style="margin:0;color:#fff;font-size:20px;">🔑 Gestor de Inventario</h2>
                </div>
                <div style="padding:32px;background:#f8fafc;text-align:center;">
                    <p style="font-size:16px;color:#1e293b;">
                        Recibimos una solicitud para restablecer la contraseña de tu cuenta.
                    </p>
                    <a href="{reset_url}"
                       style="display:inline-block;padding:14px 32px;margin:20px 0;
                              background:#4f46e5;color:#fff;text-decoration:none;
                              border-radius:8px;font-weight:bold;font-size:15px;">
                        Restablecer contraseña
                    </a>
                    <p style="font-size:13px;color:#64748b;">
                        El enlace expira en <strong>1 hora</strong>.<br>
                        Si no solicitaste este cambio, ignora este correo.
                    </p>
                </div>
                <div style="padding:16px;text-align:center;background:#f1f5f9;">
                    <p style="font-size:11px;color:#94a3b8;margin:0;">
                        Enviado desde: {SENDER_EMAIL}
                    </p>
                </div>
            </div>
        """
    }

    headers = {
        "accept": "application/json",
        "api-key": API_KEY,
        "content-type": "application/json"
    }

    try:
        response = requests.post("https://api.brevo.com/v3/smtp/email",
                                 json=payload, headers=headers, timeout=10)
        if response.status_code in (200, 201):
            logging.info(f"[BREVO] OK — correo enviado a {destinatario}")
            return True, None
        else:
            detail = response.text[:500]
            logging.error(f"[BREVO] HTTP {response.status_code} — {detail}")
            return False, f"HTTP {response.status_code}: {detail}"
    except Exception as e:
        logging.error(f"[BREVO] Excepción: {e}")
        return False, str(e)

# ==============================================================================
# HISTORIAL DE ACTIVIDAD / AUDIT LOG
# ==============================================================================

def log_actividad(accion, entidad, descripcion, entidad_id=None):
    """Añade una entrada al audit log. Debe llamarse ANTES del db.session.commit()."""
    try:
        org_id = current_user.organizacion_id if current_user.is_authenticated else None
        if not org_id:
            return
        entrada = AuditLog(
            usuario_id=current_user.id if current_user.is_authenticated else None,
            organizacion_id=org_id,
            accion=accion,
            entidad=entidad,
            entidad_id=entidad_id,
            descripcion=descripcion,
        )
        db.session.add(entrada)
    except Exception:
        pass  # El logging nunca debe romper el flujo principal

# ==============================================================================
# SISTEMA DE NOTIFICACIONES WHATSAPP (Meta Cloud API)
# ==============================================================================

def _send_whatsapp_message(to_number, body):
    """Envía un mensaje de texto vía Meta WhatsApp Cloud API."""
    token    = os.environ.get('WHATSAPP_TOKEN')
    phone_id = os.environ.get('WHATSAPP_PHONE_NUMBER_ID')
    if not token or not phone_id:
        return False
    numero = to_number.replace('+', '').replace(' ', '').replace('-', '')
    url     = f"https://graph.facebook.com/v19.0/{phone_id}/messages"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    payload = {
        "messaging_product": "whatsapp",
        "to": numero,
        "type": "text",
        "text": {"body": body, "preview_url": False}
    }
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=10)
        return resp.status_code == 200
    except Exception as e:
        print(f"[WhatsApp] Error al enviar: {e}")
        return False


def _webpush_http_status(ex):
    """Extrae el HTTP status de un WebPushException (ex.response puede ser None en pywebpush)."""
    if ex.response is not None:
        return ex.response.status_code
    import re
    m = re.search(r'(\d{3})', str(ex))
    return int(m.group(1)) if m else None

# Códigos HTTP del push service que indican suscripción inválida/caducada → borrar
_PUSH_STALE_CODES = {400, 403, 404, 410}


def enviar_push_notificacion(org_id, titulo, cuerpo, url='/dashboard'):
    """Envía una Web Push Notification a todos los suscriptores activos de la organización."""
    vapid_private = os.environ.get('VAPID_PRIVATE_KEY')
    vapid_email   = os.environ.get('VAPID_CLAIMS_EMAIL', 'notifications@inventario.app')
    if not vapid_private:
        app.logger.debug('[Push] VAPID_PRIVATE_KEY no configurada — push omitido')
        return
    try:
        from pywebpush import webpush, WebPushException
        subs = PushSubscription.query.filter_by(organizacion_id=org_id).all()
        if not subs:
            return
        payload = json.dumps({'title': titulo, 'body': cuerpo, 'url': url,
                              'icon': '/static/icons/icon-192.png'})
        to_delete = []
        for sub in subs:
            try:
                webpush(
                    subscription_info=json.loads(sub.subscription_json),
                    data=payload,
                    vapid_private_key=vapid_private,
                    vapid_claims={"sub": f"mailto:{vapid_email}"}
                )
            except WebPushException as ex:
                code = _webpush_http_status(ex)
                app.logger.warning(f'[Push] WebPushException HTTP {code} sub_id={sub.id}: {ex}')
                if code in _PUSH_STALE_CODES:
                    to_delete.append(sub)
        for sub in to_delete:
            db.session.delete(sub)
        if to_delete:
            db.session.commit()
    except ImportError:
        app.logger.error('[Push] pywebpush no instalado')
    except Exception as e:
        app.logger.error(f'[Push] Error inesperado: {e}')


def check_and_alert_stock_bajo(org_id, almacen_id):
    """
    Verifica si hay productos bajo mínimo en el almacén dado y,
    si la organización tiene número de WhatsApp configurado, envía una alerta.
    """
    try:
        org     = Organizacion.query.get(org_id)
        almacen = Almacen.query.get(almacen_id)
        if not org or not almacen or not org.whatsapp_notify:
            return

        items_bajo = Stock.query.filter(
            Stock.almacen_id == almacen_id,
            Stock.stock_minimo != None,
            Stock.stock_minimo > 0,
            Stock.cantidad < Stock.stock_minimo
        ).all()

        if not items_bajo:
            return

        lineas = [f"⚠️ *ALERTA DE STOCK BAJO*\n",
                  f"🏢 *{org.nombre}*",
                  f"🏪 Almacén: {almacen.nombre}\n"]

        for item in items_bajo[:10]:
            lineas.append(
                f"• *{item.producto.nombre}*  "
                f"Stock: {item.cantidad} / Mín: {item.stock_minimo}"
            )

        if len(items_bajo) > 10:
            lineas.append(f"\n...y {len(items_bajo) - 10} productos más.")

        lineas.append(f"\n_{now_mx().strftime('%d/%m/%Y %H:%M')}_")
        _send_whatsapp_message(org.whatsapp_notify, "\n".join(lineas))

        # Push notification (independiente del WhatsApp)
        nombres = [i.producto.nombre for i in items_bajo[:3]]
        extra = f' y {len(items_bajo)-3} más' if len(items_bajo) > 3 else ''
        enviar_push_notificacion(
            org_id=org_id,
            titulo=f'⚠️ Stock bajo — {almacen.nombre}',
            cuerpo=', '.join(nombres) + extra,
            url='/dashboard'
        )

    except Exception as e:
        print(f"[WhatsApp] Error en check_and_alert_stock_bajo: {e}")


# 2. Modifica la función original de generación de token
def send_reset_email(user):
    app_actual = current_app._get_current_object()
    s = URLSafeTimedSerializer(app_actual.config['SECRET_KEY'])
    token = s.dumps(user.email, salt='password-reset-salt')
    
    reset_url = url_for('reset_password', token=token, _external=True)
    
    # Lanzar el hilo (Thread) para que la página cargue rápido y el correo se envíe de fondo
    Thread(target=enviar_correo_api, args=(user.email, reset_url)).start()
      
# =============================================
# NUEVAS RUTAS PARA ETIQUETAS
# =============================================

@app.route('/configuracion/etiquetas', methods=['GET', 'POST'])
@login_required
@check_permission('perm_view_management')
def configurar_etiqueta_diseno():
    org = Organizacion.query.get_or_404(current_user.organizacion_id)
    if request.method == 'POST':
        fuentes_validas = {'Inter','Roboto','Montserrat','Poppins','Oswald','CenturyGothic'}
        estilos_validos = {'moderno','bold','minimalista','dark','color'}
        f = request.form.get('fuente', 'Inter')
        e = request.form.get('estilo', 'moderno')
        org.etiqueta_fuente       = f if f in fuentes_validas else 'Inter'
        org.etiqueta_color_fondo  = request.form.get('color_fondo', '#FFFFFF')[:7]
        org.etiqueta_color_texto  = request.form.get('color_texto', '#1a1a1a')[:7]
        org.etiqueta_color_sku    = request.form.get('color_sku',   '#1f4e79')[:7]
        org.etiqueta_estilo       = e if e in estilos_validos else 'moderno'
        org.etiqueta_mostrar_logo = 'mostrar_logo' in request.form
        db.session.commit()
        flash('Diseño de etiquetas guardado.', 'success')
        return redirect(url_for('configurar_etiqueta_diseno'))
    return render_template('etiqueta_personalizar.html', org=org)


@app.route('/configuracion/excel', methods=['GET', 'POST'])
@login_required
@check_permission('perm_view_management')
def configurar_excel_diseno():
    org = Organizacion.query.get_or_404(current_user.organizacion_id)
    if request.method == 'POST':
        fuentes_validas = {'Calibri', 'Arial', 'Trebuchet MS', 'Times New Roman'}
        f = request.form.get('excel_fuente', 'Calibri')
        org.excel_fuente         = f if f in fuentes_validas else 'Calibri'
        org.excel_color_header   = request.form.get('excel_color_header', '#1f4e79')[:7]
        org.excel_color_accent   = request.form.get('excel_color_accent', '#dbeafe')[:7]
        org.excel_mostrar_logo   = 'excel_mostrar_logo'   in request.form
        org.excel_mostrar_id     = 'excel_mostrar_id'     in request.form
        org.excel_mostrar_oc     = 'excel_mostrar_oc'     in request.form
        org.excel_mostrar_origen = 'excel_mostrar_origen' in request.form
        db.session.commit()
        flash('Diseño de Excel guardado. ✓', 'success')
        return redirect(url_for('configurar_excel_diseno'))
    return render_template('excel_config.html', org=org)


@app.route('/producto/<int:id>/etiqueta/configurar')
@login_required
@check_permission('perm_view_dashboard')
def configurar_etiqueta(id):
    """ Muestra el formulario para elegir tamaño de etiqueta. """
    producto = get_item_or_404(Producto, id)
    
    # Atrapamos de qué almacén viene el usuario (por la URL)
    almacen_seleccionado = request.args.get('almacen_id', type=int)
    
    # La ubicación específica ahora se maneja dinámicamente en la plantilla HTML
    # iterando sobre producto.stocks, así que ya no necesitamos buscarla aquí.
    
    org = Organizacion.query.get(current_user.organizacion_id)
    return render_template('etiqueta_config.html',
                           producto=producto,
                           almacen_seleccionado=almacen_seleccionado,
                           org=org)

@app.route('/producto/<int:id>/etiqueta/generar', methods=['POST'])
@login_required
@check_permission('perm_view_dashboard')
def generar_etiqueta_personalizada(id):
    """Genera etiqueta JPG usando la configuración de diseño de la organización."""
    producto = get_item_or_404(Producto, id)
    org = Organizacion.query.get(current_user.organizacion_id)

    almacen_id = request.form.get('almacen_id')
    ubicacion = "N/A"
    if almacen_id:
        st = Stock.query.filter_by(producto_id=id, almacen_id=almacen_id).first()
        if st and st.ubicacion:
            ubicacion = st.ubicacion

    tamano = request.form.get('tamano', '1x3')
    DPI = 300

    # ── Tamaño del lienzo ────────────────────────────────────────────────────
    if tamano == '1.75x4':
        width_px, height_px = int(4 * DPI), int(1.75 * DPI)
        fs_nombre, fs_codigo, fs_ubic = 75, 95, 45
        qr_box, margin, gap = 13, 30, 30
    else:
        width_px, height_px = int(3 * DPI), int(1 * DPI)
        fs_nombre, fs_codigo, fs_ubic = 50, 65, 35
        qr_box, margin, gap = 8, 20, 20

    # ── Config de la org (con fallbacks) ─────────────────────────────────────
    fuente       = getattr(org, 'etiqueta_fuente',       None) or 'Inter'
    color_fondo  = getattr(org, 'etiqueta_color_fondo',  None) or '#FFFFFF'
    color_texto  = getattr(org, 'etiqueta_color_texto',  None) or '#1a1a1a'
    color_sku    = getattr(org, 'etiqueta_color_sku',    None) or '#1f4e79'
    estilo       = getattr(org, 'etiqueta_estilo',       None) or 'moderno'
    mostrar_logo = getattr(org, 'etiqueta_mostrar_logo', True)

    # Estilo modifica tamaños de fuente
    if estilo == 'bold':
        fs_nombre = int(fs_nombre * 1.18)
        fs_codigo = int(fs_codigo * 1.18)
    elif estilo == 'compacto':
        fs_nombre = int(fs_nombre * 0.82)
        fs_codigo = int(fs_codigo * 0.82)
        fs_ubic   = int(fs_ubic   * 0.82)

    def hex2rgb(h):
        h = h.lstrip('#')
        try:
            return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            return (255, 255, 255)

    # ── Carga de fuentes ──────────────────────────────────────────────────────
    FONT_MAP = {
        'Inter':         ('Inter-Regular.ttf',      'Inter-Bold.ttf'),
        'Roboto':        ('Roboto-Regular.ttf',      'Roboto-Bold.ttf'),
        'Montserrat':    ('Montserrat-Regular.ttf',  'Montserrat-Bold.ttf'),
        'Poppins':       ('Poppins-Regular.ttf',     'Poppins-Bold.ttf'),
        'Oswald':        ('Oswald-Regular.ttf',      'Oswald-Bold.ttf'),
        'CenturyGothic': ('CenturyGothic.ttf',       'CenturyGothic-Bold.ttf'),
    }
    reg_file, bold_file = FONT_MAP.get(fuente, ('Inter-Regular.ttf', 'Inter-Bold.ttf'))
    fonts_dir = os.path.join(app.root_path, 'static', 'fonts')

    def _font(filename, size):
        path = os.path.join(fonts_dir, filename)
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                pass
        for fb in ('arial.ttf', 'ArialMT.ttf', 'DejaVuSans.ttf'):
            try:
                return ImageFont.truetype(fb, size)
            except Exception:
                pass
        return ImageFont.load_default()

    fnt_nombre = _font(reg_file,  fs_nombre)
    fnt_codigo = _font(bold_file, fs_codigo)
    fnt_ubic   = _font(reg_file,  fs_ubic)

    # ── Crear lienzo ──────────────────────────────────────────────────────────
    img = Image.new('RGB', (width_px, height_px), color=hex2rgb(color_fondo))
    d = ImageDraw.Draw(img)

    # Borde sutil para estilo minimalista
    if estilo == 'minimalista':
        d.rectangle([(3, 3), (width_px - 4, height_px - 4)],
                    outline=hex2rgb('#cbd5e1'), width=3)

    # ── QR (abajo-derecha) ────────────────────────────────────────────────────
    is_dark = hex2rgb(color_fondo)[0] < 128
    qr_fg = hex2rgb('#FFFFFF') if is_dark else hex2rgb('#000000')
    qr_bg = hex2rgb(color_fondo)
    qr_wrapper = qrcode.QRCode(box_size=qr_box, border=0)
    qr_wrapper.add_data(producto.codigo)
    qr_wrapper.make(fit=True)
    qr_img = qr_wrapper.make_image(fill_color=qr_fg, back_color=qr_bg).convert('RGB')
    qr_w, qr_h = qr_img.size
    x_qr = int(width_px - qr_w - margin)
    y_qr = int(height_px - qr_h - margin)
    img.paste(qr_img, (x_qr, y_qr))

    # ── Texto ─────────────────────────────────────────────────────────────────
    cur_y = margin
    qr_top = y_qr

    def _max_w(cur_y, fsize):
        return (int(width_px - margin * 2) if (cur_y + fsize) < qr_top
                else int(x_qr - margin - gap))

    def _truncate(txt, fnt, max_w):
        while d.textlength(txt + '…', font=fnt) > max_w and txt:
            txt = txt[:-1]
        return txt + '…' if len(txt) < len(producto.nombre if fnt == fnt_nombre else txt) else txt

    # Nombre
    nom = producto.nombre
    mw = _max_w(cur_y, fs_nombre)
    while d.textlength(nom + '…', font=fnt_nombre) > mw and nom:
        nom = nom[:-1]
    if nom != producto.nombre:
        nom += '…'
    d.text((margin, cur_y), nom, font=fnt_nombre, fill=hex2rgb(color_texto))
    cur_y += fs_nombre + 5

    # SKU
    cod = producto.codigo
    mw2 = _max_w(cur_y, fs_codigo)
    while d.textlength(cod, font=fnt_codigo) > mw2 and cod:
        cod = cod[:-1]
    d.text((margin, cur_y), cod, font=fnt_codigo, fill=hex2rgb(color_sku))
    cur_y += fs_codigo + 5

    # Ubicación
    ubic_txt = f"UBIC: {ubicacion}" if ubicacion and ubicacion != "N/A" else f"ID: {producto.id}"
    mw3 = _max_w(cur_y, fs_ubic)
    while d.textlength(ubic_txt, font=fnt_ubic) > mw3 and ubic_txt:
        ubic_txt = ubic_txt[:-1]
    d.text((margin, cur_y), ubic_txt, font=fnt_ubic, fill=hex2rgb(color_texto))
    cur_y += fs_ubic + 10

    # ── Imagen del producto ───────────────────────────────────────────────────
    if mostrar_logo and producto.imagen_url:
        avail_h = int(height_px - cur_y - margin)
        avail_w = int(x_qr - margin - 10)
        if avail_h > 20:
            path_img = os.path.join(app.config['UPLOAD_FOLDER'], producto.imagen_url)
            if os.path.exists(path_img):
                try:
                    prod_img = Image.open(path_img)
                    prod_img.thumbnail((avail_w, avail_h))
                    img.paste(prod_img, (margin, cur_y))
                except Exception:
                    pass

    # ── Exportar ──────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    img.save(buffer, 'JPEG', quality=100)
    buffer.seek(0)
    filename = f"Etiqueta_{secure_filename(producto.nombre)}_{tamano}.jpg"
    return send_file(buffer, mimetype='image/jpeg', as_attachment=True, download_name=filename)

# --- Rutas Principales (Dashboard) ---

@app.route('/sw.js')
def service_worker():
    """Sirve el SW desde la raíz con el header que permite scope '/'."""
    from flask import send_from_directory
    resp = make_response(send_from_directory('static', 'sw.js'))
    resp.headers['Content-Type'] = 'application/javascript'
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

@app.route('/offline')
def offline_page():
    return render_template('offline.html')


@app.route('/.well-known/assetlinks.json')
def assetlinks():
    """Vincula el dominio web con la app Android (TWA / Play Store)."""
    import json as _json
    data = current_app.config.get('ASSETLINKS', [])
    return current_app.response_class(
        _json.dumps(data, indent=2),
        mimetype='application/json'
    )


@app.route('/')
@login_required
def index():
    """ Dashboard Principal (Multiusuario) con Alertas por Almacén. """
    
    alertas_agrupadas = {}
    pending_map = {}
    
    if current_user.rol == 'super_admin':
        # (El Super Admin ve todo, agrupado por organización)
        pass
    elif current_user.organizacion_id:
        org_id = current_user.organizacion_id
        
        # 1. Encontrar todos los items de stock BAJO de esta organización
        alertas_crudas = db.session.query(Stock).join(Almacen).join(Producto).filter(
            Almacen.organizacion_id == org_id,
            Stock.cantidad < Stock.stock_minimo
        ).all()

        # 2. Encontrar OCs pendientes (borrador o enviada) de esta organización
        ordenes_pendientes = db.session.query(
            OrdenCompraDetalle.producto_id, 
            OrdenCompra.id, 
            User.username,
            OrdenCompra.estado,
            OrdenCompra.almacen_id
        ).join(
            OrdenCompra, OrdenCompraDetalle.orden_id == OrdenCompra.id
        ).join(
            User, OrdenCompra.creador_id == User.id
        ).filter(
            OrdenCompra.estado.in_(['borrador', 'enviada']),
            OrdenCompra.organizacion_id == org_id
        ).all()

        # 3. Convertir OCs pendientes en un mapa de búsqueda rápida
        for prod_id, orden_id, username, estado, alm_id in ordenes_pendientes:
            pending_map[(prod_id, alm_id)] = {
                'orden_id': orden_id, 
                'username': username,
                'estado': estado
            }
            
        # 4. Agrupar las alertas por (Almacén, Proveedor)
        alertas_agrupadas = defaultdict(list)
        for item_stock in alertas_crudas:
            if (item_stock.producto_id, item_stock.almacen_id) in pending_map:
                continue
                
            if item_stock.producto.proveedor:
                key = (item_stock.almacen_id, item_stock.almacen.nombre, 
                       item_stock.producto.proveedor_id, item_stock.producto.proveedor.nombre)
                alertas_agrupadas[key].append(item_stock)
            else:
                key = (item_stock.almacen_id, item_stock.almacen.nombre, 0, "Proveedor no asignado")
                alertas_agrupadas[key].append(item_stock)

    return render_template('index.html',
                           alertas_agrupadas=alertas_agrupadas,
                           pending_map=pending_map,
                           now=now_mx())

@app.route('/dashboard')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def dashboard():
    """ 
    Página del Dashboard de Inventario (Filtros y Tabla).
    MODIFICADO para Multi-Almacén.
    """
    
    if current_user.rol == 'super_admin':
        almacenes = Almacen.query.all()
    else:
        almacenes = Almacen.query.filter_by(organizacion_id=current_user.organizacion_id).order_by(Almacen.nombre).all()

    almacen_id_solicitado = request.args.get('almacen_id', type=int)
    almacen_seleccionado = None

    if almacen_id_solicitado:
        if current_user.rol == 'super_admin':
            almacen_seleccionado = Almacen.query.get(almacen_id_solicitado)
        else:
            almacen_seleccionado = Almacen.query.filter_by(id=almacen_id_solicitado, organizacion_id=current_user.organizacion_id).first()
    
    if not almacen_seleccionado and almacenes:
        almacen_seleccionado = almacenes[0]
    
    if almacen_seleccionado:
        items_stock = db.session.query(Stock).filter_by(almacen_id=almacen_seleccionado.id).join(Producto).order_by(Producto.nombre).all()
    else:
        items_stock = []

    # Valorización: suma (cantidad * precio_unitario) por almacén
    valor_almacen = sum(
        (item.cantidad or 0) * (item.producto.precio_unitario or 0)
        for item in items_stock
    )
    items_por_valor = sorted(
        items_stock,
        key=lambda x: (x.cantidad or 0) * (x.producto.precio_unitario or 0),
        reverse=True
    )[:10]

    # Valorización total de todos los almacenes de la organización
    if current_user.rol == 'super_admin':
        valor_total_org = db.session.query(
            db.func.sum(Stock.cantidad * Producto.precio_unitario)
        ).join(Producto, Stock.producto_id == Producto.id).scalar() or 0
    else:
        valor_total_org = db.session.query(
            db.func.sum(Stock.cantidad * Producto.precio_unitario)
        ).join(Producto, Stock.producto_id == Producto.id).join(
            Almacen, Stock.almacen_id == Almacen.id
        ).filter(Almacen.organizacion_id == current_user.organizacion_id).scalar() or 0

    if current_user.rol == 'super_admin':
        categorias = Categoria.query.all()
        proveedores = Proveedor.query.all()
    else:
        org_id = current_user.organizacion_id
        categorias = Categoria.query.filter_by(organizacion_id=org_id).all()
        proveedores = Proveedor.query.filter_by(organizacion_id=org_id).all()

    # --- KPIs de Rotación (por almacén seleccionado) ---
    kpis_rotacion = {}
    if almacen_seleccionado:
        ahora = now_mx()
        hace_30d = ahora - timedelta(days=30)
        hace_60d = ahora - timedelta(days=60)

        def _sum_salidas(almacen_id, desde, hasta=None):
            q = db.session.query(
                db.func.sum(db.func.abs(Movimiento.cantidad))
            ).filter(
                Movimiento.almacen_id == almacen_id,
                Movimiento.tipo == 'salida',
                Movimiento.fecha >= desde
            )
            if hasta:
                q = q.filter(Movimiento.fecha < hasta)
            return q.scalar() or 0

        salidas_30d = _sum_salidas(almacen_seleccionado.id, hace_30d)
        salidas_prev_30d = _sum_salidas(almacen_seleccionado.id, hace_60d, hace_30d)

        stock_total_uds = sum(item.cantidad for item in items_stock)
        tasa_diaria = salidas_30d / 30 if salidas_30d > 0 else 0
        dias_stock = round(stock_total_uds / tasa_diaria) if tasa_diaria > 0 else None

        if salidas_prev_30d > 0:
            tendencia_pct = round((salidas_30d - salidas_prev_30d) / salidas_prev_30d * 100, 1)
        else:
            tendencia_pct = None

        # Top 5 productos con más salidas en los últimos 30 días
        top_movers_raw = db.session.query(
            Movimiento.producto_id,
            db.func.sum(db.func.abs(Movimiento.cantidad)).label('total_salidas')
        ).filter(
            Movimiento.almacen_id == almacen_seleccionado.id,
            Movimiento.tipo == 'salida',
            Movimiento.fecha >= hace_30d
        ).group_by(Movimiento.producto_id
        ).order_by(db.func.sum(db.func.abs(Movimiento.cantidad)).desc()).limit(5).all()

        prod_map = {item.producto_id: item for item in items_stock}
        top_movers = []
        for row in top_movers_raw:
            stock_item = prod_map.get(row.producto_id)
            if stock_item:
                top_movers.append({
                    'nombre': stock_item.producto.nombre,
                    'codigo': stock_item.producto.codigo,
                    'salidas': int(row.total_salidas),
                    'stock': stock_item.cantidad,
                })

        kpis_rotacion = {
            'salidas_30d': salidas_30d,
            'salidas_prev_30d': salidas_prev_30d,
            'tendencia_pct': tendencia_pct,
            'dias_stock': dias_stock,
            'tasa_diaria': round(tasa_diaria, 1),
            'top_movers': top_movers,
        }

    return render_template('dashboard.html',
                           items_stock=items_stock,
                           almacenes=almacenes,
                           almacen_seleccionado=almacen_seleccionado,
                           categorias=categorias,
                           proveedores=proveedores,
                           valor_almacen=valor_almacen,
                           valor_total_org=valor_total_org,
                           items_por_valor=items_por_valor,
                           kpis_rotacion=kpis_rotacion)

# --- Rutas de Productos ---

@app.route('/api/alertas/stock-bajo')
@login_required
def api_alertas_stock_bajo():
    org_id = current_user.organizacion_id
    items = db.session.query(Stock).join(
        Almacen, Stock.almacen_id == Almacen.id
    ).join(Producto, Stock.producto_id == Producto.id).filter(
        Almacen.organizacion_id == org_id,
        Stock.stock_minimo > 0,
        Stock.cantidad < Stock.stock_minimo
    ).order_by(Stock.cantidad.asc()).limit(10).all()

    return jsonify({
        'count': len(items),
        'items': [{
            'nombre':     item.producto.nombre,
            'sku':        item.producto.codigo,
            'cantidad':   item.cantidad,
            'minimo':     item.stock_minimo,
            'almacen':    item.almacen.nombre,
            'producto_id': item.producto_id,
        } for item in items]
    })


@app.route('/api/productos/buscar')
@login_required
def api_buscar_productos():
    """
    API para buscar productos por Nombre o SKU dinámicamente.
    Retorna JSON para ser consumido por JavaScript.
    """
    query = request.args.get('q', '').strip()
    
    if not query:
        return jsonify([])

    # Buscamos coincidencias en Nombre O Código (SKU)
    # Usamos ilike para que no importen mayúsculas/minúsculas
    productos = Producto.query.filter(
        (Producto.nombre.ilike(f'%{query}%')) | 
        (Producto.codigo.ilike(f'%{query}%'))
    ).filter_by(organizacion_id=current_user.organizacion_id).limit(10).all()

    resultados = []
    for p in productos:
        resultados.append({
            'id': p.id,
            'texto_mostrar': f"{p.nombre} (SKU: {p.codigo})", # Lo que se ve en la lista
            'nombre': p.nombre,
            'codigo': p.codigo,
            'precio': p.precio_unitario
        })
    
    return jsonify(resultados)


@app.route('/api/stock/buscar')
@login_required
def api_stock_buscar():
    """Busca ítems de stock por nombre o SKU, devuelve contexto de almacén."""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    org_id = current_user.organizacion_id
    items = (
        db.session.query(Stock, Producto, Almacen)
        .join(Producto, Stock.producto_id == Producto.id)
        .join(Almacen, Stock.almacen_id == Almacen.id)
        .filter(Producto.organizacion_id == org_id)
        .filter(
            (Producto.nombre.ilike(f'%{q}%')) |
            (Producto.codigo.ilike(f'%{q}%'))
        )
        .order_by(Producto.nombre)
        .limit(8)
        .all()
    )
    return jsonify([{
        'stock_id': s.id,
        'nombre':   p.nombre,
        'codigo':   p.codigo,
        'almacen':  a.nombre,
        'cantidad': s.cantidad
    } for s, p, a in items])


@app.route('/api/ajuste/rapido', methods=['POST'])
@login_required
@check_permission('perm_edit_management')
def api_ajuste_rapido():
    """Aplica un ajuste rápido (+/-) a un ítem de stock via AJAX."""
    data     = request.get_json(silent=True) or {}
    stock_id = data.get('stock_id')
    tipo     = data.get('tipo', 'entrada')
    motivo   = (data.get('motivo') or '').strip()
    try:
        cantidad = int(data.get('cantidad', 0))
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Cantidad inválida'}), 400

    if not stock_id:
        return jsonify({'ok': False, 'error': 'stock_id requerido'}), 400
    if cantidad < 1:
        return jsonify({'ok': False, 'error': 'La cantidad debe ser ≥ 1'}), 400
    if not motivo:
        return jsonify({'ok': False, 'error': 'El motivo es obligatorio'}), 400
    if tipo not in ('entrada', 'salida'):
        return jsonify({'ok': False, 'error': 'Tipo inválido'}), 400

    org_id = current_user.organizacion_id
    stock  = Stock.query.get(stock_id)
    if not stock:
        return jsonify({'ok': False, 'error': 'Stock no encontrado'}), 404

    producto = Producto.query.get(stock.producto_id)
    if not producto or producto.organizacion_id != org_id:
        return jsonify({'ok': False, 'error': 'Sin acceso'}), 403

    delta         = cantidad if tipo == 'entrada' else -cantidad
    nueva_cantidad = stock.cantidad + delta
    if nueva_cantidad < 0:
        return jsonify({'ok': False, 'error': f'Stock insuficiente (actual: {stock.cantidad})'}), 400

    stock.cantidad = nueva_cantidad
    tipo_mov = 'ajuste-entrada' if tipo == 'entrada' else 'ajuste-salida'
    signo    = '+' if tipo == 'entrada' else '-'

    db.session.add(Movimiento(
        producto_id=stock.producto_id,
        cantidad=delta,
        tipo=tipo_mov,
        fecha=now_mx(),
        motivo=f'Ajuste Rápido: {motivo}',
        almacen_id=stock.almacen_id,
        organizacion_id=org_id
    ))
    log_actividad('ajuste', 'producto',
                  f'Ajuste rápido {signo}{cantidad} uds — {motivo}',
                  entidad_id=stock.producto_id)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500

    if tipo == 'salida':
        check_and_alert_stock_bajo(org_id, stock.almacen_id)

    return jsonify({
        'ok': True,
        'mensaje': f'Ajuste {signo}{cantidad} uds aplicado. Nuevo stock: {nueva_cantidad}.'
    })


@app.route('/productos/importar/template')
@login_required
@check_permission('perm_edit_management')
def descargar_template_importacion():
    """Descarga un archivo Excel de ejemplo para importación masiva de productos."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ["nombre*", "codigo_sku*", "precio_unitario", "categoria", "proveedor", "unidades_por_caja"]
    header_fill   = PatternFill("solid", fgColor="4F46E5")
    header_font   = Font(bold=True, color="FFFFFF")
    example_fill  = PatternFill("solid", fgColor="F0F0FF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_fill and header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    examples = [
        ["Clavos 2 pulgadas", "CLV-2IN-100", 45.50, "Ferretería", "Proveedor Central", 100],
        ["Pintura Blanca 1L",  "PIN-BL-001",  89.00, "Pinturas",   "Distribuidora ABC", 12],
        ["Lija Grano 120",     "LIJ-120",     12.00, "",           "",                   50],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = example_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    ws_info = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("INSTRUCCIONES DE IMPORTACIÓN", True),
        ("", False),
        ("Columnas obligatorias (marcadas con *):", True),
        ("  • nombre*        → Nombre del producto", False),
        ("  • codigo_sku*    → Código único (SKU). Si ya existe se omite.", False),
        ("", False),
        ("Columnas opcionales:", True),
        ("  • precio_unitario → Número decimal. Default: 0", False),
        ("  • categoria       → Nombre exacto. Si no existe se crea automáticamente.", False),
        ("  • proveedor       → Nombre exacto. Si no existe se crea automáticamente.", False),
        ("  • unidades_por_caja → Número entero. Default: 1", False),
        ("", False),
        ("NOTAS:", True),
        ("  • Elimina las filas de ejemplo antes de importar.", False),
        ("  • No modifiques los nombres de las columnas.", False),
        ("  • Puedes importar .xlsx o .csv", False),
    ]
    for row_idx, (text, bold) in enumerate(instrucciones, 1):
        cell = ws_info.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold)
    ws_info.column_dimensions["A"].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="template_importacion_productos.xlsx",
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/productos/importar', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def importar_productos():
    """Importación masiva de productos desde CSV o Excel."""
    import io, csv
    from openpyxl import load_workbook

    org_id = current_user.organizacion_id
    resultados = None

    if request.method == 'POST':
        archivo = request.files.get('archivo')
        if not archivo or archivo.filename == '':
            flash('Selecciona un archivo para importar.', 'danger')
            return redirect(url_for('importar_productos'))

        ext = archivo.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('csv', 'xlsx'):
            flash('Solo se aceptan archivos .csv o .xlsx', 'danger')
            return redirect(url_for('importar_productos'))

        try:
            # --- Leer filas ---
            filas = []
            if ext == 'xlsx':
                wb   = load_workbook(io.BytesIO(archivo.read()), data_only=True)
                ws   = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    flash('El archivo está vacío.', 'danger')
                    return redirect(url_for('importar_productos'))
                headers = [str(h).strip().lower() if h else '' for h in rows[0]]
                filas   = [dict(zip(headers, row)) for row in rows[1:]]
            else:
                content = archivo.read().decode('utf-8-sig')
                reader  = csv.DictReader(io.StringIO(content))
                headers = [h.strip().lower() for h in (reader.fieldnames or [])]
                filas   = [{k.strip().lower(): v for k, v in row.items()} for row in reader]

            col = lambda row, *keys: next((str(row.get(k) or '').strip() for k in keys if row.get(k) not in (None, '')), '')

            importados, omitidos, errores = 0, 0, []

            for idx, fila in enumerate(filas, 2):
                nombre = col(fila, 'nombre', 'name')
                codigo = col(fila, 'codigo_sku', 'codigo', 'sku', 'code')

                if not nombre or not codigo:
                    if any(v for v in fila.values() if str(v or '').strip()):
                        errores.append(f"Fila {idx}: 'nombre' y 'codigo_sku' son obligatorios.")
                    continue

                if Producto.query.filter_by(codigo=codigo, organizacion_id=org_id).first():
                    omitidos += 1
                    continue

                # Precio
                try:
                    precio = float(col(fila, 'precio_unitario', 'precio') or 0)
                except ValueError:
                    precio = 0.0

                # Unidades por caja
                try:
                    upc = int(col(fila, 'unidades_por_caja', 'unidades') or 1)
                except ValueError:
                    upc = 1

                # Categoría (crea si no existe)
                cat_nombre = col(fila, 'categoria', 'category')
                categoria  = None
                if cat_nombre:
                    categoria = Categoria.query.filter_by(nombre=cat_nombre, organizacion_id=org_id).first()
                    if not categoria:
                        categoria = Categoria(nombre=cat_nombre, organizacion_id=org_id)
                        db.session.add(categoria)
                        db.session.flush()

                # Proveedor (crea si no existe)
                prov_nombre = col(fila, 'proveedor', 'supplier', 'proveedor_nombre')
                proveedor   = None
                if prov_nombre:
                    proveedor = Proveedor.query.filter_by(nombre=prov_nombre, organizacion_id=org_id).first()
                    if not proveedor:
                        proveedor = Proveedor(nombre=prov_nombre, organizacion_id=org_id)
                        db.session.add(proveedor)
                        db.session.flush()

                producto = Producto(
                    nombre         = nombre,
                    codigo         = codigo,
                    precio_unitario= precio,
                    categoria_id   = categoria.id if categoria else None,
                    proveedor_id   = proveedor.id if proveedor else None,
                    unidades_por_caja = upc,
                    organizacion_id= org_id
                )
                db.session.add(producto)
                importados += 1

            if importados > 0:
                log_actividad('importar', 'producto', f'Importación masiva: {importados} producto(s) creados, {omitidos} omitidos (SKU duplicado)')
            db.session.commit()
            resultados = {'importados': importados, 'omitidos': omitidos, 'errores': errores}

        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el archivo: {e}', 'danger')

    return render_template('importar_productos.html', titulo='Importar Productos', resultados=resultados)


@app.route('/producto/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission #(Si usas este decorador, mantenlo)
@check_permission('perm_edit_management') #(Si usas este decorador, mantenlo)
def nuevo_producto():
    """ 
    Crea un nuevo producto. 
    CORREGIDO: Manejo de strings vacíos en la conversión a float.
    """
    org_id = current_user.organizacion_id
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).all()
    categorias = Categoria.query.filter_by(organizacion_id=org_id).all()
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).all() 
    
    if request.method == 'POST':
        imagen_filename = None
            
        def repoblar_formulario_con_error():
            # El uso de 'or 0.0' previene el error si el campo llega vacío
            costo_val = request.form.get('costo_estandar')
            precio_unitario_float = float(costo_val) if costo_val and costo_val.strip() else 0.0

            producto_temporal = Producto(
                nombre=request.form.get('nombre'),
                codigo=request.form.get('codigo'),
                categoria_id=int(request.form.get('categoria_id') or 0) or None,
                precio_unitario=precio_unitario_float, 
                proveedor_id=int(request.form.get('proveedor_id') or 0) or None,
                unidades_por_caja=int(request.form.get('unidades_por_caja') or 1),
                organizacion_id=org_id
            )
            producto_temporal.costo_estandar = producto_temporal.precio_unitario
            
            return render_template('producto_form.html', 
                                   titulo="Nuevo Producto", 
                                   proveedores=proveedores,
                                   categorias=categorias,
                                   almacenes=almacenes, 
                                   producto=producto_temporal)
            
        if 'imagen' in request.files:
            file = request.files['imagen']
            if file.filename != '' and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                imagen_filename = filename
            elif file.filename != '' and not allowed_file(file.filename):
                flash('Tipo de archivo de imagen no permitido.', 'danger')
                return repoblar_formulario_con_error()

        if not imagen_filename:
            ai_fn = secure_filename(request.form.get('ai_imagen_filename', '').strip())
            if ai_fn:
                ai_path = os.path.join(app.config['UPLOAD_FOLDER'], ai_fn)
                if os.path.isfile(ai_path):
                    imagen_filename = ai_fn

        try:
            # CORRECCIÓN VITAL: Validar string vacío antes de float()
            costo_raw = request.form.get('costo_estandar')
            precio_final = float(costo_raw) if costo_raw and costo_raw.strip() else 0.0

            nuevo_prod = Producto(
                nombre=request.form['nombre'],
                codigo=request.form['codigo'],
                categoria_id=request.form.get('categoria_id') or None,
                precio_unitario=precio_final,
                imagen_url=imagen_filename,
                proveedor_id=request.form.get('proveedor_id') or None,
                unidades_por_caja=int(request.form.get('unidades_por_caja', 1)), 
                organizacion_id=current_user.organizacion_id,
                enlace_proveedor=request.form.get('enlace_proveedor')
            )
            db.session.add(nuevo_prod)
            db.session.flush()

            cantidad_inicial = int(request.form.get('cantidad_inicial') or 0)
            almacen_inicial_id = int(request.form.get('almacen_inicial_id') or 0)
            ubicacion_inicial = request.form.get('ubicacion_inicial')

            almacen_seleccionado = None
            if almacen_inicial_id > 0:
                almacen_seleccionado = Almacen.query.filter_by(id=almacen_inicial_id, organizacion_id=org_id).first()

            if almacen_seleccionado:
                nuevo_stock = Stock(
                    producto_id=nuevo_prod.id,
                    almacen_id=almacen_seleccionado.id,
                    cantidad=cantidad_inicial, 
                    stock_minimo=int(request.form.get('stock_minimo') or 5),
                    stock_maximo=int(request.form.get('stock_maximo') or 100),
                    ubicacion=ubicacion_inicial
                )
                db.session.add(nuevo_stock)

                if cantidad_inicial > 0:
                    movimiento_inicial = Movimiento(
                        producto_id=nuevo_prod.id,
                        cantidad=cantidad_inicial,
                        tipo='entrada-inicial',
                        fecha=now_mx(),
                        motivo='Stock Inicial (Creación)',
                        almacen_id=almacen_inicial_id,
                        organizacion_id=org_id
                    )
                    db.session.add(movimiento_inicial)
                
            log_actividad('crear', 'producto', f'Producto creado: {nuevo_prod.nombre} (SKU: {nuevo_prod.codigo})', entidad_id=nuevo_prod.id)
            db.session.commit()
            flash('Producto creado exitosamente.', 'success')

            if almacen_seleccionado:
                 return redirect(url_for('gestionar_inventario_almacen', id=almacen_seleccionado.id))
            return redirect(url_for('dashboard'))
        
        except IntegrityError as e:
            db.session.rollback()
            if "producto_codigo_key" in str(e) or "UNIQUE constraint failed" in str(e):
                flash('Error: El Código (SKU) ya existe.', 'danger')
            else:
                _flash_err('Error de base de datos al guardar el producto.', e)
            return repoblar_formulario_con_error()
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error inesperado: {e}', 'danger')
            return repoblar_formulario_con_error()
            
    return render_template('producto_form.html', 
                           titulo="Nuevo Producto", 
                           proveedores=proveedores,
                           categorias=categorias,
                           almacenes=almacenes,
                           producto=None)
 
@app.route('/producto/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@check_permission('perm_edit_management')
def editar_producto(id):
    """ 
    Edita un producto.
    CORREGIDO: Previene error 'could not convert string to float' en campos vacíos.
    """
    producto = Producto.query.get_or_404(id)
    org_id = producto.organizacion_id
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).all()
    categorias = Categoria.query.filter_by(organizacion_id=org_id).all()

    almacen_id = request.args.get('almacen_id', type=int)
    stock_item = None
    if almacen_id:
        stock_item = Stock.query.filter_by(producto_id=producto.id, almacen_id=almacen_id).first()

    if request.method == 'POST':
        try:
            if 'imagen' in request.files:
                file = request.files['imagen']
                if file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    producto.imagen_url = filename

            if not producto.imagen_url or request.files.get('imagen', '').filename == '':
                ai_fn = secure_filename(request.form.get('ai_imagen_filename', '').strip())
                if ai_fn:
                    ai_path = os.path.join(app.config['UPLOAD_FOLDER'], ai_fn)
                    if os.path.isfile(ai_path):
                        producto.imagen_url = ai_fn

            producto.nombre = request.form['nombre']
            producto.codigo = request.form['codigo']
            producto.categoria_id = request.form.get('categoria_id') or None
            producto.proveedor_id = request.form.get('proveedor_id') or None
            producto.unidades_por_caja = int(request.form.get('unidades_por_caja') or 1)
            producto.enlace_proveedor = request.form.get('enlace_proveedor')
            
            # CORRECCIÓN AQUÍ: Manejar el valor vacío de costo_estandar
            costo_raw = request.form.get('costo_estandar')
            producto.precio_unitario = float(costo_raw) if costo_raw and costo_raw.strip() else 0.0

            if stock_item:
                stock_item.stock_minimo = int(request.form.get('stock_minimo') or 0)
                stock_item.stock_maximo = int(request.form.get('stock_maximo') or 0)
                stock_item.cantidad = int(request.form.get('cantidad') or 0)
                stock_item.ubicacion = request.form.get('ubicacion')

            log_actividad('editar', 'producto', f'Producto editado: {producto.nombre} (SKU: {producto.codigo})', entidad_id=producto.id)
            db.session.commit()
            flash('Producto actualizado exitosamente', 'success')
            return redirect(url_for('gestionar_inventario_almacen', id=almacen_id) if almacen_id else url_for('dashboard'))

        except Exception as e:
            db.session.rollback()
            _flash_err('Error al actualizar producto. Intenta de nuevo.', e)

    # Preparar visualización del costo para el template
    producto.costo_estandar = producto.precio_unitario
    return render_template('producto_form.html', 
                           titulo="Editar Producto", 
                           producto=producto, 
                           proveedores=proveedores, 
                           categorias=categorias, 
                           stock_item=stock_item)
  
@app.route('/producto/<int:id>/etiqueta')
@login_required
@check_permission('perm_view_dashboard')
def generar_etiqueta(id):
    producto = get_item_or_404(Producto, id)
    try:
        buffer = io.BytesIO()
        label_width = 4 * inch
        label_height = 2.5 * inch
        c = canvas.Canvas(buffer, pagesize=(label_width, label_height))
        qr_img = qrcode.make(producto.codigo)
        qr_img_path = io.BytesIO()
        qr_img.save(qr_img_path, format='PNG')
        qr_img_path.seek(0)
        qr_para_pdf = ImageReader(qr_img_path)
        c.drawImage(qr_para_pdf, label_width - (1.6 * inch), 0.6 * inch,
                    width=(1.4 * inch), height=(1.4 * inch), preserveAspectRatio=True)
        text_x = 0.25 * inch
        text_y = label_height - (0.5 * inch)
        c.setFont('Helvetica-Bold', 12)
        c.drawString(text_x, text_y, producto.nombre[:25])
        c.setFont('Helvetica', 10)
        c.drawString(text_x, text_y - (0.3 * inch), f"SKU: {producto.codigo}")
        c.setFont('Helvetica', 10)
        c.drawString(text_x, text_y - (0.6 * inch), f"Precio: ${producto.precio_unitario:.2f}")
        if producto.imagen_url:
            img_path = os.path.join(app.config['UPLOAD_FOLDER'], producto.imagen_url)
            if os.path.exists(img_path):
                try:
                    prod_img = ImageReader(img_path)
                    c.drawImage(prod_img, 0.1 * inch, 0.2 * inch,
                                width=1.5 * inch, height=1.0 * inch,
                                preserveAspectRatio=True)
                except Exception as img_err:
                    print(f"Error al dibujar imagen en PDF: {img_err}")
        c.showPage()
        c.save()
        buffer.seek(0)
        nombre_base = secure_filename(producto.nombre) 
        fecha_str = now_mx().strftime("%Y-%m-%d")
        nombre_archivo = f"{nombre_base}_{fecha_str}.pdf"
        return send_file(
            buffer,
            as_attachment=False, 
            download_name=nombre_archivo,
            mimetype='application/pdf'
        )
    except Exception as e:
        flash(f'Error al generar etiqueta: {e}', 'danger')
        return redirect(url_for('index'))

@app.route('/producto/<int:id>/historial')
@login_required
@check_permission('perm_view_dashboard')
def historial_producto(id):
    """ Muestra el Kardex (Movimientos) Y el Stock Actual por Almacén. """
    producto = get_item_or_404(Producto, id)
    
    # --- 1. OBTENER STOCK ACTUAL ---
    # Buscamos dónde está este producto (en qué almacenes y qué cantidad)
    stocks_actuales = Stock.query.filter_by(producto_id=id).join(Almacen).order_by(Almacen.nombre).all()
    
    # Calcular el total global para mostrarlo en grande
    total_global = sum(s.cantidad for s in stocks_actuales)
    
    # --- 2. OBTENER MOVIMIENTOS ---
    movimientos_query = Movimiento.query.filter_by(producto_id=id).order_by(Movimiento.fecha.desc())
    
    if current_user.rol != 'super_admin':
        movimientos_query = movimientos_query.filter(Movimiento.organizacion_id == current_user.organizacion_id)
        
    movimientos = movimientos_query.all()
    
    # Agrupar movimientos por almacén para la vista
    historial_por_almacen = defaultdict(list)
    for m in movimientos:
        # Usamos una consulta segura por si el almacén fue borrado (aunque no debería pasar con cascade)
        alm_nombre = Almacen.query.get(m.almacen_id).nombre if m.almacen_id else "Sin Almacén"
        historial_por_almacen[alm_nombre].append(m)
    
    return render_template('historial_producto.html', 
                           producto=producto, 
                           historial_por_almacen=historial_por_almacen,
                           stocks_actuales=stocks_actuales, # <-- DATO NUEVO
                           total_global=total_global)       # <-- DATO NUEVO

# --- Rutas de Categorías ---

@app.route('/categorias')
@login_required
@check_org_permission
@check_permission('perm_view_management')
def lista_categorias():
    if current_user.rol == 'super_admin':
        categorias = Categoria.query.all()
    else:
        categorias = Categoria.query.filter_by(organizacion_id=current_user.organizacion_id).all()
    return render_template('categorias.html', categorias=categorias)

@app.route('/categoria/nueva', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def nueva_categoria():
    if request.method == 'POST':
        try:
            nueva_cat = Categoria(
                nombre=request.form['nombre'],
                descripcion=request.form.get('descripcion'),
                organizacion_id=current_user.organizacion_id
            )
            db.session.add(nueva_cat)
            db.session.commit()
            flash('Categoría creada exitosamente', 'success')
            return redirect(url_for('lista_categorias'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear la categoría (quizás el nombre ya existe): {e}', 'danger')
            
    return render_template('categoria_form.html', titulo="Nueva Categoría", categoria=None)

@app.route('/categoria/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@check_permission('perm_edit_management')
def editar_categoria(id):
    categoria = get_item_or_404(Categoria, id)
    
    if request.method == 'POST':
        try:
            categoria.nombre = request.form['nombre']
            categoria.descripcion = request.form.get('descripcion')
            db.session.commit()
            flash('Categoría actualizada exitosamente', 'success')
            return redirect(url_for('lista_categorias'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la categoría: {e}', 'danger')

    return render_template('categoria_form.html', 
                           titulo="Editar Categoría", 
                           categoria=categoria)

@app.route('/categoria/eliminar/<int:id>', methods=['POST'])
@login_required
@check_permission('perm_edit_management')
def eliminar_categoria(id):
    categoria_a_eliminar = get_item_or_404(Categoria, id)
    
    try:
        org_id = categoria_a_eliminar.organizacion_id
        productos_afectados = Producto.query.filter_by(categoria_id=categoria_a_eliminar.id, organizacion_id=org_id).all()
        
        for producto in productos_afectados:
            producto.categoria_id = None
        
        db.session.delete(categoria_a_eliminar)
        db.session.commit()
        
        flash(f'Categoría "{categoria_a_eliminar.nombre}" eliminada. Los productos asociados fueron des-asignados.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la categoría: {e}', 'danger')

    return redirect(url_for('lista_categorias'))

# --- Rutas de Proveedores ---

@app.route('/proveedores')
@login_required
@check_org_permission
@check_permission('perm_view_management')
def lista_proveedores():
    if current_user.rol == 'super_admin':
        proveedores = Proveedor.query.all()
    else:
        proveedores = Proveedor.query.filter_by(organizacion_id=current_user.organizacion_id).all()
    return render_template('proveedores.html', proveedores=proveedores)

@app.route('/proveedor/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def nuevo_proveedor():
    if request.method == 'POST':
        try:
            nuevo_prov = Proveedor(
                nombre=request.form['nombre'],
                contacto_email=request.form.get('contacto_email'),
                contacto_telefono=request.form.get('contacto_telefono'),
                organizacion_id=current_user.organizacion_id
            )
            db.session.add(nuevo_prov)
            db.session.commit()
            flash('Proveedor creado exitosamente', 'success')
            return redirect(url_for('lista_proveedores'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear proveedor: {e}', 'danger')
            
    return render_template('proveedor_form.html', titulo="Nuevo Proveedor", proveedor=None)

@app.route('/proveedor/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@check_permission('perm_edit_management')
def editar_proveedor(id):
    proveedor = get_item_or_404(Proveedor, id)
    
    if request.method == 'POST':
        try:
            proveedor.nombre = request.form['nombre']
            proveedor.contacto_email = request.form.get('contacto_email')
            proveedor.contacto_telefono = request.form.get('contacto_telefono')
            
            db.session.commit()
            flash('Proveedor actualizado exitosamente', 'success')
            return redirect(url_for('lista_proveedores'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el proveedor: {e}', 'danger')

    return render_template('proveedor_form.html', 
                           titulo="Editar Proveedor", 
                           proveedor=proveedor)

# --- Rutas de Almacenes ---

@app.route('/almacenes')
@login_required
@admin_required # Solo Admins y Super Admins pueden gestionar almacenes
def lista_almacenes():
    """ Muestra la lista de almacenes de la organización. """
    if current_user.rol == 'super_admin':
        almacenes = Almacen.query.all()
    else:
        almacenes = Almacen.query.filter_by(organizacion_id=current_user.organizacion_id).all()
        
    return render_template('almacenes.html', 
                           almacenes=almacenes,
                           titulo="Gestionar Almacenes")

@app.route('/almacen/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def nuevo_almacen():
    """ Formulario para crear un nuevo almacén. """
    if request.method == 'POST':
        try:
            nuevo_alm = Almacen(
                nombre=request.form['nombre'],
                ubicacion=request.form.get('ubicacion'),
                organizacion_id=current_user.organizacion_id
            )
            db.session.add(nuevo_alm)
            db.session.commit()
            flash('Almacén creado exitosamente', 'success')
            return redirect(url_for('lista_almacenes'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el almacén: {e}', 'danger')
            
    return render_template('almacen_form.html', 
                           titulo="Nuevo Almacén", 
                           almacen=None)

@app.route('/almacen/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_almacen(id):
    """ Edita un almacén existente. """
    almacen = get_item_or_404(Almacen, id)
    
    if request.method == 'POST':
        try:
            almacen.nombre = request.form['nombre']
            almacen.ubicacion = request.form.get('ubicacion')
            db.session.commit()
            flash('Almacén actualizado exitosamente', 'success')
            return redirect(url_for('lista_almacenes'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el almacén: {e}', 'danger')

    return render_template('almacen_form.html', 
                           titulo="Editar Almacén", 
                           almacen=almacen)

@app.route('/almacen/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_almacen(id):
    # Solo administradores
    if current_user.rol not in ['super_admin', 'admin']:
        flash('No tienes permiso para eliminar almacenes.', 'danger')
        return redirect(url_for('lista_almacenes'))

    almacen = Almacen.query.get_or_404(id)

    # Opcional: Validar que esté vacío antes de borrar
    # stocks_activos = Stock.query.filter_by(almacen_id=id).filter(Stock.cantidad > 0).count()
    # if stocks_activos > 0:
    #     flash(f'No puedes eliminar "{almacen.nombre}" porque aún tiene productos con stock.', 'warning')
    #     return redirect(url_for('lista_almacenes'))

    try:
        db.session.delete(almacen)
        db.session.commit()
        flash('Almacén eliminado correctamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el almacén (puede tener datos relacionados): {e}', 'danger')

    return redirect(url_for('lista_almacenes'))

@app.route('/almacen/<int:id>/inventario', methods=['GET', 'POST'])
@login_required
# @admin_required # (O usa @check_permission si es lo que usas en tu sistema)
def gestionar_inventario_almacen(id):
    almacen = get_item_or_404(Almacen, id)
    org_id = almacen.organizacion_id

    if request.method == 'POST':
        try:
            producto_id = int(request.form.get('producto_id'))
            ubicacion = request.form.get('ubicacion')
            # 1. CAPTURAR CANTIDAD (Calculada por el frontend: Cajas * Factor)
            cantidad = float(request.form.get('cantidad', 0))

            if not producto_id:
                raise Exception("No se seleccionó un producto.")

            stock_existente = Stock.query.filter_by(
                almacen_id=id, 
                producto_id=producto_id
            ).first()
            
            if stock_existente:
                flash('Ese producto ya está registrado en este almacén.', 'warning')
            else:
                # 2. CREAR STOCK CON CANTIDAD INICIAL Y ORG_ID
                nuevo_stock = Stock(
                    producto_id=producto_id,
                    almacen_id=id,
                    cantidad=cantidad, # <-- Usamos la cantidad del form
                    stock_minimo=5,
                    stock_maximo=100,
                    ubicacion=ubicacion
                )
                db.session.add(nuevo_stock)

                # 3. REGISTRAR MOVIMIENTO INICIAL (Si aplica)
                if cantidad > 0:
                    movimiento = Movimiento(
                        producto_id=producto_id,
                        cantidad=cantidad,
                        tipo='entrada-inicial',
                        fecha=now_mx(),
                        motivo='Stock Inicial (Alta Manual en Almacén)',
                        almacen_id=id,
                        organizacion_id=org_id
                    )
                    db.session.add(movimiento)

                db.session.commit()
                flash(f'Producto añadido al almacén con stock {cantidad}.', 'success')
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al añadir producto: {e}', 'danger')
        
        return redirect(url_for('gestionar_inventario_almacen', id=id))

    # LÓGICA GET
    productos_en_stock_ids = [s.producto_id for s in almacen.stocks]
    productos_catalogo = Producto.query.filter_by(organizacion_id=org_id).all()
    
    # Filtrar solo productos que NO están en este almacén
    productos_para_anadir = [
        p for p in productos_catalogo if p.id not in productos_en_stock_ids
    ]
    
    productos_para_anadir_json = []
    for p in productos_para_anadir:
        productos_para_anadir_json.append({
            "id": p.id,
            "nombre": p.nombre,
            "codigo": p.codigo,
            # 4. ENVIAR FACTOR DE EMPAQUE (Para la calculadora JS)
            "unidades_por_caja": int(p.unidades_por_caja) if p.unidades_por_caja and p.unidades_por_caja > 0 else 1
        })
    
    return render_template('almacen_inventario.html',
                           titulo=f"Inventario de {almacen.nombre}",
                           almacen=almacen,
                           productos_para_anadir_json=productos_para_anadir_json)

# --- FUNCIÓN NUEVA: eliminar_producto_de_almacen ---
@app.route('/almacen/stock/eliminar/<int:id>', methods=['POST'])
@login_required
@admin_required
def eliminar_producto_de_almacen(id):
    """ 
    Elimina un producto de un almacén específico (borra el registro de Stock).
    El producto sigue existiendo en el catálogo global.
    """
    stock_item = Stock.query.get_or_404(id)
    almacen_id = stock_item.almacen_id
    
    if stock_item.almacen.organizacion_id != current_user.organizacion_id:
        flash('No tienes permiso para realizar esta acción.', 'danger')
        return redirect(url_for('lista_almacenes'))

    try:
        nombre_prod = stock_item.producto.nombre
        nombre_alm = stock_item.almacen.nombre
        
        db.session.delete(stock_item)
        db.session.commit()
        
        flash(f'Producto "{nombre_prod}" eliminado de "{nombre_alm}".', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar del almacén: {e}', 'danger')

    return redirect(url_for('gestionar_inventario_almacen', id=almacen_id))

# ========================
# NUEVAS RUTAS: GESTIÓN DE HUÉRFANOS
# ========================

@app.route('/productos/sin-almacen')
@login_required
@check_org_permission
@check_permission('perm_view_management')
def lista_productos_sin_almacen():
    """ Muestra productos que existen en catálogo pero NO en ningún almacén (Stock). """
    org_id = current_user.organizacion_id
    
    # 1. Subconsulta: IDs de productos que SÍ tienen al menos un registro de Stock
    # (Un producto tiene stock si existe una fila en la tabla 'stock' vinculada a él)
    ids_con_stock = db.session.query(Stock.producto_id).join(Almacen).filter(
        Almacen.organizacion_id == org_id
    ).distinct()
    
    # 2. Consulta Principal: Productos de la org que NO están en la lista anterior
    productos_huerfanos = Producto.query.filter(
        Producto.organizacion_id == org_id,
        ~Producto.id.in_(ids_con_stock) # El símbolo ~ niega la condición (NOT IN)
    ).all()
    
    # Necesitamos los almacenes para el dropdown de asignación
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).all()
    
    return render_template('productos_sin_almacen.html',
                           titulo="Productos Sin Asignar",
                           productos=productos_huerfanos,
                           almacenes=almacenes)

@app.route('/producto/asignar-rapido', methods=['POST'])
@login_required
@check_permission('perm_edit_management')
def asignar_producto_rapido():
    """ Asigna un producto a un almacén rápidamente desde la lista de huérfanos. """
    try:
        producto_id = int(request.form.get('producto_id'))
        almacen_id = int(request.form.get('almacen_id'))
        
        if not producto_id or not almacen_id:
            raise Exception("Datos incompletos.")

        producto = Producto.query.get_or_404(producto_id)
        
        # Verificar si ya existe (doble check de seguridad)
        existe = Stock.query.filter_by(producto_id=producto_id, almacen_id=almacen_id).first()
        if existe:
            flash(f'El producto ya estaba en ese almacén.', 'warning')
        else:
            # Crear Stock inicial en 0
            nuevo_stock = Stock(
                producto_id=producto_id,
                almacen_id=almacen_id,
                cantidad=0,
                stock_minimo=5,
                stock_maximo=100
            )
            db.session.add(nuevo_stock)
            db.session.commit()
            flash(f'Producto "{producto.nombre}" asignado correctamente.', 'success')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al asignar producto: {e}', 'danger')
        
    return redirect(url_for('lista_productos_sin_almacen'))

#<---------SALIDA DE PRODUCTOS (REESCRITO PARA MULTI-ALMACÉN)----------->

@app.route('/salidas')
@login_required
@check_org_permission
@check_permission('perm_do_salidas')
def historial_salidas():
    """ Muestra el historial de Hojas de Salida Diarias (Multiusuario). """
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    ahora = now_mx()
    if not mes: mes = ahora.month
    if not ano: ano = ahora.year
    meses_lista = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'), 
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'), 
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    if current_user.rol == 'super_admin':
        query = Salida.query
    else:
        query = Salida.query.filter_by(organizacion_id=current_user.organizacion_id)

    query = query.filter(
        extract('month', Salida.fecha) == mes,
        extract('year', Salida.fecha) == ano
    )
    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(Salida.fecha.desc()).paginate(page=page, per_page=12, error_out=False)

    # --- Analytics del período ---
    org_id   = current_user.organizacion_id
    base_mov = Movimiento.query.filter(
        Movimiento.organizacion_id == org_id,
        Movimiento.tipo == 'salida',
        extract('month', Movimiento.fecha) == mes,
        extract('year',  Movimiento.fecha) == ano,
    )

    total_unidades = db.session.query(
        db.func.sum(db.func.abs(Movimiento.cantidad))
    ).filter(
        Movimiento.organizacion_id == org_id,
        Movimiento.tipo == 'salida',
        extract('month', Movimiento.fecha) == mes,
        extract('year',  Movimiento.fecha) == ano,
    ).scalar() or 0

    # Top 5 productos por unidades despachadas
    top_productos_raw = db.session.query(
        Movimiento.producto_id,
        db.func.sum(db.func.abs(Movimiento.cantidad)).label('uds')
    ).filter(
        Movimiento.organizacion_id == org_id,
        Movimiento.tipo == 'salida',
        extract('month', Movimiento.fecha) == mes,
        extract('year',  Movimiento.fecha) == ano,
    ).group_by(Movimiento.producto_id
    ).order_by(db.func.sum(db.func.abs(Movimiento.cantidad)).desc()).limit(5).all()

    top_productos = []
    for row in top_productos_raw:
        prod = Producto.query.get(row.producto_id)
        if prod:
            top_productos.append({'nombre': prod.nombre, 'sku': prod.codigo, 'uds': int(row.uds)})

    # Unidades por día (para la gráfica)
    import calendar
    dias_mes = calendar.monthrange(ano, mes)[1]
    daily_raw = db.session.query(
        db.func.extract('day', Movimiento.fecha).label('dia'),
        db.func.sum(db.func.abs(Movimiento.cantidad)).label('uds')
    ).filter(
        Movimiento.organizacion_id == org_id,
        Movimiento.tipo == 'salida',
        extract('month', Movimiento.fecha) == mes,
        extract('year',  Movimiento.fecha) == ano,
    ).group_by('dia').all()

    daily_map = {int(r.dia): int(r.uds) for r in daily_raw}
    chart_labels = list(range(1, dias_mes + 1))
    chart_data   = [daily_map.get(d, 0) for d in chart_labels]

    # Almacén más activo
    almacen_top_raw = db.session.query(
        Salida.almacen_id,
        db.func.count(Salida.id).label('total')
    ).filter(
        Salida.organizacion_id == org_id,
        extract('month', Salida.fecha) == mes,
        extract('year',  Salida.fecha) == ano,
    ).group_by(Salida.almacen_id
    ).order_by(db.func.count(Salida.id).desc()).first()

    almacen_top = None
    if almacen_top_raw:
        a = Almacen.query.get(almacen_top_raw.almacen_id)
        if a:
            almacen_top = {'nombre': a.nombre, 'total': almacen_top_raw.total}

    return render_template('salidas.html',
                           salidas=pagination.items,
                           pagination=pagination,
                           meses_lista=meses_lista,
                           mes_seleccionado=mes,
                           ano_seleccionado=ano,
                           total_unidades=total_unidades,
                           top_productos=top_productos,
                           almacen_top=almacen_top,
                           chart_labels=chart_labels,
                           chart_data=chart_data)

@app.route('/salida/<int:id>')
@login_required
@check_permission('perm_do_salidas')
def ver_salida(id):
    """ Muestra el detalle de una Hoja de Salida Diaria (Multiusuario). """
    salida = get_item_or_404(Salida, id)
    # Ordenamos los movimientos por hora para verlos cronológicamente
    movimientos = salida.movimientos.order_by(Movimiento.fecha.asc()).all()
    
    return render_template('salida_detalle.html', 
                           salida=salida, 
                           movimientos=movimientos,
                           titulo=f"Salida del {salida.fecha.strftime('%Y-%m-%d')}")

@app.route('/salida', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_do_salidas')
def registrar_salida():
    """ 
    AÑADE items a la Hoja de Salida del día de hoy.
    (MODIFICADO para Multi-Almacén)
    """
    org_id = current_user.organizacion_id
    
    # --- LÓGICA DE ALMACÉN ---
    almacen_id_solicitado = request.args.get('almacen_id', type=int)
    almacenes_org = Almacen.query.filter_by(organizacion_id=org_id).all()
    
    almacen_seleccionado = None
    if almacen_id_solicitado:
        almacen_seleccionado = Almacen.query.get(almacen_id_solicitado)
        # Chequeo de seguridad
        if not almacen_seleccionado or almacen_seleccionado.organizacion_id != org_id:
            flash('Permiso denegado para ese almacén.', 'danger')
            return redirect(url_for('historial_salidas'))
    
    if not almacenes_org:
        flash('No se pueden registrar salidas porque no hay almacenes creados.', 'warning')
        return redirect(url_for('index'))
    
    if not almacen_seleccionado:
        return render_template('seleccionar_almacen.html',
                               titulo="Seleccionar Almacén de Origen",
                               almacenes=almacenes_org,
                               destino_ruta='registrar_salida') # Ruta a la que volver

    # --- LÓGICA DE BUSCAR-O-CREAR LA HOJA DIARIA ---
    today = now_mx().date()
    salida_del_dia = Salida.query.filter_by(
        fecha=today, 
        organizacion_id=org_id,
        almacen_id=almacen_seleccionado.id # <-- Filtro por almacén
    ).first()

    if not salida_del_dia:
        salida_del_dia = Salida(
            fecha=today,
            creador_id=current_user.id,
            organizacion_id=org_id,
            almacen_id=almacen_seleccionado.id # <-- Asignar almacén
        )
        db.session.add(salida_del_dia)
        db.session.flush()

    # --- LÓGICA DE PRODUCTOS ---
    # Filtramos solo productos que tienen stock EN ESE ALMACÉN
    productos_en_almacen = db.session.query(Producto).join(Stock).filter(
        Stock.almacen_id == almacen_seleccionado.id,
        Producto.organizacion_id == org_id,
        Stock.cantidad > 0 # Solo mostrar productos que TENGAN stock
    ).all()
    
    productos_lista = []
    for p in productos_en_almacen:
        # Obtenemos el stock específico de ESE almacén
        stock_item = Stock.query.filter_by(producto_id=p.id, almacen_id=almacen_seleccionado.id).first()
        productos_lista.append({
            'id': p.id,
            'nombre': p.nombre,
            'codigo': p.codigo,
            'stock_actual': stock_item.cantidad if stock_item else 0
        })

    if request.method == 'POST':
        try:
            productos_ids = request.form.getlist('producto_id[]')
            cantidades = request.form.getlist('cantidad[]')
            motivos = request.form.getlist('motivo[]') # <-- AHORA ES UNA LISTA

            if not productos_ids:
                flash('Debes añadir al menos un producto a la salida.', 'danger')
                return redirect(url_for('registrar_salida', almacen_id=almacen_seleccionado.id))

            # --- 1. FASE DE VALIDACIÓN ---
            productos_para_actualizar = [] 
            for i in range(len(productos_ids)):
                prod_id = productos_ids[i]
                cant_str = cantidades[i]
                
                if not prod_id or not cant_str: continue
                cantidad_salida = int(cant_str)
                
                # Buscamos el item de stock específico
                stock_item = Stock.query.filter_by(producto_id=prod_id, almacen_id=almacen_seleccionado.id).first()
                
                if not stock_item:
                    flash(f'Error: Producto no válido.', 'danger')
                    db.session.rollback()
                    return render_template('salida_form.html', titulo=f"Registrar Salida de: {almacen_seleccionado.nombre}", productos=productos_lista, salida_id=salida_del_dia.id, almacen=almacen_seleccionado)
                if cantidad_salida <= 0:
                    flash('Todas las cantidades deben ser positivas.', 'danger')
                    db.session.rollback() 
                    return render_template('salida_form.html', titulo=f"Registrar Salida de: {almacen_seleccionado.nombre}", productos=productos_lista, salida_id=salida_del_dia.id, almacen=almacen_seleccionado)
                
                # --- VALIDACIÓN CRÍTICA DE STOCK (Ahora en la tabla Stock) ---
                if stock_item.cantidad < cantidad_salida:
                    flash(f'Error: Stock insuficiente para "{stock_item.producto.nombre}". Stock actual: {stock_item.cantidad}', 'danger')
                    db.session.rollback()
                    return render_template('salida_form.html', titulo=f"Registrar Salida de: {almacen_seleccionado.nombre}", productos=productos_lista, salida_id=salida_del_dia.id, almacen=almacen_seleccionado)
                
                productos_para_actualizar.append((stock_item, cantidad_salida, motivos[i]))

            # --- 2. FASE DE EJECUCIÓN ---
            for stock_item, cantidad_salida, motivo_item in productos_para_actualizar:
                
                # 1. Actualizar el stock del item
                stock_item.cantidad -= cantidad_salida
                db.session.add(stock_item)
                
                # 2. Registrar el movimiento VINCULADO
                movimiento = Movimiento(
                    producto_id=stock_item.producto_id,
                    cantidad= -cantidad_salida, # Negativo
                    tipo='salida',
                    fecha=now_mx(), # <-- Hora exacta
                    motivo=motivo_item, # <-- Motivo por item
                    salida=salida_del_dia, # <-- Vinculamos a la hoja diaria
                    almacen_id=almacen_seleccionado.id, # <-- ESTAMPAR ID
                    organizacion_id=org_id
                )
                db.session.add(movimiento)
            
            total_uds = sum(v[1] for v in productos_para_actualizar)
            log_actividad('salida', 'salida', f'Salida registrada: {len(productos_para_actualizar)} producto(s), {total_uds} uds — Almacén: {almacen_seleccionado.nombre}', entidad_id=salida_del_dia.id)
            db.session.commit()
            flash(f'Se añadieron {len(productos_para_actualizar)} items a la salida del día.', 'success')
            check_and_alert_stock_bajo(org_id, almacen_seleccionado.id)
            return redirect(url_for('ver_salida', id=salida_del_dia.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar la salida: {e}', 'danger')
    
    return render_template('salida_form.html', 
                           titulo=f"Registrar Salida de: {almacen_seleccionado.nombre}", 
                           productos=productos_lista,
                           salida_id=salida_del_dia.id, # Pasamos el ID para el botón "Ver Hoja de Hoy"
                           almacen=almacen_seleccionado)

@app.route('/movimiento/<int:id>/eliminar', methods=['POST'])
@login_required
@check_permission('perm_do_salidas')
def eliminar_movimiento_salida(id):
    """ 
    Elimina un SOLO item (Movimiento) de una hoja de salida 
    y REVIERTE el stock.
    """
    movimiento = get_item_or_404(Movimiento, id)
    
    if movimiento.tipo != 'salida':
        flash('Error: Solo se pueden eliminar items de salida.', 'danger')
        return redirect(url_for('historial_salidas'))
        
    salida_id_redirect = movimiento.salida_id
    
    # 1. GUARDAR EL NOMBRE ANTES DE BORRAR (SOLUCIÓN AL ERROR)
    nombre_producto = movimiento.producto.nombre

    try:
        # --- LÓGICA MODIFICADA ---
        # Buscamos el item de stock específico
        stock_item = Stock.query.filter_by(
            producto_id=movimiento.producto_id, 
            almacen_id=movimiento.almacen_id
        ).first()
        cantidad_a_devolver = abs(movimiento.cantidad)
        
        # 1. Revertir el stock
        if stock_item:
            stock_item.cantidad += cantidad_a_devolver
            db.session.add(stock_item)
        else:
            # Si el stock no existe, lo creamos (caso raro)
            stock_item = Stock(
                producto_id=movimiento.producto_id,
                almacen_id=movimiento.almacen_id,
                cantidad=cantidad_a_devolver,
                organizacion_id=movimiento.organizacion_id # Aseguramos la org
            )
            db.session.add(stock_item)
        
        # 2. Registrar el ajuste (para auditoría)
        mov_ajuste = Movimiento(
            producto_id=movimiento.producto_id,
            cantidad=cantidad_a_devolver,
            tipo='ajuste-entrada',
            fecha=now_mx(),
            motivo=f'Corrección/Eliminación de item (Salida #{salida_id_redirect})',
            almacen_id=movimiento.almacen_id,
            organizacion_id=movimiento.organizacion_id
        )
        db.session.add(mov_ajuste)
        
        # 3. Eliminar el movimiento de salida original
        db.session.delete(movimiento)
        
        db.session.commit()
        
        # 4. USAR LA VARIABLE GUARDADA PARA EL MENSAJE
        flash(f'Item "{nombre_producto}" eliminado. Stock revertido.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el item: {e}', 'danger')

    # Si la hoja de salida todavía existe, redirige a ella
    if salida_id_redirect and Salida.query.get(salida_id_redirect):
        return redirect(url_for('ver_salida', id=salida_id_redirect))
    # Si era el último item, la hoja se borró (por la cascada), 
    # así que redirigimos al historial
    return redirect(url_for('historial_salidas'))

# ── HELPERS DE PDF (compartidos por todos los generadores) ──────────────────

def _pdf_estilos(org):
    """Devuelve (fuente, color_primario, color_secundario) listos para ReportLab."""
    fuente = org.tipo_letra if org.tipo_letra in ['Helvetica', 'Times-Roman', 'Courier'] else 'Helvetica'
    c_pri = colors.HexColor(org.color_primario)  if org.color_primario  else colors.HexColor('#333333')
    c_sec = colors.HexColor(org.color_secundario) if org.color_secundario else colors.HexColor('#f1f5f9')
    return fuente, c_pri, c_sec

def _pdf_header(story, org, styles):
    """Añade encabezado de marca (logo + nombre + RFC + correo) y barra de color."""
    fuente, c_pri, _ = _pdf_estilos(org)

    s_brand = ParagraphStyle('_Brand', fontName=f'{fuente}-Bold', fontSize=20, leading=22, textColor=colors.black, spaceAfter=2)
    s_sub   = ParagraphStyle('_Sub',   fontName=fuente, fontSize=10, leading=12, textColor=colors.gray)
    s_meta  = ParagraphStyle('_Meta',  fontName=fuente, fontSize=8,  leading=10, textColor=colors.HexColor('#64748b'))

    logo_el = []
    if org.logo_url:
        logo_path = os.path.join(app.config['UPLOAD_FOLDER'], org.logo_url)
        if os.path.exists(logo_path):
            img = ReportLabImage(logo_path)
            max_h = 1.0 * inch
            img.drawHeight = max_h
            img.drawWidth  = max_h * (img.imageWidth / float(img.imageHeight))
            logo_el.append(img)

    text_el = [Paragraph(org.header_titulo or org.nombre, s_brand)]
    if org.header_subtitulo:
        text_el.append(Paragraph(org.header_subtitulo, s_sub))

    meta_parts = []
    if org.rfc:             meta_parts.append(f'RFC: {org.rfc}')
    if org.correo_empresa:  meta_parts.append(org.correo_empresa)
    if org.telefono:        meta_parts.append(org.telefono)
    if org.direccion:       meta_parts.append(org.direccion)
    if meta_parts:
        text_el.append(Paragraph(' · '.join(meta_parts), s_meta))

    if logo_el:
        t_hdr = Table([[logo_el, text_el]], colWidths=[1.5*inch, 4.7*inch])
    else:
        t_hdr = Table([[text_el]], colWidths=[6.2*inch])
    t_hdr.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(t_hdr)
    story.append(Table([['']], colWidths=[6.2*inch], rowHeights=[2],
                       style=TableStyle([('BACKGROUND', (0,0), (-1,-1), c_pri)])))
    story.append(Spacer(1, 0.2*inch))

def _pdf_footer(story, org, doc_url=None):
    """Añade bloque de pie de página (footer_texto + fecha generación + QR opcional)."""
    fuente, c_pri, _ = _pdf_estilos(org)
    s_footer = ParagraphStyle('_Foot', fontName=fuente, fontSize=8, textColor=colors.HexColor('#64748b'), alignment=TA_CENTER)

    story.append(Spacer(1, 0.3*inch))
    story.append(Table([['']], colWidths=[6.2*inch], rowHeights=[1],
                       style=TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#e2e8f0'))])))
    story.append(Spacer(1, 0.1*inch))

    pie_parts = []
    if org.footer_texto:
        pie_parts.append(org.footer_texto)
    pie_parts.append(f"Generado el {now_mx().strftime('%d/%m/%Y a las %H:%M')} · {org.nombre}")
    story.append(Paragraph('<br/>'.join(pie_parts), s_footer))

    if org.pdf_mostrar_qr and doc_url:
        try:
            qr = qrcode.QRCode(version=1, box_size=4, border=2)
            qr.add_data(doc_url)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color='black', back_color='white')
            qr_buf = io.BytesIO()
            qr_img.save(qr_buf, format='PNG')
            qr_buf.seek(0)
            rl_qr = ReportLabImage(qr_buf)
            rl_qr.drawWidth  = 0.7 * inch
            rl_qr.drawHeight = 0.7 * inch
            story.append(Spacer(1, 0.1*inch))
            t_qr = Table([[rl_qr]], colWidths=[6.2*inch])
            t_qr.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
            story.append(t_qr)
        except Exception:
            pass

def _pdf_row_styles(data_len, c_sec):
    """Devuelve estilos de filas alternas usando color_secundario."""
    styles = []
    for i in range(1, data_len):
        bg = c_sec if i % 2 == 0 else colors.white
        styles.append(('BACKGROUND', (0, i), (-1, i), bg))
    return styles

# ─────────────────────────────────────────────────────────────────────────────

@app.route('/salida/<int:id>/pdf')
@login_required
@check_permission('perm_do_salidas')
def generar_salida_pdf(id):
    salida = get_item_or_404(Salida, id)
    org = Organizacion.query.get(salida.organizacion_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=inch, leftMargin=inch,
                            topMargin=0.5*inch, bottomMargin=inch)
    story = []
    styles = getSampleStyleSheet()
    fuente, c_pri, c_sec = _pdf_estilos(org)

    s_normal = ParagraphStyle('SNorm',  fontName=fuente, fontSize=10, leading=12)
    s_bold   = ParagraphStyle('SBold',  fontName=f'{fuente}-Bold', fontSize=10, leading=12)
    s_brand  = ParagraphStyle('SBrand', fontName=f'{fuente}-Bold', fontSize=18, leading=20, textColor=colors.black)
    s_th     = ParagraphStyle('STH',    fontName=f'{fuente}-Bold', fontSize=9, textColor=colors.white, alignment=TA_CENTER)
    s_cell   = ParagraphStyle('SCell',  fontName=fuente, fontSize=9, leading=11)
    s_cellr  = ParagraphStyle('SCellR', fontName=fuente, fontSize=9, leading=11, alignment=TA_RIGHT)

    _pdf_header(story, org, styles)

    estado_color = '#DC2626' if salida.estado == 'cancelada' else '#059669'
    info_izq = [
        Paragraph('<b>ALMACÉN:</b>', s_normal),
        Paragraph(salida.almacen.nombre, s_bold),
        Paragraph(f'Fecha: {salida.fecha.strftime("%d/%m/%Y")}', s_normal),
        Paragraph(f'Creada por: {salida.creador.username}', s_normal),
    ]
    info_der = [
        Paragraph(f'<b>SALIDA #{salida.id}</b>', s_brand),
        Paragraph(f'<font color="{estado_color}"><b>{salida.estado.upper()}</b></font>', s_bold),
    ]
    t_info = Table([[info_izq, info_der]], colWidths=[3.5*inch, 2.7*inch])
    t_info.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('LEFTPADDING', (0,0), (-1,-1), 0)]))
    story.append(t_info)
    story.append(Spacer(1, 0.25*inch))

    data = [[
        Paragraph('Producto', s_th), Paragraph('SKU', s_th),
        Paragraph('Motivo', s_th),   Paragraph('Cantidad', s_th),
    ]]
    total_items = 0
    for mov in salida.movimientos.order_by(Movimiento.fecha.asc()).all():
        cant = abs(mov.cantidad)
        total_items += cant
        data.append([
            Paragraph(mov.producto.nombre, s_cell),
            Paragraph(mov.producto.codigo, s_cell),
            Paragraph(mov.motivo or '—', s_cell),
            Paragraph(str(cant), s_cellr),
        ])
    data.append(['', '', Paragraph('TOTAL UNIDADES:', ParagraphStyle('STotL', fontName=f'{fuente}-Bold', fontSize=10, alignment=TA_RIGHT)),
                 Paragraph(str(total_items), ParagraphStyle('STotV', fontName=f'{fuente}-Bold', fontSize=11, alignment=TA_RIGHT, textColor=c_pri))])

    t_mov = Table(data, colWidths=[2.8*inch, 1.2*inch, 1.4*inch, 0.8*inch], repeatRows=1)
    row_bgs = _pdf_row_styles(len(data) - 1, c_sec)
    t_mov.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  c_pri),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('GRID',          (0,0),  (-1,-2), 0.5, colors.HexColor('#DEE2E6')),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (3,0),  (3,-1),  'RIGHT'),
        ('TOPPADDING',    (0,0),  (-1,-1), 6),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 6),
        ('SPAN',          (0,-1), (1,-1)),
        ('LINEABOVE',     (0,-1), (-1,-1), 1, colors.HexColor('#DEE2E6')),
        ('BOX',           (2,-1), (3,-1),  0.5, colors.HexColor('#DEE2E6')),
    ] + row_bgs))
    story.append(t_mov)

    _pdf_footer(story, org)
    doc.build(story)
    buffer.seek(0)
    filename = f"Salida-{salida.id}_{salida.fecha.strftime('%Y-%m-%d')}.pdf"
    return send_file(buffer, as_attachment=False, download_name=filename, mimetype='application/pdf')
    
# --- RUTAS DE ÓRDENES DE COMPRA (OC) ---

@app.route('/ordenes')
@login_required
@check_org_permission
@check_permission('perm_create_oc_standard')
def lista_ordenes():
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    prov_id = request.args.get('proveedor_id', type=int)
    
    ahora = now_mx()
    if not mes: mes = ahora.month
    if not ano: ano = ahora.year

    meses_lista = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'), 
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'), 
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    if current_user.rol == 'super_admin':
        proveedores = Proveedor.query.order_by(Proveedor.nombre).all()
        query = OrdenCompra.query
    else:
        org_id = current_user.organizacion_id
        proveedores = Proveedor.query.filter_by(organizacion_id=org_id).order_by(Proveedor.nombre).all()
        query = OrdenCompra.query.filter_by(organizacion_id=org_id)

    query = query.filter(extract('month', OrdenCompra.fecha_creacion) == mes)
    query = query.filter(extract('year', OrdenCompra.fecha_creacion) == ano)

    if prov_id and prov_id != 0:
        query = query.filter_by(proveedor_id=prov_id)

    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(OrdenCompra.fecha_creacion.desc()).paginate(page=page, per_page=12, error_out=False)

    return render_template('ordenes.html',
                           ordenes=pagination.items,
                           pagination=pagination,
                           proveedores=proveedores,
                           meses_lista=meses_lista,
                           mes_seleccionado=mes,
                           ano_seleccionado=ano,
                           prov_seleccionado=prov_id or 0)

@app.route('/orden/nueva', methods=['POST'])
@login_required
@check_org_permission
@check_permission('perm_create_oc_standard')
def nueva_orden():
    try:
        ids_productos_a_ordenar = request.form.getlist('producto_id')
        almacen_id = request.form.get('almacen_id', type=int)
        
        if not ids_productos_a_ordenar or not almacen_id:
            flash('Error en la solicitud de alerta.', 'danger')
            return redirect(url_for('index'))

        productos = Producto.query.filter(Producto.id.in_(ids_productos_a_ordenar)).all()
        
        if current_user.rol != 'super_admin':
            for p in productos:
                if p.organizacion_id != current_user.organizacion_id:
                    flash('Error: Intento de ordenar un producto no válido.', 'danger')
                    return redirect(url_for('index'))

        proveedor_id_comun = productos[0].proveedor_id
        if not all(p.proveedor_id == proveedor_id_comun for p in productos):
            flash('Error: Los productos seleccionados deben ser del mismo proveedor.', 'danger')
            return redirect(url_for('index'))

        nueva_oc = OrdenCompra(
            proveedor_id=proveedor_id_comun,
            estado='borrador',
            creador_id=current_user.id,
            organizacion_id=current_user.organizacion_id,
            almacen_id=almacen_id
        )
        db.session.add(nueva_oc)
        
        for prod in productos:
            stock_item = Stock.query.filter_by(producto_id=prod.id, almacen_id=almacen_id).first()
            if stock_item:
                cantidad_sugerida = stock_item.stock_maximo - stock_item.cantidad
            else:
                cantidad_sugerida = 5
            
            cantidad_final = max(1, cantidad_sugerida)
            
            # NUEVO: Calculamos las cajas basadas en el empaque
            factor_empaque = getattr(prod, 'unidades_por_caja', 1) or 1
            cajas_calculadas = cantidad_final / factor_empaque
            
            # Usamos costo en lugar de precio_unitario por si tu base de datos usa costo
            costo_unitario = getattr(prod, 'precio_unitario', getattr(prod, 'costo', 0))
            
            detalle = OrdenCompraDetalle(
                orden=nueva_oc,
                producto_id=prod.id,
                cantidad_solicitada=cantidad_final,
                costo_unitario_estimado=costo_unitario,
                cajas=cajas_calculadas # <- Aquí guardamos las cajas
            )
            db.session.add(detalle)
        
        db.session.commit()
        flash('Nueva Orden de Compra generada en "Borrador".', 'success')
        return redirect(url_for('lista_ordenes'))

    except Exception as e:
        db.session.rollback()
        flash(f'Error al generar la orden: {e}', 'danger')
        return redirect(url_for('index'))
        
@app.route('/ordenes/recibir/<int:id>', methods=['POST'])
@login_required
def recibir_orden(id):
    """
    Recibe una OC y actualiza el Stock del Almacén correspondiente.
    Genera movimientos de entrada (Kardex).
    """
    # Usamos la consulta estándar compatible con tu versión de Flask
    orden = OrdenCompra.query.get_or_404(id)
    
    if orden.estado == 'recibida':
        flash('Esta orden ya fue recibida anteriormente.', 'warning')
        return redirect(url_for('lista_ordenes'))

    # Validación CRÍTICA: La orden debe tener un almacén destino
    if not orden.almacen_id:
        flash('Error: La orden no tiene un almacén asignado. No se puede ingresar el stock.', 'danger')
        return redirect(url_for('lista_ordenes'))

    try:
        org_id = orden.organizacion_id
        
        # Iteramos sobre los detalles (Variables originales del código que me diste)
        for detalle in orden.detalles:
            producto = detalle.producto
            cantidad = detalle.cantidad_solicitada
            
            # 1. ACTUALIZAR STOCK DEL ALMACÉN
            # Buscamos si el producto ya existe en ESTE almacén específico
            stock_item = Stock.query.filter_by(
                producto_id=producto.id,
                almacen_id=orden.almacen_id
            ).first()

            if stock_item:
                # Si existe, sumamos
                stock_item.cantidad += cantidad
                db.session.add(stock_item) # Aseguramos persistencia
            else:
                # Si no existe, lo creamos en este almacén
                nuevo_stock = Stock(
                    producto_id=producto.id,
                    almacen_id=orden.almacen_id,
                    cantidad=cantidad,
                    stock_minimo=5,  # Valores por defecto
                    stock_maximo=100
                )
                db.session.add(nuevo_stock)

            # 2. REGISTRAR MOVIMIENTO
            # CORRECCIÓN: Se agrega 'almacen_id' obligatorio para evitar NotNullViolation
            movimiento = Movimiento(
                producto_id=producto.id,
                cantidad=cantidad,
                tipo='entrada',
                fecha=now_mx(),
                motivo=f'Recepción de OC #{orden.id}',
                orden_compra_id=orden.id,
                organizacion_id=org_id,
                almacen_id=orden.almacen_id  # <--- ¡ESTA LÍNEA ES LA SOLUCIÓN!
            )
            db.session.add(movimiento)
            
            # Opcional: Si aún usas el contador global en Producto, lo actualizamos también
            if hasattr(producto, 'cantidad_stock'):
                producto.cantidad_stock = (producto.cantidad_stock or 0) + cantidad
                db.session.add(producto)
        
        # 3. Finalizar Orden
        orden.estado = 'recibida'
        orden.fecha_recepcion = now_mx()
        db.session.add(orden)

        log_actividad('recibir_oc', 'orden_compra', f'OC #{orden.id} recibida — {len(orden.detalles)} producto(s) ingresados al almacén {orden.almacen.nombre}', entidad_id=orden.id)
        db.session.commit()
        flash(f'¡Orden recibida! Stock ingresado correctamente al almacén: {orden.almacen.nombre}', 'success')
        enviar_push_notificacion(
            org_id=org_id,
            titulo='📦 OC Recibida',
            cuerpo=f'OC #{orden.id} de {orden.proveedor.nombre} — {len(orden.detalles)} producto(s) ingresados a {orden.almacen.nombre}.',
            url=url_for('ver_orden', id=orden.id)
        )
        
    except Exception as e:
        db.session.rollback()
        # Mejoramos el mensaje de error para que sea más legible si vuelve a pasar
        flash(f'Error al recibir la orden: {str(e)}', 'danger')
        print(f"DEBUG ERROR RECIBIR: {e}") 
    
    return redirect(url_for('lista_ordenes'))

@app.route('/orden/<int:id>/enviar', methods=['POST'])
@login_required
@check_permission('perm_create_oc_standard')
def enviar_orden(id):
    orden = OrdenCompra.query.filter_by(
        id=id, 
        organizacion_id=current_user.organizacion_id
    ).first_or_404()

    if orden.estado in ['borrador', 'Pendiente']:
        try:
            orden.estado = 'enviada'
            db.session.commit()
            flash('Orden marcada como "Enviada".', 'info')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
    
    return redirect(url_for('lista_ordenes'))

@app.route('/orden/<int:id>/pdf')
@login_required
@check_permission('perm_create_oc_standard')
def generar_oc_pdf(id):
    orden = OrdenCompra.query.filter_by(id=id, organizacion_id=current_user.organizacion_id).first_or_404()
    org = orden.organizacion
    proveedor = orden.proveedor

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=inch, leftMargin=inch,
                            topMargin=0.5*inch, bottomMargin=inch)
    story = []
    styles = getSampleStyleSheet()
    fuente, c_pri, c_sec = _pdf_estilos(org)

    s_normal = ParagraphStyle('OCNorm',  fontName=fuente, fontSize=10, leading=12)
    s_bold   = ParagraphStyle('OCBold',  fontName=f'{fuente}-Bold', fontSize=10, leading=12)
    s_brand  = ParagraphStyle('OCBrand', fontName=f'{fuente}-Bold', fontSize=18, leading=20, textColor=colors.black)
    s_th     = ParagraphStyle('OCTH',    fontName=f'{fuente}-Bold', fontSize=10, textColor=colors.white, alignment=TA_CENTER)
    s_totlbl = ParagraphStyle('OCTotL',  fontName=f'{fuente}-Bold', fontSize=11, alignment=TA_RIGHT)
    s_totval = ParagraphStyle('OCTotV',  fontName=f'{fuente}-Bold', fontSize=11, alignment=TA_RIGHT)

    _pdf_header(story, org, styles)

    p_email    = getattr(proveedor, 'contacto_email', getattr(proveedor, 'correo', '-'))
    p_tel      = getattr(proveedor, 'contacto_telefono', getattr(proveedor, 'telefono', '-'))
    p_contacto = getattr(proveedor, 'nombre_contacto', getattr(proveedor, 'contacto', '-'))

    info_proveedor = [
        Paragraph("<b>PROVEEDOR:</b>", s_normal),
        Paragraph(proveedor.nombre, s_bold),
        Paragraph(f"Contacto: {p_contacto}", s_normal),
        Paragraph(f"Email: {p_email}", s_normal),
        Paragraph(f"Tel: {p_tel}", s_normal),
    ]
    info_orden = [
        Paragraph(f"<b>ORDEN DE COMPRA #{orden.id}</b>", s_brand),
        Paragraph(f"<b>Fecha:</b> {orden.fecha_creacion.strftime('%d/%m/%Y')}", s_normal),
        Paragraph(f"<b>Estado:</b> {orden.estado.upper()}", s_normal),
        Paragraph(f"<b>Almacén:</b> {orden.almacen.nombre if orden.almacen else 'General'}", s_normal),
    ]
    t_info = Table([[info_proveedor, info_orden]], colWidths=[3.5*inch, 2.7*inch])
    t_info.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('LEFTPADDING', (0,0), (-1,-1), 0)]))
    story.append(t_info)
    story.append(Spacer(1, 0.2*inch))

    data_table = [[
        Paragraph("Producto / SKU", s_th), Paragraph("Cajas", s_th),
        Paragraph("Unidades", s_th),       Paragraph("Costo U.", s_th),
        Paragraph("Subtotal", s_th),
    ]]
    total_general = 0
    for detalle in orden.detalles:
        subtotal = detalle.cantidad_solicitada * detalle.costo_unitario_estimado
        total_general += subtotal
        factor_empaque = getattr(detalle.producto, 'unidades_por_caja', 1) or 1
        cajas = getattr(detalle, 'cajas', 0)
        enlace_url = getattr(detalle, 'enlace_proveedor', None) or getattr(detalle.producto, 'enlace_proveedor', None)
        desc = f"<b>{detalle.producto.nombre}</b><br/>SKU: {detalle.producto.codigo}<br/><font color='gray' size='8'>Empaque: {factor_empaque} ud(s)</font>"
        if enlace_url:
            display_url = (enlace_url[:50] + '...') if len(enlace_url) > 53 else enlace_url
            desc += f"<br/><font color='blue' size='7'>{display_url}</font>"
        data_table.append([
            Paragraph(desc, s_normal),
            Paragraph(f"{cajas:g}" if cajas else "0", s_normal),
            Paragraph(str(int(detalle.cantidad_solicitada)), s_normal),
            Paragraph(f"${detalle.costo_unitario_estimado:,.2f}", s_normal),
            Paragraph(f"${subtotal:,.2f}", s_normal),
        ])
    data_table.append(['', '', '', Paragraph("TOTAL:", s_totlbl), Paragraph(f"${total_general:,.2f}", s_totval)])

    t_productos = Table(data_table, colWidths=[2.4*inch, 0.8*inch, 0.8*inch, 1.0*inch, 1.2*inch], repeatRows=1)
    row_bgs = _pdf_row_styles(len(data_table) - 1, c_sec)
    t_productos.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  c_pri),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('ALIGN',         (0,0),  (-1,0),  'CENTER'),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('GRID',          (0,0),  (-1,-2), 0.5, colors.HexColor('#DEE2E6')),
        ('ALIGN',         (1,1),  (2,-2),  'CENTER'),
        ('ALIGN',         (3,1),  (-1,-1), 'RIGHT'),
        ('TOPPADDING',    (0,0),  (-1,-1), 6),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 6),
        ('SPAN',          (0,-1), (2,-1)),
        ('LINEABOVE',     (0,-1), (-1,-1), 1, colors.HexColor('#DEE2E6')),
        ('BACKGROUND',    (3,-1), (-1,-1), colors.whitesmoke),
        ('BOX',           (3,-1), (4,-1),  0.5, colors.HexColor('#DEE2E6')),
    ] + row_bgs))
    story.append(t_productos)

    _pdf_footer(story, org)
    doc.build(story)
    buffer.seek(0)
    filename = f"OC_{orden.id}_{secure_filename(org.nombre)}.pdf"
    return send_file(buffer, as_attachment=False, download_name=filename, mimetype='application/pdf')

@app.route('/orden/nueva/manual', methods=['GET', 'POST'])
@login_required
@check_permission('perm_create_oc_standard')
def nueva_orden_manual():
    """ Crea una nueva Orden de Compra manualmente. """
    
    if request.method == 'POST':
        try:
            proveedor_id = request.form.get('proveedor_id')
            almacen_id = request.form.get('almacen_id') # Nuevo campo Multi-Almacén
            
            if not proveedor_id:
                flash("Debes seleccionar un proveedor.", "warning")
                return redirect(request.url)

            # 1. Crear la Cabecera de la Orden
            nueva_orden = OrdenCompra(
                proveedor_id=proveedor_id,
                organizacion_id=current_user.organizacion_id,
                creador_id=current_user.id,
                estado='borrador', # <-- CORRECCIÓN: Debe nacer como borrador
                almacen_id=almacen_id if almacen_id else None # Guardamos el almacén destino
            )
            db.session.add(nueva_orden)
            db.session.flush() # Para obtener el ID de la orden antes de seguir

            # 2. Procesar las líneas de productos
            productos_ids = request.form.getlist('producto_id[]')
            cantidades = request.form.getlist('cantidad[]')
            costos = request.form.getlist('costo[]')
            cajas_lista = request.form.getlist('cajas[]') # <-- Capturamos las cajas
            enlaces_lista = request.form.getlist('enlace[]') # <-- NUEVO: Capturamos los enlaces

            for i in range(len(productos_ids)):
                if productos_ids[i] and float(cantidades[i]) > 0:
                    
                    # Extraer el valor de las cajas de forma segura
                    try:
                        cajas_val = float(cajas_lista[i])
                    except (IndexError, ValueError, TypeError):
                        cajas_val = 0.0
                        
                    # NUEVO: Extraer el enlace de forma segura
                    try:
                        enlace_val = enlaces_lista[i]
                    except IndexError:
                        enlace_val = ''
                    
                    detalle = OrdenCompraDetalle( 
                        orden_id=nueva_orden.id,
                        producto_id=productos_ids[i],
                        cantidad_solicitada=float(cantidades[i]),
                        costo_unitario_estimado=float(costos[i]),
                        cajas=cajas_val, # <-- Guardamos las cajas
                        enlace_proveedor=enlace_val # <-- CORRECCIÓN: Guardamos en enlace_proveedor
                    )
                    db.session.add(detalle)

            db.session.commit()
            flash(f"Orden #{nueva_orden.id} creada exitosamente en estado borrador.", "success")
            return redirect(url_for('lista_ordenes'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error al crear orden: {e}", "danger")
            return redirect(request.url)

    # --- MÉTODO GET: Renderizar el formulario ---
    org_id = current_user.organizacion_id
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).all()
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).all()

    # Convertir productos a diccionario para que tojson funcione
    productos_query = Producto.query.filter_by(organizacion_id=org_id).all()
    productos_lista = []
    for p in productos_query:
        productos_lista.append({
            'id': p.id,
            'nombre': p.nombre,
            'codigo': p.codigo,
            'precio_unitario': getattr(p, 'precio_unitario', getattr(p, 'costo', 0)),
            'proveedor_id': p.proveedor_id,
            'unidades_por_caja': getattr(p, 'unidades_por_caja', 1),
            'enlace': getattr(p, 'enlace_proveedor', '') # <-- CORRECCIÓN: Leemos enlace_proveedor
        })
    
    return render_template('orden_form.html', 
                           titulo="Nueva Orden de Compra",
                           orden=None,
                           proveedores=proveedores,
                           productos=productos_lista,
                           almacenes=almacenes)

@app.route('/orden/<int:id>')
@login_required
@check_permission('perm_view_dashboard')
def ver_orden(id):
    """ Muestra el detalle de una Orden de Compra (Solo lectura). """
    orden = OrdenCompra.query.filter_by(
        id=id, 
        organizacion_id=current_user.organizacion_id
    ).first_or_404()
    return render_template('orden_detalle.html', 
                           orden=orden, 
                           titulo=f"Detalle de Orden #{orden.id}")

@app.route('/orden/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@check_permission('perm_create_oc_standard')
def editar_orden(id):
    orden = OrdenCompra.query.filter_by(
        id=id, 
        organizacion_id=current_user.organizacion_id
    ).first_or_404()

    if orden.estado != 'borrador': # <-- CORRECCIÓN: Ajustado a 'borrador'
        flash('Solo se pueden editar órdenes en estado "Borrador".', 'warning')
        return redirect(url_for('ver_orden', id=id))

    org_id = orden.organizacion_id
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).all()
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).all()
    
    # NUEVO: Lista de productos construida como diccionarios (igual que en crear manual)
    productos_query = Producto.query.filter_by(organizacion_id=org_id).all()
    productos_lista = []
    for p in productos_query:
        productos_lista.append({
            'id': p.id,
            'nombre': p.nombre,
            'codigo': p.codigo,
            'precio_unitario': getattr(p, 'precio_unitario', getattr(p, 'costo', 0)),
            'proveedor_id': p.proveedor_id,
            'unidades_por_caja': getattr(p, 'unidades_por_caja', 1), # <-- Aseguramos que el frontend sepa cuantas cajas
            'enlace': getattr(p, 'enlace_proveedor', '') # <-- CORRECCIÓN: Leemos enlace_proveedor
        })
    
    if request.method == 'POST':
        try:
            orden.proveedor_id = request.form.get('proveedor_id')
            
            # Borramos detalles anteriores
            OrdenCompraDetalle.query.filter_by(orden_id=orden.id).delete()
            
            # Capturamos del formulario
            productos_ids = request.form.getlist('producto_id[]')
            cantidades = request.form.getlist('cantidad[]')
            costos = request.form.getlist('costo[]')
            cajas_lista = request.form.getlist('cajas[]')
            enlaces_lista = request.form.getlist('enlace[]') # <-- ¡NUEVO! Capturamos enlaces editados

            if not productos_ids:
                 flash('La orden debe tener al menos un producto.', 'danger')
                 db.session.rollback()
                 return render_template('orden_form.html',
                                        titulo=f"Editar Orden de Compra #{orden.id}",
                                        proveedores=proveedores,
                                        productos=productos_lista,
                                        almacenes=almacenes,
                                        orden=orden)
            
            # Llenamos los nuevos detalles usando el loop seguro por índices
            for i in range(len(productos_ids)):
                prod_id = productos_ids[i]
                cant = cantidades[i]
                cost = costos[i]

                if not prod_id or not cant or not cost:
                    continue 
                
                # Extraer el valor de cajas protegiendo contra listas vacías
                try:
                    cajas_val = float(cajas_lista[i])
                except (IndexError, ValueError, TypeError):
                    cajas_val = 0.0

                # NUEVO: Extraer el enlace de forma segura
                try:
                    enlace_val = enlaces_lista[i]
                except IndexError:
                    enlace_val = ''

                detalle = OrdenCompraDetalle(
                    orden_id=orden.id,
                    producto_id=int(prod_id),
                    cantidad_solicitada=int(cant),
                    costo_unitario_estimado=float(cost),
                    cajas=cajas_val,
                    enlace_proveedor=enlace_val # <-- Lo guardamos en BD
                )
                db.session.add(detalle)
            
            db.session.commit()
            flash('Orden de Compra actualizada exitosamente.', 'success')
            return redirect(url_for('ver_orden', id=id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la orden: {e}', 'danger')
            return render_template('orden_form.html',
                                   titulo=f"Editar Orden de Compra #{orden.id}",
                                   proveedores=proveedores,
                                   productos=productos_lista,
                                   almacenes=almacenes,
                                   orden=orden)

    return render_template('orden_form.html', 
                           titulo=f"Editar Orden de Compra #{orden.id}",
                           proveedores=proveedores,
                           productos=productos_lista, # <-- Mandamos lista de diccionarios
                           almacenes=almacenes,
                           orden=orden)

@app.route('/orden/<int:id>/cancelar', methods=['POST'])
@login_required
@check_permission('perm_create_oc_standard')
def cancelar_orden(id):
    orden = get_item_or_404(OrdenCompra, id)
    
    if orden.estado != 'borrador':
        flash('Error: Solo se pueden cancelar órdenes en estado "Borrador".', 'danger')
        return redirect(url_for('lista_ordenes'))

    try:
        orden.estado = 'cancelada'
        orden.cancelado_por_id = current_user.id
        db.session.commit()
        flash('Orden de Compra cancelada exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cancelar la orden: {e}', 'danger')
    
    return redirect(url_for('lista_ordenes'))

@app.route('/orden/<int:id>/eliminar', methods=['POST'])
@login_required
@check_permission('perm_create_oc_standard') # Usamos el mismo permiso de crear/editar
def eliminar_orden(id):
    """
    Maneja la eliminación de órdenes:
    - Si estaba Pendiente: Funciona como 'Cancelar y Eliminar' (Automático).
    - Si estaba Recibida: Funciona como 'Limpiar Historial'.
    """
    orden = OrdenCompra.query.filter_by(
        id=id, 
        organizacion_id=current_user.organizacion_id
    ).first_or_404()
    
    # Validaciones de seguridad opcionales
    # (Por ejemplo, impedir borrar si ya tiene movimientos de stock complejos asociados, 
    # aunque en este sistema simple asumimos que al borrar la OC no revertimos el stock histórico, solo borramos el papel).
    
    estado_anterior = orden.estado
    
    try:
        # Primero eliminamos los detalles para evitar errores de integridad (si no hay cascada configurada)
        OrdenCompraDetalle.query.filter_by(orden_id=orden.id).delete()
        
        # Ahora eliminamos la cabecera
        db.session.delete(orden)
        db.session.commit()
        
        if estado_anterior == 'Pendiente':
            flash(f'Orden #{id} cancelada y eliminada correctamente.', 'success')
        else:
            flash(f'Orden #{id} eliminada del historial.', 'info')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la orden: {e}', 'danger')
        
    return redirect(url_for('lista_ordenes'))

# =============================================
# RUTAS PARA OC DE PROYECTO
# =============================================

@app.route('/proyectos-oc')
@login_required
@check_org_permission
@check_permission('perm_create_oc_proyecto')
def lista_proyectos_oc():
    org_id = current_user.organizacion_id
    if current_user.rol == 'super_admin':
        query = ProyectoOC.query
    else:
        query = ProyectoOC.query.filter_by(organizacion_id=org_id)

    # Filtros
    mes  = request.args.get('mes',  type=int)
    ano  = request.args.get('ano',  type=int)
    estado_filtro = request.args.get('estado', '')

    if mes:
        query = query.filter(extract('month', ProyectoOC.fecha_creacion) == mes)
    if ano:
        query = query.filter(extract('year',  ProyectoOC.fecha_creacion) == ano)
    if estado_filtro:
        query = query.filter(ProyectoOC.estado == estado_filtro)

    proyectos_oc = query.order_by(ProyectoOC.fecha_creacion.desc()).all()
    proyectos    = ProyectoOC.query.filter_by(
        organizacion_id=org_id).with_entities(
        ProyectoOC.id, ProyectoOC.nombre_proyecto).distinct().all() \
        if current_user.rol != 'super_admin' else \
        ProyectoOC.query.with_entities(ProyectoOC.id, ProyectoOC.nombre_proyecto).all()

    return render_template('proyecto_oc_lista.html',
                           proyectos_oc=proyectos_oc,
                           proyectos=proyectos,
                           mes_sel=mes, ano_sel=ano, estado_sel=estado_filtro,
                           titulo="OC de Proyectos")

@app.route('/proyecto-oc/<int:id>')
@login_required
@check_permission('perm_create_oc_proyecto')
def ver_proyecto_oc(id):
    proyecto_oc = get_item_or_404(ProyectoOC, id)
    solicitud_pendiente = SolicitudAprobacion.query.filter_by(
        entidad_tipo='proyecto_oc',
        entidad_id=proyecto_oc.id,
        estado='pendiente'
    ).first()
    return render_template('proyecto_oc_detalle.html',
                           proyecto_oc=proyecto_oc,
                           solicitud_pendiente=solicitud_pendiente,
                           titulo=f"OC Proyecto #{proyecto_oc.id} — {proyecto_oc.nombre_proyecto}")

@app.route('/proyecto-oc/nueva', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_create_oc_proyecto')
def nuevo_proyecto_oc():
    org_id = current_user.organizacion_id
    
    # 1. Preparar Productos para JS
    productos_query = Producto.query.filter_by(organizacion_id=org_id).all()
    productos_lista = []
    for p in productos_query:
        productos_lista.append({
            'id': p.id,
            'nombre': p.nombre,
            'codigo': p.codigo,
            'precio_unitario': p.precio_unitario or getattr(p, 'costo', 0) or 0,
        })
        
    # 2. Preparar Proveedores para JS
    proveedores_query = Proveedor.query.filter_by(organizacion_id=org_id).order_by(Proveedor.nombre).all()
    proveedores_lista = [{'id': p.id, 'nombre': p.nombre} for p in proveedores_query]

    # 3. Obtener Almacenes para Jinja
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).all()

    if request.method == 'POST':
        try:
            nombre_proyecto = request.form.get('nombre_proyecto')
            almacen_id = request.form.get('almacen_id')
            
            # Validación de campo obligatorio
            if not nombre_proyecto:
                flash("El nombre del proyecto es obligatorio.", "danger")
                return render_template('proyecto_oc_form.html', 
                                       titulo="Crear OC de Proyecto",
                                       productos=productos_lista,
                                       proveedores=proveedores_lista,
                                       almacenes=almacenes,
                                       orden=None)

            # Crear la Cabecera de la Orden de Proyecto
            nueva_orden = ProyectoOC(
                nombre_proyecto=nombre_proyecto,
                creador_id=current_user.id,
                organizacion_id=org_id,
                almacen_id=int(almacen_id) if (almacen_id and almacen_id.isdigit()) else None,
                estado='borrador'
            )
            db.session.add(nueva_orden)
            db.session.flush() 

            # Capturar listas del formulario con nombres EXACTOS del nuevo HTML
            tipos = request.form.getlist('tipo_item[]')
            productos_existentes_ids = request.form.getlist('producto_id_existente[]')
            descripciones_nuevas = request.form.getlist('descripcion_nuevo[]')
            cantidades = request.form.getlist('cantidad[]')
            costos = request.form.getlist('costo_unitario[]')
            proveedores_sugeridos = request.form.getlist('proveedor_sugerido[]')
            enlaces = request.form.getlist('enlace_proveedor[]')
            comentarios = request.form.getlist('comentarios_detalle[]')

            for i in range(len(tipos)):
                # Saltar líneas vacías o en cero
                if not cantidades[i] or float(cantidades[i]) <= 0:
                    continue 

                detalle = ProyectoOCDetalle(
                    proyecto_oc_id=nueva_orden.id,
                    cantidad=float(cantidades[i]),
                    costo_unitario=float(costos[i]) if costos[i] else 0.0,
                    proveedor_sugerido=proveedores_sugeridos[i] if i < len(proveedores_sugeridos) else None,
                    enlace_proveedor=enlaces[i] if i < len(enlaces) else None,
                    comentarios_detalle=comentarios[i] if i < len(comentarios) else None,
                    descripcion_nuevo=descripciones_nuevas[i] if i < len(descripciones_nuevas) else "Sin descripción"
                )
                
                # Asignar el ID de producto si venía del catálogo
                if tipos[i] == 'existente':
                    pid_raw = productos_existentes_ids[i] if i < len(productos_existentes_ids) else '0'
                    if pid_raw.isdigit() and int(pid_raw) > 0:
                        detalle.producto_id = int(pid_raw)
                
                db.session.add(detalle)

            db.session.commit()
            flash(f'OC de Proyecto #{nueva_orden.id} creada exitosamente.', 'success')
            return redirect(url_for('lista_proyectos_oc'))

        except Exception as e:
            db.session.rollback()
            print(f"ERROR OC PROYECTO: {e}")
            _flash_err('Error al guardar la OC de proyecto. Intenta de nuevo.', e)
    
    return render_template('proyecto_oc_form.html', 
                           titulo="Crear OC de Proyecto",
                           productos=productos_lista,
                           proveedores=proveedores_lista,
                           almacenes=almacenes,
                           orden=None)


@app.route('/proyecto-oc/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@check_permission('perm_create_oc_proyecto')
def editar_proyecto_oc(id):
    proyecto_oc = get_item_or_404(ProyectoOC, id)
    org_id = proyecto_oc.organizacion_id

    if proyecto_oc.estado != 'borrador':
        flash('Solo se pueden editar Órdenes de Proyecto en estado "Borrador".', 'warning')
        return redirect(url_for('ver_proyecto_oc', id=id))

    if request.method == 'POST':
        try:
            proyecto_oc.nombre_proyecto = request.form.get('nombre_proyecto')
            almacen_id_val = request.form.get('almacen_id', type=int)
            proyecto_oc.almacen_id = almacen_id_val if almacen_id_val else None

            # Borrar detalles viejos y crear los nuevos
            ProyectoOCDetalle.query.filter_by(proyecto_oc_id=id).delete()
            
            # Capturar listas del formulario
            tipos = request.form.getlist('tipo_item[]')
            productos_existentes_ids = request.form.getlist('producto_id_existente[]') 
            descripciones_nuevas = request.form.getlist('descripcion_nuevo[]')
            cantidades = request.form.getlist('cantidad[]')
            costos = request.form.getlist('costo_unitario[]')
            proveedores_sugeridos = request.form.getlist('proveedor_sugerido[]')
            enlaces = request.form.getlist('enlace_proveedor[]')
            comentarios = request.form.getlist('comentarios_detalle[]')

            for i in range(len(tipos)):
                if not cantidades[i] or float(cantidades[i]) <= 0: 
                    continue 

                detalle = ProyectoOCDetalle(
                    proyecto_oc_id=id,
                    cantidad=float(cantidades[i]),
                    costo_unitario=float(costos[i]) if costos[i] else 0.0,
                    proveedor_sugerido=proveedores_sugeridos[i] if i < len(proveedores_sugeridos) else None,
                    enlace_proveedor=enlaces[i] if i < len(enlaces) else None,
                    comentarios_detalle=comentarios[i] if i < len(comentarios) else None,
                    descripcion_nuevo=descripciones_nuevas[i] if i < len(descripciones_nuevas) else "Sin descripción"
                )

                if tipos[i] == 'existente':
                    prod_id_val = int(productos_existentes_ids[i]) if productos_existentes_ids[i].isdigit() else 0
                    if prod_id_val > 0: 
                        detalle.producto_id = prod_id_val
                
                db.session.add(detalle)

            db.session.commit()
            flash(f'OC de Proyecto #{proyecto_oc.id} actualizada.', 'success')
            return redirect(url_for('ver_proyecto_oc', id=id))

        except Exception as e:
            db.session.rollback()
            print(f"ERROR EDITAR OC PROYECTO: {e}")
            _flash_err('Error al actualizar la OC de Proyecto. Intenta de nuevo.', e)
            return redirect(url_for('editar_proyecto_oc', id=id))
    
    # --- GET: Preparar datos ---
    productos_query = Producto.query.filter_by(organizacion_id=org_id).all()
    productos_lista = []
    for p in productos_query:
        productos_lista.append({
            'id': p.id, 'nombre': p.nombre, 'codigo': p.codigo, 'precio_unitario': getattr(p, 'precio_unitario', getattr(p, 'costo', 0))
        })
        
    proveedores_query = Proveedor.query.filter_by(organizacion_id=org_id).order_by(Proveedor.nombre).all()
    proveedores_lista = [{'id': p.id, 'nombre': p.nombre} for p in proveedores_query]

    almacenes = Almacen.query.filter_by(organizacion_id=org_id).all()
    
    # IMPORTANTE: Enviar TODOS los campos a JS para que cargue la edición perfecta
    detalles_json = []
    for d in proyecto_oc.detalles:
        detalles_json.append({
            'tipo': 'existente' if d.producto_id else 'nuevo',
            'producto_id': d.producto_id,
            'descripcion_nuevo': d.descripcion_nuevo,
            'cantidad': d.cantidad,
            'costo_unitario': d.costo_unitario,
            'proveedor_sugerido': d.proveedor_sugerido,
            'enlace_proveedor': d.enlace_proveedor,
            'comentarios_detalle': d.comentarios_detalle
        })
    
    return render_template('proyecto_oc_form.html', 
                           titulo=f"Editar OC de Proyecto #{proyecto_oc.id}",
                           productos=productos_lista,
                           proveedores=proveedores_lista,
                           almacenes=almacenes,
                           proyecto_oc=proyecto_oc,
                           detalles_json=detalles_json)


@app.route('/proyecto-oc/<int:id>/solicitar-aprobacion', methods=['POST'])
@login_required
@check_permission('perm_create_oc_proyecto')
def solicitar_aprobacion_oc(id):
    """Envía la OC de Proyecto a revisión de un administrador."""
    proyecto_oc = get_item_or_404(ProyectoOC, id)

    if proyecto_oc.estado != 'borrador':
        flash('Solo se puede solicitar aprobación desde estado Borrador.', 'danger')
        return redirect(url_for('ver_proyecto_oc', id=id))

    if not proyecto_oc.detalles:
        flash('La OC debe tener al menos un ítem antes de solicitar aprobación.', 'warning')
        return redirect(url_for('ver_proyecto_oc', id=id))

    try:
        proyecto_oc.estado = 'pendiente_aprobacion'
        solicitud = SolicitudAprobacion(
            entidad_tipo    = 'proyecto_oc',
            entidad_id      = proyecto_oc.id,
            solicitante_id  = current_user.id,
            organizacion_id = proyecto_oc.organizacion_id,
        )
        db.session.add(solicitud)
        log_actividad('solicitar_aprobacion', 'proyecto_oc',
                      f'OC Proyecto #{proyecto_oc.id} enviada a aprobación por {current_user.username}.',
                      entidad_id=proyecto_oc.id)
        db.session.commit()
        enviar_push_notificacion(
            org_id=proyecto_oc.organizacion_id,
            titulo='Aprobación requerida',
            cuerpo=f'{current_user.username} solicita aprobar OC-PROY-{proyecto_oc.id}: {proyecto_oc.nombre_proyecto}',
            url=f'/aprobaciones'
        )
        flash('Solicitud de aprobación enviada. Un administrador revisará la OC.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {e}', 'danger')

    return redirect(url_for('ver_proyecto_oc', id=id))


@app.route('/proyecto-oc/<int:id>/enviar', methods=['POST'])
@login_required
@check_permission('perm_create_oc_proyecto')
def enviar_proyecto_oc(id):
    """Marca la OC de Proyecto como enviada al proveedor (requiere estado aprobada o admin en borrador)."""
    proyecto_oc = get_item_or_404(ProyectoOC, id)
    es_admin = current_user.rol in ['super_admin', 'admin']

    estados_validos = ['aprobada']
    if es_admin:
        estados_validos.append('borrador')

    if proyecto_oc.estado not in estados_validos:
        flash('La OC debe estar aprobada antes de enviarse al proveedor.', 'danger')
        return redirect(url_for('ver_proyecto_oc', id=id))

    try:
        proyecto_oc.estado      = 'enviada'
        proyecto_oc.fecha_envio = now_mx()
        log_actividad('enviar', 'proyecto_oc',
                      f'OC de Proyecto #{proyecto_oc.id} "{proyecto_oc.nombre_proyecto}" marcada como enviada.',
                      entidad_id=proyecto_oc.id)
        db.session.commit()
        flash(f'OC #{proyecto_oc.id} marcada como enviada al proveedor.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {e}', 'danger')

    return redirect(url_for('ver_proyecto_oc', id=id))


# ──────────────────────────────────────────────
# FASE E — FLUJOS DE APROBACIÓN (INBOX)
# ──────────────────────────────────────────────

@app.route('/aprobaciones')
@login_required
@check_org_permission
def lista_aprobaciones():
    """Inbox de solicitudes de aprobación para administradores."""
    if current_user.rol not in ['super_admin', 'admin']:
        flash('No tienes permiso para ver las aprobaciones.', 'danger')
        return redirect(url_for('dashboard'))

    org_id = current_user.organizacion_id
    estado_filtro = request.args.get('estado', 'pendiente')

    q = SolicitudAprobacion.query.filter_by(organizacion_id=org_id)
    if estado_filtro:
        q = q.filter_by(estado=estado_filtro)
    solicitudes = q.order_by(SolicitudAprobacion.creado_en.desc()).all()

    # Enriquecer con el objeto entidad
    items = []
    for s in solicitudes:
        ent = None
        ent_url = '#'
        if s.entidad_tipo == 'proyecto_oc':
            ent = ProyectoOC.query.get(s.entidad_id)
            ent_url = url_for('ver_proyecto_oc', id=s.entidad_id)
        items.append({'s': s, 'ent': ent, 'ent_url': ent_url})

    pendientes = SolicitudAprobacion.query.filter_by(
        organizacion_id=org_id, estado='pendiente').count()

    return render_template('aprobaciones_inbox.html',
        titulo='Aprobaciones',
        items=items,
        estado_filtro=estado_filtro,
        pendientes=pendientes,
        now=now_mx(),
    )


@app.route('/aprobacion/<int:id>/aprobar', methods=['POST'])
@login_required
@check_org_permission
def aprobar_solicitud(id):
    if current_user.rol not in ['super_admin', 'admin']:
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('dashboard'))

    s = SolicitudAprobacion.query.filter_by(
        id=id, organizacion_id=current_user.organizacion_id).first_or_404()

    if s.estado != 'pendiente':
        flash('Esta solicitud ya fue resuelta.', 'warning')
        return redirect(url_for('lista_aprobaciones'))

    try:
        s.estado       = 'aprobado'
        s.aprobador_id = current_user.id
        s.resuelto_en  = now_mx()

        if s.entidad_tipo == 'proyecto_oc':
            oc = ProyectoOC.query.get(s.entidad_id)
            if oc:
                oc.estado = 'aprobada'
                log_actividad('aprobar', 'proyecto_oc',
                              f'OC Proyecto #{oc.id} aprobada por {current_user.username}.',
                              entidad_id=oc.id)

        db.session.commit()
        enviar_push_notificacion(
            org_id=current_user.organizacion_id,
            titulo='OC Aprobada',
            cuerpo=f'{current_user.username} aprobó la solicitud. Ya puede enviarse al proveedor.',
            url=f'/proyecto-oc/{s.entidad_id}' if s.entidad_tipo == 'proyecto_oc' else '/aprobaciones'
        )
        flash('Solicitud aprobada. La OC ya puede enviarse al proveedor.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {e}', 'danger')

    return redirect(url_for('lista_aprobaciones'))


@app.route('/aprobacion/<int:id>/rechazar', methods=['POST'])
@login_required
@check_org_permission
def rechazar_solicitud(id):
    if current_user.rol not in ['super_admin', 'admin']:
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('dashboard'))

    s = SolicitudAprobacion.query.filter_by(
        id=id, organizacion_id=current_user.organizacion_id).first_or_404()

    if s.estado != 'pendiente':
        flash('Esta solicitud ya fue resuelta.', 'warning')
        return redirect(url_for('lista_aprobaciones'))

    try:
        comentario     = request.form.get('comentario', '').strip()
        s.estado       = 'rechazado'
        s.aprobador_id = current_user.id
        s.resuelto_en  = now_mx()
        s.comentario   = comentario or None

        if s.entidad_tipo == 'proyecto_oc':
            oc = ProyectoOC.query.get(s.entidad_id)
            if oc:
                oc.estado = 'borrador'
                log_actividad('rechazar', 'proyecto_oc',
                              f'OC Proyecto #{oc.id} rechazada por {current_user.username}. Motivo: {comentario}',
                              entidad_id=oc.id)

        db.session.commit()
        enviar_push_notificacion(
            org_id=current_user.organizacion_id,
            titulo='OC Rechazada',
            cuerpo=f'Tu solicitud fue rechazada. {comentario[:80] if comentario else "Revisa los detalles."}',
            url=f'/proyecto-oc/{s.entidad_id}' if s.entidad_tipo == 'proyecto_oc' else '/aprobaciones'
        )
        flash('Solicitud rechazada. La OC vuelve a estado Borrador.', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {e}', 'danger')

    return redirect(url_for('lista_aprobaciones'))


@app.route('/proyecto-oc/<int:id>/recibir', methods=['GET', 'POST'])
@login_required
@check_permission('perm_create_oc_proyecto')
def recibir_proyecto_oc(id):
    """Registra la recepción física de la OC de Proyecto e ingresa stock al almacén seleccionado."""
    proyecto_oc = get_item_or_404(ProyectoOC, id)
    org_id = proyecto_oc.organizacion_id

    if proyecto_oc.estado != 'enviada':
        flash('Solo se puede registrar la recepción de una OC en estado "Enviada".', 'danger')
        return redirect(url_for('ver_proyecto_oc', id=id))

    almacenes = Almacen.query.filter_by(organizacion_id=org_id).order_by(Almacen.nombre).all()

    if request.method == 'POST':
        almacen_id_dest = request.form.get('almacen_id', type=int)
        if not almacen_id_dest:
            flash('Debes seleccionar un almacén destino.', 'danger')
            return render_template('proyecto_oc_recibir.html', proyecto_oc=proyecto_oc, almacenes=almacenes)

        almacen_dest = Almacen.query.get(almacen_id_dest)
        if not almacen_dest or almacen_dest.organizacion_id != org_id:
            flash('Almacén no válido.', 'danger')
            return render_template('proyecto_oc_recibir.html', proyecto_oc=proyecto_oc, almacenes=almacenes)

        # IDs de items del catálogo que el usuario quiere ingresar al stock
        items_a_ingresar = set(request.form.getlist('ingresar_item[]', type=int))

        try:
            items_ingresados = 0
            for detalle in proyecto_oc.detalles:
                if detalle.producto_id and detalle.producto_id in items_a_ingresar:
                    # Actualizar Stock
                    stock_item = Stock.query.filter_by(
                        producto_id=detalle.producto_id,
                        almacen_id=almacen_id_dest
                    ).first()
                    if stock_item:
                        stock_item.cantidad += detalle.cantidad
                    else:
                        stock_item = Stock(
                            producto_id=detalle.producto_id,
                            almacen_id=almacen_id_dest,
                            organizacion_id=org_id,
                            cantidad=detalle.cantidad,
                            stock_minimo=5,
                            stock_maximo=100
                        )
                        db.session.add(stock_item)

                    # Registrar Movimiento (Kárdex)
                    db.session.add(Movimiento(
                        producto_id=detalle.producto_id,
                        cantidad=detalle.cantidad,
                        tipo='entrada',
                        fecha=now_mx(),
                        motivo=f'Recepción OC Proyecto #{proyecto_oc.id} — {proyecto_oc.nombre_proyecto}',
                        almacen_id=almacen_id_dest,
                        organizacion_id=org_id
                    ))
                    items_ingresados += 1

            proyecto_oc.estado          = 'recibida'
            proyecto_oc.fecha_recepcion = now_mx()
            proyecto_oc.almacen_id      = almacen_id_dest
            proyecto_oc.recibido_por_id = current_user.id

            log_actividad('recibir', 'proyecto_oc',
                f'OC Proyecto #{proyecto_oc.id} recibida en "{almacen_dest.nombre}". '
                f'{items_ingresados} producto(s) ingresados al inventario.',
                entidad_id=proyecto_oc.id)
            db.session.commit()

            flash(f'✓ OC #{proyecto_oc.id} recibida. {items_ingresados} producto(s) ingresados al inventario de "{almacen_dest.nombre}".', 'success')
            return redirect(url_for('ver_proyecto_oc', id=id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar la recepción: {e}', 'danger')

    return render_template('proyecto_oc_recibir.html', proyecto_oc=proyecto_oc, almacenes=almacenes)


@app.route('/proyecto-oc/<int:id>/pdf')
@login_required
@check_permission('perm_create_oc_proyecto')
def generar_proyecto_oc_pdf(id):
    proyecto_oc = get_item_or_404(ProyectoOC, id)
    org = Organizacion.query.get(proyecto_oc.organizacion_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=inch, leftMargin=inch,
                            topMargin=0.5*inch, bottomMargin=inch)
    story = []
    styles = getSampleStyleSheet()
    fuente, c_pri, c_sec = _pdf_estilos(org)

    s_normal = ParagraphStyle('PONorm',  fontName=fuente, fontSize=10, leading=12)
    s_bold   = ParagraphStyle('POBold',  fontName=f'{fuente}-Bold', fontSize=10, leading=12)
    s_brand  = ParagraphStyle('POBrand', fontName=f'{fuente}-Bold', fontSize=18, leading=20, textColor=colors.black)
    s_th     = ParagraphStyle('POTH',    fontName=f'{fuente}-Bold', fontSize=9, textColor=colors.white, alignment=TA_CENTER)
    s_cell   = ParagraphStyle('POCell',  fontName=fuente, fontSize=9, leading=11)
    s_cellr  = ParagraphStyle('POCellR', fontName=fuente, fontSize=9, leading=11, alignment=TA_RIGHT)
    s_totlbl = ParagraphStyle('POTotL',  fontName=f'{fuente}-Bold', fontSize=10, alignment=TA_RIGHT)
    s_totval = ParagraphStyle('POTotV',  fontName=f'{fuente}-Bold', fontSize=11, alignment=TA_RIGHT, textColor=c_pri)

    _pdf_header(story, org, styles)

    estado_color = {'borrador':'#D97706','enviada':'#0891B2','recibida':'#059669','cancelada':'#64748B'}.get(proyecto_oc.estado, '#64748B')
    info_proyecto = [
        Paragraph('<b>PROYECTO:</b>', s_normal),
        Paragraph(proyecto_oc.nombre_proyecto, s_bold),
        Paragraph(f'Creado por: {proyecto_oc.creador.username}', s_normal),
        Paragraph(f'Fecha: {proyecto_oc.fecha_creacion.strftime("%d/%m/%Y")}', s_normal),
    ]
    if proyecto_oc.fecha_envio:
        info_proyecto.append(Paragraph(f'Enviado: {proyecto_oc.fecha_envio.strftime("%d/%m/%Y")}', s_normal))
    if proyecto_oc.fecha_recepcion:
        info_proyecto.append(Paragraph(f'Recibido: {proyecto_oc.fecha_recepcion.strftime("%d/%m/%Y")}', s_normal))
    if proyecto_oc.recibido_por:
        info_proyecto.append(Paragraph(f'Recibido por: {proyecto_oc.recibido_por.username}', s_normal))

    info_oc = [
        Paragraph(f'<b>OC-PROY-{proyecto_oc.id}</b>', s_brand),
        Paragraph(f'<font color="{estado_color}"><b>{proyecto_oc.estado.upper()}</b></font>', s_bold),
    ]
    if proyecto_oc.almacen:
        info_oc.append(Paragraph(f'Almacén: {proyecto_oc.almacen.nombre}', s_normal))

    t_info = Table([[info_proyecto, info_oc]], colWidths=[3.5*inch, 2.7*inch])
    t_info.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('LEFTPADDING', (0,0), (-1,-1), 0)]))
    story.append(t_info)
    story.append(Spacer(1, 0.25*inch))

    data = [[
        Paragraph('Descripción / SKU', s_th), Paragraph('Proveedor Sug.', s_th),
        Paragraph('Cant.', s_th), Paragraph('Costo Unit.', s_th), Paragraph('Subtotal', s_th),
    ]]
    total = 0
    for d in proyecto_oc.detalles:
        if d.producto_id and d.producto:
            desc_html = f'<b>{d.producto.nombre}</b><br/><font size="8" color="gray">SKU: {d.producto.codigo}</font>'
        else:
            desc_html = f'<b>{d.descripcion_nuevo or "Sin descripción"}</b><br/><font size="8" color="gray">Artículo externo</font>'
        if d.enlace_proveedor:
            short = (d.enlace_proveedor[:45] + '...') if len(d.enlace_proveedor) > 48 else d.enlace_proveedor
            desc_html += f'<br/><font size="7" color="blue">{short}</font>'
        if d.comentarios_detalle:
            desc_html += f'<br/><font size="7" color="gray">{d.comentarios_detalle}</font>'
        sub = d.cantidad * d.costo_unitario
        total += sub
        data.append([
            Paragraph(desc_html, s_cell), Paragraph(d.proveedor_sugerido or '—', s_cell),
            Paragraph(str(d.cantidad), s_cellr), Paragraph(f'${d.costo_unitario:,.2f}', s_cellr),
            Paragraph(f'${sub:,.2f}', s_cellr),
        ])
    data.append(['', '', '', Paragraph('TOTAL ESTIMADO:', s_totlbl), Paragraph(f'${total:,.2f}', s_totval)])

    t_art = Table(data, colWidths=[2.6*inch, 1.4*inch, 0.5*inch, 0.9*inch, 0.9*inch], repeatRows=1)
    row_bgs = _pdf_row_styles(len(data) - 1, c_sec)
    t_art.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  c_pri),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('GRID',          (0,0),  (-1,-2), 0.5, colors.HexColor('#DEE2E6')),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (2,0),  (-1,-1), 'RIGHT'),
        ('TOPPADDING',    (0,0),  (-1,-1), 6),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 6),
        ('SPAN',          (0,-1), (2,-1)),
        ('LINEABOVE',     (0,-1), (-1,-1), 1, colors.HexColor('#DEE2E6')),
        ('BOX',           (3,-1), (4,-1),  0.5, colors.HexColor('#DEE2E6')),
    ] + row_bgs))
    story.append(t_art)

    _pdf_footer(story, org)
    doc.build(story)
    buffer.seek(0)
    filename = f"OC-Proyecto-{proyecto_oc.id}_{proyecto_oc.fecha_creacion.strftime('%Y-%m-%d')}.pdf"
    return send_file(buffer, as_attachment=False, download_name=filename, mimetype='application/pdf')
    
@app.route('/proyecto-oc/<int:id>/cancelar', methods=['POST'])
@login_required
@check_permission('perm_create_oc_proyecto')
def cancelar_proyecto_oc(id):
    """Cancela una OC de Proyecto (soft delete — cambia estado, no borra el registro)."""
    proyecto_oc = get_item_or_404(ProyectoOC, id)

    if proyecto_oc.estado in ['recibida', 'cancelada']:
        flash('No se puede cancelar una orden ya recibida o previamente cancelada.', 'danger')
        return redirect(url_for('ver_proyecto_oc', id=id))

    try:
        proyecto_oc.estado = 'cancelada'
        log_actividad('cancelar', 'proyecto_oc', f'OC de Proyecto #{proyecto_oc.id} "{proyecto_oc.nombre_proyecto}" cancelada.', entidad_id=proyecto_oc.id)
        db.session.commit()
        flash(f'OC de Proyecto #{proyecto_oc.id} cancelada.', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cancelar la orden: {e}', 'danger')

    return redirect(url_for('lista_proyectos_oc'))


@app.route('/proyecto-oc/exportar.xlsx')
@login_required
@check_org_permission
@check_permission('perm_create_oc_proyecto')
def exportar_proyectos_oc_excel():
    org_id = current_user.organizacion_id
    estado_f = request.args.get('estado')
    mes_f    = request.args.get('mes', type=int)
    ano_f    = request.args.get('ano', type=int)

    query = ProyectoOC.query.filter_by(organizacion_id=org_id)
    if estado_f:
        query = query.filter(ProyectoOC.estado == estado_f)
    if mes_f:
        query = query.filter(extract('month', ProyectoOC.fecha_creacion) == mes_f)
    if ano_f:
        query = query.filter(extract('year', ProyectoOC.fecha_creacion) == ano_f)
    proyectos = query.order_by(ProyectoOC.fecha_creacion.desc()).all()

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen de OC ───────────────────────────────────────────────
    ws = wb.active
    ws.title = 'OC de Proyectos'

    COLOR_HDR  = 'FF4F46E5'
    COLOR_ALT  = 'FFF0F4FF'
    COLOR_TOT  = 'FFDBEAFE'

    h_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    b_font = Font(name='Calibri', size=10)
    t_font = Font(name='Calibri', size=11, bold=True)

    h_fill  = PatternFill('solid', fgColor=COLOR_HDR)
    alt_fill= PatternFill('solid', fgColor=COLOR_ALT)
    tot_fill= PatternFill('solid', fgColor=COLOR_TOT)

    thin = Side(style='thin', color='FFBFDBFE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdrs = ['ID', 'Proyecto', 'Estado', 'Creado Por', 'Fecha Creación',
            'Fecha Envío', 'Fecha Recepción', 'Almacén Destino',
            'Recibido Por', 'Artículos', 'Total Estimado (MXN)']
    ws.append(hdrs)
    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(1, col)
        cell.font   = h_font
        cell.fill   = h_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    estado_labels = {'borrador': 'Borrador', 'enviada': 'Enviada',
                     'recibida': 'Recibida', 'cancelada': 'Cancelada'}

    for i, poc in enumerate(proyectos, 2):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        row = [
            poc.id,
            poc.nombre_proyecto,
            estado_labels.get(poc.estado, poc.estado),
            poc.creador.username,
            poc.fecha_creacion.strftime('%d/%m/%Y'),
            poc.fecha_envio.strftime('%d/%m/%Y') if poc.fecha_envio else '—',
            poc.fecha_recepcion.strftime('%d/%m/%Y') if poc.fecha_recepcion else '—',
            poc.almacen.nombre if poc.almacen else '—',
            poc.recibido_por.username if poc.recibido_por else '—',
            len(poc.detalles),
            poc.costo_total,
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(i, col)
            cell.font   = b_font
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if col == 11:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fila de total
    last = len(proyectos) + 2
    ws.cell(last, 10, 'TOTAL:').font = t_font
    ws.cell(last, 10).alignment = Alignment(horizontal='right')
    total_val = ws.cell(last, 11, sum(p.costo_total for p in proyectos))
    total_val.font = t_font
    total_val.number_format = '"$"#,##0.00'
    total_val.alignment = Alignment(horizontal='right')
    for col in range(1, 12):
        ws.cell(last, col).fill = tot_fill

    col_widths = [6, 30, 12, 16, 16, 16, 18, 22, 16, 10, 22]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    # ── Hoja 2: Detalle de artículos ───────────────────────────────────────
    ws2 = wb.create_sheet('Artículos')
    hdrs2 = ['OC ID', 'Proyecto', 'Estado OC', 'Tipo', 'Artículo / Descripción',
             'SKU', 'Proveedor Sug.', 'Cantidad', 'Costo Unit.', 'Subtotal']
    ws2.append(hdrs2)
    for col, h in enumerate(hdrs2, 1):
        cell = ws2.cell(1, col)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    row_idx = 2
    for poc in proyectos:
        for d in poc.detalles:
            if d.producto_id and d.producto:
                tipo  = 'Catálogo'
                nombre = d.producto.nombre
                sku   = d.producto.codigo
            else:
                tipo  = 'Externo'
                nombre = d.descripcion_nuevo or 'Sin descripción'
                sku   = '—'
            sub = d.cantidad * d.costo_unitario
            row2 = [poc.id, poc.nombre_proyecto, estado_labels.get(poc.estado, poc.estado),
                    tipo, nombre, sku, d.proveedor_sugerido or '—',
                    d.cantidad, d.costo_unitario, sub]
            ws2.append(row2)
            fill2 = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col in range(1, len(row2) + 1):
                cell = ws2.cell(row_idx, col)
                cell.font   = b_font
                cell.fill   = fill2
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                if col in (9, 10):
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            row_idx += 1

    col_widths2 = [6, 28, 12, 10, 30, 14, 20, 10, 14, 14]
    for col, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha_str = now_mx().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f'OC-Proyectos_{fecha_str}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# --- RUTAS DE REPORTES ---

# ──────────────────────────────────────────────
# DASHBOARD FINANCIERO UNIFICADO
# ──────────────────────────────────────────────

@app.route('/finanzas')
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def finanzas_dashboard():
    org_id = current_user.organizacion_id
    ahora  = now_mx()
    mes_actual = ahora.month
    ano_actual = ahora.year

    # Mes anterior
    mes_ant = mes_actual - 1 if mes_actual > 1 else 12
    ano_ant = ano_actual if mes_actual > 1 else ano_actual - 1

    def _sum_gastos(mes, ano):
        return db.session.query(db.func.coalesce(db.func.sum(Gasto.monto), 0.0)).filter(
            Gasto.organizacion_id == org_id,
            db.extract('month', Gasto.fecha) == mes,
            db.extract('year',  Gasto.fecha) == ano
        ).scalar()

    def _sum_ocs(mes, ano):
        return db.session.query(
            db.func.coalesce(db.func.sum(
                OrdenCompraDetalle.cantidad_solicitada * OrdenCompraDetalle.costo_unitario_estimado
            ), 0.0)
        ).join(OrdenCompra).filter(
            OrdenCompra.organizacion_id == org_id,
            OrdenCompra.estado == 'recibida',
            db.extract('month', OrdenCompra.fecha_recepcion) == mes,
            db.extract('year',  OrdenCompra.fecha_recepcion) == ano
        ).scalar()

    def _sum_servicios(mes, ano):
        return db.session.query(db.func.coalesce(db.func.sum(PagoServicio.monto), 0.0)).join(Servicio).filter(
            Servicio.organizacion_id == org_id,
            PagoServicio.estado == 'pagado',
            db.extract('month', PagoServicio.fecha_pago) == mes,
            db.extract('year',  PagoServicio.fecha_pago) == ano
        ).scalar()

    # KPIs mes actual
    gastos_mes      = _sum_gastos(mes_actual, ano_actual)
    ocs_mes         = _sum_ocs(mes_actual, ano_actual)
    servicios_mes   = _sum_servicios(mes_actual, ano_actual)
    total_mes       = gastos_mes + ocs_mes + servicios_mes

    # KPI mes anterior (para variación %)
    total_mes_ant   = _sum_gastos(mes_ant, ano_ant) + _sum_ocs(mes_ant, ano_ant) + _sum_servicios(mes_ant, ano_ant)
    variacion       = ((total_mes - total_mes_ant) / total_mes_ant * 100) if total_mes_ant > 0 else None

    # OCs comprometidas (estado=enviada, aún no recibidas)
    ocs_en_transito = db.session.query(
        db.func.coalesce(db.func.sum(
            OrdenCompraDetalle.cantidad_solicitada * OrdenCompraDetalle.costo_unitario_estimado
        ), 0.0)
    ).join(OrdenCompra).filter(
        OrdenCompra.organizacion_id == org_id,
        OrdenCompra.estado == 'enviada'
    ).scalar()

    # Servicios por vencer próximos 15 días
    limite_aviso        = ahora.date() + timedelta(days=15)
    servicios_por_vencer = (
        PagoServicio.query.join(Servicio)
        .filter(Servicio.organizacion_id == org_id,
                PagoServicio.estado == 'pendiente',
                PagoServicio.fecha_vencimiento <= limite_aviso)
        .order_by(PagoServicio.fecha_vencimiento)
        .all()
    )
    monto_por_vencer = sum(p.monto for p in servicios_por_vencer)

    # Gastos por categoría (mes actual)
    gastos_por_cat = db.session.query(
        db.func.coalesce(Gasto.categoria, 'Sin categoría'),
        db.func.sum(Gasto.monto)
    ).filter(
        Gasto.organizacion_id == org_id,
        db.extract('month', Gasto.fecha) == mes_actual,
        db.extract('year',  Gasto.fecha) == ano_actual
    ).group_by(Gasto.categoria).all()

    # Últimas transacciones unificadas
    ultimos_gastos = (Gasto.query
        .filter_by(organizacion_id=org_id)
        .order_by(Gasto.fecha.desc()).limit(8).all())

    ultimas_ocs = (OrdenCompra.query
        .filter_by(organizacion_id=org_id, estado='recibida')
        .order_by(OrdenCompra.fecha_recepcion.desc()).limit(8).all())

    ultimos_pagos = (PagoServicio.query.join(Servicio)
        .filter(Servicio.organizacion_id == org_id,
                PagoServicio.estado == 'pagado')
        .order_by(PagoServicio.fecha_pago.desc()).limit(8).all())

    # Mezcla y ordena por fecha descendente
    transacciones = []
    for g in ultimos_gastos:
        transacciones.append({
            'fecha': g.fecha.date() if hasattr(g.fecha, 'date') else g.fecha,
            'tipo': 'Gasto',
            'descripcion': g.descripcion,
            'categoria': g.categoria or 'Sin categoría',
            'monto': g.monto,
            'badge_class': 'badge-borrador',
            'icon': 'bi-cash-coin',
        })
    for oc in ultimas_ocs:
        transacciones.append({
            'fecha': oc.fecha_recepcion.date() if oc.fecha_recepcion and hasattr(oc.fecha_recepcion, 'date') else oc.fecha_recepcion,
            'tipo': 'Compra',
            'descripcion': f'OC #{oc.id} — {oc.proveedor.nombre}',
            'categoria': 'Inventario',
            'monto': oc.costo_total,
            'badge_class': 'badge-recibida',
            'icon': 'bi-cart-check-fill',
        })
    for p in ultimos_pagos:
        transacciones.append({
            'fecha': p.fecha_pago,
            'tipo': 'Servicio',
            'descripcion': p.servicio.nombre,
            'categoria': p.servicio.tipo.capitalize() if p.servicio.tipo else 'Servicio',
            'monto': p.monto,
            'badge_class': 'badge-enviada',
            'icon': 'bi-lightning-charge-fill',
        })
    transacciones.sort(key=lambda x: x['fecha'] if x['fecha'] else date.min, reverse=True)
    transacciones = transacciones[:15]

    return render_template('finanzas_dashboard.html',
        titulo='Dashboard Financiero',
        total_mes=total_mes,
        gastos_mes=gastos_mes,
        ocs_mes=ocs_mes,
        servicios_mes=servicios_mes,
        ocs_en_transito=ocs_en_transito,
        monto_por_vencer=monto_por_vencer,
        variacion=variacion,
        total_mes_ant=total_mes_ant,
        servicios_por_vencer=servicios_por_vencer,
        gastos_por_cat=gastos_por_cat,
        transacciones=transacciones,
        ahora=ahora,
        mes_actual=mes_actual,
        ano_actual=ano_actual,
        now=ahora,
    )


@app.route('/api/finanzas/mensual')
@login_required
@check_org_permission
def api_finanzas_mensual():
    org_id = current_user.organizacion_id
    ahora  = now_mx()
    labels, gastos_data, ocs_data, servicios_data = [], [], [], []

    for i in range(5, -1, -1):
        m = ahora.month - i
        y = ahora.year
        if m <= 0:
            m += 12
            y -= 1

        g = db.session.query(db.func.coalesce(db.func.sum(Gasto.monto), 0.0)).filter(
            Gasto.organizacion_id == org_id,
            db.extract('month', Gasto.fecha) == m,
            db.extract('year',  Gasto.fecha) == y
        ).scalar()

        o = db.session.query(
            db.func.coalesce(db.func.sum(
                OrdenCompraDetalle.cantidad_solicitada * OrdenCompraDetalle.costo_unitario_estimado
            ), 0.0)
        ).join(OrdenCompra).filter(
            OrdenCompra.organizacion_id == org_id,
            OrdenCompra.estado == 'recibida',
            db.extract('month', OrdenCompra.fecha_recepcion) == m,
            db.extract('year',  OrdenCompra.fecha_recepcion) == y
        ).scalar()

        s = db.session.query(db.func.coalesce(db.func.sum(PagoServicio.monto), 0.0)).join(Servicio).filter(
            Servicio.organizacion_id == org_id,
            PagoServicio.estado == 'pagado',
            db.extract('month', PagoServicio.fecha_pago) == m,
            db.extract('year',  PagoServicio.fecha_pago) == y
        ).scalar()

        import calendar
        labels.append(calendar.month_abbr[m] + f' {y}')
        gastos_data.append(round(float(g), 2))
        ocs_data.append(round(float(o), 2))
        servicios_data.append(round(float(s), 2))

    return jsonify({'labels': labels, 'gastos': gastos_data, 'ocs': ocs_data, 'servicios': servicios_data})


@app.route('/reportes')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def reportes():
    org_id = current_user.organizacion_id
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).order_by(Almacen.nombre).all()
    return render_template('reportes.html', titulo='Reportes', almacenes=almacenes, now=now_mx())


@app.route('/reportes/inventario.xlsx')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def exportar_inventario_excel():
    org_id = current_user.organizacion_id
    almacen_id = request.args.get('almacen_id', type=int)

    if almacen_id:
        almacen = Almacen.query.filter_by(id=almacen_id, organizacion_id=org_id).first_or_404()
        items = db.session.query(Stock).filter_by(almacen_id=almacen_id).join(Producto).order_by(Producto.nombre).all()
        nombre_almacen = almacen.nombre
    else:
        items = db.session.query(Stock).join(
            Almacen, Stock.almacen_id == Almacen.id
        ).filter(Almacen.organizacion_id == org_id).join(Producto).order_by(Producto.nombre).all()
        nombre_almacen = 'Todos los Almacenes'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    h_font  = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    h_fill  = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    h_align = Alignment(horizontal='center', vertical='center')
    b_font  = Font(name='Arial', size=10)
    thin    = Side(border_style='thin', color='DEE2E6')
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    mxn     = NamedStyle(name='mxn_inv', number_format='"$"#,##0.00')
    if 'mxn_inv' not in wb.named_styles:
        wb.add_named_style(mxn)

    ws.merge_cells('A1:J1')
    ws['A1'].value = f"Inventario — {nombre_almacen} — {now_mx().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font  = Font(name='Arial', size=13, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    headers = ['SKU', 'Producto', 'Categoría', 'Proveedor', 'Stock', 'Mín.', 'Máx.', 'Estado', 'Precio Unit. MXN', 'Valor Total MXN']
    ws.append(headers)
    for cell in ws[2]:
        cell.font  = h_font
        cell.fill  = h_fill
        cell.alignment = h_align
        cell.border = border
    ws.row_dimensions[2].height = 20

    estado_label = {'bajo': 'Bajo Mínimo', 'exceso': 'Exceso', 'ok': 'Óptimo'}
    valor_total = 0

    for item in items:
        valor = (item.cantidad or 0) * (item.producto.precio_unitario or 0)
        valor_total += valor
        ws.append([
            item.producto.codigo,
            item.producto.nombre,
            item.producto.categoria.nombre if item.producto.categoria else '',
            item.producto.proveedor.nombre if item.producto.proveedor else '',
            item.cantidad,
            item.stock_minimo,
            item.stock_maximo,
            estado_label.get(item.estado_stock, ''),
            item.producto.precio_unitario or 0,
            valor,
        ])
        r = ws.max_row
        for col in range(1, 11):
            ws.cell(row=r, column=col).font   = b_font
            ws.cell(row=r, column=col).border = border
        estado_cell = ws.cell(row=r, column=8)
        if item.estado_stock == 'bajo':
            estado_cell.font = Font(name='Arial', size=10, color='DC2626', bold=True)
        elif item.estado_stock == 'exceso':
            estado_cell.font = Font(name='Arial', size=10, color='0891B2', bold=True)
        else:
            estado_cell.font = Font(name='Arial', size=10, color='059669', bold=True)
        ws.cell(row=r, column=9).style  = 'mxn_inv'
        ws.cell(row=r, column=10).style = 'mxn_inv'

    tr = ws.max_row + 1
    ws.cell(row=tr, column=9).value     = 'VALOR TOTAL (MXN)'
    ws.cell(row=tr, column=9).font      = Font(name='Arial', size=11, bold=True)
    ws.cell(row=tr, column=9).alignment = Alignment(horizontal='right')
    ws.cell(row=tr, column=10).value    = valor_total
    ws.cell(row=tr, column=10).style    = 'mxn_inv'
    ws.cell(row=tr, column=10).font     = Font(name='Arial', size=11, bold=True)

    for i, w in enumerate([15, 32, 18, 22, 9, 7, 7, 14, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"Inventario_{nombre_almacen.replace(' ', '_')}_{now_mx().strftime('%Y%m%d')}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True)


@app.route('/reportes/movimientos.xlsx')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def exportar_movimientos_excel():
    org_id     = current_user.organizacion_id
    almacen_id = request.args.get('almacen_id', type=int)
    desde_str  = request.args.get('desde', '')
    hasta_str  = request.args.get('hasta', '')
    tipo_f     = request.args.get('tipo', '')
    ahora      = now_mx()

    try:
        fecha_desde = datetime.strptime(desde_str, '%Y-%m-%d') if desde_str else ahora.replace(day=1, hour=0, minute=0, second=0)
        fecha_hasta = datetime.strptime(hasta_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59) if hasta_str else ahora
    except ValueError:
        fecha_desde = ahora.replace(day=1, hour=0, minute=0, second=0)
        fecha_hasta = ahora

    q = Movimiento.query.filter_by(organizacion_id=org_id).filter(
        Movimiento.fecha >= fecha_desde,
        Movimiento.fecha <= fecha_hasta
    )
    if almacen_id:
        q = q.filter(Movimiento.almacen_id == almacen_id)
    if tipo_f:
        q = q.filter(Movimiento.tipo == tipo_f)
    movimientos = q.order_by(Movimiento.fecha.desc()).all()

    almacen_map = {a.id: a.nombre for a in Almacen.query.filter_by(organizacion_id=org_id).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    h_font  = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    h_fill  = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    b_font  = Font(name='Arial', size=10)
    thin    = Side(border_style='thin', color='DEE2E6')
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    rango_str = f"{fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    ws.merge_cells('A1:H1')
    ws['A1'].value = f"Historial de Movimientos — {rango_str}"
    ws['A1'].font  = Font(name='Arial', size=13, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    headers = ['Fecha', 'Hora', 'Tipo', 'Producto', 'SKU', 'Cantidad', 'Motivo', 'Almacén']
    ws.append(headers)
    for cell in ws[2]:
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
    ws.row_dimensions[2].height = 20

    tipo_labels = {
        'entrada': 'Entrada', 'entrada-inicial': 'Stock Inicial',
        'salida': 'Salida', 'ajuste-entrada': 'Ajuste (+)', 'ajuste-salida': 'Ajuste (-)',
    }

    for mov in movimientos:
        ws.append([
            mov.fecha.strftime('%d/%m/%Y'),
            mov.fecha.strftime('%H:%M'),
            tipo_labels.get(mov.tipo, mov.tipo),
            mov.producto.nombre if mov.producto else '',
            mov.producto.codigo if mov.producto else '',
            mov.cantidad,
            mov.motivo,
            almacen_map.get(mov.almacen_id, ''),
        ])
        r = ws.max_row
        for col in range(1, 9):
            ws.cell(row=r, column=col).font   = b_font
            ws.cell(row=r, column=col).border = border
        qty_cell = ws.cell(row=r, column=6)
        if mov.cantidad > 0:
            qty_cell.font = Font(name='Arial', size=10, color='059669', bold=True)
        elif mov.cantidad < 0:
            qty_cell.font = Font(name='Arial', size=10, color='DC2626', bold=True)

    for i, w in enumerate([12, 8, 16, 32, 15, 10, 40, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"Movimientos_{now_mx().strftime('%Y%m%d')}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True)


@app.route('/reportes/valorizacion.pdf')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def exportar_valorizacion_pdf():
    if current_user.rol not in ['super_admin', 'admin']:
        flash('Solo los administradores pueden exportar reportes de valorización.', 'danger')
        return redirect(url_for('reportes'))

    org_id     = current_user.organizacion_id
    almacen_id = request.args.get('almacen_id', type=int)
    org        = Organizacion.query.get(org_id)

    if almacen_id:
        almacen = Almacen.query.filter_by(id=almacen_id, organizacion_id=org_id).first_or_404()
        items   = db.session.query(Stock).filter_by(almacen_id=almacen_id).join(Producto).order_by(Producto.nombre).all()
        nombre_almacen = almacen.nombre
    else:
        items = db.session.query(Stock).join(
            Almacen, Stock.almacen_id == Almacen.id
        ).filter(Almacen.organizacion_id == org_id).join(Producto).order_by(Producto.nombre).all()
        nombre_almacen = 'Todos los Almacenes'

    items_sorted = sorted(items, key=lambda x: (x.cantidad or 0) * (x.producto.precio_unitario or 0), reverse=True)
    valor_total  = sum((i.cantidad or 0) * (i.producto.precio_unitario or 0) for i in items)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    story  = []
    styles = getSampleStyleSheet()

    fuente, primary, c_sec = _pdf_estilos(org)
    light_gray = colors.HexColor('#F8F9FA')
    mid_gray   = colors.HexColor('#DEE2E6')

    s_title  = ParagraphStyle('RPTitle',  fontName=f'{fuente}-Bold', fontSize=18, textColor=primary, spaceAfter=4)
    s_sub    = ParagraphStyle('RPSub',    fontName=fuente,           fontSize=10, textColor=colors.HexColor('#6B7280'), spaceAfter=2)
    s_cell   = ParagraphStyle('RPCell',   fontName=fuente,           fontSize=8,  leading=10)
    s_cellb  = ParagraphStyle('RPCellB',  fontName=f'{fuente}-Bold', fontSize=8,  leading=10)
    s_cellr  = ParagraphStyle('RPCellR',  fontName=fuente,           fontSize=8,  leading=10, alignment=TA_RIGHT)
    s_cellbr = ParagraphStyle('RPCellBR', fontName=f'{fuente}-Bold', fontSize=8,  leading=10, alignment=TA_RIGHT)
    s_big    = ParagraphStyle('RPBig',    fontName=f'{fuente}-Bold', fontSize=14, textColor=primary)

    _pdf_header(story, org, styles)
    story.append(Paragraph("Reporte de Valorización de Inventario", s_title))
    story.append(Paragraph(f"Almacén: {nombre_almacen}", s_sub))
    story.append(Spacer(1, 0.2*inch))

    # Resumen
    resumen = [
        [Paragraph('<b>Total Productos</b>', s_cellb),
         Paragraph('<b>Valor Total (MXN)</b>', s_cellb),
         Paragraph('<b>Almacén</b>', s_cellb)],
        [Paragraph(str(len(items)), s_big),
         Paragraph(f"$ {valor_total:,.2f}", s_big),
         Paragraph(nombre_almacen, s_cell)],
    ]
    t_res = Table(resumen, colWidths=[2*inch, 2.8*inch, 2.2*inch])
    t_res.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), light_gray),
        ('BOX',        (0,0), (-1,-1), 1,   mid_gray),
        ('INNERGRID',  (0,0), (-1,-1), 0.5, mid_gray),
        ('PADDING',    (0,0), (-1,-1), 8),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t_res)
    story.append(Spacer(1, 0.2*inch))

    # Tabla principal
    col_w = [0.3*inch, 0.85*inch, 2.1*inch, 1.1*inch, 0.5*inch, 0.95*inch, 0.95*inch, 0.65*inch]
    hdrs  = ['#', 'SKU', 'Producto', 'Categoría', 'Stock', 'Precio Unit.', 'Valor Total', '% Total']
    data  = [[Paragraph(h, s_cellb) for h in hdrs]]

    for i, item in enumerate(items_sorted, 1):
        valor_item = (item.cantidad or 0) * (item.producto.precio_unitario or 0)
        pct        = (valor_item / valor_total * 100) if valor_total > 0 else 0
        data.append([
            Paragraph(str(i), s_cellr),
            Paragraph(item.producto.codigo, s_cell),
            Paragraph(item.producto.nombre[:40], s_cell),
            Paragraph(item.producto.categoria.nombre if item.producto.categoria else '—', s_cell),
            Paragraph(str(item.cantidad), s_cellr),
            Paragraph(f"$ {(item.producto.precio_unitario or 0):,.2f}", s_cellr),
            Paragraph(f"$ {valor_item:,.2f}", s_cellr),
            Paragraph(f"{pct:.1f}%", s_cellr),
        ])

    data.append([
        Paragraph('', s_cell), Paragraph('', s_cell), Paragraph('', s_cell), Paragraph('', s_cell),
        Paragraph('', s_cell),
        Paragraph('TOTAL', s_cellbr),
        Paragraph(f"$ {valor_total:,.2f}", s_cellbr),
        Paragraph('100%', s_cellbr),
    ])

    t_main = Table(data, colWidths=col_w, repeatRows=1)
    row_bgs = _pdf_row_styles(len(data) - 1, c_sec)
    t_main.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),  (-1,0),  primary),
        ('TEXTCOLOR',    (0,0),  (-1,0),  colors.white),
        ('BACKGROUND',   (0,-1), (-1,-1), colors.HexColor('#EEEDFC')),
        ('FONTNAME',     (0,-1), (-1,-1), f'{fuente}-Bold'),
        ('GRID',         (0,0),  (-1,-1), 0.5, mid_gray),
        ('BOX',          (0,0),  (-1,-1), 1,   mid_gray),
        ('PADDING',      (0,0),  (-1,-1), 5),
        ('VALIGN',       (0,0),  (-1,-1), 'MIDDLE'),
    ] + row_bgs))
    story.append(t_main)

    _pdf_footer(story, org)
    doc.build(story)
    buf.seek(0)
    fname = f"Valorizacion_{nombre_almacen.replace(' ','_')}_{now_mx().strftime('%Y%m%d')}.pdf"
    return send_file(buf, download_name=fname, mimetype='application/pdf', as_attachment=True)


# --- RUTAS DE HISTORIAL DE ACTIVIDAD ---

@app.route('/actividad')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def historial_actividad():
    """Timeline de actividad reciente de la organización."""
    org_id = current_user.organizacion_id
    page = request.args.get('page', 1, type=int)
    accion_filtro = request.args.get('accion', '')
    per_page = 25

    q = AuditLog.query.filter_by(organizacion_id=org_id)
    if accion_filtro:
        q = q.filter(AuditLog.accion == accion_filtro)
    pagination = q.order_by(AuditLog.fecha.desc()).paginate(page=page, per_page=per_page, error_out=False)

    acciones = db.session.query(AuditLog.accion).filter_by(organizacion_id=org_id).distinct().all()
    acciones = [a[0] for a in acciones]

    return render_template('actividad.html',
                           titulo='Historial de Actividad',
                           pagination=pagination,
                           entradas=pagination.items,
                           acciones=acciones,
                           accion_filtro=accion_filtro)


# --- RUTAS DE CONTROL DE GASTOS ---
@app.route('/gastos')
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def lista_gastos():
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    ahora = now_mx()
    if not mes: mes = ahora.month
    if not ano: ano = ahora.year
    meses_lista = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'), 
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'), 
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    if current_user.rol == 'super_admin':
        query_gastos = Gasto.query
    else:
        query_gastos = Gasto.query.filter_by(organizacion_id=current_user.organizacion_id)

    query_gastos = query_gastos.filter(
        extract('month', Gasto.fecha) == mes,
        extract('year', Gasto.fecha) == ano
    ).order_by(Gasto.fecha.desc())
    
    if current_user.rol == 'super_admin':
        total_gastos = db.session.query(db.func.sum(Gasto.monto)).filter(
            extract('month', Gasto.fecha) == mes,
            extract('year', Gasto.fecha) == ano
        ).scalar() or 0
    else:
        total_gastos = db.session.query(db.func.sum(Gasto.monto)).filter(
            Gasto.organizacion_id == current_user.organizacion_id,
            extract('month', Gasto.fecha) == mes,
            extract('year', Gasto.fecha) == ano
        ).scalar() or 0

    page = request.args.get('page', 1, type=int)
    pagination = query_gastos.paginate(page=page, per_page=15, error_out=False)

    return render_template('gastos.html',
                           gastos=pagination.items,
                           pagination=pagination,
                           total_gastos=total_gastos,
                           mes_seleccionado=mes,
                           ano_seleccionado=ano,
                           meses_lista=meses_lista)

@app.route('/gasto/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def nuevo_gasto():
    org_id = current_user.organizacion_id
    ordenes = OrdenCompra.query.filter_by(organizacion_id=org_id).order_by(OrdenCompra.fecha_creacion.desc()).all()
    centros = CentroCosto.query.filter_by(organizacion_id=org_id).order_by(CentroCosto.nombre).all()

    if request.method == 'POST':
        try:
            monto_val = float(request.form['monto'])
            if monto_val <= 0:
                flash('El monto debe ser mayor a cero.', 'danger')
                return redirect(url_for('nuevo_gasto'))
            categoria_val = request.form['categoria']
            if categoria_val not in CATEGORIAS_GASTO:
                flash('Categoría no válida.', 'danger')
                return redirect(url_for('nuevo_gasto'))
            fecha_gasto = datetime.strptime(request.form['fecha'], '%Y-%m-%d')
            oc_id = request.form.get('orden_compra_id')
            if oc_id == "": oc_id = None
            cc_id = request.form.get('centro_costo_id') or None

            nuevo_gasto = Gasto(
                descripcion=request.form['descripcion'],
                monto=monto_val,
                categoria=categoria_val,
                fecha=fecha_gasto,
                orden_compra_id=oc_id,
                centro_costo_id=cc_id,
                organizacion_id=current_user.organizacion_id
            )
            db.session.add(nuevo_gasto)
            db.session.flush()
            log_actividad('crear', 'gasto', f'Gasto registrado: {nuevo_gasto.descripcion} — ${nuevo_gasto.monto:,.2f} ({nuevo_gasto.categoria})', entidad_id=nuevo_gasto.id)
            db.session.commit()
            flash('Gasto registrado exitosamente', 'success')
            return redirect(url_for('lista_gastos'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al registrar el gasto.', e)

    return render_template('gasto_form.html',
                           titulo="Registrar Nuevo Gasto",
                           ordenes=ordenes,
                           centros=centros,
                           now=now_mx())

@app.route('/gasto/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@check_permission('perm_view_gastos')
def editar_gasto(id):
    """ Edita un gasto existente. """
    gasto = get_item_or_404(Gasto, id)
    org_id = current_user.organizacion_id
    ordenes = OrdenCompra.query.filter_by(organizacion_id=org_id).order_by(OrdenCompra.fecha_creacion.desc()).all()
    centros = CentroCosto.query.filter_by(organizacion_id=org_id).order_by(CentroCosto.nombre).all()

    if request.method == 'POST':
        try:
            monto_val = float(request.form['monto'])
            if monto_val <= 0:
                flash('El monto debe ser mayor a cero.', 'danger')
                return redirect(url_for('editar_gasto', id=id))
            categoria_val = request.form['categoria']
            if categoria_val not in CATEGORIAS_GASTO:
                flash('Categoría no válida.', 'danger')
                return redirect(url_for('editar_gasto', id=id))
            fecha_gasto = datetime.strptime(request.form['fecha'], '%Y-%m-%d')
            oc_id = request.form.get('orden_compra_id')
            if oc_id == "" or oc_id == "None":
                oc_id = None
            cc_id = request.form.get('centro_costo_id') or None

            monto_anterior = gasto.monto
            gasto.descripcion = request.form['descripcion']
            gasto.monto = monto_val
            gasto.categoria = categoria_val
            gasto.fecha = fecha_gasto
            gasto.orden_compra_id = oc_id
            gasto.centro_costo_id = cc_id

            log_actividad('editar', 'gasto',
                f'Gasto editado: {gasto.descripcion} — antes ${monto_anterior:,.2f} → ahora ${gasto.monto:,.2f} ({gasto.categoria})',
                entidad_id=gasto.id)
            db.session.commit()
            flash('Gasto actualizado exitosamente', 'success')
            return redirect(url_for('lista_gastos'))

        except Exception as e:
            db.session.rollback()
            _flash_err('Error al actualizar el gasto.', e)

    return render_template('gasto_form.html',
                           titulo="Editar Gasto",
                           ordenes=ordenes,
                           centros=centros,
                           gasto=gasto)

@app.route('/gastos/exportar_excel')
@login_required
@check_permission('perm_view_gastos')
def exportar_gastos_excel():
    ahora = now_mx()
    org   = Organizacion.query.get(current_user.organizacion_id)

    # ── Rango de meses ────────────────────────────────────────────────────────
    mes_desde = request.args.get('mes_desde', type=int)
    ano_desde = request.args.get('ano_desde', type=int)
    mes_hasta = request.args.get('mes_hasta', type=int)
    ano_hasta = request.args.get('ano_hasta', type=int)
    if not mes_desde:
        mes_desde = mes_hasta = request.args.get('mes', type=int) or ahora.month
        ano_desde = ano_hasta = request.args.get('ano', type=int) or ahora.year

    periodos = []
    y, m = ano_desde, mes_desde
    while (y < ano_hasta) or (y == ano_hasta and m <= mes_hasta):
        periodos.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    if not periodos:
        periodos = [(ano_desde, mes_desde)]

    base_query = (Gasto.query if current_user.rol == 'super_admin'
                  else Gasto.query.filter_by(organizacion_id=current_user.organizacion_id))

    # ── Configuración de diseño ───────────────────────────────────────────────
    def _argb(h):
        return 'FF' + (h or '#000000').lstrip('#').upper()

    col_hdr     = _argb(getattr(org, 'excel_color_header',  '#1f4e79'))
    col_acc     = _argb(getattr(org, 'excel_color_accent',  '#dbeafe'))
    fuente      = getattr(org, 'excel_fuente',         'Calibri') or 'Calibri'
    show_logo   = getattr(org, 'excel_mostrar_logo',   True)
    show_id     = getattr(org, 'excel_mostrar_id',     True)
    show_oc     = getattr(org, 'excel_mostrar_oc',     True)
    show_origen = getattr(org, 'excel_mostrar_origen', True)
    empresa     = (org.header_titulo or org.nombre) if org else 'Empresa'

    # ── Columnas dinámicas ────────────────────────────────────────────────────
    COLS = ['Fecha', 'Descripción', 'Categoría', 'Monto']
    if show_id:     COLS = ['ID'] + COLS
    if show_oc:     COLS.append('OC Asociada')
    if show_origen: COLS.append('Origen')
    N         = len(COLS)
    monto_idx = COLS.index('Monto') + 1   # 1-based
    last_col  = get_column_letter(N)

    # ── Estilos ───────────────────────────────────────────────────────────────
    fill_hdr    = PatternFill('solid', fgColor=col_hdr)
    fill_acc    = PatternFill('solid', fgColor=col_acc)
    fill_svc    = PatternFill('solid', fgColor='FFFFF8E1')  # ámbar muy suave

    bd_s  = Side(border_style='thin',   color='CCCCCC')
    bd_m  = Side(border_style='medium', color='888888')
    bd    = Border(left=bd_s,  right=bd_s,  top=bd_s,  bottom=bd_s)
    bd_tt = Border(left=bd_m,  right=bd_m,  top=bd_m,  bottom=bd_m)

    f_title  = Font(name=fuente, size=14, bold=True,  color='FFFFFF')
    f_sub    = Font(name=fuente, size=10,             color='FFFFFF')
    f_hdr    = Font(name=fuente, size=10, bold=True,  color='FFFFFF')
    f_normal = Font(name=fuente, size=10)
    f_bold   = Font(name=fuente, size=11, bold=True)
    f_wht    = Font(name=fuente, size=11, bold=True,  color='FFFFFF')

    a_c = Alignment(horizontal='center', vertical='center')
    a_r = Alignment(horizontal='right',  vertical='center')
    a_l = Alignment(horizontal='left',   vertical='center')
    cur_fmt = '$#,##0.00'

    # ── Logo path ─────────────────────────────────────────────────────────────
    logo_path = None
    if show_logo and org and org.logo_url:
        candidate = os.path.join(app.config['UPLOAD_FOLDER'], org.logo_url)
        if os.path.exists(candidate):
            logo_path = candidate

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _auto_width(ws, max_w=52):
        for ci, col in enumerate(ws.columns, 1):
            w = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, max_w)

    def _banner(ws, title, subtitle=''):
        ws.merge_cells(f'A1:{last_col}1')
        c = ws['A1']
        c.value, c.font, c.fill, c.alignment = title, f_title, fill_hdr, a_c
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f'A2:{last_col}2')
        c2 = ws['A2']
        c2.value, c2.font, c2.fill, c2.alignment = subtitle, f_sub, fill_hdr, a_c
        ws.row_dimensions[2].height = 16
        if logo_path:
            try:
                img = XlImage(logo_path)
                img.height, img.width = 42, 42
                ws.add_image(img, 'A1')
            except Exception:
                pass

    def _col_headers(ws, row, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font, c.fill, c.alignment, c.border = f_hdr, fill_hdr, a_c, bd
        ws.row_dimensions[row].height = 18

    def _cat_summary(ws, gastos_list):
        """Tabla de resumen por categoría al pie de cada hoja."""
        cat_totals = {}
        for g in gastos_list:
            k = g.categoria or 'Sin categoría'
            cat_totals[k] = cat_totals.get(k, 0) + g.monto
        if not cat_totals:
            return
        ws.append([])
        sr = ws.max_row + 1
        ws.merge_cells(f'A{sr}:B{sr}')
        sh = ws[f'A{sr}']
        sh.value, sh.font, sh.fill, sh.alignment = 'Resumen por Categoría', f_hdr, fill_hdr, a_c
        ws.row_dimensions[sr].height = 16
        hr = sr + 1
        for ci, txt in enumerate(['Categoría', 'Total'], 1):
            c = ws.cell(hr, ci, txt)
            c.font, c.fill, c.alignment, c.border = f_hdr, fill_hdr, (a_l if ci == 1 else a_r), bd
        for cat, tot in sorted(cat_totals.items(), key=lambda x: -x[1]):
            r = ws.max_row + 1
            c1 = ws.cell(r, 1, cat)
            c1.font, c1.border, c1.alignment = f_normal, bd, a_l
            c2 = ws.cell(r, 2, tot)
            c2.font, c2.border, c2.alignment = f_normal, bd, a_r
            c2.number_format = cur_fmt

    def _add_month_sheet(wb, year, month, table_idx):
        gastos = base_query.filter(
            extract('month', Gasto.fecha) == month,
            extract('year',  Gasto.fecha) == year
        ).order_by(Gasto.fecha.asc()).all()

        nombre_mes = datetime(year, month, 1).strftime('%B').capitalize()
        ws = wb.create_sheet(title=f"{nombre_mes[:3]} {year}")

        _banner(ws, empresa, f'Control de Gastos — {nombre_mes} {year}')
        _col_headers(ws, 3, COLS)

        total = 0.0
        for i, g in enumerate(gastos):
            origen = 'Servicio' if g.descripcion.startswith('Servicio:') else 'Manual'
            row_data = [g.fecha.date(), g.descripcion, g.categoria or '—', g.monto]
            if show_id:     row_data = [g.id] + row_data
            if show_oc:     row_data.append(g.orden_compra_id or '—')
            if show_origen: row_data.append(origen)
            ws.append(row_data)
            r = ws.max_row
            use_acc = (i % 2 == 1)
            for ci in range(1, N + 1):
                c = ws.cell(r, ci)
                c.font, c.border = f_normal, bd
                if origen == 'Servicio':
                    c.fill = fill_svc
                elif use_acc:
                    c.fill = fill_acc
            ws.cell(r, monto_idx).number_format = cur_fmt
            ws.cell(r, monto_idx).alignment     = a_r
            total += g.monto

        if gastos:
            try:
                tbl = ExcelTable(displayName=f'Gastos{table_idx}',
                                 ref=f'A3:{last_col}{ws.max_row}')
                tbl.tableStyleInfo = TableStyleInfo(
                    name='TableStyleMedium2', showRowStripes=False)
                ws.add_table(tbl)
            except Exception:
                pass

        # Fila total
        ft = ws.max_row + 1
        pre = get_column_letter(monto_idx - 1)
        ws.merge_cells(f'A{ft}:{pre}{ft}')
        c_lbl = ws.cell(ft, 1, 'Total del Mes')
        c_lbl.font, c_lbl.fill, c_lbl.alignment, c_lbl.border = f_bold, fill_acc, a_r, bd
        c_tot = ws.cell(ft, monto_idx, total)
        c_tot.number_format = cur_fmt
        c_tot.font, c_tot.fill, c_tot.alignment, c_tot.border = f_bold, fill_acc, a_r, bd

        _cat_summary(ws, gastos)
        ws.freeze_panes = 'A4'
        _auto_width(ws)
        return total, len(gastos)

    # ── Construir workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_res = None
    if len(periodos) > 1:
        ws_res = wb.create_sheet(title='Resumen', index=0)
        nm1 = datetime(periodos[0][0],  periodos[0][1],  1).strftime('%B').capitalize()
        nm2 = datetime(periodos[-1][0], periodos[-1][1], 1).strftime('%B').capitalize()
        _banner(ws_res, empresa,
                f'Auditoría de Gastos — {nm1} {periodos[0][0]} a {nm2} {periodos[-1][0]}')
        _col_headers(ws_res, 3, ['Período', 'Registros', 'Total (MXN)', 'Con Servicios'])

    totals_resumen = []
    for idx, (year, month) in enumerate(periodos, 1):
        total_mes, count_mes = _add_month_sheet(wb, year, month, idx)
        totals_resumen.append((year, month, total_mes, count_mes))

    if ws_res is not None:
        gran_total  = 0.0
        all_g_range = []
        for year, month, total_mes, count_mes in totals_resumen:
            nombre_mes = datetime(year, month, 1).strftime('%B').capitalize()
            gastos_mes = base_query.filter(
                extract('month', Gasto.fecha) == month,
                extract('year',  Gasto.fecha) == year).all()
            all_g_range.extend(gastos_mes)
            tiene_svc = any(g.descripcion.startswith('Servicio:') for g in gastos_mes)
            ws_res.append([f'{nombre_mes} {year}', count_mes, total_mes,
                           'Sí' if tiene_svc else 'No'])
            r = ws_res.max_row
            ws_res.cell(r, 3).number_format = cur_fmt
            ws_res.cell(r, 3).alignment     = a_r
            for ci in range(1, 5):
                ws_res.cell(r, ci).font   = f_normal
                ws_res.cell(r, ci).border = bd
            gran_total += total_mes

        data_end = ws_res.max_row
        gt = data_end + 1
        ws_res.cell(gt, 1, 'GRAN TOTAL').font      = f_wht
        ws_res.cell(gt, 1).fill, ws_res.cell(gt, 1).alignment = fill_hdr, a_r
        ws_res.cell(gt, 1).border = bd
        ws_res.cell(gt, 2, sum(c for _, _, _, c in totals_resumen)).font = f_wht
        ws_res.cell(gt, 2).fill,  ws_res.cell(gt, 2).border              = fill_hdr, bd
        ws_res.cell(gt, 2).alignment = a_c
        ws_res.cell(gt, 3, gran_total).number_format = cur_fmt
        ws_res.cell(gt, 3).font, ws_res.cell(gt, 3).fill   = f_wht, fill_hdr
        ws_res.cell(gt, 3).alignment, ws_res.cell(gt, 3).border = a_r, bd

        # Tabla T de categorías del período completo
        _cat_summary(ws_res, all_g_range)

        # Gráfico de barras por mes
        try:
            chart = BarChart()
            chart.type, chart.grouping = 'col', 'clustered'
            chart.title   = 'Gastos por Mes'
            chart.y_axis.title = 'Total (MXN)'
            chart.x_axis.title = 'Período'
            data_ref = Reference(ws_res, min_col=3, max_col=3,
                                 min_row=3, max_row=data_end)
            cats_ref = Reference(ws_res, min_col=1, max_col=1,
                                 min_row=4, max_row=data_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width, chart.height = 16, 10
            ws_res.add_chart(chart, 'F3')
        except Exception:
            pass

        ws_res.freeze_panes = 'A4'
        _auto_width(ws_res)

    # ── Nombre de archivo ─────────────────────────────────────────────────────
    if len(periodos) == 1:
        nom = datetime(periodos[0][0], periodos[0][1], 1).strftime('%B').capitalize()
        filename = f"Gastos_{nom}_{periodos[0][0]}.xlsx"
    else:
        n1 = datetime(periodos[0][0],  periodos[0][1],  1).strftime('%b').capitalize()
        n2 = datetime(periodos[-1][0], periodos[-1][1], 1).strftime('%b').capitalize()
        filename = f"Gastos_{n1}{periodos[0][0]}_a_{n2}{periodos[-1][0]}.xlsx"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

# ──────────────────────────────────────────────
# CENTROS DE COSTO — FASE C ERP
# ──────────────────────────────────────────────

@app.route('/centros-costo')
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def lista_centros_costo():
    org_id = current_user.organizacion_id
    centros = CentroCosto.query.filter_by(organizacion_id=org_id).order_by(CentroCosto.creado_en.desc()).all()
    return render_template('centros_costo_lista.html',
        titulo='Centros de Costo', centros=centros, now=now_mx())


@app.route('/centro-costo/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def nuevo_centro_costo():
    if request.method == 'POST':
        try:
            cc = CentroCosto(
                nombre          = request.form['nombre'].strip(),
                descripcion     = request.form.get('descripcion', '').strip() or None,
                presupuesto     = float(request.form['presupuesto']) if request.form.get('presupuesto') else None,
                organizacion_id = current_user.organizacion_id,
                creador_id      = current_user.id,
            )
            db.session.add(cc)
            db.session.commit()
            flash('Centro de costo creado.', 'success')
            return redirect(url_for('lista_centros_costo'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
    return render_template('centro_costo_form.html', titulo='Nuevo Centro de Costo', centro=None, now=now_mx())


@app.route('/centro-costo/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def editar_centro_costo(id):
    org_id = current_user.organizacion_id
    cc = CentroCosto.query.filter_by(id=id, organizacion_id=org_id).first_or_404()
    if request.method == 'POST':
        try:
            cc.nombre       = request.form['nombre'].strip()
            cc.descripcion  = request.form.get('descripcion', '').strip() or None
            cc.presupuesto  = float(request.form['presupuesto']) if request.form.get('presupuesto') else None
            db.session.commit()
            flash('Centro de costo actualizado.', 'success')
            return redirect(url_for('detalle_centro_costo', id=cc.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
    return render_template('centro_costo_form.html', titulo='Editar Centro de Costo', centro=cc, now=now_mx())


@app.route('/centro-costo/<int:id>/cerrar', methods=['POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def cerrar_centro_costo(id):
    org_id = current_user.organizacion_id
    cc = CentroCosto.query.filter_by(id=id, organizacion_id=org_id).first_or_404()
    cc.estado = 'cerrado' if cc.estado == 'activo' else 'activo'
    db.session.commit()
    flash(f'Centro de costo {"cerrado" if cc.estado == "cerrado" else "reactivado"}.', 'success')
    return redirect(url_for('detalle_centro_costo', id=cc.id))


@app.route('/centro-costo/<int:id>')
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def detalle_centro_costo(id):
    org_id = current_user.organizacion_id
    cc = CentroCosto.query.filter_by(id=id, organizacion_id=org_id).first_or_404()

    # Desglose por tipo
    total_gastos    = sum(g.monto for g in cc.gastos)
    total_servicios = sum(p.monto for p in cc.pagos_servicio if p.estado == 'pagado')
    total_facturas  = sum(f.monto for f in cc.facturas)
    total           = total_gastos + total_servicios + total_facturas

    # Gastos por categoría (horizontal bar)
    cat_map = {}
    for g in cc.gastos:
        k = g.categoria or 'Sin categoría'
        cat_map[k] = cat_map.get(k, 0) + g.monto
    cat_items = sorted(cat_map.items(), key=lambda x: -x[1])

    # Feed unificado de transacciones
    txs = []
    for g in cc.gastos:
        txs.append({'fecha': g.fecha.date() if hasattr(g.fecha,'date') else g.fecha,
                    'tipo':'Gasto','desc':g.descripcion,'cat':g.categoria or '—',
                    'monto':g.monto,'icon':'bi-cash-coin','cls':'badge-borrador'})
    for p in cc.pagos_servicio:
        txs.append({'fecha': p.fecha_pago,
                    'tipo':'Servicio','desc':p.servicio.nombre,'cat':p.servicio.tipo or '—',
                    'monto':p.monto,'icon':'bi-lightning-charge-fill','cls':'badge-enviada'})
    for f in cc.facturas:
        txs.append({'fecha': f.fecha_emision,
                    'tipo':'Factura','desc':f.numero_factura + ' — ' + f.proveedor.nombre,'cat':'Proveedor',
                    'monto':f.monto,'icon':'bi-file-earmark-text','cls':'badge-recibida'})
    txs.sort(key=lambda x: x['fecha'] if x['fecha'] else date.min, reverse=True)

    return render_template('centro_costo_detalle.html',
        titulo=cc.nombre, cc=cc,
        total=total, total_gastos=total_gastos,
        total_servicios=total_servicios, total_facturas=total_facturas,
        cat_items=cat_items, txs=txs, now=now_mx())


# ──────────────────────────────────────────────
# FASE D — PRESUPUESTOS POR CATEGORÍA
# ──────────────────────────────────────────────

def _semaforo(pct):
    """Devuelve (clase_bootstrap, etiqueta) según porcentaje gastado."""
    if pct >= 90:
        return 'danger', 'Crítico'
    if pct >= 70:
        return 'warning', 'Alerta'
    return 'success', 'OK'

def _real_por_categoria(org_id, categoria, anio, mes):
    """Suma de gastos reales de una categoría en un período dado."""
    q = Gasto.query.filter_by(organizacion_id=org_id, categoria=categoria)
    q = q.filter(extract('year', Gasto.fecha) == anio)
    if mes:
        q = q.filter(extract('month', Gasto.fecha) == mes)
    return sum(g.monto for g in q.all())

@app.route('/presupuestos')
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def lista_presupuestos():
    org_id = current_user.organizacion_id
    ahora  = now_mx()

    anio = request.args.get('anio', ahora.year, type=int)
    mes  = request.args.get('mes',  0,           type=int)   # 0 = anual

    q = Presupuesto.query.filter_by(organizacion_id=org_id, anio=anio)
    if mes:
        presupuestos = q.filter_by(mes=mes).order_by(Presupuesto.categoria).all()
    else:
        presupuestos = q.filter(Presupuesto.mes.is_(None)).order_by(Presupuesto.categoria).all()

    items = []
    for p in presupuestos:
        gastado = _real_por_categoria(org_id, p.categoria, anio, mes or None)
        pct     = min(round(gastado / p.monto * 100, 1), 999) if p.monto > 0 else 0
        cls, lbl = _semaforo(pct)
        items.append({
            'p': p,
            'gastado': gastado,
            'disponible': p.monto - gastado,
            'pct': pct,
            'pct_bar': min(pct, 100),
            'cls': cls,
            'label': lbl,
        })

    # Categorías que aún no tienen presupuesto en este período
    cats_con = {i['p'].categoria for i in items}
    cats_sin = [c for c in CATEGORIAS_GASTO if c not in cats_con]

    total_presupuestado = sum(i['p'].monto for i in items)
    total_gastado_real  = sum(i['gastado'] for i in items)
    en_riesgo           = sum(1 for i in items if i['cls'] == 'danger')
    pct_global          = min(round(total_gastado_real / total_presupuestado * 100, 1), 999) if total_presupuestado else 0

    anios_disponibles = list(range(ahora.year - 1, ahora.year + 3))

    return render_template('presupuestos_lista.html',
        titulo='Presupuestos',
        items=items,
        cats_sin=cats_sin,
        total_presupuestado=total_presupuestado,
        total_gastado=total_gastado_real,
        en_riesgo=en_riesgo,
        pct_global=pct_global,
        anio=anio, mes=mes,
        anios_disponibles=anios_disponibles,
        meses_es=MESES_ES,
        now=ahora,
    )


@app.route('/presupuesto/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def nuevo_presupuesto():
    org_id = current_user.organizacion_id
    ahora  = now_mx()

    if request.method == 'POST':
        categoria = request.form['categoria']
        anio      = int(request.form['anio'])
        mes_raw   = request.form.get('mes', '')
        mes       = int(mes_raw) if mes_raw else None
        monto     = float(request.form['monto'])

        existente = Presupuesto.query.filter_by(
            organizacion_id=org_id, categoria=categoria, anio=anio, mes=mes
        ).first()
        if existente:
            flash(f'Ya existe un presupuesto para {categoria} en ese período.', 'warning')
        else:
            p = Presupuesto(categoria=categoria, anio=anio, mes=mes,
                            monto=monto, organizacion_id=org_id)
            db.session.add(p)
            db.session.commit()
            flash('Presupuesto creado.', 'success')
            return redirect(url_for('lista_presupuestos',
                                    anio=anio, mes=mes or 0))

    return render_template('presupuesto_form.html',
        titulo='Nuevo Presupuesto',
        presupuesto=None,
        categorias=CATEGORIAS_GASTO,
        meses_es=MESES_ES,
        anio_actual=ahora.year,
        now=ahora,
    )


@app.route('/presupuesto/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def editar_presupuesto(id):
    org_id = current_user.organizacion_id
    p = Presupuesto.query.filter_by(id=id, organizacion_id=org_id).first_or_404()

    if request.method == 'POST':
        p.monto = float(request.form['monto'])
        db.session.commit()
        flash('Presupuesto actualizado.', 'success')
        return redirect(url_for('lista_presupuestos',
                                anio=p.anio, mes=p.mes or 0))

    return render_template('presupuesto_form.html',
        titulo='Editar Presupuesto',
        presupuesto=p,
        categorias=CATEGORIAS_GASTO,
        meses_es=MESES_ES,
        anio_actual=p.anio,
        now=now_mx(),
    )


@app.route('/presupuesto/<int:id>/eliminar', methods=['POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def eliminar_presupuesto(id):
    org_id = current_user.organizacion_id
    p = Presupuesto.query.filter_by(id=id, organizacion_id=org_id).first_or_404()
    anio, mes = p.anio, p.mes or 0
    db.session.delete(p)
    db.session.commit()
    flash('Presupuesto eliminado.', 'success')
    return redirect(url_for('lista_presupuestos', anio=anio, mes=mes))


# ──────────────────────────────────────────────
# CUENTAS POR PAGAR — FACTURAS DE PROVEEDORES
# ──────────────────────────────────────────────

@app.route('/cuentas-por-pagar')
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def lista_facturas():
    org_id = current_user.organizacion_id
    ahora  = now_mx()

    # Actualizar estado a 'vencido' automáticamente
    vencidas = FacturaProveedor.query.filter_by(organizacion_id=org_id, estado='pendiente').filter(
        FacturaProveedor.fecha_vencimiento < ahora.date()
    ).all()
    for f in vencidas:
        f.estado = 'vencido'
    if vencidas:
        db.session.commit()

    estado_filtro = request.args.get('estado', '')
    proveedor_filtro = request.args.get('proveedor_id', type=int, default=0)

    q = FacturaProveedor.query.filter_by(organizacion_id=org_id)
    if estado_filtro:
        q = q.filter_by(estado=estado_filtro)
    if proveedor_filtro:
        q = q.filter_by(proveedor_id=proveedor_filtro)
    facturas = q.order_by(FacturaProveedor.fecha_vencimiento.asc()).all()

    # KPIs
    todas = FacturaProveedor.query.filter_by(organizacion_id=org_id).all()
    total_pendiente = sum(f.monto for f in todas if f.estado in ('pendiente', 'vencido'))
    total_vencido   = sum(f.monto for f in todas if f.estado == 'vencido')
    total_pagado_mes = sum(
        f.monto for f in todas
        if f.estado == 'pagado'
        and f.fecha_vencimiento.month == ahora.month
        and f.fecha_vencimiento.year  == ahora.year
    )

    # Aging buckets
    aging = {'vigente': 0.0, '1-30': 0.0, '31-60': 0.0, '61-90': 0.0, '90+': 0.0}
    for f in todas:
        if f.estado != 'pagado':
            aging[f.bucket_aging] = aging.get(f.bucket_aging, 0.0) + f.monto

    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).order_by(Proveedor.nombre).all()

    return render_template('facturas_lista.html',
        titulo='Cuentas por Pagar',
        facturas=facturas,
        total_pendiente=total_pendiente,
        total_vencido=total_vencido,
        total_pagado_mes=total_pagado_mes,
        aging=aging,
        proveedores=proveedores,
        estado_filtro=estado_filtro,
        proveedor_filtro=proveedor_filtro,
        ahora=ahora,
        now=ahora,
    )


@app.route('/factura/nueva', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def nueva_factura():
    org_id = current_user.organizacion_id
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).order_by(Proveedor.nombre).all()
    ordenes = OrdenCompra.query.filter_by(organizacion_id=org_id).order_by(OrdenCompra.fecha_creacion.desc()).all()
    centros = CentroCosto.query.filter_by(organizacion_id=org_id).order_by(CentroCosto.nombre).all()

    if request.method == 'POST':
        try:
            monto_val = float(request.form['monto'])
            if monto_val <= 0:
                flash('El monto debe ser mayor a cero.', 'danger')
                return redirect(url_for('nueva_factura'))
            factura = FacturaProveedor(
                numero_factura    = request.form['numero_factura'].strip(),
                proveedor_id      = int(request.form['proveedor_id']),
                orden_compra_id   = int(request.form['orden_compra_id']) if request.form.get('orden_compra_id') else None,
                centro_costo_id   = int(request.form['centro_costo_id']) if request.form.get('centro_costo_id') else None,
                monto             = monto_val,
                fecha_emision     = date.fromisoformat(request.form['fecha_emision']),
                fecha_vencimiento = date.fromisoformat(request.form['fecha_vencimiento']),
                notas             = request.form.get('notas', '').strip() or None,
                registrado_por_id = current_user.id,
                organizacion_id   = org_id,
            )
            db.session.add(factura)
            db.session.flush()
            log_actividad('crear', 'factura',
                f'Factura registrada: #{factura.numero_factura} — ${factura.monto:,.2f}',
                entidad_id=factura.id)
            db.session.commit()
            flash('Factura registrada correctamente.', 'success')
            return redirect(url_for('lista_facturas'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al registrar la factura.', e)

    return render_template('factura_form.html',
        titulo='Nueva Factura',
        factura=None,
        proveedores=proveedores,
        ordenes=ordenes,
        centros=centros,
        now=now_mx(),
    )


@app.route('/factura/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def editar_factura(id):
    org_id  = current_user.organizacion_id
    factura = FacturaProveedor.query.filter_by(id=id, organizacion_id=org_id).first_or_404()
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).order_by(Proveedor.nombre).all()
    ordenes     = OrdenCompra.query.filter_by(organizacion_id=org_id).order_by(OrdenCompra.fecha_creacion.desc()).all()
    centros     = CentroCosto.query.filter_by(organizacion_id=org_id).order_by(CentroCosto.nombre).all()

    if request.method == 'POST':
        try:
            monto_val = float(request.form['monto'])
            if monto_val <= 0:
                flash('El monto debe ser mayor a cero.', 'danger')
                return redirect(url_for('editar_factura', id=id))
            monto_anterior = factura.monto
            factura.numero_factura    = request.form['numero_factura'].strip()
            factura.proveedor_id      = int(request.form['proveedor_id'])
            factura.orden_compra_id   = int(request.form['orden_compra_id']) if request.form.get('orden_compra_id') else None
            factura.centro_costo_id   = int(request.form['centro_costo_id']) if request.form.get('centro_costo_id') else None
            factura.monto             = monto_val
            factura.fecha_emision     = date.fromisoformat(request.form['fecha_emision'])
            factura.fecha_vencimiento = date.fromisoformat(request.form['fecha_vencimiento'])
            factura.notas             = request.form.get('notas', '').strip() or None
            log_actividad('editar', 'factura',
                f'Factura editada: #{factura.numero_factura} — antes ${monto_anterior:,.2f} → ahora ${factura.monto:,.2f}',
                entidad_id=factura.id)
            db.session.commit()
            flash('Factura actualizada.', 'success')
            return redirect(url_for('lista_facturas'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al actualizar la factura.', e)

    return render_template('factura_form.html',
        titulo='Editar Factura',
        factura=factura,
        proveedores=proveedores,
        ordenes=ordenes,
        centros=centros,
        now=now_mx(),
    )


@app.route('/factura/<int:id>/marcar-pagada', methods=['POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def marcar_factura_pagada(id):
    org_id  = current_user.organizacion_id
    factura = FacturaProveedor.query.filter_by(id=id, organizacion_id=org_id).first_or_404()
    factura.estado = 'pagado'
    log_actividad('pagar', 'factura',
        f'Factura marcada como pagada: #{factura.numero_factura} — ${factura.monto:,.2f}',
        entidad_id=factura.id)
    db.session.commit()
    flash(f'Factura #{factura.numero_factura} marcada como pagada.', 'success')
    return redirect(url_for('lista_facturas'))


@app.route('/factura/<int:id>/eliminar', methods=['POST'])
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def eliminar_factura(id):
    org_id  = current_user.organizacion_id
    factura = FacturaProveedor.query.filter_by(id=id, organizacion_id=org_id).first_or_404()
    try:
        log_actividad('eliminar', 'factura',
            f'Factura eliminada: #{factura.numero_factura} — ${factura.monto:,.2f} (estado: {factura.estado})',
            entidad_id=factura.id)
        db.session.delete(factura)
        db.session.commit()
        flash('Factura eliminada.', 'success')
    except Exception as e:
        db.session.rollback()
        _flash_err('Error al eliminar la factura.', e)
    return redirect(url_for('lista_facturas'))


# --- RUTAS DE AUTENTICACIÓN ---

@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("10 per minute; 30 per hour")
def register():
    """ Página de Registro de nuevos usuarios (MODIFICADA para códigos). """
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        
        # --- LÓGICA DE CÓDIGO DE INVITACIÓN AÑADIDA ---
        org_id_asignada = None
        rol_asignado = 'user' # Por defecto es 'user'
        
        codigo = form.codigo_invitacion.data
        if codigo:
            # Si el usuario escribió un código, lo buscamos
            org = Organizacion.query.filter_by(codigo_invitacion=codigo.upper()).first()
            
            if not org:
                # El código es inválido, detenemos el registro y mostramos error
                flash('El código de invitación no es válido.', 'danger')
                return render_template('register.html', titulo="Registro", form=form)
            else:
                # ¡Código válido! Asignamos la organización y el rol
                org_id_asignada = org.id
                rol_asignado = 'user' # Los usuarios que se unen por código son 'user'
        # --- FIN DE LÓGICA AÑADIDA ---
        
        try:
            new_user = User(
                username=form.username.data,
                email=form.email.data,
                organizacion_id=org_id_asignada, # <-- MODIFICADO
                rol=rol_asignado                 # <-- MODIFICADO
            )
            new_user.set_password(form.password.data)
            
            db.session.add(new_user)
            db.session.commit()
            
            if org_id_asignada:
                flash(f'¡Cuenta creada! Has sido añadido a la organización {org.nombre}.', 'success')
            else:
                flash(f'¡Cuenta creada! Pide a un Super Admin que te asigne a una organización.', 'success')
            
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al crear la cuenta. Intenta de nuevo.', e)

    return render_template('register.html', titulo="Registro", form=form)

@app.route('/account/delete_picture', methods=['POST'])
@login_required
def delete_picture():
    """ Elimina la foto de perfil del usuario y la revierte a 'default.jpg'. """
    if current_user.image_file != 'default.jpg':
        try:
            picture_path = os.path.join(app.root_path, 'static/uploads/profile_pics', current_user.image_file)
            if os.path.exists(picture_path):
                os.remove(picture_path)
            current_user.image_file = 'default.jpg'
            db.session.commit()
            flash('Tu foto de perfil ha sido eliminada.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al eliminar la foto: {e}', 'danger')
    
    return redirect(url_for('account'))

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute; 50 per hour")
def login():
    """ Página de Inicio de Sesión. """
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and user.check_password(form.password.data):
            login_user(user)
            next_page = request.args.get('next') 
            flash('Inicio de sesión exitoso.', 'success')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        else:
            flash('Inicio de sesión fallido. Verifica tu usuario y contraseña.', 'danger')
            
    return render_template('login.html', titulo="Inicio de Sesión", form=form)

@app.route('/logout')
@login_required
def logout():
    """ Cierra la sesión del usuario. """
    logout_user()
    flash('Has cerrado la sesión.', 'info')
    return redirect(url_for('login'))

@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit("5 per minute; 20 per hour")
def forgot_password():
    """ Página para solicitar el reseteo de contraseña. """
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = RequestResetForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user:
            send_reset_email(user)
        # Siempre mostramos el mismo mensaje para no revelar si un email está registrado (Buena práctica de seguridad)
        flash('Si existe una cuenta con ese e-mail, recibirás un correo con las instrucciones.', 'info')
        return redirect(url_for('login'))
        
    return render_template('forgot_password.html', titulo="Recuperar Contraseña", form=form)

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    """ Página para ingresar la nueva contraseña (accedida desde el e-mail). """
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    _s = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])

    try:
        email = _s.loads(token, salt='password-reset-salt', max_age=1800)
    except Exception:
        flash('El enlace de reseteo no es válido o ha expirado.', 'danger')
        return redirect(url_for('forgot_password'))

    # AUTH-02: rechazar tokens ya utilizados (single-use)
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    ya_usado = TokenUsado.query.filter_by(token_hash=token_hash).first()
    if ya_usado:
        flash('Este enlace de reseteo ya fue utilizado. Solicita uno nuevo.', 'danger')
        return redirect(url_for('forgot_password'))

    user = User.query.filter_by(email=email).first()
    if user is None:
        flash('Usuario no encontrado.', 'danger')
        return redirect(url_for('login'))

    form = ResetPasswordForm()
    if form.validate_on_submit():
        try:
            user.password_hash = generate_password_hash(form.password.data)
            # Marcar token como usado antes de commit para que sean atómicos
            expira_en = datetime.utcnow() + timedelta(seconds=1800)
            db.session.add(TokenUsado(token_hash=token_hash, expira_en=expira_en))
            db.session.commit()
            flash('¡Tu contraseña ha sido actualizada! Ya puedes iniciar sesión.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al actualizar la contraseña.', e)

    return render_template('reset_password.html', titulo="Restablecer Contraseña", form=form, token=token)
  
@app.route('/account', methods=['GET', 'POST'])
@login_required
def account():
    """ Página de configuración de la cuenta del usuario. """
    form_account = UpdateAccountForm()
    form_password = ChangePasswordForm()

    if form_account.submit_account.data and form_account.validate_on_submit():
        try:
            if form_account.picture.data:
                if current_user.image_file != 'default.jpg':
                    old_pic_path = os.path.join(app.root_path, 'static/uploads/profile_pics', current_user.image_file)
                    if os.path.exists(old_pic_path):
                        os.remove(old_pic_path)
                picture_file = save_picture(form_account.picture.data)
                current_user.image_file = picture_file
            
            current_user.username = form_account.username.data
            current_user.email = form_account.email.data
            db.session.commit()
            flash('¡Tu cuenta ha sido actualizada!', 'success')
            return redirect(url_for('account'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la cuenta: {e}', 'danger')

    if form_password.submit_password.data and form_password.validate_on_submit():
        try:
            if current_user.check_password(form_password.old_password.data):
                current_user.set_password(form_password.password.data)
                db.session.commit()
                flash('¡Tu contraseña ha sido cambiada!', 'success')
                return redirect(url_for('account'))
            else:
                flash('La contraseña actual es incorrecta.', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al cambiar la contraseña: {e}', 'danger')

    if request.method == 'GET':
        form_account.username.data = current_user.username
        form_account.email.data = current_user.email
    
    image_url = url_for('static', filename='uploads/profile_pics/' + current_user.image_file)
    
    return render_template('account.html', 
                           titulo="Configuración de Cuenta",
                           image_url=image_url,
                           form_account=form_account,
                           form_password=form_password)

@app.route('/configuracion/plantilla', methods=['GET', 'POST'])
@login_required
@admin_required
def configurar_plantilla():
    organizacion = current_user.organizacion
    
    if request.method == 'POST':
        try:
            # Logo
            if 'logo' in request.files:
                file = request.files['logo']
                if file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(f"logo_org_{organizacion.id}_{file.filename}")
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    organizacion.logo_url = filename
            if request.form.get('eliminar_logo') == '1':
                organizacion.logo_url = None

            # Identidad
            organizacion.nombre          = request.form.get('nombre', organizacion.nombre).strip()
            organizacion.header_titulo   = request.form.get('header_titulo',  '').strip() or None
            organizacion.header_subtitulo= request.form.get('header_subtitulo','').strip() or None
            organizacion.rfc             = request.form.get('rfc',    '').strip() or None
            organizacion.correo_empresa  = request.form.get('correo_empresa','').strip() or None
            organizacion.direccion       = request.form.get('direccion','').strip() or None
            organizacion.telefono        = request.form.get('telefono','').strip() or None

            # Diseño
            organizacion.color_primario  = request.form.get('color_primario',  '#333333')
            organizacion.color_secundario= request.form.get('color_secundario','#f1f5f9')
            organizacion.tipo_letra      = request.form.get('tipo_letra', 'Helvetica')

            # Documentos PDF
            organizacion.footer_texto    = request.form.get('footer_texto','').strip() or None
            organizacion.pdf_mostrar_qr  = request.form.get('pdf_mostrar_qr') == '1'

            # Notificaciones
            organizacion.whatsapp_notify = request.form.get('whatsapp_notify', '').strip() or None

            db.session.commit()
            flash('Configuración de marca actualizada.', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar: {e}', 'danger')

        return redirect(url_for('configurar_plantilla'))
        
    return render_template('plantilla_config.html', org=organizacion)

    
# --- RUTAS DEL SUPER ADMIN ---

@app.route('/superadmin', methods=['GET'])
@login_required
@super_admin_required
def super_admin():
    """ 
    Panel principal del Super Admin para gestionar
    Organizaciones y Usuarios.
    """
    organizaciones = Organizacion.query.order_by(Organizacion.nombre).all()
    usuarios = User.query.order_by(User.username).all()
    
    return render_template('super_admin.html', 
                           titulo="Super Admin Panel",
                           organizaciones=organizaciones,
                           usuarios=usuarios)

@app.route('/superadmin/organizacion/nueva', methods=['POST'])
@login_required
@super_admin_required
def nueva_organizacion():
    """ Crea una nueva organización y le genera un código de invitación. """
    nombre = request.form.get('nombre')
    if not nombre:
        flash('El nombre de la organización no puede estar vacío.', 'danger')
        return redirect(url_for('super_admin'))
        
    existente = Organizacion.query.filter_by(nombre=nombre).first()
    if existente:
        flash(f'La organización "{nombre}" ya existe.', 'warning')
        return redirect(url_for('super_admin'))
        
    try:
        # --- LÓGICA DE CÓDIGO ÚNICO AÑADIDA ---
        codigo = None
        while codigo is None or Organizacion.query.filter_by(codigo_invitacion=codigo).first():
            # Genera un código de 8 caracteres (ej: "A1b-C2dE")
            codigo = secrets.token_urlsafe(6).upper() 
        # --- FIN DE LÓGICA AÑADIDA ---

        nueva_org = Organizacion(
            nombre=nombre,
            codigo_invitacion=codigo # <-- AÑADIDO
        )
        db.session.add(nueva_org)
        db.session.commit()
        flash(f'Organización "{nombre}" creada. Código de invitación: {codigo}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear la organización: {e}', 'danger')
        
    return redirect(url_for('super_admin'))

@app.route('/superadmin/usuario/asignar/<int:user_id>', methods=['POST'])
@login_required
@super_admin_required
def asignar_usuario(user_id):
    """ Asigna un rol y una organización a un usuario. """
    user = User.query.get_or_404(user_id)
    nuevo_rol = request.form.get('rol')
    nueva_org_id = request.form.get('organizacion_id')

    if not nuevo_rol:
        flash('Error: No se seleccionó un rol.', 'danger')
        return redirect(url_for('super_admin'))

    try:
        user.rol = nuevo_rol
        
        if nueva_org_id == '0':
            user.organizacion_id = None
        else:
            user.organizacion_id = int(nueva_org_id)
        
        if user.rol == 'super_admin':
            user.organizacion_id = None
            
        db.session.commit()
        flash(f'Usuario "{user.username}" actualizado.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar el usuario: {e}', 'danger')

    return redirect(url_for('super_admin'))

@app.route('/api/ai/generar-imagen-producto')
@login_required
def ai_generar_imagen_producto():
    import uuid as _uuid
    nombre = request.args.get('nombre', '').strip()
    seed   = request.args.get('seed', '42')
    if not nombre:
        return jsonify({'error': 'Proporciona un nombre de producto'}), 400
    prompt = f"{nombre}, product photography, white background, professional studio, clean, high quality"
    poll_url = (
        f"https://image.pollinations.ai/prompt/{requests.utils.quote(prompt)}"
        f"?width=512&height=512&nologo=true&seed={seed}&model=flux"
    )
    try:
        resp = requests.get(poll_url, timeout=50)
        if not resp.ok:
            return jsonify({'error': 'Pollinations no respondió correctamente'}), 502
        filename = f"ai_{_uuid.uuid4().hex[:12]}.jpg"
        with open(os.path.join(app.config['UPLOAD_FOLDER'], filename), 'wb') as fh:
            fh.write(resp.content)
        return jsonify({'filename': filename,
                        'url': url_for('static', filename=f'uploads/{filename}')})
    except requests.Timeout:
        return jsonify({'error': 'La IA tardó demasiado, intenta de nuevo.'}), 504
    except Exception as e:
        return jsonify({'error': 'Error al generar imagen'}), 500


@app.route('/api/ai/mejorar-descripcion', methods=['POST'])
@login_required
def ai_mejorar_descripcion():
    from google import genai
    import os

    data = request.get_json()
    producto = data.get('producto', '')

    if not producto:
        return jsonify({'error': 'Producto vacío'}), 400

    API_KEY = os.environ.get("GEMINI_API_KEY")
    if not API_KEY:
        return jsonify({'error': 'IA no configurada en el servidor (falta GEMINI_API_KEY).'}), 503

    try:
        client = genai.Client(api_key=API_KEY)
        
        import json as _json
        prompt = f"""Eres un experto en compras corporativas (Procurement Manager) para una empresa en México.
El usuario necesita comprar: "{producto}"

Devuelve ÚNICAMENTE un objeto JSON válido con estos dos campos (sin markdown, sin texto extra):
{{
  "especificaciones": "especificaciones técnicas breves en 3-5 viñetas con guion, listas para pegar en una OC",
  "costo_estimado_mxn": <número entero, precio unitario realista en pesos mexicanos (MXN)>
}}

Reglas:
- especificaciones: máximo 5 líneas, tono técnico, sin saludos ni introducción
- costo_estimado_mxn: precio unitario promedio de mercado en México, solo el número sin símbolo"""

        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt
        )

        text = response.text.strip()
        if text.startswith('```'):
            text = text.split('```')[1]
            if text.startswith('json'):
                text = text[4:]
        result = _json.loads(text)
        return jsonify({
            'sugerencia':          result.get('especificaciones', ''),
            'costo_estimado_mxn':  result.get('costo_estimado_mxn', 0),
        })

    except Exception as e:
        import logging
        logging.error(f"Error AI Gemini: {e}")
        return jsonify({'error': 'No se pudo conectar con la IA en este momento.'}), 500
    
# ========================
# DIAGNÓSTICO / TEST
# ========================

@app.route('/admin/test-email')
@login_required
def test_email():
    """Diagnóstico Brevo: envía un correo de prueba y muestra el resultado real."""
    if current_user.rol not in ['super_admin', 'admin']:
        return "Acceso denegado", 403

    API_KEY      = os.environ.get("BREVO_API_KEY", "")
    SENDER_EMAIL = os.environ.get("BREVO_SENDER_EMAIL", "deinventarioc@gmail.com")
    destinatario = current_user.email
    test_url     = url_for('dashboard', _external=True)

    if not API_KEY:
        return (
            "<h3>❌ BREVO_API_KEY no está definida</h3>"
            "<p>Agrégala en <code>/etc/systemd/system/inventario.service.d/override.conf</code> "
            "con <code>Environment=BREVO_API_KEY=tu_clave</code> y haz "
            "<code>systemctl daemon-reload && systemctl restart inventario</code></p>"
        ), 500

    ok, error = enviar_correo_api(destinatario, test_url)

    if ok:
        return (
            f"<h3 style='color:green'>✅ Correo enviado correctamente</h3>"
            f"<p>Enviado a: <strong>{destinatario}</strong><br>"
            f"Remitente configurado: <strong>{SENDER_EMAIL}</strong></p>"
            f"<p>Si no llega en 2 minutos:</p>"
            f"<ol>"
            f"<li>Revisa la carpeta de <strong>Spam / No deseado</strong></li>"
            f"<li>Verifica en Brevo → <em>Settings → Senders &amp; IP → Senders</em> "
            f"que <strong>{SENDER_EMAIL}</strong> esté verificado (ícono verde)</li>"
            f"<li>Si el remitente no está verificado, Brevo puede aceptar la llamada "
            f"API pero NO entregar el correo</li>"
            f"</ol>"
        )
    else:
        return (
            f"<h3 style='color:red'>❌ Error al enviar</h3>"
            f"<p>Remitente: <strong>{SENDER_EMAIL}</strong><br>"
            f"Destinatario: <strong>{destinatario}</strong></p>"
            f"<pre style='background:#fee;padding:12px;border-radius:6px;'>{error}</pre>"
            f"<h4>Causas comunes:</h4>"
            f"<ol>"
            f"<li><strong>Sender not verified</strong> — Ve a Brevo → Settings → Senders &amp; IP → Senders "
            f"y agrega/verifica <code>{SENDER_EMAIL}</code></li>"
            f"<li><strong>API Key inválida</strong> — Ve a Brevo → Settings → API Keys y regenera la clave</li>"
            f"<li><strong>Plan gratuito agotado</strong> — Brevo Free permite 300 correos/día</li>"
            f"</ol>"
        ), 500

# ========================
# NUEVAS RUTAS DEL ADMIN
# ========================

@app.route('/admin_panel')
@login_required
@admin_required # Solo 'admin' o 'super_admin'
def admin_panel():
    """ Panel para que un Admin gestione los usuarios de SU organización. """
    
    if current_user.rol == 'super_admin':
        return redirect(url_for('super_admin'))
        
    usuarios = User.query.filter_by(
        organizacion_id=current_user.organizacion_id
    ).order_by(User.username).all()
    
    forms = {}
    for user in usuarios:
        form = AdminPermissionForm()
        form.perm_view_dashboard.data = user.perm_view_dashboard
        form.perm_view_management.data = user.perm_view_management
        form.perm_edit_management.data = user.perm_edit_management
        form.perm_create_oc_standard.data = user.perm_create_oc_standard
        form.perm_create_oc_proyecto.data = user.perm_create_oc_proyecto
        form.perm_do_salidas.data = user.perm_do_salidas
        form.perm_view_gastos.data = user.perm_view_gastos
        forms[user.id] = form
        
    return render_template('admin_panel.html', 
                           titulo="Panel de Administrador",
                           usuarios=usuarios,
                           forms=forms)

@app.route('/admin_panel/update/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def update_user_permissions(user_id):
    user_to_update = User.query.get_or_404(user_id)

    if current_user.rol == 'admin' and user_to_update.organizacion_id != current_user.organizacion_id:
        flash('No tienes permiso para editar a este usuario.', 'danger')
        return redirect(url_for('admin_panel'))

    if user_to_update.id == current_user.id and current_user.rol != 'super_admin':
        flash('No puedes editar tus propios permisos.', 'warning')
        return redirect(url_for('admin_panel'))

    form = AdminPermissionForm()
    if form.validate_on_submit():
        try:
            user_to_update.perm_view_dashboard    = form.perm_view_dashboard.data
            user_to_update.perm_view_management   = form.perm_view_management.data
            user_to_update.perm_edit_management   = form.perm_edit_management.data
            user_to_update.perm_create_oc_standard = form.perm_create_oc_standard.data
            user_to_update.perm_create_oc_proyecto = form.perm_create_oc_proyecto.data
            user_to_update.perm_do_salidas        = form.perm_do_salidas.data
            user_to_update.perm_view_gastos       = form.perm_view_gastos.data
            db.session.commit()
            flash(f'Permisos para {user_to_update.username} actualizados.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar permisos: {e}', 'danger')
    else:
        flash('Error de validación del formulario.', 'danger')

    return redirect(url_for('admin_panel'))


# ========================
# MANUAL DE USUARIO
# ========================

@app.route('/admin/manual')
@login_required
@admin_required
def manual_usuario():
    """Manual de uso del sistema, descargable como PDF por admins."""
    return render_template(
        'manual_usuario.html',
        org=current_user.organizacion,
        now=now_mx(),
    )

    return redirect(url_for('admin_panel'))


# ==============================================================================
# SYNC OFFLINE — procesa la cola de operaciones realizadas sin conexión
# ==============================================================================

@app.route('/api/sync', methods=['POST'])
@login_required
@check_org_permission
def api_sync():
    """
    Recibe una lista de operaciones offline y las ejecuta en orden.
    Responde con el resultado de cada una (ok / error).
    Estrategia estricta: valida stock antes de ejecutar.
    """
    data = request.get_json(silent=True) or {}
    operations = data.get('operations', [])
    if not isinstance(operations, list):
        return jsonify(ok=False, error='Payload inválido'), 400

    org_id   = current_user.organizacion_id
    results  = []

    for op in operations:
        op_id   = op.get('id')
        op_type = op.get('type')
        payload = op.get('payload', {})

        try:
            if op_type == 'gasto':
                result = _sync_gasto(payload, org_id)
            elif op_type == 'salida':
                result = _sync_salida(payload, org_id)
            else:
                result = {'ok': False, 'error': f'Tipo desconocido: {op_type}'}
        except Exception as e:
            db.session.rollback()
            result = {'ok': False, 'error': str(e)}

        result['id'] = op_id
        results.append(result)

    return jsonify(ok=True, results=results)


def _sync_gasto(payload, org_id):
    from datetime import datetime as _dt
    fecha_str   = payload.get('fecha')
    descripcion = payload.get('descripcion', '').strip()
    monto_str   = payload.get('monto')
    categoria   = payload.get('categoria', '').strip()
    oc_id       = payload.get('orden_compra_id') or None

    if not fecha_str or not descripcion or not monto_str or not categoria:
        return {'ok': False, 'error': 'Gasto: faltan campos obligatorios'}

    try:
        fecha = _dt.strptime(fecha_str, '%Y-%m-%d')
        monto = float(monto_str)
    except (ValueError, TypeError):
        return {'ok': False, 'error': 'Gasto: fecha o monto inválidos'}

    gasto = Gasto(
        fecha           = fecha,
        descripcion     = descripcion,
        monto           = monto,
        categoria       = categoria,
        orden_compra_id = int(oc_id) if oc_id else None,
        organizacion_id = org_id,
        creador_id      = current_user.id,
    )
    db.session.add(gasto)
    db.session.commit()
    log_actividad('crear', 'gasto', f'Gasto offline sincronizado: {descripcion} ${monto:.2f}', entidad_id=gasto.id)
    return {'ok': True}


def _sync_salida(payload, org_id):
    almacen_id = payload.get('almacen_id')
    items      = payload.get('items', [])

    if not almacen_id or not items:
        return {'ok': False, 'error': 'Salida: faltan almacén o items'}

    almacen = Almacen.query.filter_by(id=almacen_id, organizacion_id=org_id).first()
    if not almacen:
        return {'ok': False, 'error': 'Salida: almacén no válido'}

    # ── Fase de validación ───────────────────────────────────────────────
    para_ejecutar = []
    for item in items:
        prod_id  = item.get('producto_id')
        cantidad = item.get('cantidad')
        motivo   = item.get('motivo', 'Offline')

        if not prod_id or not cantidad:
            return {'ok': False, 'error': 'Salida: item con datos incompletos'}

        try:
            cantidad = int(cantidad)
        except (ValueError, TypeError):
            return {'ok': False, 'error': f'Salida: cantidad inválida para producto {prod_id}'}

        if cantidad <= 0:
            return {'ok': False, 'error': 'Salida: cantidades deben ser positivas'}

        stock_item = Stock.query.filter_by(
            producto_id=prod_id, almacen_id=almacen_id
        ).first()

        if not stock_item:
            return {'ok': False, 'error': f'Salida: producto {prod_id} sin stock en este almacén'}

        if stock_item.cantidad < cantidad:
            return {
                'ok':    False,
                'error': f'Stock insuficiente para "{stock_item.producto.nombre}": '
                         f'disponible {stock_item.cantidad}, solicitado {cantidad}',
            }

        para_ejecutar.append((stock_item, cantidad, motivo))

    # ── Fase de ejecución ────────────────────────────────────────────────
    today = now_mx().date()
    salida_del_dia = Salida.query.filter_by(
        fecha=today, organizacion_id=org_id, almacen_id=almacen_id
    ).first()
    if not salida_del_dia:
        salida_del_dia = Salida(
            fecha=today,
            creador_id=current_user.id,
            organizacion_id=org_id,
            almacen_id=almacen_id,
        )
        db.session.add(salida_del_dia)
        db.session.flush()

    for stock_item, cantidad, motivo in para_ejecutar:
        stock_item.cantidad -= cantidad
        db.session.add(stock_item)
        db.session.add(Movimiento(
            producto_id    = stock_item.producto_id,
            cantidad       = -cantidad,
            tipo           = 'salida',
            fecha          = now_mx(),
            motivo         = f'[Offline] {motivo}',
            salida         = salida_del_dia,
            almacen_id     = almacen_id,
            organizacion_id= org_id,
        ))

    db.session.commit()
    total_uds = sum(v[1] for v in para_ejecutar)
    log_actividad(
        'salida', 'salida',
        f'Salida offline sincronizada: {len(para_ejecutar)} producto(s), {total_uds} uds — {almacen.nombre}',
        entidad_id=salida_del_dia.id,
    )
    check_and_alert_stock_bajo(org_id, almacen_id)
    return {'ok': True}


@app.route('/api/permisos/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def api_toggle_permiso(user_id):
    """API para auto-guardar un permiso individual vía AJAX."""
    PERMS_VALIDOS = {
        'perm_view_dashboard', 'perm_view_management', 'perm_edit_management',
        'perm_create_oc_standard', 'perm_create_oc_proyecto',
        'perm_do_salidas', 'perm_view_gastos',
    }
    user_to_update = User.query.get_or_404(user_id)

    if current_user.rol == 'admin' and user_to_update.organizacion_id != current_user.organizacion_id:
        return jsonify(ok=False, error='Sin permiso'), 403

    if user_to_update.id == current_user.id and current_user.rol != 'super_admin':
        return jsonify(ok=False, error='No puedes editar tus propios permisos'), 403

    if user_to_update.rol != 'user':
        return jsonify(ok=False, error='Solo se pueden editar permisos de usuarios base'), 400

    data = request.get_json(silent=True) or {}
    perm  = data.get('perm')
    value = data.get('value')

    if perm not in PERMS_VALIDOS or not isinstance(value, bool):
        return jsonify(ok=False, error='Datos inválidos'), 400

    try:
        setattr(user_to_update, perm, value)
        db.session.commit()
        return jsonify(ok=True, username=user_to_update.username, perm=perm, value=value)
    except Exception as e:
        db.session.rollback()
        return jsonify(ok=False, error=str(e)), 500

# ==============================================================================
# TRANSFERENCIAS ENTRE ALMACENES
# ==============================================================================

@app.route('/transferencia/nueva', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_do_salidas')
def nueva_transferencia():
    """Mueve stock de un almacén origen a un almacén destino en una transacción atómica."""
    org_id = current_user.organizacion_id
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).order_by(Almacen.nombre).all()

    if len(almacenes) < 2:
        flash('Necesitas al menos 2 almacenes para realizar una transferencia.', 'warning')
        return redirect(url_for('lista_almacenes'))

    if request.method == 'POST':
        try:
            origen_id  = int(request.form.get('almacen_origen_id'))
            destino_id = int(request.form.get('almacen_destino_id'))
            producto_id = int(request.form.get('producto_id'))
            cantidad   = int(request.form.get('cantidad', 0))
            motivo     = request.form.get('motivo', '').strip() or 'Transferencia entre almacenes'

            if origen_id == destino_id:
                flash('El almacén de origen y destino no pueden ser el mismo.', 'danger')
                return redirect(url_for('nueva_transferencia'))
            if cantidad <= 0:
                flash('La cantidad debe ser mayor a cero.', 'danger')
                return redirect(url_for('nueva_transferencia'))

            stock_origen = Stock.query.filter_by(
                producto_id=producto_id, almacen_id=origen_id
            ).first()

            if not stock_origen or stock_origen.cantidad < cantidad:
                flash(f'Stock insuficiente en el almacén origen. Disponible: {stock_origen.cantidad if stock_origen else 0}', 'danger')
                return redirect(url_for('nueva_transferencia'))

            ref = secrets.token_hex(4).upper()

            # Descontar del origen
            stock_origen.cantidad -= cantidad

            # Sumar (o crear) en el destino
            stock_destino = Stock.query.filter_by(
                producto_id=producto_id, almacen_id=destino_id
            ).first()
            if stock_destino:
                stock_destino.cantidad += cantidad
            else:
                stock_destino = Stock(
                    producto_id=producto_id,
                    almacen_id=destino_id,
                    cantidad=cantidad,
                    stock_minimo=stock_origen.stock_minimo,
                    stock_maximo=stock_origen.stock_maximo,
                )
                db.session.add(stock_destino)

            now = now_mx()
            db.session.add(Movimiento(
                producto_id=producto_id, cantidad=-cantidad,
                tipo='transferencia-salida', fecha=now,
                motivo=f'[REF:{ref}] {motivo}',
                almacen_id=origen_id, organizacion_id=org_id
            ))
            db.session.add(Movimiento(
                producto_id=producto_id, cantidad=cantidad,
                tipo='transferencia-entrada', fecha=now,
                motivo=f'[REF:{ref}] {motivo}',
                almacen_id=destino_id, organizacion_id=org_id
            ))

            db.session.commit()
            flash(f'Transferencia REF:{ref} completada. {cantidad} unidades movidas correctamente.', 'success')
            return redirect(url_for('dashboard'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al realizar la transferencia: {e}', 'danger')

    return render_template('transferencia_form.html',
                           titulo='Nueva Transferencia de Stock',
                           almacenes=almacenes)


@app.route('/api/almacen/<int:almacen_id>/productos-con-stock')
@login_required
def api_productos_con_stock(almacen_id):
    """Retorna los productos con stock > 0 en un almacén dado (para el select dinámico)."""
    org_id = current_user.organizacion_id
    almacen = Almacen.query.filter_by(id=almacen_id, organizacion_id=org_id).first_or_404()

    items = db.session.query(Stock).filter(
        Stock.almacen_id == almacen.id,
        Stock.cantidad > 0
    ).join(Producto).order_by(Producto.nombre).all()

    return jsonify([{
        'id': s.producto_id,
        'nombre': s.producto.nombre,
        'codigo': s.producto.codigo,
        'cantidad': s.cantidad
    } for s in items])


# ==============================================================================
# AJUSTE MANUAL DE INVENTARIO
# ==============================================================================

@app.route('/ajuste/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def nuevo_ajuste():
    """Registra un ajuste físico de inventario con auditoría completa."""
    org_id = current_user.organizacion_id
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).order_by(Almacen.nombre).all()

    if request.method == 'POST':
        try:
            almacen_id  = int(request.form.get('almacen_id'))
            producto_id = int(request.form.get('producto_id'))
            cantidad_fisica = int(request.form.get('cantidad_fisica', 0))
            motivo = request.form.get('motivo', '').strip()

            if not motivo:
                flash('El motivo del ajuste es obligatorio para la auditoría.', 'danger')
                return redirect(url_for('nuevo_ajuste'))

            stock = Stock.query.filter_by(
                producto_id=producto_id, almacen_id=almacen_id
            ).first()

            if not stock:
                flash('No se encontró ese producto en el almacén seleccionado.', 'danger')
                return redirect(url_for('nuevo_ajuste'))

            diferencia = cantidad_fisica - stock.cantidad

            if diferencia == 0:
                flash('No hay diferencia entre el conteo físico y el sistema. No se realizó ningún ajuste.', 'info')
                return redirect(url_for('nuevo_ajuste'))

            tipo_mov = 'ajuste-entrada' if diferencia > 0 else 'ajuste-salida'
            stock.cantidad = cantidad_fisica

            db.session.add(Movimiento(
                producto_id=producto_id,
                cantidad=diferencia,
                tipo=tipo_mov,
                fecha=now_mx(),
                motivo=f'Ajuste Físico: {motivo}',
                almacen_id=almacen_id,
                organizacion_id=org_id
            ))
            signo = '+' if diferencia > 0 else ''
            log_actividad('ajuste', 'producto', f'Ajuste de inventario: {signo}{diferencia} uds — {motivo}', entidad_id=producto_id)
            db.session.commit()

            if diferencia < 0:
                check_and_alert_stock_bajo(org_id, almacen_id)

            flash(f'Ajuste registrado. Diferencia aplicada: {signo}{diferencia} unidades.', 'success')
            return redirect(url_for('dashboard'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar el ajuste: {e}', 'danger')

    return render_template('ajuste_form.html',
                           titulo='Ajuste Manual de Inventario',
                           almacenes=almacenes)


# ==============================================================================
# ENDPOINTS JSON PARA GRÁFICAS (CHART.JS)
# ==============================================================================

@app.route('/api/charts/movimientos-mes')
@login_required
@check_org_permission
def api_chart_movimientos_mes():
    """Retorna entradas y salidas de los últimos 6 meses para la gráfica de barras."""
    org_id = current_user.organizacion_id
    hoy = now_mx()

    labels, entradas, salidas = [], [], []

    for i in range(5, -1, -1):
        mes = (hoy.month - i - 1) % 12 + 1
        ano = hoy.year if (hoy.month - i) > 0 else hoy.year - 1

        total_entrada = db.session.query(db.func.sum(Movimiento.cantidad)).filter(
            Movimiento.organizacion_id == org_id,
            Movimiento.tipo.in_(['entrada', 'entrada-inicial', 'ajuste-entrada']),
            extract('month', Movimiento.fecha) == mes,
            extract('year', Movimiento.fecha) == ano
        ).scalar() or 0

        total_salida = abs(db.session.query(db.func.sum(Movimiento.cantidad)).filter(
            Movimiento.organizacion_id == org_id,
            Movimiento.tipo == 'salida',
            extract('month', Movimiento.fecha) == mes,
            extract('year', Movimiento.fecha) == ano
        ).scalar() or 0)

        nombre_mes = datetime(ano, mes, 1).strftime('%b %Y')
        labels.append(nombre_mes)
        entradas.append(int(total_entrada))
        salidas.append(int(total_salida))

    return jsonify({'labels': labels, 'entradas': entradas, 'salidas': salidas})


@app.route('/api/charts/estado-stock')
@login_required
@check_org_permission
def api_chart_estado_stock():
    """Retorna conteo de productos por estado (bajo/ok/exceso) para la gráfica de dona."""
    org_id = current_user.organizacion_id

    bajo = db.session.query(db.func.count(Stock.id)).join(Almacen).filter(
        Almacen.organizacion_id == org_id,
        Stock.cantidad < Stock.stock_minimo
    ).scalar() or 0

    exceso = db.session.query(db.func.count(Stock.id)).join(Almacen).filter(
        Almacen.organizacion_id == org_id,
        Stock.cantidad > Stock.stock_maximo
    ).scalar() or 0

    ok = db.session.query(db.func.count(Stock.id)).join(Almacen).filter(
        Almacen.organizacion_id == org_id,
        Stock.cantidad >= Stock.stock_minimo,
        Stock.cantidad <= Stock.stock_maximo
    ).scalar() or 0

    return jsonify({'bajo': int(bajo), 'ok': int(ok), 'exceso': int(exceso)})


@app.route('/api/charts/top-productos')
@login_required
@check_org_permission
def api_chart_top_productos():
    """Retorna los 8 productos con más salidas en los últimos 30 días."""
    org_id = current_user.organizacion_id
    desde = now_mx().replace(day=1)

    resultados = db.session.query(
        Producto.nombre,
        db.func.sum(db.func.abs(Movimiento.cantidad)).label('total')
    ).join(Movimiento, Movimiento.producto_id == Producto.id).filter(
        Movimiento.organizacion_id == org_id,
        Movimiento.tipo == 'salida',
        Movimiento.fecha >= desde
    ).group_by(Producto.nombre).order_by(db.desc('total')).limit(8).all()

    return jsonify({
        'labels': [r.nombre[:25] for r in resultados],
        'valores': [int(r.total) for r in resultados]
    })


@app.route('/api/dashboard/actividad-reciente')
@login_required
@check_org_permission
def api_actividad_reciente():
    """Retorna los últimos 10 movimientos de la organización para el feed del dashboard."""
    org_id = current_user.organizacion_id
    movs = (
        Movimiento.query
        .filter_by(organizacion_id=org_id)
        .order_by(Movimiento.fecha.desc())
        .limit(10)
        .all()
    )
    TIPO_META = {
        'entrada':         {'icon': 'bi-box-arrow-in-down', 'color': '#10b981', 'label': 'Entrada'},
        'entrada-inicial': {'icon': 'bi-database-add',      'color': '#3b82f6', 'label': 'Stock Inicial'},
        'salida':          {'icon': 'bi-box-arrow-right',   'color': '#ef4444', 'label': 'Salida'},
        'ajuste-entrada':  {'icon': 'bi-plus-circle',       'color': '#8b5cf6', 'label': 'Ajuste (+)'},
        'ajuste-salida':   {'icon': 'bi-dash-circle',       'color': '#f59e0b', 'label': 'Ajuste (-)'},
        'transferencia-entrada': {'icon': 'bi-arrow-left-right', 'color': '#06b6d4', 'label': 'Transferencia (+)'},
        'transferencia-salida':  {'icon': 'bi-arrow-left-right', 'color': '#64748b', 'label': 'Transferencia (-)'},
    }
    resultado = []
    for m in movs:
        meta = TIPO_META.get(m.tipo, {'icon': 'bi-arrow-repeat', 'color': '#64748b', 'label': m.tipo})
        almacen_nombre = m.almacen.nombre if m.almacen else '—'
        resultado.append({
            'id':       m.id,
            'tipo':     m.tipo,
            'label':    meta['label'],
            'icon':     meta['icon'],
            'color':    meta['color'],
            'cantidad': abs(m.cantidad),
            'signo':    '+' if m.cantidad >= 0 else '−',
            'producto': m.producto.nombre if m.producto else '—',
            'almacen':  almacen_nombre,
            'motivo':   m.motivo or '',
            'fecha':    m.fecha.strftime('%d/%m %H:%M'),
        })
    return jsonify(resultado)


# ==============================================================================
# WEB PUSH NOTIFICATIONS — API
# ==============================================================================

@app.route('/api/push/vapid-key')
@login_required
def api_vapid_key():
    return jsonify({'publicKey': os.environ.get('VAPID_PUBLIC_KEY', '')})


@app.route('/api/push/subscribe', methods=['POST'])
@login_required
def api_push_subscribe():
    data = request.get_json(silent=True)
    if not data or 'endpoint' not in data:
        return jsonify({'error': 'datos inválidos'}), 400
    endpoint = data['endpoint']
    existing = PushSubscription.query.filter_by(endpoint=endpoint).first()
    if existing:
        existing.subscription_json = json.dumps(data)
        existing.user_id = current_user.id
    else:
        db.session.add(PushSubscription(
            user_id=current_user.id,
            organizacion_id=current_user.organizacion_id,
            endpoint=endpoint,
            subscription_json=json.dumps(data)
        ))
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/push/unsubscribe', methods=['POST'])
@login_required
def api_push_unsubscribe():
    data = request.get_json(silent=True)
    if not data or 'endpoint' not in data:
        return jsonify({'error': 'datos inválidos'}), 400
    sub = PushSubscription.query.filter_by(endpoint=data['endpoint']).first()
    if sub:
        db.session.delete(sub)
        db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/push/test', methods=['POST'])
@login_required
def api_push_test():
    """Envía una notificación de prueba al usuario actual para verificar la configuración."""
    subs = PushSubscription.query.filter_by(user_id=current_user.id).all()
    if not subs:
        return jsonify({'ok': False, 'error': 'Sin suscripción activa — activa las notificaciones primero'}), 400
    vapid_private = os.environ.get('VAPID_PRIVATE_KEY')
    if not vapid_private:
        return jsonify({'ok': False, 'error': 'VAPID_PRIVATE_KEY no configurada en el servidor'}), 503
    try:
        from pywebpush import webpush, WebPushException
        vapid_email = os.environ.get('VAPID_CLAIMS_EMAIL', 'notifications@inventario.app')
        payload = json.dumps({'title': 'Prueba de notificación', 'body': 'Las notificaciones push están funcionando.', 'url': '/dashboard'})
        sent, errors, to_delete = 0, [], []
        for sub in subs:
            try:
                webpush(
                    subscription_info=json.loads(sub.subscription_json),
                    data=payload,
                    vapid_private_key=vapid_private,
                    vapid_claims={"sub": f"mailto:{vapid_email}"}
                )
                sent += 1
            except WebPushException as ex:
                code = _webpush_http_status(ex)
                errors.append(f'HTTP {code}: {ex}')
                if code in _PUSH_STALE_CODES:
                    to_delete.append(sub)
        for sub in to_delete:
            db.session.delete(sub)
        if to_delete:
            db.session.commit()
        if sent > 0:
            return jsonify({'ok': True, 'sent': sent})
        return jsonify({'ok': False, 'error': '; '.join(errors) or 'Sin suscripciones válidas'}), 500
    except ImportError:
        return jsonify({'ok': False, 'error': 'pywebpush no instalado en el servidor'}), 503
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# --- Manejadores de Error ---

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500

# ==============================================================================
# SERVICIOS — Control de pagos recurrentes (Agua, Luz, Gas, etc.)
# ==============================================================================

TIPOS_SERVICIO = {
    'luz':      ('bi-lightning-charge-fill', '#f59e0b', 'Electricidad / Luz'),
    'agua':     ('bi-droplet-fill',          '#3b82f6', 'Agua'),
    'gas':      ('bi-fire',                  '#ef4444', 'Gas'),
    'internet': ('bi-wifi',                  '#8b5cf6', 'Internet'),
    'telefono': ('bi-telephone-fill',        '#10b981', 'Teléfono'),
    'renta':    ('bi-building',              '#64748b', 'Renta'),
    'otro':     ('bi-receipt',               '#94a3b8', 'Otro'),
}

_TIPO_A_CATEGORIA_GASTO = {
    'luz':       'Energía Eléctrica',
    'agua':      'Agua y Drenaje',
    'gas':       'Gas',
    'internet':  'Internet',
    'telefono':  'Telefonía',
    'renta':     'Renta',
    'limpieza':  'Limpieza',
    'otro':      'Servicios',
}

def _registrar_gasto_servicio(pago):
    """Crea un Gasto automáticamente al marcar un PagoServicio como pagado."""
    s = pago.servicio
    if not s:
        return
    categoria = _TIPO_A_CATEGORIA_GASTO.get(s.tipo or 'otro', 'Servicios')
    from datetime import datetime as _dt
    fecha_dt = _dt.combine(pago.fecha_pago, _dt.min.time())
    gasto = Gasto(
        descripcion=f"Servicio: {s.nombre}",
        monto=pago.monto,
        categoria=categoria,
        fecha=fecha_dt,
        organizacion_id=s.organizacion_id,
    )
    db.session.add(gasto)


def _actualizar_estados_pagos(org_id):
    """Marca como 'vencido' los pagos pendientes con fecha_vencimiento ya pasada."""
    hoy = now_mx().date()
    serv_ids = db.session.query(Servicio.id).filter_by(organizacion_id=org_id).subquery()
    PagoServicio.query.filter(
        PagoServicio.servicio_id.in_(serv_ids),
        PagoServicio.estado == 'pendiente',
        PagoServicio.fecha_vencimiento < hoy
    ).update({'estado': 'vencido'}, synchronize_session=False)
    db.session.commit()


@app.route('/servicios')
@login_required
def lista_servicios():
    _actualizar_estados_pagos(current_user.organizacion_id)
    hoy = now_mx().date()
    servicios = Servicio.query.filter_by(
        organizacion_id=current_user.organizacion_id, activo=True
    ).order_by(Servicio.nombre).all()

    vencidos = PagoServicio.query.join(Servicio).filter(
        Servicio.organizacion_id == current_user.organizacion_id,
        PagoServicio.estado == 'vencido'
    ).count()
    proximos = PagoServicio.query.join(Servicio).filter(
        Servicio.organizacion_id == current_user.organizacion_id,
        PagoServicio.estado == 'pendiente',
        PagoServicio.fecha_vencimiento <= hoy + timedelta(days=7)
    ).count()
    gasto_mes = db.session.query(db.func.sum(PagoServicio.monto)).join(Servicio).filter(
        Servicio.organizacion_id == current_user.organizacion_id,
        PagoServicio.estado == 'pagado',
        db.func.extract('year',  PagoServicio.fecha_pago) == hoy.year,
        db.func.extract('month', PagoServicio.fecha_pago) == hoy.month,
    ).scalar() or 0

    return render_template('servicios_lista.html',
        servicios=servicios, tipos=TIPOS_SERVICIO,
        vencidos=vencidos, proximos=proximos,
        gasto_mes=gasto_mes, hoy=hoy)


@app.route('/servicios/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_servicio():
    if request.method == 'POST':
        s = Servicio(
            nombre           = request.form['nombre'].strip(),
            tipo             = request.form.get('tipo', 'otro'),
            proveedor_nombre = request.form.get('proveedor_nombre', '').strip() or None,
            numero_contrato  = request.form.get('numero_contrato', '').strip() or None,
            dia_vencimiento  = int(request.form['dia_vencimiento']) if request.form.get('dia_vencimiento') else None,
            dias_aviso       = int(request.form.get('dias_aviso', 5)),
            notas            = request.form.get('notas', '').strip() or None,
            organizacion_id  = current_user.organizacion_id,
        )
        db.session.add(s)
        db.session.commit()
        flash(f'Servicio "{s.nombre}" registrado.', 'success')
        return redirect(url_for('lista_servicios'))
    return render_template('servicio_form.html', servicio=None, tipos=TIPOS_SERVICIO, accion='nuevo')


@app.route('/servicios/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_servicio(id):
    s = Servicio.query.filter_by(id=id, organizacion_id=current_user.organizacion_id).first_or_404()
    if request.method == 'POST':
        s.nombre           = request.form['nombre'].strip()
        s.tipo             = request.form.get('tipo', 'otro')
        s.proveedor_nombre = request.form.get('proveedor_nombre', '').strip() or None
        s.numero_contrato  = request.form.get('numero_contrato', '').strip() or None
        s.dia_vencimiento  = int(request.form['dia_vencimiento']) if request.form.get('dia_vencimiento') else None
        s.dias_aviso       = int(request.form.get('dias_aviso', 5))
        s.notas            = request.form.get('notas', '').strip() or None
        db.session.commit()
        flash('Servicio actualizado.', 'success')
        return redirect(url_for('detalle_servicio', id=s.id))
    return render_template('servicio_form.html', servicio=s, tipos=TIPOS_SERVICIO, accion='editar')


@app.route('/servicios/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_servicio(id):
    if current_user.rol not in ['super_admin', 'admin']:
        flash('Sin permiso para eliminar servicios.', 'danger')
        return redirect(url_for('lista_servicios'))
    s = Servicio.query.filter_by(id=id, organizacion_id=current_user.organizacion_id).first_or_404()
    nombre = s.nombre
    db.session.delete(s)
    db.session.commit()
    flash(f'Servicio "{nombre}" eliminado.', 'success')
    return redirect(url_for('lista_servicios'))


@app.route('/servicios/<int:id>')
@login_required
def detalle_servicio(id):
    _actualizar_estados_pagos(current_user.organizacion_id)
    s    = Servicio.query.filter_by(id=id, organizacion_id=current_user.organizacion_id).first_or_404()
    hoy  = now_mx().date()
    info = TIPOS_SERVICIO.get(s.tipo, TIPOS_SERVICIO['otro'])
    pagados = [p for p in s.pagos if p.estado == 'pagado'][:6]
    promedio = (sum(p.monto for p in pagados) / len(pagados)) if pagados else 0
    return render_template('servicio_detalle.html',
        servicio=s, info=info, hoy=hoy, promedio=promedio, tipos=TIPOS_SERVICIO)


@app.route('/servicios/<int:id>/pago/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_pago_servicio(id):
    import calendar
    s = Servicio.query.filter_by(id=id, organizacion_id=current_user.organizacion_id).first_or_404()
    centros = CentroCosto.query.filter_by(organizacion_id=current_user.organizacion_id).order_by(CentroCosto.nombre).all()
    if request.method == 'POST':
        monto_val = float(request.form['monto'])
        if monto_val <= 0:
            flash('El monto debe ser mayor a cero.', 'danger')
            return redirect(url_for('nuevo_pago_servicio', id=id))
        p = PagoServicio(
            servicio_id       = s.id,
            monto             = monto_val,
            fecha_vencimiento = datetime.strptime(request.form['fecha_vencimiento'], '%Y-%m-%d').date(),
            notas             = request.form.get('notas', '').strip() or None,
            centro_costo_id   = int(request.form['centro_costo_id']) if request.form.get('centro_costo_id') else None,
            registrado_por_id = current_user.id,
        )
        if request.form.get('fecha_pago'):
            p.fecha_pago = datetime.strptime(request.form['fecha_pago'], '%Y-%m-%d').date()
            p.estado = 'pagado'
        db.session.add(p)
        db.session.flush()  # para obtener p.id antes del commit
        log_actividad('crear', 'pago_servicio',
            f'Pago registrado — {s.nombre}: ${p.monto:,.2f} (estado: {p.estado})',
            entidad_id=p.id)
        if p.estado == 'pagado':
            _registrar_gasto_servicio(p)
        # Guardar comprobante si se subió
        comp = request.files.get('comprobante')
        if comp and comp.filename:
            ext = secure_filename(comp.filename).rsplit('.', 1)[-1].lower()
            if ext in ('jpg', 'jpeg', 'png', 'pdf', 'webp'):
                carpeta = os.path.join(app.config['UPLOAD_FOLDER'], 'comprobantes')
                os.makedirs(carpeta, exist_ok=True)
                nombre = f"comp_{p.id}_{secrets.token_hex(6)}.{ext}"
                comp.save(os.path.join(carpeta, nombre))
                p.comprobante_url = nombre
        db.session.commit()
        if p.estado == 'pagado':
            enviar_push_notificacion(
                org_id=s.organizacion_id,
                titulo=f'✅ Pago registrado — {s.nombre}',
                cuerpo=f'${p.monto:,.2f} MXN · {p.fecha_pago.strftime("%d/%m/%Y")}',
                url=f'/servicios/{s.id}'
            )
        flash('Pago registrado.', 'success')
        return redirect(url_for('detalle_servicio', id=s.id))
    hoy = now_mx().date()
    fecha_sugerida = ''
    if s.dia_vencimiento:
        mes  = hoy.month if hoy.day < s.dia_vencimiento else (hoy.month % 12 + 1)
        anio = hoy.year  if mes >= hoy.month else hoy.year + 1
        dia  = min(s.dia_vencimiento, calendar.monthrange(anio, mes)[1])
        fecha_sugerida = f'{anio}-{mes:02d}-{dia:02d}'
    return render_template('pago_servicio_form.html', servicio=s, fecha_sugerida=fecha_sugerida, centros=centros)


@app.route('/servicios/pago/<int:id>/marcar-pagado', methods=['POST'])
@login_required
def marcar_pago_pagado(id):
    p = PagoServicio.query.join(Servicio).filter(
        PagoServicio.id == id,
        Servicio.organizacion_id == current_user.organizacion_id
    ).first_or_404()
    fecha_str  = request.form.get('fecha_pago')
    p.fecha_pago = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else now_mx().date()
    p.estado = 'pagado'
    log_actividad('pagar', 'pago_servicio',
        f'Pago marcado como pagado — {p.servicio.nombre}: ${p.monto:,.2f}',
        entidad_id=p.id)
    _registrar_gasto_servicio(p)
    db.session.commit()
    enviar_push_notificacion(
        org_id=p.servicio.organizacion_id,
        titulo=f'✅ Pago registrado — {p.servicio.nombre}',
        cuerpo=f'${p.monto:,.2f} MXN · {p.fecha_pago.strftime("%d/%m/%Y")}',
        url=f'/servicios/{p.servicio_id}'
    )
    flash('Pago marcado como pagado. Gasto registrado automáticamente. ✓', 'success')
    return redirect(url_for('detalle_servicio', id=p.servicio_id))


@app.route('/servicios/pago/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_pago_servicio(id):
    p = PagoServicio.query.join(Servicio).filter(
        PagoServicio.id == id,
        Servicio.organizacion_id == current_user.organizacion_id
    ).first_or_404()
    serv_id = p.servicio_id
    try:
        log_actividad('eliminar', 'pago_servicio',
            f'Pago eliminado — {p.servicio.nombre}: ${p.monto:,.2f} (estado: {p.estado})',
            entidad_id=p.id)
        # Borrar comprobante si existe
        if p.comprobante_url:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], 'comprobantes', p.comprobante_url))
            except OSError:
                pass
        db.session.delete(p)
        db.session.commit()
        flash('Registro de pago eliminado.', 'success')
    except Exception as e:
        db.session.rollback()
        _flash_err('Error al eliminar el pago.', e)
    return redirect(url_for('detalle_servicio', id=serv_id))


@app.route('/api/servicios/ocr-recibo', methods=['POST'])
@login_required
def api_ocr_recibo():
    """Recibe imagen o PDF de un recibo y devuelve monto y fecha extraídos por OCR."""
    if 'archivo' not in request.files:
        return jsonify({'error': 'No se recibió ningún archivo.'}), 400
    archivo = request.files['archivo']
    if not archivo.filename:
        return jsonify({'error': 'Archivo vacío.'}), 400

    ext = archivo.filename.rsplit('.', 1)[-1].lower() if '.' in archivo.filename else ''
    if ext not in ('jpg', 'jpeg', 'png', 'webp', 'pdf'):
        return jsonify({'error': 'Formato no soportado. Usa JPG, PNG o PDF.'}), 400

    try:
        import pytesseract
        from PIL import Image
        import io as _io

        # Rutas explícitas para gunicorn (PATH reducido en systemd)
        pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'
        TESSERACT_CONFIG = '--oem 1 --psm 6'  # LSTM engine, bloque uniforme de texto

        contenido = archivo.read()

        if ext == 'pdf':
            try:
                from pdf2image import convert_from_bytes
                # 150 DPI es suficiente para recibos y procesa ~2x más rápido que 200
                paginas = convert_from_bytes(
                    contenido, first_page=1, last_page=1, dpi=150,
                    poppler_path='/usr/bin'
                )
                texto = '\n'.join(
                    pytesseract.image_to_string(p, lang='spa', config=TESSERACT_CONFIG)
                    for p in paginas
                )
            except ImportError:
                return jsonify({'error': 'pdf2image no instalado en el servidor.'}), 503
        else:
            img = Image.open(_io.BytesIO(contenido))
            # Convertir a escala de grises mejora velocidad y precisión
            img = img.convert('L')
            if img.width < 1200:
                factor = 1200 / img.width
                img = img.resize((int(img.width * factor), int(img.height * factor)), Image.LANCZOS)
            texto = pytesseract.image_to_string(img, lang='spa', config=TESSERACT_CONFIG)

        from servicios_ocr import analizar_recibo
        resultado = analizar_recibo(texto)
        return jsonify(resultado)

    except ImportError:
        return jsonify({'error': 'Tesseract / pytesseract no instalado en el servidor.'}), 503
    except Exception as e:
        current_app.logger.error(f'OCR recibo: {e}')
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500


@app.errorhandler(403)
def forbidden(e):
    return render_template('403.html'), 403


# --- Inicialización ---
if __name__ == '__main__':
    modo_debug = os.environ.get('FLASK_DEBUG', 'False') == 'True'
    app.run(host='0.0.0.0', port=5000, debug=modo_debug)
