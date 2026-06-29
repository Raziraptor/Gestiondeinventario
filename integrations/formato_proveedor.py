"""
Generación de archivos de OC en formato personalizado por proveedor.

Cada proveedor puede tener un FormatoProveedor con columnas mapeadas
desde campos internos (ej: "producto.codigo") a headers propios (ej: "Item Number").
Soporta salida en CSV y XLSX.
"""

import io
import csv
from decimal import Decimal
from datetime import datetime

# Campos internos disponibles para mapear. Tupla (id_campo, etiqueta_legible).
CAMPOS_DISPONIBLES = [
    ('producto.codigo',          'SKU / Código'),
    ('producto.nombre',          'Nombre del Producto'),
    ('producto.hd_sku',          'HD SKU'),
    ('producto.precio_unitario', 'Precio Unitario (catálogo)'),
    ('cantidad_solicitada',      'Cantidad'),
    ('costo_unitario_estimado',  'Costo Unitario Estimado'),
    ('cajas',                    'Cajas'),
    ('subtotal',                 'Subtotal'),
    ('enlace_proveedor',         'Enlace al Proveedor'),
    ('orden.id',                 'ID de Orden'),
    ('orden.fecha_creacion',     'Fecha de Orden'),
    ('orden.proveedor.nombre',   'Nombre del Proveedor'),
]

_CAMPO_IDS = {c for c, _ in CAMPOS_DISPONIBLES}


def _resolver_campo(campo: str, detalle, orden) -> str:
    """Resuelve un campo dot-notation contra el detalle o la orden."""
    partes = campo.split('.')
    if partes[0] == 'orden':
        obj = orden
        resto = partes[1:]
    else:
        obj = detalle
        resto = partes

    for attr in resto:
        if obj is None:
            return ''
        obj = getattr(obj, attr, None)

    if obj is None:
        return ''
    if isinstance(obj, datetime):
        return obj.strftime('%Y-%m-%d')
    if isinstance(obj, Decimal):
        return str(obj)
    return str(obj)


def generar_archivo(orden, formato) -> tuple[bytes, str, list[str]]:
    """
    Genera el archivo de OC en el formato del proveedor.

    Returns:
        (bytes, mimetype, omitidos)

    Raises:
        ValueError: si formato.columnas está vacío.
    """
    columnas = formato.columnas or []
    if not columnas:
        raise ValueError('El formato no tiene columnas configuradas.')

    campos  = [c['campo']  for c in columnas]
    headers = [c['header'] for c in columnas]

    # Validar campos (ignorar columnas con campo desconocido para no reventar)
    campos_validos  = [c for c in campos  if c in _CAMPO_IDS]
    headers_validos = [h for c, h in zip(campos, headers) if c in _CAMPO_IDS]
    if not campos_validos:
        raise ValueError('Ninguna columna configurada tiene un campo válido.')

    omitidos = []
    filas = []
    for detalle in orden.detalles:
        if not detalle.producto:
            omitidos.append(f'Detalle #{detalle.id} sin producto')
            continue
        fila = [_resolver_campo(c, detalle, orden) for c in campos_validos]
        filas.append(fila)

    tipo = (formato.tipo_archivo or 'xlsx').lower()

    if tipo == 'csv':
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers_validos)
        writer.writerows(filas)
        return buf.getvalue().encode('utf-8'), 'text/csv; charset=utf-8', omitidos

    # XLSX por defecto
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Orden de Compra'

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, header in enumerate(headers_validos, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    for row_idx, fila in enumerate(filas, 2):
        for col_idx, valor in enumerate(fila, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    buf = io.BytesIO()
    wb.save(buf)
    return (buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            omitidos)
