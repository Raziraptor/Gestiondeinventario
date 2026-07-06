"""
Blueprint inventory — productos, categorías, proveedores, almacenes, salidas,
transferencias y ajustes de inventario.

Rutas:
  configurar_etiqueta_diseno, configurar_excel_diseno, configurar_etiqueta,
  generar_etiqueta_personalizada, descargar_template_importacion, importar_productos,
  nuevo_producto, editar_producto, generar_etiqueta, historial_producto,
  lista_categorias, nueva_categoria, editar_categoria, eliminar_categoria,
  lista_proveedores, nuevo_proveedor, editar_proveedor, guardar_integracion_hd,
  lista_almacenes, nuevo_almacen, editar_almacen, eliminar_almacen,
  gestionar_inventario_almacen, eliminar_producto_de_almacen,
  lista_productos_sin_almacen, asignar_producto_rapido,
  historial_salidas, ver_salida, registrar_salida, eliminar_movimiento_salida,
  generar_salida_pdf, nueva_transferencia, nuevo_ajuste
"""

import io
import os
import csv
import uuid
import calendar
from collections import defaultdict

from flask import (
    render_template, request, redirect, url_for, flash, send_file,
    make_response, jsonify, current_app,
)
from flask_login import login_required, current_user
from sqlalchemy import extract
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from PIL import Image as PILImage, ImageDraw, ImageFont
import qrcode

from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    Image as ReportLabImage,
)

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from . import inventory_bp
from app.extensions import db
from app.helpers import (
    now_mx, _flash_err, check_org_permission, check_permission,
    get_item_or_404, admin_required, allowed_file, save_picture, log_actividad,
)
from app.models import (
    Producto, Stock, Movimiento, Categoria, Proveedor, Almacen,
    Organizacion, Salida, AuditLog, ProveedorIntegracion, FormatoProveedor,
)
from integrations.formato_proveedor import CAMPOS_DISPONIBLES


# ==============================================================================
# HELPERS LOCALES
# ==============================================================================

def _check_and_alert_stock_bajo(org_id, almacen_id):
    try:
        from app.blueprints.api.routes import check_and_alert_stock_bajo
        check_and_alert_stock_bajo(org_id, almacen_id)
    except Exception:
        pass


# ── PDF helpers ───────────────────────────────────────────────────────────────

def _pdf_estilos(org):
    fuente = org.tipo_letra if org.tipo_letra in ['Helvetica', 'Times-Roman', 'Courier'] else 'Helvetica'
    c_pri = colors.HexColor(org.color_primario)   if org.color_primario   else colors.HexColor('#333333')
    c_sec = colors.HexColor(org.color_secundario) if org.color_secundario else colors.HexColor('#f1f5f9')
    return fuente, c_pri, c_sec


def _pdf_bold(fuente):
    return 'Times-Bold' if fuente == 'Times-Roman' else f'{fuente}-Bold'


def _pdf_header(story, org, styles):
    fuente, c_pri, _ = _pdf_estilos(org)
    s_brand = ParagraphStyle('_Brand', fontName=_pdf_bold(fuente), fontSize=20, leading=22, textColor=colors.black, spaceAfter=2)
    s_sub   = ParagraphStyle('_Sub',   fontName=fuente, fontSize=10, leading=12, textColor=colors.gray)
    s_meta  = ParagraphStyle('_Meta',  fontName=fuente, fontSize=8,  leading=10, textColor=colors.HexColor('#64748b'))

    logo_el = []
    if org.logo_url:
        logo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], org.logo_url)
        if os.path.exists(logo_path):
            img = ReportLabImage(logo_path)
            max_h = 1.0 * inch
            img.drawHeight = max_h
            img.drawWidth  = max_h * (img.imageWidth / float(img.imageHeight))
            logo_el.append(img)

    text_el = [Paragraph(org.header_titulo or org.nombre, s_brand)]
    if org.header_subtitulo:
        text_el.append(Paragraph(org.header_subtitulo, s_sub))

    meta_parts = []
    if org.rfc:             meta_parts.append(f'RFC: {org.rfc}')
    if org.correo_empresa:  meta_parts.append(org.correo_empresa)
    if org.telefono:        meta_parts.append(org.telefono)
    if org.direccion:       meta_parts.append(org.direccion)
    if meta_parts:
        text_el.append(Paragraph(' · '.join(meta_parts), s_meta))

    if logo_el:
        t_hdr = Table([[logo_el, text_el]], colWidths=[1.5*inch, 4.7*inch])
    else:
        t_hdr = Table([[text_el]], colWidths=[6.2*inch])
    t_hdr.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(t_hdr)
    story.append(Table([['']], colWidths=[6.2*inch], rowHeights=[2],
                       style=TableStyle([('BACKGROUND', (0,0), (-1,-1), c_pri)])))
    story.append(Spacer(1, 0.2*inch))


def _pdf_footer(story, org, doc_url=None):
    import qrcode as _qrcode
    fuente, _, _ = _pdf_estilos(org)
    s_footer = ParagraphStyle('_Foot', fontName=fuente, fontSize=8,
                               textColor=colors.HexColor('#64748b'), alignment=TA_CENTER)
    story.append(Spacer(1, 0.3*inch))
    story.append(Table([['']], colWidths=[6.2*inch], rowHeights=[1],
                       style=TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#e2e8f0'))])))
    story.append(Spacer(1, 0.1*inch))

    pie_parts = []
    if org.footer_texto:
        pie_parts.append(org.footer_texto)
    pie_parts.append(f"Generado el {now_mx().strftime('%d/%m/%Y a las %H:%M')} · {org.nombre}")
    story.append(Paragraph('<br/>'.join(pie_parts), s_footer))

    if org.pdf_mostrar_qr and doc_url:
        try:
            qr = _qrcode.QRCode(version=1, box_size=4, border=2)
            qr.add_data(doc_url)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color='black', back_color='white')
            qr_buf = io.BytesIO()
            qr_img.save(qr_buf, format='PNG')
            qr_buf.seek(0)
            rl_qr = ReportLabImage(qr_buf)
            rl_qr.drawWidth  = 0.7 * inch
            rl_qr.drawHeight = 0.7 * inch
            story.append(Spacer(1, 0.1*inch))
            t_qr = Table([[rl_qr]], colWidths=[6.2*inch])
            t_qr.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
            story.append(t_qr)
        except Exception:
            pass


def _pdf_row_styles(data_len, c_sec):
    styles = []
    for i in range(1, data_len):
        bg = c_sec if i % 2 == 0 else colors.white
        styles.append(('BACKGROUND', (0, i), (-1, i), bg))
    return styles


# ==============================================================================
# CONFIGURACIÓN DE ETIQUETAS Y EXCEL
# ==============================================================================

@inventory_bp.route('/configuracion/etiquetas', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_management')
def configurar_etiqueta_diseno():
    org = Organizacion.query.get_or_404(current_user.organizacion_id)
    if request.method == 'POST':
        fuentes_validas = {'Inter','Roboto','Montserrat','Poppins','Oswald','CenturyGothic'}
        estilos_validos = {'moderno','bold','minimalista','dark','color'}
        f = request.form.get('fuente', 'Inter')
        e = request.form.get('estilo', 'moderno')
        org.etiqueta_fuente       = f if f in fuentes_validas else 'Inter'
        org.etiqueta_color_fondo  = request.form.get('color_fondo', '#FFFFFF')[:7]
        org.etiqueta_color_texto  = request.form.get('color_texto', '#1a1a1a')[:7]
        org.etiqueta_color_sku    = request.form.get('color_sku',   '#1f4e79')[:7]
        org.etiqueta_estilo       = e if e in estilos_validos else 'moderno'
        org.etiqueta_mostrar_logo = 'mostrar_logo' in request.form
        db.session.commit()
        flash('Diseño de etiquetas guardado.', 'success')
        return redirect(url_for('inventory.configurar_etiqueta_diseno'))
    return render_template('etiqueta_personalizar.html', org=org)


@inventory_bp.route('/configuracion/excel', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_view_management')
def configurar_excel_diseno():
    org = Organizacion.query.get_or_404(current_user.organizacion_id)
    if request.method == 'POST':
        fuentes_validas = {'Calibri', 'Arial', 'Trebuchet MS', 'Times New Roman'}
        f = request.form.get('excel_fuente', 'Calibri')
        org.excel_fuente         = f if f in fuentes_validas else 'Calibri'
        org.excel_color_header   = request.form.get('excel_color_header', '#1f4e79')[:7]
        org.excel_color_accent   = request.form.get('excel_color_accent', '#dbeafe')[:7]
        org.excel_mostrar_logo   = 'excel_mostrar_logo'   in request.form
        org.excel_mostrar_id     = 'excel_mostrar_id'     in request.form
        org.excel_mostrar_oc     = 'excel_mostrar_oc'     in request.form
        org.excel_mostrar_origen = 'excel_mostrar_origen' in request.form
        db.session.commit()
        flash('Diseño de Excel guardado. ✓', 'success')
        return redirect(url_for('inventory.configurar_excel_diseno'))
    return render_template('excel_config.html', org=org)


@inventory_bp.route('/producto/<int:id>/etiqueta/configurar')
@login_required
@check_permission('perm_view_dashboard')
def configurar_etiqueta(id):
    producto = get_item_or_404(Producto, id)
    almacen_seleccionado = request.args.get('almacen_id', type=int)
    org = Organizacion.query.get(current_user.organizacion_id)
    return render_template('etiqueta_config.html',
                           producto=producto,
                           almacen_seleccionado=almacen_seleccionado,
                           org=org)


@inventory_bp.route('/producto/<int:id>/etiqueta/generar', methods=['POST'])
@login_required
@check_permission('perm_view_dashboard')
def generar_etiqueta_personalizada(id):
    producto = get_item_or_404(Producto, id)
    org      = Organizacion.query.get(current_user.organizacion_id)

    almacen_id = request.form.get('almacen_id')
    ubicacion  = "N/A"
    if almacen_id:
        st = Stock.query.filter_by(producto_id=id, almacen_id=almacen_id).first()
        if st and st.ubicacion:
            ubicacion = st.ubicacion

    tamano = request.form.get('tamano', '1x3')
    DPI    = 300

    if tamano == '1.75x4':
        width_px, height_px = int(4 * DPI), int(1.75 * DPI)
        fs_nombre, fs_codigo, fs_ubic = 75, 95, 45
        qr_box, margin, gap = 13, 30, 30
    else:
        width_px, height_px = int(3 * DPI), int(1 * DPI)
        fs_nombre, fs_codigo, fs_ubic = 50, 65, 35
        qr_box, margin, gap = 8, 20, 20

    fuente       = getattr(org, 'etiqueta_fuente',       None) or 'Inter'
    color_fondo  = getattr(org, 'etiqueta_color_fondo',  None) or '#FFFFFF'
    color_texto  = getattr(org, 'etiqueta_color_texto',  None) or '#1a1a1a'
    color_sku    = getattr(org, 'etiqueta_color_sku',    None) or '#1f4e79'
    estilo       = getattr(org, 'etiqueta_estilo',       None) or 'moderno'
    mostrar_logo = getattr(org, 'etiqueta_mostrar_logo', True)

    if estilo == 'bold':
        fs_nombre = int(fs_nombre * 1.18)
        fs_codigo = int(fs_codigo * 1.18)
    elif estilo == 'compacto':
        fs_nombre = int(fs_nombre * 0.82)
        fs_codigo = int(fs_codigo * 0.82)
        fs_ubic   = int(fs_ubic   * 0.82)

    def hex2rgb(h):
        h = h.lstrip('#')
        try:
            return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            return (255, 255, 255)

    FONT_MAP = {
        'Inter':         ('Inter-Regular.ttf',     'Inter-Bold.ttf'),
        'Roboto':        ('Roboto-Regular.ttf',     'Roboto-Bold.ttf'),
        'Montserrat':    ('Montserrat-Regular.ttf', 'Montserrat-Bold.ttf'),
        'Poppins':       ('Poppins-Regular.ttf',    'Poppins-Bold.ttf'),
        'Oswald':        ('Oswald-Regular.ttf',     'Oswald-Bold.ttf'),
        'CenturyGothic': ('CenturyGothic.ttf',      'CenturyGothic-Bold.ttf'),
    }
    reg_file, bold_file = FONT_MAP.get(fuente, ('Inter-Regular.ttf', 'Inter-Bold.ttf'))
    fonts_dir = os.path.join(current_app.root_path, 'static', 'fonts')

    def _font(filename, size):
        path = os.path.join(fonts_dir, filename)
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                pass
        for fb in ('arial.ttf', 'ArialMT.ttf', 'DejaVuSans.ttf'):
            try:
                return ImageFont.truetype(fb, size)
            except Exception:
                pass
        return ImageFont.load_default()

    fnt_nombre = _font(reg_file,  fs_nombre)
    fnt_codigo = _font(bold_file, fs_codigo)
    fnt_ubic   = _font(reg_file,  fs_ubic)

    img = PILImage.new('RGB', (width_px, height_px), color=hex2rgb(color_fondo))
    d   = ImageDraw.Draw(img)

    if estilo == 'minimalista':
        d.rectangle([(3, 3), (width_px - 4, height_px - 4)],
                    outline=hex2rgb('#cbd5e1'), width=3)

    is_dark = hex2rgb(color_fondo)[0] < 128
    qr_fg   = hex2rgb('#FFFFFF') if is_dark else hex2rgb('#000000')
    qr_bg   = hex2rgb(color_fondo)
    qr_wrapper = qrcode.QRCode(box_size=qr_box, border=0)
    qr_wrapper.add_data(producto.codigo)
    qr_wrapper.make(fit=True)
    qr_img = qr_wrapper.make_image(fill_color=qr_fg, back_color=qr_bg).convert('RGB')
    qr_w, qr_h = qr_img.size
    x_qr = int(width_px - qr_w - margin)
    y_qr = int(height_px - qr_h - margin)
    img.paste(qr_img, (x_qr, y_qr))

    cur_y = margin
    qr_top = y_qr

    def _max_w(cur_y, fsize):
        return (int(width_px - margin * 2) if (cur_y + fsize) < qr_top
                else int(x_qr - margin - gap))

    nom = producto.nombre
    mw  = _max_w(cur_y, fs_nombre)
    while d.textlength(nom + '…', font=fnt_nombre) > mw and nom:
        nom = nom[:-1]
    if nom != producto.nombre:
        nom += '…'
    d.text((margin, cur_y), nom, font=fnt_nombre, fill=hex2rgb(color_texto))
    cur_y += fs_nombre + 5

    cod = producto.codigo
    mw2 = _max_w(cur_y, fs_codigo)
    while d.textlength(cod, font=fnt_codigo) > mw2 and cod:
        cod = cod[:-1]
    d.text((margin, cur_y), cod, font=fnt_codigo, fill=hex2rgb(color_sku))
    cur_y += fs_codigo + 5

    ubic_txt = f"UBIC: {ubicacion}" if ubicacion and ubicacion != "N/A" else f"ID: {producto.id}"
    mw3 = _max_w(cur_y, fs_ubic)
    while d.textlength(ubic_txt, font=fnt_ubic) > mw3 and ubic_txt:
        ubic_txt = ubic_txt[:-1]
    d.text((margin, cur_y), ubic_txt, font=fnt_ubic, fill=hex2rgb(color_texto))
    cur_y += fs_ubic + 10

    if mostrar_logo and producto.imagen_url:
        avail_h = int(height_px - cur_y - margin)
        avail_w = int(x_qr - margin - 10)
        if avail_h > 20:
            path_img = os.path.join(current_app.config['UPLOAD_FOLDER'], producto.imagen_url)
            if os.path.exists(path_img):
                try:
                    prod_img = PILImage.open(path_img)
                    prod_img.thumbnail((avail_w, avail_h))
                    img.paste(prod_img, (margin, cur_y))
                except Exception:
                    pass

    buffer   = io.BytesIO()
    img.save(buffer, 'JPEG', quality=100)
    buffer.seek(0)
    filename = f"Etiqueta_{secure_filename(producto.nombre)}_{tamano}.jpg"
    return send_file(buffer, mimetype='image/jpeg', as_attachment=True, download_name=filename)


# ==============================================================================
# IMPORTACIÓN DE PRODUCTOS
# ==============================================================================

@inventory_bp.route('/productos/importar/template')
@login_required
@check_permission('perm_edit_management')
def descargar_template_importacion():
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers      = ["nombre*", "codigo_sku*", "precio_unitario", "categoria", "proveedor", "unidades_por_caja"]
    header_fill  = PatternFill("solid", fgColor="4F46E5")
    header_font  = Font(bold=True, color="FFFFFF")
    example_fill = PatternFill("solid", fgColor="F0F0FF")

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    examples = [
        ["Clavos 2 pulgadas", "CLV-2IN-100", 45.50, "Ferretería",  "Proveedor Central",  100],
        ["Pintura Blanca 1L",  "PIN-BL-001",  89.00, "Pinturas",    "Distribuidora ABC",   12],
        ["Lija Grano 120",     "LIJ-120",     12.00, "",            "",                    50],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).fill = example_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    ws_info = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("INSTRUCCIONES DE IMPORTACIÓN", True),
        ("", False),
        ("Columnas obligatorias (marcadas con *):", True),
        ("  • nombre*        → Nombre del producto", False),
        ("  • codigo_sku*    → Código único (SKU). Si ya existe se omite.", False),
        ("", False),
        ("Columnas opcionales:", True),
        ("  • precio_unitario → Número decimal. Default: 0", False),
        ("  • categoria       → Nombre exacto. Si no existe se crea automáticamente.", False),
        ("  • proveedor       → Nombre exacto. Si no existe se crea automáticamente.", False),
        ("  • unidades_por_caja → Número entero. Default: 1", False),
        ("", False),
        ("NOTAS:", True),
        ("  • Elimina las filas de ejemplo antes de importar.", False),
        ("  • No modifiques los nombres de las columnas.", False),
        ("  • Puedes importar .xlsx o .csv", False),
    ]
    for row_idx, (text, bold) in enumerate(instrucciones, 1):
        cell      = ws_info.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold)
    ws_info.column_dimensions["A"].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="template_importacion_productos.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@inventory_bp.route('/productos/importar', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def importar_productos():
    org_id     = current_user.organizacion_id
    resultados = None

    if request.method == 'POST':
        archivo = request.files.get('archivo')
        if not archivo or archivo.filename == '':
            flash('Selecciona un archivo para importar.', 'danger')
            return redirect(url_for('inventory.importar_productos'))

        ext = archivo.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('csv', 'xlsx'):
            flash('Solo se aceptan archivos .csv o .xlsx', 'danger')
            return redirect(url_for('inventory.importar_productos'))

        try:
            filas = []
            if ext == 'xlsx':
                wb_imp = load_workbook(io.BytesIO(archivo.read()), data_only=True)
                ws_imp = wb_imp.active
                rows   = list(ws_imp.iter_rows(values_only=True))
                if not rows:
                    flash('El archivo está vacío.', 'danger')
                    return redirect(url_for('inventory.importar_productos'))
                headers = [str(h).strip().lower() if h else '' for h in rows[0]]
                filas   = [dict(zip(headers, row)) for row in rows[1:]]
            else:
                content = archivo.read().decode('utf-8-sig')
                reader  = csv.DictReader(io.StringIO(content))
                headers = [h.strip().lower() for h in (reader.fieldnames or [])]
                filas   = [{k.strip().lower(): v for k, v in row.items()} for row in reader]

            col = lambda row, *keys: next(
                (str(row.get(k) or '').strip() for k in keys if row.get(k) not in (None, '')), ''
            )

            importados, omitidos, errores = 0, 0, []

            # Pre-load lookups to avoid per-row queries
            categorias_dict = {c.nombre.lower(): c for c in Categoria.query.filter_by(organizacion_id=org_id).all()}
            proveedores_dict = {p.nombre.lower(): p for p in Proveedor.query.filter_by(organizacion_id=org_id).all()}
            existing_skus = set(
                row[0] for row in Producto.query.filter(
                    Producto.organizacion_id == org_id, Producto.codigo.isnot(None)
                ).with_entities(Producto.codigo).all()
            )

            for idx, fila in enumerate(filas, 2):
                nombre = col(fila, 'nombre', 'name')
                codigo = col(fila, 'codigo_sku', 'codigo', 'sku', 'code')

                if not nombre or not codigo:
                    if any(v for v in fila.values() if str(v or '').strip()):
                        errores.append(f"Fila {idx}: 'nombre' y 'codigo_sku' son obligatorios.")
                    continue

                if codigo in existing_skus:
                    omitidos += 1
                    continue

                try:
                    precio = float(col(fila, 'precio_unitario', 'precio') or 0)
                except ValueError:
                    precio = 0.0
                try:
                    upc = int(col(fila, 'unidades_por_caja', 'unidades') or 1)
                except ValueError:
                    upc = 1

                cat_nombre = col(fila, 'categoria', 'category')
                categoria  = None
                if cat_nombre:
                    categoria = categorias_dict.get(cat_nombre.lower())
                    if not categoria:
                        categoria = Categoria(nombre=cat_nombre, organizacion_id=org_id)
                        db.session.add(categoria)
                        db.session.flush()
                        categorias_dict[cat_nombre.lower()] = categoria

                prov_nombre = col(fila, 'proveedor', 'supplier', 'proveedor_nombre')
                proveedor   = None
                if prov_nombre:
                    proveedor = proveedores_dict.get(prov_nombre.lower())
                    if not proveedor:
                        proveedor = Proveedor(nombre=prov_nombre, organizacion_id=org_id)
                        db.session.add(proveedor)
                        db.session.flush()
                        proveedores_dict[prov_nombre.lower()] = proveedor

                db.session.add(Producto(
                    nombre=nombre, codigo=codigo, precio_unitario=precio,
                    categoria_id=categoria.id if categoria else None,
                    proveedor_id=proveedor.id if proveedor else None,
                    unidades_por_caja=upc, organizacion_id=org_id,
                ))
                existing_skus.add(codigo)
                importados += 1

            if importados > 0:
                log_actividad('importar', 'producto',
                              f'Importación masiva: {importados} producto(s) creados, {omitidos} omitidos (SKU duplicado)')
            db.session.commit()
            resultados = {'importados': importados, 'omitidos': omitidos, 'errores': errores}

        except Exception as e:
            db.session.rollback()
            _flash_err('Error al procesar el archivo.', e)

    return render_template('importar_productos.html', titulo='Importar Productos', resultados=resultados)


# ==============================================================================
# PRODUCTOS
# ==============================================================================

@inventory_bp.route('/producto/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def nuevo_producto():
    org_id      = current_user.organizacion_id
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).all()
    categorias  = Categoria.query.filter_by(organizacion_id=org_id).all()
    almacenes   = Almacen.query.filter_by(organizacion_id=org_id).all()

    if request.method == 'POST':
        imagen_filename = None

        def repoblar():
            costo_val = request.form.get('costo_estandar')
            precio_tmp = float(costo_val) if costo_val and costo_val.strip() else 0.0
            producto_temporal = Producto(
                nombre=request.form.get('nombre'), codigo=request.form.get('codigo'),
                categoria_id=int(request.form.get('categoria_id') or 0) or None,
                precio_unitario=precio_tmp,
                proveedor_id=int(request.form.get('proveedor_id') or 0) or None,
                unidades_por_caja=int(request.form.get('unidades_por_caja') or 1),
                organizacion_id=org_id,
            )
            producto_temporal.costo_estandar = precio_tmp
            return render_template('producto_form.html', titulo="Nuevo Producto",
                                   proveedores=proveedores, categorias=categorias,
                                   almacenes=almacenes, producto=producto_temporal)

        if 'imagen' in request.files:
            file = request.files['imagen']
            if file.filename != '' and allowed_file(file.filename):
                ext = secure_filename(file.filename).rsplit('.', 1)[-1].lower()
                filename = f"{uuid.uuid4().hex}.{ext}"
                file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                imagen_filename = filename
            elif file.filename != '' and not allowed_file(file.filename):
                flash('Tipo de archivo de imagen no permitido.', 'danger')
                return repoblar()

        if not imagen_filename:
            ai_fn = secure_filename(request.form.get('ai_imagen_filename', '').strip())
            if ai_fn:
                ai_path = os.path.join(current_app.config['UPLOAD_FOLDER'], ai_fn)
                if os.path.isfile(ai_path):
                    imagen_filename = ai_fn

        try:
            costo_raw   = request.form.get('costo_estandar')
            precio_final = float(costo_raw) if costo_raw and costo_raw.strip() else 0.0

            nuevo_prod = Producto(
                nombre=request.form['nombre'], codigo=request.form['codigo'],
                categoria_id=request.form.get('categoria_id') or None,
                precio_unitario=precio_final, imagen_url=imagen_filename,
                proveedor_id=request.form.get('proveedor_id') or None,
                unidades_por_caja=int(request.form.get('unidades_por_caja', 1)),
                organizacion_id=current_user.organizacion_id,
                enlace_proveedor=request.form.get('enlace_proveedor'),
            )
            db.session.add(nuevo_prod)
            db.session.flush()

            cantidad_inicial   = int(request.form.get('cantidad_inicial') or 0)
            almacen_inicial_id = int(request.form.get('almacen_inicial_id') or 0)
            ubicacion_inicial  = request.form.get('ubicacion_inicial')

            almacen_seleccionado = None
            if almacen_inicial_id > 0:
                almacen_seleccionado = Almacen.query.filter_by(id=almacen_inicial_id, organizacion_id=org_id).first()

            if almacen_seleccionado:
                db.session.add(Stock(
                    producto_id=nuevo_prod.id, almacen_id=almacen_seleccionado.id,
                    cantidad=cantidad_inicial,
                    stock_minimo=int(request.form.get('stock_minimo') or 5),
                    stock_maximo=int(request.form.get('stock_maximo') or 100),
                    ubicacion=ubicacion_inicial,
                ))
                if cantidad_inicial > 0:
                    db.session.add(Movimiento(
                        producto_id=nuevo_prod.id, cantidad=cantidad_inicial,
                        tipo='entrada-inicial', fecha=now_mx(),
                        motivo='Stock Inicial (Creación)',
                        almacen_id=almacen_inicial_id, organizacion_id=org_id,
                    ))

            log_actividad('crear', 'producto',
                          f'Producto creado: {nuevo_prod.nombre} (SKU: {nuevo_prod.codigo})',
                          entidad_id=nuevo_prod.id)
            db.session.commit()
            flash('Producto creado exitosamente.', 'success')

            if almacen_seleccionado:
                return redirect(url_for('inventory.gestionar_inventario_almacen', id=almacen_seleccionado.id))
            return redirect(url_for('main.dashboard'))

        except IntegrityError as e:
            db.session.rollback()
            if "producto_codigo_key" in str(e) or "UNIQUE constraint failed" in str(e):
                flash('Error: El Código (SKU) ya existe.', 'danger')
            else:
                _flash_err('Error de base de datos al guardar el producto.', e)
            return repoblar()
        except Exception as e:
            db.session.rollback()
            _flash_err('Error inesperado.', e)
            return repoblar()

    return render_template('producto_form.html', titulo="Nuevo Producto",
                           proveedores=proveedores, categorias=categorias,
                           almacenes=almacenes, producto=None)


@inventory_bp.route('/producto/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@check_permission('perm_edit_management')
def editar_producto(id):
    producto    = get_item_or_404(Producto, id)
    org_id      = current_user.organizacion_id if current_user.rol != 'super_admin' else producto.organizacion_id
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id).all()
    categorias  = Categoria.query.filter_by(organizacion_id=org_id).all()

    almacen_id = request.args.get('almacen_id', type=int)
    stock_item = None
    if almacen_id:
        stock_item = Stock.query.filter_by(producto_id=producto.id, almacen_id=almacen_id).first()

    if request.method == 'POST':
        try:
            if 'imagen' in request.files:
                file = request.files['imagen']
                if file.filename != '' and allowed_file(file.filename):
                    ext = secure_filename(file.filename).rsplit('.', 1)[-1].lower()
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                    producto.imagen_url = filename

            if not producto.imagen_url or request.files.get('imagen', '').filename == '':
                ai_fn = secure_filename(request.form.get('ai_imagen_filename', '').strip())
                if ai_fn:
                    ai_path = os.path.join(current_app.config['UPLOAD_FOLDER'], ai_fn)
                    if os.path.isfile(ai_path):
                        producto.imagen_url = ai_fn

            producto.nombre             = request.form['nombre']
            producto.codigo             = request.form['codigo']
            producto.categoria_id       = request.form.get('categoria_id') or None
            producto.proveedor_id       = request.form.get('proveedor_id') or None
            producto.unidades_por_caja  = int(request.form.get('unidades_por_caja') or 1)
            producto.enlace_proveedor   = request.form.get('enlace_proveedor')

            costo_raw = request.form.get('costo_estandar')
            producto.precio_unitario = float(costo_raw) if costo_raw and costo_raw.strip() else 0.0

            if stock_item:
                stock_item.stock_minimo = int(request.form.get('stock_minimo') or 0)
                stock_item.stock_maximo = int(request.form.get('stock_maximo') or 0)
                stock_item.cantidad     = int(request.form.get('cantidad') or 0)
                stock_item.ubicacion    = request.form.get('ubicacion')

            log_actividad('editar', 'producto',
                          f'Producto editado: {producto.nombre} (SKU: {producto.codigo})',
                          entidad_id=producto.id)
            db.session.commit()
            flash('Producto actualizado exitosamente', 'success')
            return redirect(
                url_for('inventory.gestionar_inventario_almacen', id=almacen_id)
                if almacen_id else url_for('main.dashboard')
            )

        except Exception as e:
            db.session.rollback()
            _flash_err('Error al actualizar producto. Intenta de nuevo.', e)

    producto.costo_estandar = producto.precio_unitario
    return render_template('producto_form.html', titulo="Editar Producto",
                           producto=producto, proveedores=proveedores,
                           categorias=categorias, stock_item=stock_item)


@inventory_bp.route('/producto/<int:id>/etiqueta')
@login_required
@check_permission('perm_view_dashboard')
def generar_etiqueta(id):
    producto = get_item_or_404(Producto, id)
    try:
        buffer      = io.BytesIO()
        label_width = 4 * inch
        label_height = 2.5 * inch
        c = rl_canvas.Canvas(buffer, pagesize=(label_width, label_height))

        qr_img      = qrcode.make(producto.codigo)
        qr_img_path = io.BytesIO()
        qr_img.save(qr_img_path, format='PNG')
        qr_img_path.seek(0)
        qr_para_pdf = ImageReader(qr_img_path)
        c.drawImage(qr_para_pdf, label_width - (1.6 * inch), 0.6 * inch,
                    width=(1.4 * inch), height=(1.4 * inch), preserveAspectRatio=True)

        text_x = 0.25 * inch
        text_y = label_height - (0.5 * inch)
        c.setFont('Helvetica-Bold', 12)
        c.drawString(text_x, text_y, producto.nombre[:25])
        c.setFont('Helvetica', 10)
        c.drawString(text_x, text_y - (0.3 * inch), f"SKU: {producto.codigo}")
        c.drawString(text_x, text_y - (0.6 * inch), f"Precio: ${producto.precio_unitario:.2f}")

        if producto.imagen_url:
            img_path = os.path.join(current_app.config['UPLOAD_FOLDER'], producto.imagen_url)
            if os.path.exists(img_path):
                try:
                    prod_img = ImageReader(img_path)
                    c.drawImage(prod_img, 0.1 * inch, 0.2 * inch,
                                width=1.5 * inch, height=1.0 * inch,
                                preserveAspectRatio=True)
                except Exception:
                    pass

        c.showPage()
        c.save()
        buffer.seek(0)
        nombre_base    = secure_filename(producto.nombre)
        fecha_str      = now_mx().strftime("%Y-%m-%d")
        nombre_archivo = f"{nombre_base}_{fecha_str}.pdf"
        return send_file(buffer, as_attachment=False, download_name=nombre_archivo, mimetype='application/pdf')
    except Exception as e:
        _flash_err('Error al generar etiqueta.', e)
        return redirect(url_for('main.index'))


@inventory_bp.route('/producto/<int:id>/historial')
@login_required
@check_permission('perm_view_dashboard')
def historial_producto(id):
    producto = get_item_or_404(Producto, id)

    stocks_actuales = Stock.query.filter_by(producto_id=id).join(Almacen).filter(
        Almacen.organizacion_id == current_user.organizacion_id
    ).order_by(Almacen.nombre).all()

    total_global = sum(s.cantidad for s in stocks_actuales)

    movimientos_query = Movimiento.query.filter_by(producto_id=id).options(
        joinedload(Movimiento.almacen)
    ).order_by(Movimiento.fecha.desc())
    if current_user.rol != 'super_admin':
        movimientos_query = movimientos_query.filter(
            Movimiento.organizacion_id == current_user.organizacion_id
        )
    movimientos = movimientos_query.all()

    historial_por_almacen = defaultdict(list)
    for m in movimientos:
        alm_nombre = m.almacen.nombre if m.almacen_id and m.almacen else "Sin Almacén"
        historial_por_almacen[alm_nombre].append(m)

    return render_template('historial_producto.html',
                           producto=producto,
                           historial_por_almacen=historial_por_almacen,
                           stocks_actuales=stocks_actuales,
                           total_global=total_global)


# ==============================================================================
# CATEGORÍAS
# ==============================================================================

@inventory_bp.route('/categorias')
@login_required
@check_org_permission
@check_permission('perm_view_management')
def lista_categorias():
    if current_user.rol == 'super_admin':
        categorias = Categoria.query.all()
    else:
        categorias = Categoria.query.filter_by(organizacion_id=current_user.organizacion_id).all()
    return render_template('categorias.html', categorias=categorias)


@inventory_bp.route('/categoria/nueva', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def nueva_categoria():
    if request.method == 'POST':
        try:
            db.session.add(Categoria(
                nombre=request.form['nombre'],
                descripcion=request.form.get('descripcion'),
                organizacion_id=current_user.organizacion_id,
            ))
            db.session.commit()
            flash('Categoría creada exitosamente', 'success')
            return redirect(url_for('inventory.lista_categorias'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al crear la categoría.', e)
    return render_template('categoria_form.html', titulo="Nueva Categoría", categoria=None)


@inventory_bp.route('/categoria/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@check_permission('perm_edit_management')
def editar_categoria(id):
    categoria = get_item_or_404(Categoria, id)
    if request.method == 'POST':
        try:
            categoria.nombre      = request.form['nombre']
            categoria.descripcion = request.form.get('descripcion')
            db.session.commit()
            flash('Categoría actualizada exitosamente', 'success')
            return redirect(url_for('inventory.lista_categorias'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al actualizar la categoría.', e)
    return render_template('categoria_form.html', titulo="Editar Categoría", categoria=categoria)


@inventory_bp.route('/categoria/eliminar/<int:id>', methods=['POST'])
@login_required
@check_permission('perm_edit_management')
def eliminar_categoria(id):
    categoria = get_item_or_404(Categoria, id)
    try:
        org_id = categoria.organizacion_id
        for p in Producto.query.filter_by(categoria_id=categoria.id, organizacion_id=org_id).all():
            p.categoria_id = None
        db.session.delete(categoria)
        db.session.commit()
        flash(f'Categoría "{categoria.nombre}" eliminada. Los productos asociados fueron des-asignados.', 'success')
    except Exception as e:
        db.session.rollback()
        _flash_err('Error al eliminar la categoría.', e)
    return redirect(url_for('inventory.lista_categorias'))


# ==============================================================================
# PROVEEDORES
# ==============================================================================

@inventory_bp.route('/proveedores')
@login_required
@check_org_permission
@check_permission('perm_view_management')
def lista_proveedores():
    if current_user.rol == 'super_admin':
        proveedores = Proveedor.query.all()
    else:
        proveedores = Proveedor.query.filter_by(organizacion_id=current_user.organizacion_id).all()
    return render_template('proveedores.html', proveedores=proveedores)


@inventory_bp.route('/proveedor/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def nuevo_proveedor():
    if request.method == 'POST':
        try:
            db.session.add(Proveedor(
                nombre=request.form['nombre'],
                contacto_email=request.form.get('contacto_email'),
                contacto_telefono=request.form.get('contacto_telefono'),
                organizacion_id=current_user.organizacion_id,
            ))
            db.session.commit()
            flash('Proveedor creado exitosamente', 'success')
            return redirect(url_for('inventory.lista_proveedores'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al crear proveedor.', e)
    return render_template('proveedor_form.html', titulo="Nuevo Proveedor", proveedor=None)


@inventory_bp.route('/proveedor/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def editar_proveedor(id):
    proveedor = get_item_or_404(Proveedor, id)
    if request.method == 'POST':
        try:
            proveedor.nombre             = request.form['nombre']
            proveedor.contacto_email     = request.form.get('contacto_email')
            proveedor.contacto_telefono  = request.form.get('contacto_telefono')
            db.session.commit()
            flash('Proveedor actualizado exitosamente', 'success')
            return redirect(url_for('inventory.lista_proveedores'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al actualizar el proveedor.', e)

    org_id = current_user.organizacion_id
    integracion = ProveedorIntegracion.query.filter_by(
        proveedor_id=proveedor.id, organizacion_id=org_id
    ).first()
    formato_oc = FormatoProveedor.query.filter_by(
        proveedor_id=proveedor.id, organizacion_id=org_id
    ).first()
    return render_template('proveedor_form.html', titulo="Editar Proveedor",
                           proveedor=proveedor, integracion=integracion,
                           formato_oc=formato_oc, campos_disponibles=CAMPOS_DISPONIBLES)


@inventory_bp.route('/proveedor/<int:id>/integracion-hd', methods=['POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def guardar_integracion_hd(id):
    proveedor = get_item_or_404(Proveedor, id)
    org_id    = proveedor.organizacion_id

    if current_user.rol not in ('super_admin', 'admin'):
        flash('Solo administradores pueden configurar integraciones.', 'danger')
        return redirect(url_for('inventory.editar_proveedor', id=id))

    if not os.environ.get('FERNET_KEY'):
        flash('FERNET_KEY no configurada en el servidor. Contacta al administrador del sistema.', 'danger')
        return redirect(url_for('inventory.editar_proveedor', id=id))

    activo   = request.form.get('hd_activo') == '1'
    usuario  = request.form.get('hd_usuario', '').strip()
    password = request.form.get('hd_password', '').strip()

    integracion = ProveedorIntegracion.query.filter_by(
        proveedor_id=proveedor.id, organizacion_id=org_id
    ).first()

    try:
        if integracion is None:
            if not usuario or not password:
                flash('Usuario y contraseña son requeridos para activar la integración.', 'warning')
                return redirect(url_for('inventory.editar_proveedor', id=id))
            integracion = ProveedorIntegracion(
                proveedor_id=proveedor.id, organizacion_id=org_id,
                tipo='homedepot', activo=activo,
            )
            integracion.credenciales = {'usuario': usuario, 'password': password}
            db.session.add(integracion)
        else:
            integracion.activo = activo
            if usuario or password:
                creds_actuales = integracion.credenciales
                integracion.credenciales = {
                    'usuario':  usuario  or creds_actuales.get('usuario',  ''),
                    'password': password or creds_actuales.get('password', ''),
                }
        db.session.commit()
        flash('Integración con Home Depot Pro guardada.', 'success')
    except ValueError as e:
        db.session.rollback()
        flash(str(e), 'danger')
    except Exception as e:
        db.session.rollback()
        _flash_err('Error al guardar la integración.', e)

    return redirect(url_for('inventory.editar_proveedor', id=id))


@inventory_bp.route('/proveedor/<int:id>/formato-oc', methods=['POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def guardar_formato_oc(id):
    proveedor = get_item_or_404(Proveedor, id)
    org_id    = proveedor.organizacion_id

    if current_user.rol not in ('super_admin', 'admin'):
        flash('Solo administradores pueden configurar formatos de OC.', 'danger')
        return redirect(url_for('inventory.editar_proveedor', id=id))

    campos_validos = {c for c, _ in CAMPOS_DISPONIBLES}

    campos_form  = request.form.getlist('formato_campo[]')
    headers_form = request.form.getlist('formato_header[]')

    columnas = []
    for campo, header in zip(campos_form, headers_form):
        campo  = campo.strip()
        header = header.strip()
        if campo in campos_validos and header:
            columnas.append({'campo': campo, 'header': header})

    activo         = request.form.get('formato_activo') == '1'
    tipo_archivo   = request.form.get('formato_tipo', 'xlsx')
    nombre_archivo = request.form.get('formato_nombre', 'OC-{id}').strip() or 'OC-{id}'

    if tipo_archivo not in ('xlsx', 'csv'):
        tipo_archivo = 'xlsx'

    fmt = FormatoProveedor.query.filter_by(
        proveedor_id=proveedor.id, organizacion_id=org_id
    ).first()

    try:
        if fmt is None:
            fmt = FormatoProveedor(
                proveedor_id=proveedor.id, organizacion_id=org_id,
            )
            db.session.add(fmt)
        fmt.activo        = activo
        fmt.tipo_archivo  = tipo_archivo
        fmt.nombre_archivo = nombre_archivo
        fmt.columnas      = columnas  # asignación directa, no mutación in-place
        db.session.commit()
        flash('Formato de OC guardado.', 'success')
    except Exception as e:
        db.session.rollback()
        _flash_err('Error al guardar el formato de OC.', e)

    return redirect(url_for('inventory.editar_proveedor', id=id))


@inventory_bp.route('/proveedor/<int:id>/formato-oc/eliminar', methods=['POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def eliminar_formato_oc(id):
    proveedor = get_item_or_404(Proveedor, id)
    org_id    = proveedor.organizacion_id

    if current_user.rol not in ('super_admin', 'admin'):
        flash('Solo administradores pueden eliminar formatos de OC.', 'danger')
        return redirect(url_for('inventory.editar_proveedor', id=id))

    fmt = FormatoProveedor.query.filter_by(
        proveedor_id=proveedor.id, organizacion_id=org_id
    ).first()

    if fmt:
        try:
            db.session.delete(fmt)
            db.session.commit()
            flash('Formato de OC eliminado.', 'success')
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al eliminar el formato de OC.', e)
    else:
        flash('No hay formato de OC configurado para este proveedor.', 'warning')

    return redirect(url_for('inventory.editar_proveedor', id=id))


# ==============================================================================
# ALMACENES
# ==============================================================================

@inventory_bp.route('/almacenes')
@login_required
@admin_required
def lista_almacenes():
    if current_user.rol == 'super_admin':
        almacenes = Almacen.query.order_by(Almacen.id).all()
    else:
        almacenes = Almacen.query.filter_by(
            organizacion_id=current_user.organizacion_id
        ).order_by(Almacen.id).all()
    return render_template('almacenes.html', almacenes=almacenes, titulo="Gestionar Almacenes")


@inventory_bp.route('/almacen/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def nuevo_almacen():
    if request.method == 'POST':
        try:
            db.session.add(Almacen(
                nombre=request.form['nombre'],
                ubicacion=request.form.get('ubicacion'),
                organizacion_id=current_user.organizacion_id,
            ))
            db.session.commit()
            flash('Almacén creado exitosamente', 'success')
            return redirect(url_for('inventory.lista_almacenes'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al crear el almacén.', e)
    return render_template('almacen_form.html', titulo="Nuevo Almacén", almacen=None)


@inventory_bp.route('/almacen/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_almacen(id):
    almacen = get_item_or_404(Almacen, id)
    if request.method == 'POST':
        try:
            almacen.nombre    = request.form['nombre']
            almacen.ubicacion = request.form.get('ubicacion')
            db.session.commit()
            flash('Almacén actualizado exitosamente', 'success')
            return redirect(url_for('inventory.lista_almacenes'))
        except Exception as e:
            db.session.rollback()
            _flash_err('Error al actualizar el almacén.', e)
    return render_template('almacen_form.html', titulo="Editar Almacén", almacen=almacen)


@inventory_bp.route('/almacen/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_almacen(id):
    if current_user.rol not in ['super_admin', 'admin']:
        flash('No tienes permiso para eliminar almacenes.', 'danger')
        return redirect(url_for('inventory.lista_almacenes'))

    almacen = Almacen.query.filter_by(id=id, organizacion_id=current_user.organizacion_id).first_or_404()
    try:
        db.session.delete(almacen)
        db.session.commit()
        flash('Almacén eliminado correctamente.', 'success')
    except Exception as e:
        db.session.rollback()
        _flash_err('Error al eliminar el almacén.', e)
    return redirect(url_for('inventory.lista_almacenes'))


@inventory_bp.route('/almacen/<int:id>/inventario', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def gestionar_inventario_almacen(id):
    almacen = get_item_or_404(Almacen, id)
    org_id  = almacen.organizacion_id

    if request.method == 'POST':
        try:
            producto_id = int(request.form.get('producto_id'))
            ubicacion   = request.form.get('ubicacion')
            cantidad    = float(request.form.get('cantidad', 0))

            if not producto_id:
                raise Exception("No se seleccionó un producto.")

            if Stock.query.filter_by(almacen_id=id, producto_id=producto_id).first():
                flash('Ese producto ya está registrado en este almacén.', 'warning')
            else:
                db.session.add(Stock(
                    producto_id=producto_id, almacen_id=id,
                    cantidad=cantidad, stock_minimo=5, stock_maximo=100, ubicacion=ubicacion,
                ))
                if cantidad > 0:
                    db.session.add(Movimiento(
                        producto_id=producto_id, cantidad=cantidad, tipo='entrada-inicial',
                        fecha=now_mx(), motivo='Stock Inicial (Alta Manual en Almacén)',
                        almacen_id=id, organizacion_id=org_id,
                    ))
                db.session.commit()
                flash(f'Producto añadido al almacén con stock {cantidad}.', 'success')

        except Exception as e:
            db.session.rollback()
            _flash_err('Error al añadir producto.', e)

        return redirect(url_for('inventory.gestionar_inventario_almacen', id=id))

    productos_en_stock_ids = [s.producto_id for s in almacen.stocks]
    productos_para_anadir_json = [
        {
            "id": p.id, "nombre": p.nombre, "codigo": p.codigo,
            "unidades_por_caja": int(p.unidades_por_caja) if p.unidades_por_caja and p.unidades_por_caja > 0 else 1,
        }
        for p in Producto.query.filter_by(organizacion_id=org_id).all()
        if p.id not in productos_en_stock_ids
    ]

    return render_template('almacen_inventario.html',
                           titulo=f"Inventario de {almacen.nombre}",
                           almacen=almacen,
                           productos_para_anadir_json=productos_para_anadir_json)


@inventory_bp.route('/almacen/stock/eliminar/<int:id>', methods=['POST'])
@login_required
@admin_required
def eliminar_producto_de_almacen(id):
    stock_item = db.session.query(Stock).join(Almacen).filter(
        Stock.id == id,
        Almacen.organizacion_id == current_user.organizacion_id,
    ).first_or_404()
    almacen_id = stock_item.almacen_id
    try:
        nombre_prod = stock_item.producto.nombre
        nombre_alm  = stock_item.almacen.nombre
        db.session.delete(stock_item)
        db.session.commit()
        flash(f'Producto "{nombre_prod}" eliminado de "{nombre_alm}".', 'success')
    except Exception as e:
        db.session.rollback()
        _flash_err('Error al eliminar del almacén.', e)
    return redirect(url_for('inventory.gestionar_inventario_almacen', id=almacen_id))


@inventory_bp.route('/productos/sin-almacen')
@login_required
@check_org_permission
@check_permission('perm_view_management')
def lista_productos_sin_almacen():
    org_id = current_user.organizacion_id
    ids_con_stock = db.session.query(Stock.producto_id).join(Almacen).filter(
        Almacen.organizacion_id == org_id
    ).distinct()
    productos_huerfanos = Producto.query.filter(
        Producto.organizacion_id == org_id,
        ~Producto.id.in_(ids_con_stock),
    ).all()
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).all()
    return render_template('productos_sin_almacen.html',
                           titulo="Productos Sin Asignar",
                           productos=productos_huerfanos,
                           almacenes=almacenes)


@inventory_bp.route('/producto/asignar-rapido', methods=['POST'])
@login_required
@check_permission('perm_edit_management')
def asignar_producto_rapido():
    try:
        producto_id = int(request.form.get('producto_id'))
        almacen_id  = int(request.form.get('almacen_id'))

        if not producto_id or not almacen_id:
            raise Exception("Datos incompletos.")

        producto = get_item_or_404(Producto, producto_id)
        org_id   = current_user.organizacion_id
        if not Almacen.query.filter_by(id=almacen_id, organizacion_id=org_id).first():
            flash('Almacén no autorizado.', 'danger')
            return redirect(url_for('inventory.lista_productos_sin_almacen'))

        if Stock.query.filter_by(producto_id=producto_id, almacen_id=almacen_id).first():
            flash('El producto ya estaba en ese almacén.', 'warning')
        else:
            db.session.add(Stock(
                producto_id=producto_id, almacen_id=almacen_id,
                cantidad=0, stock_minimo=5, stock_maximo=100,
            ))
            db.session.commit()
            flash(f'Producto "{producto.nombre}" asignado correctamente.', 'success')

    except Exception as e:
        db.session.rollback()
        _flash_err('Error al asignar producto.', e)

    return redirect(url_for('inventory.lista_productos_sin_almacen'))


# ==============================================================================
# SALIDAS
# ==============================================================================

@inventory_bp.route('/salidas')
@login_required
@check_org_permission
@check_permission('perm_do_salidas')
def historial_salidas():
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    ahora = now_mx()
    if not mes: mes = ahora.month
    if not ano: ano = ahora.year

    meses_lista = [
        (1,'Enero'),(2,'Febrero'),(3,'Marzo'),(4,'Abril'),
        (5,'Mayo'),(6,'Junio'),(7,'Julio'),(8,'Agosto'),
        (9,'Septiembre'),(10,'Octubre'),(11,'Noviembre'),(12,'Diciembre'),
    ]

    query = Salida.query if current_user.rol == 'super_admin' else \
            Salida.query.filter_by(organizacion_id=current_user.organizacion_id)
    query = query.filter(
        extract('month', Salida.fecha) == mes,
        extract('year',  Salida.fecha) == ano,
    )
    page       = request.args.get('page', 1, type=int)
    pagination = query.options(
        joinedload(Salida.almacen)
    ).order_by(Salida.fecha.desc()).paginate(page=page, per_page=12, error_out=False)

    org_id = current_user.organizacion_id
    total_unidades = db.session.query(
        db.func.sum(db.func.abs(Movimiento.cantidad))
    ).filter(
        Movimiento.organizacion_id == org_id, Movimiento.tipo == 'salida',
        extract('month', Movimiento.fecha) == mes,
        extract('year',  Movimiento.fecha) == ano,
    ).scalar() or 0

    top_productos_raw = db.session.query(
        Movimiento.producto_id,
        db.func.sum(db.func.abs(Movimiento.cantidad)).label('uds'),
    ).filter(
        Movimiento.organizacion_id == org_id, Movimiento.tipo == 'salida',
        extract('month', Movimiento.fecha) == mes,
        extract('year',  Movimiento.fecha) == ano,
    ).group_by(Movimiento.producto_id
    ).order_by(db.func.sum(db.func.abs(Movimiento.cantidad)).desc()).limit(5).all()

    top_prod_ids = [row.producto_id for row in top_productos_raw]
    top_prod_map = {p.id: p for p in Producto.query.filter(Producto.id.in_(top_prod_ids)).all()} if top_prod_ids else {}
    top_productos = []
    for row in top_productos_raw:
        prod = top_prod_map.get(row.producto_id)
        if prod:
            top_productos.append({'nombre': prod.nombre, 'sku': prod.codigo, 'uds': int(row.uds)})

    dias_mes   = calendar.monthrange(ano, mes)[1]
    daily_raw  = db.session.query(
        db.func.extract('day', Movimiento.fecha).label('dia'),
        db.func.sum(db.func.abs(Movimiento.cantidad)).label('uds'),
    ).filter(
        Movimiento.organizacion_id == org_id, Movimiento.tipo == 'salida',
        extract('month', Movimiento.fecha) == mes,
        extract('year',  Movimiento.fecha) == ano,
    ).group_by('dia').all()

    daily_map   = {int(r.dia): int(r.uds) for r in daily_raw}
    chart_labels = list(range(1, dias_mes + 1))
    chart_data   = [daily_map.get(d, 0) for d in chart_labels]

    almacen_top_raw = db.session.query(
        Salida.almacen_id, db.func.count(Salida.id).label('total'),
    ).filter(
        Salida.organizacion_id == org_id,
        extract('month', Salida.fecha) == mes,
        extract('year',  Salida.fecha) == ano,
    ).group_by(Salida.almacen_id
    ).order_by(db.func.count(Salida.id).desc()).first()

    almacen_top = None
    if almacen_top_raw:
        a = Almacen.query.get(almacen_top_raw.almacen_id)
        if a:
            almacen_top = {'nombre': a.nombre, 'total': almacen_top_raw.total}

    return render_template('salidas.html',
                           salidas=pagination.items, pagination=pagination,
                           meses_lista=meses_lista, mes_seleccionado=mes, ano_seleccionado=ano,
                           total_unidades=total_unidades, top_productos=top_productos,
                           almacen_top=almacen_top, chart_labels=chart_labels, chart_data=chart_data)


@inventory_bp.route('/salida/<int:id>')
@login_required
@check_permission('perm_do_salidas')
def ver_salida(id):
    salida      = get_item_or_404(Salida, id)
    movimientos = salida.movimientos.options(joinedload(Movimiento.producto)).order_by(Movimiento.fecha.asc()).all()
    return render_template('salida_detalle.html', salida=salida, movimientos=movimientos,
                           titulo=f"Salida del {salida.fecha.strftime('%Y-%m-%d')}")


@inventory_bp.route('/salida', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_do_salidas')
def registrar_salida():
    org_id = current_user.organizacion_id

    almacen_id_solicitado = request.args.get('almacen_id', type=int)
    almacenes_org         = Almacen.query.filter_by(organizacion_id=org_id).all()

    almacen_seleccionado = None
    if almacen_id_solicitado:
        almacen_seleccionado = Almacen.query.get(almacen_id_solicitado)
        if not almacen_seleccionado or almacen_seleccionado.organizacion_id != org_id:
            flash('Permiso denegado para ese almacén.', 'danger')
            return redirect(url_for('inventory.historial_salidas'))

    if not almacenes_org:
        flash('No se pueden registrar salidas porque no hay almacenes creados.', 'warning')
        return redirect(url_for('main.index'))

    if not almacen_seleccionado:
        return render_template('seleccionar_almacen.html',
                               titulo="Seleccionar Almacén de Origen",
                               almacenes=almacenes_org,
                               destino_ruta='inventory.registrar_salida')

    today         = now_mx().date()
    salida_del_dia = Salida.query.filter_by(
        fecha=today, organizacion_id=org_id, almacen_id=almacen_seleccionado.id
    ).first()

    if not salida_del_dia:
        salida_del_dia = Salida(
            fecha=today, creador_id=current_user.id,
            organizacion_id=org_id, almacen_id=almacen_seleccionado.id,
        )
        db.session.add(salida_del_dia)
        db.session.flush()

    productos_lista = []
    for p in db.session.query(Producto).join(Stock).filter(
        Stock.almacen_id == almacen_seleccionado.id,
        Producto.organizacion_id == org_id,
        Stock.cantidad > 0,
    ).all():
        stock_item = Stock.query.filter_by(producto_id=p.id, almacen_id=almacen_seleccionado.id).first()
        productos_lista.append({
            'id': p.id, 'nombre': p.nombre, 'codigo': p.codigo,
            'stock_actual': stock_item.cantidad if stock_item else 0,
        })

    if request.method == 'POST':
        try:
            productos_ids = request.form.getlist('producto_id[]')
            cantidades    = request.form.getlist('cantidad[]')
            motivos       = request.form.getlist('motivo[]')

            if not productos_ids:
                flash('Debes añadir al menos un producto a la salida.', 'danger')
                return redirect(url_for('inventory.registrar_salida', almacen_id=almacen_seleccionado.id))

            productos_para_actualizar = []
            for i in range(len(productos_ids)):
                prod_id  = productos_ids[i]
                cant_str = cantidades[i]
                if not prod_id or not cant_str:
                    continue
                cantidad_salida = int(cant_str)
                stock_item = Stock.query.filter_by(
                    producto_id=prod_id, almacen_id=almacen_seleccionado.id
                ).first()

                if not stock_item:
                    flash('Error: Producto no válido.', 'danger')
                    db.session.rollback()
                    return render_template('salida_form.html',
                                           titulo=f"Registrar Salida de: {almacen_seleccionado.nombre}",
                                           productos=productos_lista, salida_id=salida_del_dia.id,
                                           almacen=almacen_seleccionado)
                if cantidad_salida <= 0:
                    flash('Todas las cantidades deben ser positivas.', 'danger')
                    db.session.rollback()
                    return render_template('salida_form.html',
                                           titulo=f"Registrar Salida de: {almacen_seleccionado.nombre}",
                                           productos=productos_lista, salida_id=salida_del_dia.id,
                                           almacen=almacen_seleccionado)
                if stock_item.cantidad < cantidad_salida:
                    flash(f'Error: Stock insuficiente para "{stock_item.producto.nombre}". Stock actual: {stock_item.cantidad}', 'danger')
                    db.session.rollback()
                    return render_template('salida_form.html',
                                           titulo=f"Registrar Salida de: {almacen_seleccionado.nombre}",
                                           productos=productos_lista, salida_id=salida_del_dia.id,
                                           almacen=almacen_seleccionado)
                productos_para_actualizar.append((stock_item, cantidad_salida, motivos[i]))

            for stock_item, cantidad_salida, motivo_item in productos_para_actualizar:
                stock_item.cantidad -= cantidad_salida
                db.session.add(stock_item)
                db.session.add(Movimiento(
                    producto_id=stock_item.producto_id, cantidad=-cantidad_salida,
                    tipo='salida', fecha=now_mx(), motivo=motivo_item,
                    salida=salida_del_dia, almacen_id=almacen_seleccionado.id,
                    organizacion_id=org_id,
                ))

            total_uds = sum(v[1] for v in productos_para_actualizar)
            log_actividad('salida', 'salida',
                          f'Salida registrada: {len(productos_para_actualizar)} producto(s), '
                          f'{total_uds} uds — Almacén: {almacen_seleccionado.nombre}',
                          entidad_id=salida_del_dia.id)
            db.session.commit()
            flash(f'Se añadieron {len(productos_para_actualizar)} items a la salida del día.', 'success')
            _check_and_alert_stock_bajo(org_id, almacen_seleccionado.id)
            return redirect(url_for('inventory.ver_salida', id=salida_del_dia.id))

        except Exception as e:
            db.session.rollback()
            _flash_err('Error al registrar la salida.', e)

    return render_template('salida_form.html',
                           titulo=f"Registrar Salida de: {almacen_seleccionado.nombre}",
                           productos=productos_lista,
                           salida_id=salida_del_dia.id,
                           almacen=almacen_seleccionado)


@inventory_bp.route('/movimiento/<int:id>/eliminar', methods=['POST'])
@login_required
@check_permission('perm_do_salidas')
def eliminar_movimiento_salida(id):
    movimiento = get_item_or_404(Movimiento, id)

    if movimiento.tipo != 'salida':
        flash('Error: Solo se pueden eliminar items de salida.', 'danger')
        return redirect(url_for('inventory.historial_salidas'))

    salida_id_redirect = movimiento.salida_id
    nombre_producto    = movimiento.producto.nombre

    try:
        stock_item = Stock.query.filter_by(
            producto_id=movimiento.producto_id, almacen_id=movimiento.almacen_id
        ).first()
        cantidad_a_devolver = abs(movimiento.cantidad)

        if stock_item:
            stock_item.cantidad += cantidad_a_devolver
            db.session.add(stock_item)
        else:
            db.session.add(Stock(
                producto_id=movimiento.producto_id,
                almacen_id=movimiento.almacen_id,
                cantidad=cantidad_a_devolver,
                organizacion_id=movimiento.organizacion_id,
            ))

        db.session.add(Movimiento(
            producto_id=movimiento.producto_id,
            cantidad=cantidad_a_devolver,
            tipo='ajuste-entrada',
            fecha=now_mx(),
            motivo=f'Corrección/Eliminación de item (Salida #{salida_id_redirect})',
            almacen_id=movimiento.almacen_id,
            organizacion_id=movimiento.organizacion_id,
        ))
        db.session.delete(movimiento)
        db.session.commit()
        flash(f'Item "{nombre_producto}" eliminado. Stock revertido.', 'success')

    except Exception as e:
        db.session.rollback()
        _flash_err('Error al eliminar el item.', e)

    if salida_id_redirect and Salida.query.get(salida_id_redirect):
        return redirect(url_for('inventory.ver_salida', id=salida_id_redirect))
    return redirect(url_for('inventory.historial_salidas'))


@inventory_bp.route('/salida/<int:id>/pdf')
@login_required
@check_permission('perm_do_salidas')
def generar_salida_pdf(id):
    salida = get_item_or_404(Salida, id)
    org    = Organizacion.query.get(salida.organizacion_id)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               rightMargin=inch, leftMargin=inch,
                               topMargin=0.5*inch, bottomMargin=inch)
    story  = []
    styles = getSampleStyleSheet()
    fuente, c_pri, c_sec = _pdf_estilos(org)

    s_normal = ParagraphStyle('SNorm',  fontName=fuente, fontSize=10, leading=12)
    s_bold   = ParagraphStyle('SBold',  fontName=_pdf_bold(fuente), fontSize=10, leading=12)
    s_brand  = ParagraphStyle('SBrand', fontName=_pdf_bold(fuente), fontSize=18, leading=20, textColor=colors.black)
    s_th     = ParagraphStyle('STH',    fontName=_pdf_bold(fuente), fontSize=9, textColor=colors.white, alignment=TA_CENTER)
    s_cell   = ParagraphStyle('SCell',  fontName=fuente, fontSize=9, leading=11)
    s_cellr  = ParagraphStyle('SCellR', fontName=fuente, fontSize=9, leading=11, alignment=TA_RIGHT)

    _pdf_header(story, org, styles)

    estado_color = '#DC2626' if salida.estado == 'cancelada' else '#059669'
    info_izq = [
        Paragraph('<b>ALMACÉN:</b>', s_normal),
        Paragraph(salida.almacen.nombre, s_bold),
        Paragraph(f'Fecha: {salida.fecha.strftime("%d/%m/%Y")}', s_normal),
        Paragraph(f'Creada por: {salida.creador.username}', s_normal),
    ]
    info_der = [
        Paragraph(f'<b>SALIDA #{salida.id}</b>', s_brand),
        Paragraph(f'<font color="{estado_color}"><b>{salida.estado.upper()}</b></font>', s_bold),
    ]
    t_info = Table([[info_izq, info_der]], colWidths=[3.5*inch, 2.7*inch])
    t_info.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('LEFTPADDING', (0,0), (-1,-1), 0)]))
    story.append(t_info)
    story.append(Spacer(1, 0.25*inch))

    data = [[
        Paragraph('Producto', s_th), Paragraph('SKU', s_th),
        Paragraph('Motivo', s_th),   Paragraph('Cantidad', s_th),
    ]]
    total_items = 0
    for mov in salida.movimientos.options(joinedload(Movimiento.producto)).order_by(Movimiento.fecha.asc()).all():
        cant = abs(mov.cantidad)
        total_items += cant
        data.append([
            Paragraph(mov.producto.nombre, s_cell),
            Paragraph(mov.producto.codigo, s_cell),
            Paragraph(mov.motivo or '—', s_cell),
            Paragraph(str(cant), s_cellr),
        ])
    s_totl = ParagraphStyle('STotL', fontName=_pdf_bold(fuente), fontSize=10, alignment=TA_RIGHT)
    s_totv = ParagraphStyle('STotV', fontName=_pdf_bold(fuente), fontSize=11, alignment=TA_RIGHT, textColor=c_pri)
    data.append(['', '', Paragraph('TOTAL UNIDADES:', s_totl), Paragraph(str(total_items), s_totv)])

    t_mov = Table(data, colWidths=[2.8*inch, 1.2*inch, 1.4*inch, 0.8*inch], repeatRows=1)
    row_bgs = _pdf_row_styles(len(data) - 1, c_sec)
    t_mov.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  c_pri),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('GRID',          (0,0),  (-1,-2), 0.5, colors.HexColor('#DEE2E6')),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (3,0),  (3,-1),  'RIGHT'),
        ('TOPPADDING',    (0,0),  (-1,-1), 6),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 6),
        ('SPAN',          (0,-1), (1,-1)),
        ('LINEABOVE',     (0,-1), (-1,-1), 1, colors.HexColor('#DEE2E6')),
        ('BOX',           (2,-1), (3,-1),  0.5, colors.HexColor('#DEE2E6')),
    ] + row_bgs))
    story.append(t_mov)

    _pdf_footer(story, org)
    doc.build(story)
    buffer.seek(0)
    filename = f"Salida-{salida.id}_{salida.fecha.strftime('%Y-%m-%d')}.pdf"
    return send_file(buffer, as_attachment=False, download_name=filename, mimetype='application/pdf')


# ==============================================================================
# TRANSFERENCIA ENTRE ALMACENES
# ==============================================================================

@inventory_bp.route('/transferencia/nueva', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_do_salidas')
def nueva_transferencia():
    import secrets as _secrets
    org_id    = current_user.organizacion_id
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).order_by(Almacen.nombre).all()

    if len(almacenes) < 2:
        flash('Necesitas al menos 2 almacenes para realizar una transferencia.', 'warning')
        return redirect(url_for('inventory.lista_almacenes'))

    if request.method == 'POST':
        try:
            origen_id   = int(request.form.get('almacen_origen_id'))
            destino_id  = int(request.form.get('almacen_destino_id'))
            producto_id = int(request.form.get('producto_id'))
            cantidad    = int(request.form.get('cantidad', 0))
            motivo      = request.form.get('motivo', '').strip() or 'Transferencia entre almacenes'

            if origen_id == destino_id:
                flash('El almacén de origen y destino no pueden ser el mismo.', 'danger')
                return redirect(url_for('inventory.nueva_transferencia'))
            if not Almacen.query.filter_by(id=origen_id, organizacion_id=org_id).first() or \
               not Almacen.query.filter_by(id=destino_id, organizacion_id=org_id).first():
                flash('Almacén no autorizado.', 'danger')
                return redirect(url_for('inventory.nueva_transferencia'))
            if cantidad <= 0:
                flash('La cantidad debe ser mayor a cero.', 'danger')
                return redirect(url_for('inventory.nueva_transferencia'))

            stock_origen = Stock.query.filter_by(producto_id=producto_id, almacen_id=origen_id).first()
            if not stock_origen or stock_origen.cantidad < cantidad:
                flash(f'Stock insuficiente en el almacén origen. Disponible: {stock_origen.cantidad if stock_origen else 0}', 'danger')
                return redirect(url_for('inventory.nueva_transferencia'))

            ref = _secrets.token_hex(4).upper()
            stock_origen.cantidad -= cantidad

            stock_destino = Stock.query.filter_by(producto_id=producto_id, almacen_id=destino_id).first()
            if stock_destino:
                stock_destino.cantidad += cantidad
            else:
                db.session.add(Stock(
                    producto_id=producto_id, almacen_id=destino_id, cantidad=cantidad,
                    stock_minimo=stock_origen.stock_minimo, stock_maximo=stock_origen.stock_maximo,
                ))

            now = now_mx()
            db.session.add(Movimiento(
                producto_id=producto_id, cantidad=-cantidad, tipo='transferencia-salida',
                fecha=now, motivo=f'[REF:{ref}] {motivo}',
                almacen_id=origen_id, organizacion_id=org_id,
            ))
            db.session.add(Movimiento(
                producto_id=producto_id, cantidad=cantidad, tipo='transferencia-entrada',
                fecha=now, motivo=f'[REF:{ref}] {motivo}',
                almacen_id=destino_id, organizacion_id=org_id,
            ))

            db.session.commit()
            flash(f'Transferencia REF:{ref} completada. {cantidad} unidades movidas correctamente.', 'success')
            return redirect(url_for('main.dashboard'))

        except Exception as e:
            db.session.rollback()
            _flash_err('Error al realizar la transferencia.', e)

    return render_template('transferencia_form.html',
                           titulo='Nueva Transferencia de Stock', almacenes=almacenes)


# ==============================================================================
# AJUSTE MANUAL DE INVENTARIO
# ==============================================================================

@inventory_bp.route('/ajuste/nuevo', methods=['GET', 'POST'])
@login_required
@check_org_permission
@check_permission('perm_edit_management')
def nuevo_ajuste():
    org_id    = current_user.organizacion_id
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).order_by(Almacen.nombre).all()

    if request.method == 'POST':
        try:
            almacen_id      = int(request.form.get('almacen_id'))
            producto_id     = int(request.form.get('producto_id'))
            cantidad_fisica = int(request.form.get('cantidad_fisica', 0))
            motivo          = request.form.get('motivo', '').strip()

            if not motivo:
                flash('El motivo del ajuste es obligatorio para la auditoría.', 'danger')
                return redirect(url_for('inventory.nuevo_ajuste'))
            if not Almacen.query.filter_by(id=almacen_id, organizacion_id=org_id).first():
                flash('Almacén no autorizado.', 'danger')
                return redirect(url_for('inventory.nuevo_ajuste'))

            stock = Stock.query.filter_by(producto_id=producto_id, almacen_id=almacen_id).first()
            if not stock:
                flash('No se encontró ese producto en el almacén seleccionado.', 'danger')
                return redirect(url_for('inventory.nuevo_ajuste'))

            diferencia = cantidad_fisica - stock.cantidad
            if diferencia == 0:
                flash('No hay diferencia entre el conteo físico y el sistema. No se realizó ningún ajuste.', 'info')
                return redirect(url_for('inventory.nuevo_ajuste'))

            tipo_mov = 'ajuste-entrada' if diferencia > 0 else 'ajuste-salida'
            stock.cantidad = cantidad_fisica

            db.session.add(Movimiento(
                producto_id=producto_id, cantidad=diferencia, tipo=tipo_mov,
                fecha=now_mx(), motivo=f'Ajuste Físico: {motivo}',
                almacen_id=almacen_id, organizacion_id=org_id,
            ))
            signo = '+' if diferencia > 0 else ''
            log_actividad('ajuste', 'producto',
                          f'Ajuste de inventario: {signo}{diferencia} uds — {motivo}',
                          entidad_id=producto_id)
            db.session.commit()

            if diferencia < 0:
                _check_and_alert_stock_bajo(org_id, almacen_id)

            flash(f'Ajuste registrado. Diferencia aplicada: {signo}{diferencia} unidades.', 'success')
            return redirect(url_for('main.dashboard'))

        except Exception as e:
            db.session.rollback()
            _flash_err('Error al registrar el ajuste.', e)

    return render_template('ajuste_form.html',
                           titulo='Ajuste Manual de Inventario', almacenes=almacenes)
