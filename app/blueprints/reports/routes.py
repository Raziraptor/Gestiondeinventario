# ==============================================================================
# Blueprint: reports
# Rutas de reportes y exportaciones (Excel / PDF)
# ==============================================================================

import os
import io
from datetime import datetime
from decimal import Decimal

from flask import (
    render_template, request, redirect, url_for, flash,
    send_file, make_response, current_app,
)
from flask_login import login_required, current_user
from sqlalchemy import extract
from sqlalchemy.orm import joinedload, contains_eager, selectinload

# ReportLab — PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, Image as ReportLabImage,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# openpyxl — Excel
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, NamedStyle, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XlImage

# QR (opcional)
try:
    import qrcode
except ImportError:
    qrcode = None  # type: ignore

# Internal
from app.extensions import db
from app.models import (
    Almacen, Producto, Stock, Movimiento,
    Gasto, Organizacion,
    ProyectoOC, ProyectoOCDetalle,
)
from app.helpers import (
    now_mx, _flash_err,
    check_org_permission, admin_required, check_permission,
    get_item_or_404, MESES_ES,
)
from . import reports_bp


# ==============================================================================
# PDF HELPER FUNCTIONS
# ==============================================================================

def _pdf_estilos(org):
    """Devuelve (fuente, color_primario, color_secundario) listos para ReportLab."""
    fuente = org.tipo_letra if org.tipo_letra in ['Helvetica', 'Times-Roman', 'Courier'] else 'Helvetica'
    c_pri = colors.HexColor(org.color_primario)  if org.color_primario  else colors.HexColor('#333333')
    c_sec = colors.HexColor(org.color_secundario) if org.color_secundario else colors.HexColor('#f1f5f9')
    return fuente, c_pri, c_sec


def _pdf_bold(fuente):
    # Times-Roman usa 'Times-Bold', no 'Times-Roman-Bold' (que no existe en ReportLab)
    return 'Times-Bold' if fuente == 'Times-Roman' else f'{fuente}-Bold'


def _pdf_header(story, org, styles):
    """Añade encabezado de marca (logo + nombre + RFC + correo) y barra de color."""
    fuente, c_pri, _ = _pdf_estilos(org)

    s_brand = ParagraphStyle('_Brand', fontName=_pdf_bold(fuente), fontSize=20, leading=22, textColor=colors.black, spaceAfter=2)
    s_sub   = ParagraphStyle('_Sub',   fontName=fuente, fontSize=10, leading=12, textColor=colors.gray)
    s_meta  = ParagraphStyle('_Meta',  fontName=fuente, fontSize=8,  leading=10, textColor=colors.HexColor('#64748b'))

    logo_el = []
    if org.logo_url:
        logo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], org.logo_url)
        if os.path.exists(logo_path):
            img = ReportLabImage(logo_path)
            max_h = 1.0 * inch
            img.drawHeight = max_h
            img.drawWidth  = max_h * (img.imageWidth / float(img.imageHeight))
            logo_el.append(img)

    text_el = [Paragraph(org.header_titulo or org.nombre, s_brand)]
    if org.header_subtitulo:
        text_el.append(Paragraph(org.header_subtitulo, s_sub))

    meta_parts = []
    if org.rfc:             meta_parts.append(f'RFC: {org.rfc}')
    if org.correo_empresa:  meta_parts.append(org.correo_empresa)
    if org.telefono:        meta_parts.append(org.telefono)
    if org.direccion:       meta_parts.append(org.direccion)
    if meta_parts:
        text_el.append(Paragraph(' · '.join(meta_parts), s_meta))

    if logo_el:
        t_hdr = Table([[logo_el, text_el]], colWidths=[1.5*inch, 4.7*inch])
    else:
        t_hdr = Table([[text_el]], colWidths=[6.2*inch])
    t_hdr.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(t_hdr)
    story.append(Table([['']], colWidths=[6.2*inch], rowHeights=[2],
                       style=TableStyle([('BACKGROUND', (0,0), (-1,-1), c_pri)])))
    story.append(Spacer(1, 0.2*inch))


def _pdf_footer(story, org, doc_url=None):
    """Añade bloque de pie de página (footer_texto + fecha generación + QR opcional)."""
    fuente, c_pri, _ = _pdf_estilos(org)
    s_footer = ParagraphStyle('_Foot', fontName=fuente, fontSize=8, textColor=colors.HexColor('#64748b'), alignment=TA_CENTER)

    story.append(Spacer(1, 0.3*inch))
    story.append(Table([['']], colWidths=[6.2*inch], rowHeights=[1],
                       style=TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#e2e8f0'))])))
    story.append(Spacer(1, 0.1*inch))

    pie_parts = []
    if org.footer_texto:
        pie_parts.append(org.footer_texto)
    pie_parts.append(f"Generado el {now_mx().strftime('%d/%m/%Y a las %H:%M')} · {org.nombre}")
    story.append(Paragraph('<br/>'.join(pie_parts), s_footer))

    if org.pdf_mostrar_qr and doc_url and qrcode is not None:
        try:
            qr = qrcode.QRCode(version=1, box_size=4, border=2)
            qr.add_data(doc_url)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color='black', back_color='white')
            qr_buf = io.BytesIO()
            qr_img.save(qr_buf, format='PNG')
            qr_buf.seek(0)
            rl_qr = ReportLabImage(qr_buf)
            rl_qr.drawWidth  = 0.7 * inch
            rl_qr.drawHeight = 0.7 * inch
            story.append(Spacer(1, 0.1*inch))
            t_qr = Table([[rl_qr]], colWidths=[6.2*inch])
            t_qr.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
            story.append(t_qr)
        except Exception:
            pass


def _pdf_row_styles(data_len, c_sec):
    """Devuelve estilos de filas alternas usando color_secundario."""
    styles = []
    for i in range(1, data_len):
        bg = c_sec if i % 2 == 0 else colors.white
        styles.append(('BACKGROUND', (0, i), (-1, i), bg))
    return styles


# ==============================================================================
# ROUTES
# ==============================================================================

@reports_bp.route('/reportes')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def reportes():
    org_id = current_user.organizacion_id
    almacenes = Almacen.query.filter_by(organizacion_id=org_id).order_by(Almacen.nombre).all()
    return render_template('reportes.html', titulo='Reportes', almacenes=almacenes, now=now_mx())


@reports_bp.route('/reportes/inventario.xlsx')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def exportar_inventario_excel():
    org_id = current_user.organizacion_id
    almacen_id = request.args.get('almacen_id', type=int)

    if almacen_id:
        almacen = Almacen.query.filter_by(id=almacen_id, organizacion_id=org_id).first_or_404()
        items = (
            db.session.query(Stock)
            .filter_by(almacen_id=almacen_id)
            .join(Producto)
            .options(contains_eager(Stock.producto)
                     .options(joinedload(Producto.categoria), joinedload(Producto.proveedor)))
            .order_by(Producto.nombre)
            .all()
        )
        nombre_almacen = almacen.nombre
    else:
        items = (
            db.session.query(Stock)
            .join(Almacen, Stock.almacen_id == Almacen.id)
            .filter(Almacen.organizacion_id == org_id)
            .join(Producto)
            .options(contains_eager(Stock.producto)
                     .options(joinedload(Producto.categoria), joinedload(Producto.proveedor)))
            .order_by(Producto.nombre)
            .all()
        )
        nombre_almacen = 'Todos los Almacenes'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    h_font  = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    h_fill  = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    h_align = Alignment(horizontal='center', vertical='center')
    b_font  = Font(name='Arial', size=10)
    thin    = Side(border_style='thin', color='DEE2E6')
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    mxn     = NamedStyle(name='mxn_inv', number_format='"$"#,##0.00')
    if 'mxn_inv' not in wb.named_styles:
        wb.add_named_style(mxn)

    ws.merge_cells('A1:J1')
    ws['A1'].value = f"Inventario — {nombre_almacen} — {now_mx().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font  = Font(name='Arial', size=13, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    headers = ['SKU', 'Producto', 'Categoría', 'Proveedor', 'Stock', 'Mín.', 'Máx.', 'Estado', 'Precio Unit. MXN', 'Valor Total MXN']
    ws.append(headers)
    for cell in ws[2]:
        cell.font  = h_font
        cell.fill  = h_fill
        cell.alignment = h_align
        cell.border = border
    ws.row_dimensions[2].height = 20

    estado_label = {'bajo': 'Bajo Mínimo', 'exceso': 'Exceso', 'ok': 'Óptimo'}
    valor_total = 0

    for item in items:
        valor = (item.cantidad or 0) * (item.producto.precio_unitario or 0)
        valor_total += valor
        ws.append([
            item.producto.codigo,
            item.producto.nombre,
            item.producto.categoria.nombre if item.producto.categoria else '',
            item.producto.proveedor.nombre if item.producto.proveedor else '',
            item.cantidad,
            item.stock_minimo,
            item.stock_maximo,
            estado_label.get(item.estado_stock, ''),
            item.producto.precio_unitario or 0,
            valor,
        ])
        r = ws.max_row
        for col in range(1, 11):
            ws.cell(row=r, column=col).font   = b_font
            ws.cell(row=r, column=col).border = border
        estado_cell = ws.cell(row=r, column=8)
        if item.estado_stock == 'bajo':
            estado_cell.font = Font(name='Arial', size=10, color='DC2626', bold=True)
        elif item.estado_stock == 'exceso':
            estado_cell.font = Font(name='Arial', size=10, color='0891B2', bold=True)
        else:
            estado_cell.font = Font(name='Arial', size=10, color='059669', bold=True)
        ws.cell(row=r, column=9).style  = 'mxn_inv'
        ws.cell(row=r, column=10).style = 'mxn_inv'

    tr = ws.max_row + 1
    ws.cell(row=tr, column=9).value     = 'VALOR TOTAL (MXN)'
    ws.cell(row=tr, column=9).font      = Font(name='Arial', size=11, bold=True)
    ws.cell(row=tr, column=9).alignment = Alignment(horizontal='right')
    ws.cell(row=tr, column=10).value    = valor_total
    ws.cell(row=tr, column=10).style    = 'mxn_inv'
    ws.cell(row=tr, column=10).font     = Font(name='Arial', size=11, bold=True)

    for i, w in enumerate([15, 32, 18, 22, 9, 7, 7, 14, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"Inventario_{nombre_almacen.replace(' ', '_')}_{now_mx().strftime('%Y%m%d')}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True)


@reports_bp.route('/reportes/movimientos.xlsx')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def exportar_movimientos_excel():
    org_id     = current_user.organizacion_id
    almacen_id = request.args.get('almacen_id', type=int)
    desde_str  = request.args.get('desde', '')
    hasta_str  = request.args.get('hasta', '')
    tipo_f     = request.args.get('tipo', '')
    ahora      = now_mx()

    try:
        fecha_desde = datetime.strptime(desde_str, '%Y-%m-%d') if desde_str else ahora.replace(day=1, hour=0, minute=0, second=0)
        fecha_hasta = datetime.strptime(hasta_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59) if hasta_str else ahora
    except ValueError:
        fecha_desde = ahora.replace(day=1, hour=0, minute=0, second=0)
        fecha_hasta = ahora

    q = Movimiento.query.filter_by(organizacion_id=org_id).filter(
        Movimiento.fecha >= fecha_desde,
        Movimiento.fecha <= fecha_hasta
    )
    if almacen_id:
        q = q.filter(Movimiento.almacen_id == almacen_id)
    if tipo_f:
        q = q.filter(Movimiento.tipo == tipo_f)
    movimientos = q.options(joinedload(Movimiento.producto)).order_by(Movimiento.fecha.desc()).all()

    almacen_map = {a.id: a.nombre for a in Almacen.query.filter_by(organizacion_id=org_id).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    h_font  = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    h_fill  = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    b_font  = Font(name='Arial', size=10)
    thin    = Side(border_style='thin', color='DEE2E6')
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    rango_str = f"{fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    ws.merge_cells('A1:H1')
    ws['A1'].value = f"Historial de Movimientos — {rango_str}"
    ws['A1'].font  = Font(name='Arial', size=13, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    headers = ['Fecha', 'Hora', 'Tipo', 'Producto', 'SKU', 'Cantidad', 'Motivo', 'Almacén']
    ws.append(headers)
    for cell in ws[2]:
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
    ws.row_dimensions[2].height = 20

    tipo_labels = {
        'entrada': 'Entrada', 'entrada-inicial': 'Stock Inicial',
        'salida': 'Salida', 'ajuste-entrada': 'Ajuste (+)', 'ajuste-salida': 'Ajuste (-)',
    }

    for mov in movimientos:
        ws.append([
            mov.fecha.strftime('%d/%m/%Y'),
            mov.fecha.strftime('%H:%M'),
            tipo_labels.get(mov.tipo, mov.tipo),
            mov.producto.nombre if mov.producto else '',
            mov.producto.codigo if mov.producto else '',
            mov.cantidad,
            mov.motivo,
            almacen_map.get(mov.almacen_id, ''),
        ])
        r = ws.max_row
        for col in range(1, 9):
            ws.cell(row=r, column=col).font   = b_font
            ws.cell(row=r, column=col).border = border
        qty_cell = ws.cell(row=r, column=6)
        if mov.cantidad > 0:
            qty_cell.font = Font(name='Arial', size=10, color='059669', bold=True)
        elif mov.cantidad < 0:
            qty_cell.font = Font(name='Arial', size=10, color='DC2626', bold=True)

    for i, w in enumerate([12, 8, 16, 32, 15, 10, 40, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"Movimientos_{now_mx().strftime('%Y%m%d')}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True)


@reports_bp.route('/reportes/valorizacion.pdf')
@login_required
@check_org_permission
@check_permission('perm_view_dashboard')
def exportar_valorizacion_pdf():
    if current_user.rol not in ['super_admin', 'admin']:
        flash('Solo los administradores pueden exportar reportes de valorización.', 'danger')
        return redirect(url_for('reports.reportes'))

    org_id     = current_user.organizacion_id
    almacen_id = request.args.get('almacen_id', type=int)
    org        = Organizacion.query.get(org_id)

    if almacen_id:
        almacen = Almacen.query.filter_by(id=almacen_id, organizacion_id=org_id).first_or_404()
        items = (
            db.session.query(Stock)
            .filter_by(almacen_id=almacen_id)
            .join(Producto)
            .options(contains_eager(Stock.producto)
                     .options(joinedload(Producto.categoria), joinedload(Producto.proveedor)))
            .order_by(Producto.nombre)
            .all()
        )
        nombre_almacen = almacen.nombre
    else:
        items = (
            db.session.query(Stock)
            .join(Almacen, Stock.almacen_id == Almacen.id)
            .filter(Almacen.organizacion_id == org_id)
            .join(Producto)
            .options(contains_eager(Stock.producto)
                     .options(joinedload(Producto.categoria), joinedload(Producto.proveedor)))
            .order_by(Producto.nombre)
            .all()
        )
        nombre_almacen = 'Todos los Almacenes'

    items_sorted = sorted(items, key=lambda x: (x.cantidad or 0) * (x.producto.precio_unitario or 0), reverse=True)
    valor_total  = sum((i.cantidad or 0) * (i.producto.precio_unitario or 0) for i in items)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    story  = []
    styles = getSampleStyleSheet()

    fuente, primary, c_sec = _pdf_estilos(org)
    light_gray = colors.HexColor('#F8F9FA')
    mid_gray   = colors.HexColor('#DEE2E6')

    s_title  = ParagraphStyle('RPTitle',  fontName=_pdf_bold(fuente), fontSize=18, textColor=primary, spaceAfter=4)
    s_sub    = ParagraphStyle('RPSub',    fontName=fuente,           fontSize=10, textColor=colors.HexColor('#6B7280'), spaceAfter=2)
    s_cell   = ParagraphStyle('RPCell',   fontName=fuente,           fontSize=8,  leading=10)
    s_cellb  = ParagraphStyle('RPCellB',  fontName=_pdf_bold(fuente), fontSize=8,  leading=10)
    s_cellr  = ParagraphStyle('RPCellR',  fontName=fuente,           fontSize=8,  leading=10, alignment=TA_RIGHT)
    s_cellbr = ParagraphStyle('RPCellBR', fontName=_pdf_bold(fuente), fontSize=8,  leading=10, alignment=TA_RIGHT)
    s_big    = ParagraphStyle('RPBig',    fontName=_pdf_bold(fuente), fontSize=14, textColor=primary)

    _pdf_header(story, org, styles)
    story.append(Paragraph("Reporte de Valorización de Inventario", s_title))
    story.append(Paragraph(f"Almacén: {nombre_almacen}", s_sub))
    story.append(Spacer(1, 0.2*inch))

    # Resumen
    resumen = [
        [Paragraph('<b>Total Productos</b>', s_cellb),
         Paragraph('<b>Valor Total (MXN)</b>', s_cellb),
         Paragraph('<b>Almacén</b>', s_cellb)],
        [Paragraph(str(len(items)), s_big),
         Paragraph(f"$ {valor_total:,.2f}", s_big),
         Paragraph(nombre_almacen, s_cell)],
    ]
    t_res = Table(resumen, colWidths=[2*inch, 2.8*inch, 2.2*inch])
    t_res.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), light_gray),
        ('BOX',        (0,0), (-1,-1), 1,   mid_gray),
        ('INNERGRID',  (0,0), (-1,-1), 0.5, mid_gray),
        ('PADDING',    (0,0), (-1,-1), 8),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t_res)
    story.append(Spacer(1, 0.2*inch))

    # Tabla principal
    col_w = [0.3*inch, 0.85*inch, 2.1*inch, 1.1*inch, 0.5*inch, 0.95*inch, 0.95*inch, 0.65*inch]
    hdrs  = ['#', 'SKU', 'Producto', 'Categoría', 'Stock', 'Precio Unit.', 'Valor Total', '% Total']
    data  = [[Paragraph(h, s_cellb) for h in hdrs]]

    for i, item in enumerate(items_sorted, 1):
        valor_item = (item.cantidad or 0) * (item.producto.precio_unitario or 0)
        pct        = (valor_item / valor_total * 100) if valor_total > 0 else 0
        data.append([
            Paragraph(str(i), s_cellr),
            Paragraph(item.producto.codigo, s_cell),
            Paragraph(item.producto.nombre[:40], s_cell),
            Paragraph(item.producto.categoria.nombre if item.producto.categoria else '—', s_cell),
            Paragraph(str(item.cantidad), s_cellr),
            Paragraph(f"$ {(item.producto.precio_unitario or 0):,.2f}", s_cellr),
            Paragraph(f"$ {valor_item:,.2f}", s_cellr),
            Paragraph(f"{pct:.1f}%", s_cellr),
        ])

    data.append([
        Paragraph('', s_cell), Paragraph('', s_cell), Paragraph('', s_cell), Paragraph('', s_cell),
        Paragraph('', s_cell),
        Paragraph('TOTAL', s_cellbr),
        Paragraph(f"$ {valor_total:,.2f}", s_cellbr),
        Paragraph('100%', s_cellbr),
    ])

    t_main = Table(data, colWidths=col_w, repeatRows=1)
    row_bgs = _pdf_row_styles(len(data) - 1, c_sec)
    t_main.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),  (-1,0),  primary),
        ('TEXTCOLOR',    (0,0),  (-1,0),  colors.white),
        ('BACKGROUND',   (0,-1), (-1,-1), colors.HexColor('#EEEDFC')),
        ('FONTNAME',     (0,-1), (-1,-1), _pdf_bold(fuente)),
        ('GRID',         (0,0),  (-1,-1), 0.5, mid_gray),
        ('BOX',          (0,0),  (-1,-1), 1,   mid_gray),
        ('PADDING',      (0,0),  (-1,-1), 5),
        ('VALIGN',       (0,0),  (-1,-1), 'MIDDLE'),
    ] + row_bgs))
    story.append(t_main)

    _pdf_footer(story, org)
    doc.build(story)
    buf.seek(0)
    fname = f"Valorizacion_{nombre_almacen.replace(' ','_')}_{now_mx().strftime('%Y%m%d')}.pdf"
    return send_file(buf, download_name=fname, mimetype='application/pdf', as_attachment=True)


@reports_bp.route('/proyecto-oc/exportar.xlsx')
@login_required
@check_org_permission
@check_permission('perm_create_oc_proyecto')
def exportar_proyectos_oc_excel():
    org_id = current_user.organizacion_id
    estado_f = request.args.get('estado')
    mes_f    = request.args.get('mes', type=int)
    ano_f    = request.args.get('ano', type=int)

    query = ProyectoOC.query.filter_by(organizacion_id=org_id).options(
        joinedload(ProyectoOC.almacen),
        joinedload(ProyectoOC.creador),
        joinedload(ProyectoOC.recibido_por),
        selectinload(ProyectoOC.detalles).joinedload(ProyectoOCDetalle.producto),
    )
    if estado_f:
        query = query.filter(ProyectoOC.estado == estado_f)
    if mes_f:
        query = query.filter(extract('month', ProyectoOC.fecha_creacion) == mes_f)
    if ano_f:
        query = query.filter(extract('year', ProyectoOC.fecha_creacion) == ano_f)
    proyectos = query.order_by(ProyectoOC.fecha_creacion.desc()).all()

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen de OC ───────────────────────────────────────────────
    ws = wb.active
    ws.title = 'OC de Proyectos'

    COLOR_HDR  = 'FF4F46E5'
    COLOR_ALT  = 'FFF0F4FF'
    COLOR_TOT  = 'FFDBEAFE'

    h_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    b_font = Font(name='Calibri', size=10)
    t_font = Font(name='Calibri', size=11, bold=True)

    h_fill  = PatternFill('solid', fgColor=COLOR_HDR)
    alt_fill= PatternFill('solid', fgColor=COLOR_ALT)
    tot_fill= PatternFill('solid', fgColor=COLOR_TOT)

    thin = Side(style='thin', color='FFBFDBFE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdrs = ['ID', 'Proyecto', 'Estado', 'Creado Por', 'Fecha Creación',
            'Fecha Envío', 'Fecha Recepción', 'Almacén Destino',
            'Recibido Por', 'Artículos', 'Total Estimado (MXN)']
    ws.append(hdrs)
    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(1, col)
        cell.font   = h_font
        cell.fill   = h_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    estado_labels = {'borrador': 'Borrador', 'enviada': 'Enviada',
                     'recibida': 'Recibida', 'cancelada': 'Cancelada'}

    for i, poc in enumerate(proyectos, 2):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        row = [
            poc.id,
            poc.nombre_proyecto,
            estado_labels.get(poc.estado, poc.estado),
            poc.creador.username,
            poc.fecha_creacion.strftime('%d/%m/%Y'),
            poc.fecha_envio.strftime('%d/%m/%Y') if poc.fecha_envio else '—',
            poc.fecha_recepcion.strftime('%d/%m/%Y') if poc.fecha_recepcion else '—',
            poc.almacen.nombre if poc.almacen else '—',
            poc.recibido_por.username if poc.recibido_por else '—',
            len(poc.detalles),
            poc.costo_total,
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(i, col)
            cell.font   = b_font
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if col == 11:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fila de total
    last = len(proyectos) + 2
    ws.cell(last, 10, 'TOTAL:').font = t_font
    ws.cell(last, 10).alignment = Alignment(horizontal='right')
    total_val = ws.cell(last, 11, sum(p.costo_total for p in proyectos))
    total_val.font = t_font
    total_val.number_format = '"$"#,##0.00'
    total_val.alignment = Alignment(horizontal='right')
    for col in range(1, 12):
        ws.cell(last, col).fill = tot_fill

    col_widths = [6, 30, 12, 16, 16, 16, 18, 22, 16, 10, 22]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    # ── Hoja 2: Detalle de artículos ───────────────────────────────────────
    ws2 = wb.create_sheet('Artículos')
    hdrs2 = ['OC ID', 'Proyecto', 'Estado OC', 'Tipo', 'Artículo / Descripción',
             'SKU', 'Proveedor Sug.', 'Cantidad', 'Costo Unit.', 'Subtotal']
    ws2.append(hdrs2)
    for col, h in enumerate(hdrs2, 1):
        cell = ws2.cell(1, col)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    row_idx = 2
    for poc in proyectos:
        for d in poc.detalles:
            if d.producto_id and d.producto:
                tipo  = 'Catálogo'
                nombre = d.producto.nombre
                sku   = d.producto.codigo
            else:
                tipo  = 'Externo'
                nombre = d.descripcion_nuevo or 'Sin descripción'
                sku   = '—'
            sub = d.cantidad * d.costo_unitario
            row2 = [poc.id, poc.nombre_proyecto, estado_labels.get(poc.estado, poc.estado),
                    tipo, nombre, sku, d.proveedor_sugerido or '—',
                    d.cantidad, d.costo_unitario, sub]
            ws2.append(row2)
            fill2 = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col in range(1, len(row2) + 1):
                cell = ws2.cell(row_idx, col)
                cell.font   = b_font
                cell.fill   = fill2
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                if col in (9, 10):
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            row_idx += 1

    col_widths2 = [6, 28, 12, 10, 30, 14, 20, 10, 14, 14]
    for col, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha_str = now_mx().strftime('%Y-%m-%d')
    return send_file(buf, as_attachment=True,
                     download_name=f'OC-Proyectos_{fecha_str}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/gastos/exportar_excel')
@login_required
@check_org_permission
@check_permission('perm_view_gastos')
def exportar_gastos_excel():
    ahora = now_mx()
    org   = Organizacion.query.get(current_user.organizacion_id)

    # ── Rango de meses ────────────────────────────────────────────────────────
    mes_desde = request.args.get('mes_desde', type=int)
    ano_desde = request.args.get('ano_desde', type=int)
    mes_hasta = request.args.get('mes_hasta', type=int)
    ano_hasta = request.args.get('ano_hasta', type=int)
    if not mes_desde:
        mes_desde = mes_hasta = request.args.get('mes', type=int) or ahora.month
        ano_desde = ano_hasta = request.args.get('ano', type=int) or ahora.year

    periodos = []
    y, m = ano_desde, mes_desde
    while (y < ano_hasta) or (y == ano_hasta and m <= mes_hasta):
        periodos.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    if not periodos:
        periodos = [(ano_desde, mes_desde)]

    base_query = (Gasto.query if current_user.rol == 'super_admin'
                  else Gasto.query.filter_by(organizacion_id=current_user.organizacion_id))

    # ── Configuración de diseño ───────────────────────────────────────────────
    def _argb(h):
        return 'FF' + (h or '#000000').lstrip('#').upper()

    col_hdr     = _argb(getattr(org, 'excel_color_header',  '#1f4e79'))
    col_acc     = _argb(getattr(org, 'excel_color_accent',  '#dbeafe'))
    fuente      = getattr(org, 'excel_fuente',         'Calibri') or 'Calibri'
    show_logo   = getattr(org, 'excel_mostrar_logo',   True)
    show_id     = getattr(org, 'excel_mostrar_id',     True)
    show_oc     = getattr(org, 'excel_mostrar_oc',     True)
    show_origen = getattr(org, 'excel_mostrar_origen', True)
    empresa     = (org.header_titulo or org.nombre) if org else 'Empresa'

    # ── Columnas dinámicas ────────────────────────────────────────────────────
    COLS = ['Fecha', 'Descripción', 'Categoría', 'Monto']
    if show_id:     COLS = ['ID'] + COLS
    if show_oc:     COLS.append('OC Asociada')
    if show_origen: COLS.append('Origen')
    N         = len(COLS)
    monto_idx = COLS.index('Monto') + 1   # 1-based
    last_col  = get_column_letter(N)

    # ── Estilos ───────────────────────────────────────────────────────────────
    fill_hdr    = PatternFill('solid', fgColor=col_hdr)
    fill_acc    = PatternFill('solid', fgColor=col_acc)
    fill_svc    = PatternFill('solid', fgColor='FFFFF8E1')  # ámbar muy suave

    bd_s  = Side(border_style='thin',   color='CCCCCC')
    bd_m  = Side(border_style='medium', color='888888')
    bd    = Border(left=bd_s,  right=bd_s,  top=bd_s,  bottom=bd_s)
    bd_tt = Border(left=bd_m,  right=bd_m,  top=bd_m,  bottom=bd_m)

    f_title  = Font(name=fuente, size=14, bold=True,  color='FFFFFF')
    f_sub    = Font(name=fuente, size=10,             color='FFFFFF')
    f_hdr    = Font(name=fuente, size=10, bold=True,  color='FFFFFF')
    f_normal = Font(name=fuente, size=10)
    f_bold   = Font(name=fuente, size=11, bold=True)
    f_wht    = Font(name=fuente, size=11, bold=True,  color='FFFFFF')

    a_c = Alignment(horizontal='center', vertical='center')
    a_r = Alignment(horizontal='right',  vertical='center')
    a_l = Alignment(horizontal='left',   vertical='center')
    cur_fmt = '$#,##0.00'

    # ── Logo path ─────────────────────────────────────────────────────────────
    logo_path = None
    if show_logo and org and org.logo_url:
        candidate = os.path.join(current_app.config['UPLOAD_FOLDER'], org.logo_url)
        if os.path.exists(candidate):
            logo_path = candidate

    # ── Helpers internos ──────────────────────────────────────────────────────
    def _auto_width(ws, max_w=52):
        for ci, col in enumerate(ws.columns, 1):
            w = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, max_w)

    def _banner(ws, title, subtitle=''):
        ws.merge_cells(f'A1:{last_col}1')
        c = ws['A1']
        c.value, c.font, c.fill, c.alignment = title, f_title, fill_hdr, a_c
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f'A2:{last_col}2')
        c2 = ws['A2']
        c2.value, c2.font, c2.fill, c2.alignment = subtitle, f_sub, fill_hdr, a_c
        ws.row_dimensions[2].height = 16
        if logo_path:
            try:
                img = XlImage(logo_path)
                img.height, img.width = 42, 42
                ws.add_image(img, 'A1')
            except Exception:
                pass

    def _col_headers(ws, row, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font, c.fill, c.alignment, c.border = f_hdr, fill_hdr, a_c, bd
        ws.row_dimensions[row].height = 18

    def _cat_summary(ws, gastos_list):
        """Tabla de resumen por categoría al pie de cada hoja."""
        cat_totals = {}
        for g in gastos_list:
            k = g.categoria or 'Sin categoría'
            cat_totals[k] = cat_totals.get(k, 0) + g.monto
        if not cat_totals:
            return
        ws.append([])
        sr = ws.max_row + 1
        ws.merge_cells(f'A{sr}:B{sr}')
        sh = ws[f'A{sr}']
        sh.value, sh.font, sh.fill, sh.alignment = 'Resumen por Categoría', f_hdr, fill_hdr, a_c
        ws.row_dimensions[sr].height = 16
        hr = sr + 1
        for ci, txt in enumerate(['Categoría', 'Total'], 1):
            c = ws.cell(hr, ci, txt)
            c.font, c.fill, c.alignment, c.border = f_hdr, fill_hdr, (a_l if ci == 1 else a_r), bd
        for cat, tot in sorted(cat_totals.items(), key=lambda x: -x[1]):
            r = ws.max_row + 1
            c1 = ws.cell(r, 1, cat)
            c1.font, c1.border, c1.alignment = f_normal, bd, a_l
            c2 = ws.cell(r, 2, tot)
            c2.font, c2.border, c2.alignment = f_normal, bd, a_r
            c2.number_format = cur_fmt

    def _add_month_sheet(wb, year, month, table_idx):
        gastos = base_query.filter(
            extract('month', Gasto.fecha) == month,
            extract('year',  Gasto.fecha) == year
        ).order_by(Gasto.fecha.asc()).all()

        nombre_mes = datetime(year, month, 1).strftime('%B').capitalize()
        ws = wb.create_sheet(title=f"{nombre_mes[:3]} {year}")

        _banner(ws, empresa, f'Control de Gastos — {nombre_mes} {year}')
        _col_headers(ws, 3, COLS)

        total = 0.0
        for i, g in enumerate(gastos):
            origen = 'Servicio' if g.descripcion.startswith('Servicio:') else 'Manual'
            row_data = [g.fecha.date(), g.descripcion, g.categoria or '—', g.monto]
            if show_id:     row_data = [g.id] + row_data
            if show_oc:     row_data.append(g.orden_compra_id or '—')
            if show_origen: row_data.append(origen)
            ws.append(row_data)
            r = ws.max_row
            use_acc = (i % 2 == 1)
            for ci in range(1, N + 1):
                c = ws.cell(r, ci)
                c.font, c.border = f_normal, bd
                if origen == 'Servicio':
                    c.fill = fill_svc
                elif use_acc:
                    c.fill = fill_acc
            ws.cell(r, monto_idx).number_format = cur_fmt
            ws.cell(r, monto_idx).alignment     = a_r
            total += g.monto

        if gastos:
            try:
                tbl = ExcelTable(displayName=f'Gastos{table_idx}',
                                 ref=f'A3:{last_col}{ws.max_row}')
                tbl.tableStyleInfo = TableStyleInfo(
                    name='TableStyleMedium2', showRowStripes=False)
                ws.add_table(tbl)
            except Exception:
                pass

        # Fila total
        ft = ws.max_row + 1
        pre = get_column_letter(monto_idx - 1)
        ws.merge_cells(f'A{ft}:{pre}{ft}')
        c_lbl = ws.cell(ft, 1, 'Total del Mes')
        c_lbl.font, c_lbl.fill, c_lbl.alignment, c_lbl.border = f_bold, fill_acc, a_r, bd
        c_tot = ws.cell(ft, monto_idx, total)
        c_tot.number_format = cur_fmt
        c_tot.font, c_tot.fill, c_tot.alignment, c_tot.border = f_bold, fill_acc, a_r, bd

        _cat_summary(ws, gastos)
        ws.freeze_panes = 'A4'
        _auto_width(ws)
        return total, len(gastos)

    # ── Construir workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_res = None
    if len(periodos) > 1:
        ws_res = wb.create_sheet(title='Resumen', index=0)
        nm1 = datetime(periodos[0][0],  periodos[0][1],  1).strftime('%B').capitalize()
        nm2 = datetime(periodos[-1][0], periodos[-1][1], 1).strftime('%B').capitalize()
        _banner(ws_res, empresa,
                f'Auditoría de Gastos — {nm1} {periodos[0][0]} a {nm2} {periodos[-1][0]}')
        _col_headers(ws_res, 3, ['Período', 'Registros', 'Total (MXN)', 'Con Servicios'])

    totals_resumen = []
    for idx, (year, month) in enumerate(periodos, 1):
        total_mes, count_mes = _add_month_sheet(wb, year, month, idx)
        totals_resumen.append((year, month, total_mes, count_mes))

    if ws_res is not None:
        gran_total  = 0.0
        all_g_range = []
        for year, month, total_mes, count_mes in totals_resumen:
            nombre_mes = datetime(year, month, 1).strftime('%B').capitalize()
            gastos_mes = base_query.filter(
                extract('month', Gasto.fecha) == month,
                extract('year',  Gasto.fecha) == year).all()
            all_g_range.extend(gastos_mes)
            tiene_svc = any(g.descripcion.startswith('Servicio:') for g in gastos_mes)
            ws_res.append([f'{nombre_mes} {year}', count_mes, total_mes,
                           'Sí' if tiene_svc else 'No'])
            r = ws_res.max_row
            ws_res.cell(r, 3).number_format = cur_fmt
            ws_res.cell(r, 3).alignment     = a_r
            for ci in range(1, 5):
                ws_res.cell(r, ci).font   = f_normal
                ws_res.cell(r, ci).border = bd
            gran_total += total_mes

        data_end = ws_res.max_row
        gt = data_end + 1
        ws_res.cell(gt, 1, 'GRAN TOTAL').font      = f_wht
        ws_res.cell(gt, 1).fill, ws_res.cell(gt, 1).alignment = fill_hdr, a_r
        ws_res.cell(gt, 1).border = bd
        ws_res.cell(gt, 2, sum(c for _, _, _, c in totals_resumen)).font = f_wht
        ws_res.cell(gt, 2).fill,  ws_res.cell(gt, 2).border              = fill_hdr, bd
        ws_res.cell(gt, 2).alignment = a_c
        ws_res.cell(gt, 3, gran_total).number_format = cur_fmt
        ws_res.cell(gt, 3).font, ws_res.cell(gt, 3).fill   = f_wht, fill_hdr
        ws_res.cell(gt, 3).alignment, ws_res.cell(gt, 3).border = a_r, bd

        # Tabla de categorías del período completo
        _cat_summary(ws_res, all_g_range)

        # Gráfico de barras por mes
        try:
            chart = BarChart()
            chart.type, chart.grouping = 'col', 'clustered'
            chart.title   = 'Gastos por Mes'
            chart.y_axis.title = 'Total (MXN)'
            chart.x_axis.title = 'Período'
            data_ref = Reference(ws_res, min_col=3, max_col=3,
                                 min_row=3, max_row=data_end)
            cats_ref = Reference(ws_res, min_col=1, max_col=1,
                                 min_row=4, max_row=data_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width, chart.height = 16, 10
            ws_res.add_chart(chart, 'F3')
        except Exception:
            pass

        ws_res.freeze_panes = 'A4'
        _auto_width(ws_res)

    # ── Nombre de archivo ─────────────────────────────────────────────────────
    if len(periodos) == 1:
        nom = datetime(periodos[0][0], periodos[0][1], 1).strftime('%B').capitalize()
        filename = f"Gastos_{nom}_{periodos[0][0]}.xlsx"
    else:
        n1 = datetime(periodos[0][0],  periodos[0][1],  1).strftime('%b').capitalize()
        n2 = datetime(periodos[-1][0], periodos[-1][1], 1).strftime('%b').capitalize()
        filename = f"Gastos_{n1}{periodos[0][0]}_a_{n2}{periodos[-1][0]}.xlsx"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
